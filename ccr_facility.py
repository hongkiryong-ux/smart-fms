# -*- coding: utf-8 -*-
"""중앙관제실(설비) 운영일보 — 1일·월보·년보·특이사항."""
from __future__ import annotations

import io
import json
import re
from copy import deepcopy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models import Building, CcrFacilityArchive, CcrFacilityDaily

ROOT = Path(__file__).resolve().parent
SCHEMA_PATH = ROOT / "resources" / "ccr_facility_schema.json"
TEMPLATE_XLSX = ROOT / "resources" / "ccr_facility_template.xlsx"

_schema_cache: dict | None = None


def load_schema() -> dict:
    global _schema_cache
    if _schema_cache is None:
        _schema_cache = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))
    return _schema_cache


def is_ccr_facility_building(building: Building | None) -> bool:
    if not building:
        return False
    name = (building.name or "").strip()
    target = load_schema().get("building_name", "중앙관제실(설비)")
    return name == target


def _parse_num(raw: Any) -> float | None:
    if raw is None or raw == "":
        return None
    try:
        return float(str(raw).replace(",", "").strip())
    except ValueError:
        return None


def parse_time_range_hours(raw: str) -> float | None:
    """08:00~20:00 → 12.0, 단일 숫자도 허용."""
    s = (raw or "").strip()
    if not s:
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})\s*[~\-–—]\s*(\d{1,2}):(\d{2})$", s)
    if m:
        h1, m1, h2, m2 = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
        start = h1 + m1 / 60
        end = h2 + m2 / 60
        if end < start:
            end += 24
        return round(end - start, 1)
    n = _parse_num(s)
    return round(n, 1) if n is not None else None


def _sum_hrs(*vals: Any) -> float | None:
    nums = [parse_time_range_hours(str(v)) if ":" in str(v or "") else _parse_num(v) for v in vals]
    nums = [n for n in nums if n is not None]
    return round(sum(nums), 1) if nums else None


def _fmt_num(n: float | None) -> str:
    if n is None:
        return ""
    if abs(n - round(n)) < 0.05:
        return str(int(round(n)))
    return str(round(n, 1))


def _empty_shift_row(shifts: int) -> dict:
    row: dict[str, Any] = {}
    for i in range(1, shifts + 1):
        row[f"s{i}_time"] = ""
        row[f"s{i}_hr"] = ""
    row["daily"] = ""
    row["monthly"] = ""
    row["prev_day"] = ""
    return row


def empty_daily_payload() -> dict[str, Any]:
    schema = load_schema()
    s2_fields = {f["id"]: "" for f in schema["section2"]["fields"]}
    data: dict[str, Any] = {
        "s1": {
            "heat": {"prev": "", "today": "", "daily": "", "monthly": "", "prev_manual": False},
            "flow": {"prev": "", "today": "", "daily": "", "monthly": "", "prev_manual": False},
            "power": {
                "prev": "",
                "today": "",
                "daily": "",
                "monthly": "",
                "prev_monthly": "",
                "prev_manual": False,
            },
            "peak": {"time": "", "load": ""},
        },
        "s2": {r["id"]: dict(s2_fields) for r in schema["section2"]["rows"]},
        "s3": {
            r["id"]: _empty_shift_row(schema["section3"]["shifts"])
            for r in schema["section3"]["rows"]
        },
        "s4": {
            u: {"s1": "", "s2": "", "s3": "", "daily": "", "prev_day": "", "monthly": ""}
            for u in schema["section4"]["units"]
        },
        "s5": {
            r["id"]: {
                **_empty_shift_row(schema["section5"]["shifts"]),
                "steam": "",
                "notes": "",
            }
            for r in schema["section5"]["rows"]
        },
        "s6": {
            "op": {
                r["id"]: _empty_shift_row(schema["section6"]["shifts"])
                for r in schema["section6"]["operation_rows"]
            },
            "chiller": {
                r["id"]: {f["id"]: "" for f in schema["section6"]["chiller_fields"]}
                for r in schema["section6"]["chiller_rows"]
            },
        },
    }
    return data


def _calc_meter(prev: str, today: str, prev_monthly: str = "") -> tuple[str, str]:
    p, t = _parse_num(prev), _parse_num(today)
    daily = ""
    if p is not None and t is not None:
        daily = _fmt_num(t - p)
    pm = _parse_num(prev_monthly)
    monthly = ""
    d = _parse_num(daily)
    if pm is not None and d is not None:
        monthly = _fmt_num(pm + d)
    elif d is not None:
        monthly = daily
    return daily, monthly


def _apply_shift_calc(row: dict, shifts: int, time_to_hr: bool = True) -> None:
    hrs = []
    for i in range(1, shifts + 1):
        tkey, hkey = f"s{i}_time", f"s{i}_hr"
        if time_to_hr and row.get(tkey):
            hr = parse_time_range_hours(str(row[tkey]))
            if hr is not None:
                row[hkey] = _fmt_num(hr)
        h = _parse_num(row.get(hkey))
        if h is not None:
            hrs.append(h)
    if hrs:
        row["daily"] = _fmt_num(sum(hrs))
    pd = _parse_num(row.get("prev_day"))
    d = _parse_num(row.get("daily"))
    if pd is not None and d is not None:
        row["monthly"] = _fmt_num(pd + d)
    elif d is not None:
        row["monthly"] = row["daily"]


def _apply_s4_calc(unit: dict) -> None:
    hrs = [_parse_num(unit.get(f"s{i}")) for i in (1, 2, 3)]
    hrs = [h for h in hrs if h is not None]
    if hrs:
        unit["daily"] = _fmt_num(sum(hrs))
    pd = _parse_num(unit.get("prev_day"))
    d = _parse_num(unit.get("daily"))
    if pd is not None and d is not None:
        unit["monthly"] = _fmt_num(pd + d)
    elif d is not None:
        unit["monthly"] = unit["daily"]


def recompute_daily(data: dict) -> dict:
    data = deepcopy(data or {})
    s1 = data.setdefault("s1", empty_daily_payload()["s1"])

    for mid in ("heat", "flow"):
        m = s1.setdefault(mid, {})
        daily, monthly = _calc_meter(m.get("prev", ""), m.get("today", ""), "")
        m["daily"] = daily
        m["monthly"] = monthly

    pm = s1.setdefault("power", {})
    daily, monthly = _calc_meter(pm.get("prev", ""), pm.get("today", ""), pm.get("prev_monthly", ""))
    pm["daily"] = daily
    pm["monthly"] = monthly

    schema = load_schema()
    for rid, row in data.setdefault("s3", {}).items():
        _apply_shift_calc(row, schema["section3"]["shifts"])
    for uid, unit in data.setdefault("s4", {}).items():
        _apply_s4_calc(unit)
    for rid, row in data.setdefault("s5", {}).items():
        _apply_shift_calc(row, schema["section5"]["shifts"])
    for rid, row in data.setdefault("s6", {}).get("op", {}).items():
        _apply_shift_calc(row, schema["section6"]["shifts"])
    return data


async def get_or_create_daily(
    session: AsyncSession, building_id: int, log_date: date
) -> CcrFacilityDaily:
    row = (
        await session.execute(
            select(CcrFacilityDaily).where(
                CcrFacilityDaily.building_id == building_id,
                CcrFacilityDaily.log_date == log_date,
            )
        )
    ).scalar_one_or_none()
    if row:
        return row
    row = CcrFacilityDaily(building_id=building_id, log_date=log_date, data=empty_daily_payload())
    session.add(row)
    await session.flush()
    return row


async def _fetch_day_data(
    session: AsyncSession, building_id: int, log_date: date
) -> dict:
    row = (
        await session.execute(
            select(CcrFacilityDaily).where(
                CcrFacilityDaily.building_id == building_id,
                CcrFacilityDaily.log_date == log_date,
            )
        )
    ).scalar_one_or_none()
    return (row.data or {}) if row else {}


async def sync_prev_values(
    session: AsyncSession, building_id: int, log_date: date, data: dict
) -> dict:
    """전일 금일지침·월누계·전일누계 자동 반영."""
    data = recompute_daily(data)
    prev_date = log_date - timedelta(days=1)
    prev = await _fetch_day_data(session, building_id, prev_date)
    if not prev:
        return data

    prev = recompute_daily(prev)
    s1 = data.setdefault("s1", {})
    prev_s1 = prev.get("s1", {})

    for mid in ("heat", "flow", "power"):
        m = s1.setdefault(mid, {})
        if not m.get("prev_manual"):
            m["prev"] = (prev_s1.get(mid) or {}).get("today", "")
    pm = s1.setdefault("power", {})
    if not pm.get("prev_manual"):
        pm["prev_monthly"] = (prev_s1.get("power") or {}).get("monthly", "")

    for mid in ("heat", "flow"):
        m = s1.setdefault(mid, {})
        prev_m = prev_s1.get(mid) or {}
        prev_monthly_val = prev_m.get("monthly", "")
        daily, monthly = _calc_meter(m.get("prev", ""), m.get("today", ""), prev_monthly_val)
        m["daily"] = daily
        m["monthly"] = monthly

    pm = s1.setdefault("power", {})
    daily, monthly = _calc_meter(
        pm.get("prev", ""), pm.get("today", ""), pm.get("prev_monthly", "")
    )
    pm["daily"] = daily
    pm["monthly"] = monthly

    schema = load_schema()
    for sec_key, rows_key, shifts in (
        ("s3", schema["section3"]["rows"], schema["section3"]["shifts"]),
        ("s5", schema["section5"]["rows"], schema["section5"]["shifts"]),
    ):
        for r in rows_key:
            rid = r["id"]
            row = data.setdefault(sec_key, {}).setdefault(rid, _empty_shift_row(shifts))
            prev_row = (prev.get(sec_key) or {}).get(rid, {})
            row["prev_day"] = prev_row.get("monthly", "") or prev_row.get("daily", "")

    for uid in schema["section4"]["units"]:
        unit = data.setdefault("s4", {}).setdefault(uid, {})
        prev_unit = (prev.get("s4") or {}).get(uid, {})
        unit["prev_day"] = prev_unit.get("monthly", "") or prev_unit.get("daily", "")

    for r in schema["section6"]["operation_rows"]:
        rid = r["id"]
        row = data.setdefault("s6", {}).setdefault("op", {}).setdefault(
            rid, _empty_shift_row(schema["section6"]["shifts"])
        )
        prev_row = (prev.get("s6") or {}).get("op", {}).get(rid, {})
        row["prev_day"] = prev_row.get("monthly", "") or prev_row.get("daily", "")

    return recompute_daily(data)


def merge_daily_save(existing: dict, posted: dict) -> dict:
    out = deepcopy(existing or empty_daily_payload())
    for key, val in (posted or {}).items():
        if isinstance(val, dict) and isinstance(out.get(key), dict):
            _deep_merge(out[key], val)
        else:
            out[key] = val
    return recompute_daily(out)


def _deep_merge(base: dict, patch: dict) -> None:
    for k, v in patch.items():
        if isinstance(v, dict) and isinstance(base.get(k), dict):
            _deep_merge(base[k], v)
        else:
            base[k] = v


async def propagate_to_next_day(
    session: AsyncSession, building_id: int, log_date: date, data: dict
) -> None:
    nxt = log_date + timedelta(days=1)
    row = await get_or_create_daily(session, building_id, nxt)
    nxt_data = row.data or empty_daily_payload()
    synced = await sync_prev_values(session, building_id, nxt, nxt_data)
    row.data = synced
    row.updated_at = datetime.utcnow()


def compute_monthly_report(
    building_id: int,
    year: int,
    month: int,
    daily_rows: list[CcrFacilityDaily],
) -> dict[str, Any]:
    schema = load_schema()
    by_date = {r.log_date: recompute_daily(r.data or {}) for r in daily_rows}
    import calendar

    last = calendar.monthrange(year, month)[1]
    days = []
    for d in range(1, last + 1):
        ld = date(year, month, d)
        dd = by_date.get(ld, {})
        s1 = dd.get("s1", {})
        days.append(
            {
                "day": d,
                "date": ld.isoformat(),
                "heat_daily": (s1.get("heat") or {}).get("daily", ""),
                "flow_daily": (s1.get("flow") or {}).get("daily", ""),
                "power_daily": (s1.get("power") or {}).get("daily", ""),
                "heat_monthly": (s1.get("heat") or {}).get("monthly", ""),
                "flow_monthly": (s1.get("flow") or {}).get("monthly", ""),
                "power_monthly": (s1.get("power") or {}).get("monthly", ""),
                "s3": {r["id"]: (dd.get("s3") or {}).get(r["id"], {}) for r in schema["section3"]["rows"]},
                "s5": {r["id"]: (dd.get("s5") or {}).get(r["id"], {}) for r in schema["section5"]["rows"]},
                "s6_op": {
                    r["id"]: (dd.get("s6") or {}).get("op", {}).get(r["id"], {})
                    for r in schema["section6"]["operation_rows"]
                },
            }
        )
    return {"year": year, "month": month, "days": days, "schema": schema}


async def fetch_yearly_report_data(
    session: AsyncSession, building_id: int, year: int
) -> dict[str, Any]:
    rows = (
        await session.execute(
            select(CcrFacilityDaily).where(
                CcrFacilityDaily.building_id == building_id,
                CcrFacilityDaily.log_date >= date(year, 1, 1),
                CcrFacilityDaily.log_date <= date(year, 12, 31),
            )
        )
    ).scalars().all()
    months = []
    for m in range(1, 13):
        month_rows = [r for r in rows if r.log_date.month == m]
        rep = compute_monthly_report(building_id, year, m, month_rows)
        heat_sum = sum(_parse_num(d["heat_daily"]) or 0 for d in rep["days"])
        flow_sum = sum(_parse_num(d["flow_daily"]) or 0 for d in rep["days"])
        power_sum = sum(_parse_num(d["power_daily"]) or 0 for d in rep["days"])
        months.append(
            {
                "month": m,
                "heat_usage": _fmt_num(heat_sum) if heat_sum else "",
                "flow_usage": _fmt_num(flow_sum) if flow_sum else "",
                "power_usage": _fmt_num(power_sum) if power_sum else "",
            }
        )
    return {"year": year, "months": months}


async def fetch_notes_list(
    session: AsyncSession, building_id: int, year: int, month: int
) -> dict[str, Any]:
    import calendar

    last = calendar.monthrange(year, month)[1]
    rows = (
        await session.execute(
            select(CcrFacilityDaily).where(
                CcrFacilityDaily.building_id == building_id,
                CcrFacilityDaily.log_date >= date(year, month, 1),
                CcrFacilityDaily.log_date <= date(year, month, last),
            )
        )
    ).scalars().all()
    schema = load_schema()
    entries = []
    for row in sorted(rows, key=lambda r: r.log_date):
        data = row.data or {}
        items = []
        for s2r in schema["section2"]["rows"]:
            notes = ((data.get("s2") or {}).get(s2r["id"]) or {}).get("notes", "")
            if (notes or "").strip():
                items.append({"time": s2r["time"], "text": notes.strip(), "section": "가압펌프"})
        for s5r in schema["section5"]["rows"]:
            notes = ((data.get("s5") or {}).get(s5r["id"]) or {}).get("notes", "")
            if (notes or "").strip():
                items.append({"time": s5r["time"], "text": notes.strip(), "section": "보일러"})
        if items:
            entries.append(
                {"date": row.log_date.isoformat(), "day": row.log_date.day, "items": items}
            )
    return {"entries": entries}


def parse_daily_form(form) -> dict:
    data = empty_daily_payload()
    items = form.multi_items() if hasattr(form, "multi_items") else form.items()
    for key, val in items:
        if not isinstance(key, str) or not key.startswith("f__"):
            continue
        raw = (val or "").strip() if isinstance(val, str) else str(val or "")
        parts = key.split("__")
        if len(parts) < 3:
            continue
        _, section = parts[0], parts[1]
        if section == "s1":
            meter, field = parts[2], parts[3]
            if field == "prev_manual":
                data["s1"].setdefault(meter, {})["prev_manual"] = raw == "1"
            else:
                data["s1"].setdefault(meter, {})[field] = raw
        elif section == "s2":
            row_id, field = parts[2], parts[3]
            data["s2"].setdefault(row_id, {})[field] = raw
        elif section == "s3":
            row_id, field = parts[2], parts[3]
            data["s3"].setdefault(row_id, {})[field] = raw
        elif section == "s4":
            unit, field = parts[2], parts[3]
            data["s4"].setdefault(unit, {})[field] = raw
        elif section == "s5":
            row_id, field = parts[2], parts[3]
            data["s5"].setdefault(row_id, {})[field] = raw
        elif section == "s6op":
            row_id, field = parts[2], parts[3]
            data["s6"]["op"].setdefault(row_id, {})[field] = raw
        elif section == "s6ch":
            row_id, field = parts[2], parts[3]
            data["s6"]["chiller"].setdefault(row_id, {})[field] = raw
    return data


def export_daily_to_excel(data: dict, log_date: date) -> bytes:
    data = recompute_daily(data)
    if TEMPLATE_XLSX.exists():
        wb = load_workbook(TEMPLATE_XLSX)
        ws = wb.active
    else:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "설비"

    s1 = data.get("s1", {})
    heat = s1.get("heat", {})
    flow = s1.get("flow", {})
    power = s1.get("power", {})
    peak = s1.get("peak", {})
    ws["B9"] = heat.get("prev")
    ws["C9"] = heat.get("today")
    ws["D9"] = heat.get("daily")
    ws["E9"] = heat.get("monthly")
    ws["F9"] = flow.get("prev")
    ws["G9"] = flow.get("today")
    ws["H9"] = flow.get("daily")
    ws["I9"] = flow.get("monthly")
    ws["J9"] = power.get("prev")
    ws["K9"] = power.get("today")
    ws["L9"] = power.get("daily")
    ws["M9"] = power.get("monthly")
    ws["N9"] = power.get("prev_monthly")
    ws["O9"] = peak.get("time")
    ws["P9"] = peak.get("load")

    s2 = data.get("s2", {})
    for r_idx, rid in ((15, "day"), (16, "night")):
        row = s2.get(rid, {})
        ws.cell(r_idx, 6, row.get("supply_temp"))
        ws.cell(r_idx, 8, row.get("supply_pressure"))
        ws.cell(r_idx, 10, row.get("return_temp"))
        ws.cell(r_idx, 12, row.get("return_pressure"))
        ws.cell(r_idx, 14, row.get("notes"))

    s3_rows = {"housing": 21, "baegun": 22, "dongbaek": 23, "welfare": 24}
    for rid, r in s3_rows.items():
        row = (data.get("s3") or {}).get(rid, {})
        ws.cell(r, 4, row.get("s1_time"))
        ws.cell(r, 6, row.get("s1_hr"))
        ws.cell(r, 7, row.get("s2_time"))
        ws.cell(r, 9, row.get("s2_hr"))
        ws.cell(r, 10, row.get("s3_time"))
        ws.cell(r, 12, row.get("s3_hr"))
        ws.cell(r, 13, row.get("daily"))
        ws.cell(r, 14, row.get("monthly"))
        ws.cell(r, 15, row.get("prev_day"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def archive_daily_excel(session: AsyncSession, building_id: int, log_date: date) -> None:
    row = (
        await session.execute(
            select(CcrFacilityDaily).where(
                CcrFacilityDaily.building_id == building_id,
                CcrFacilityDaily.log_date == log_date,
            )
        )
    ).scalar_one_or_none()
    if not row:
        return
    xbytes = export_daily_to_excel(row.data or {}, log_date)
    arch = (
        await session.execute(
            select(CcrFacilityArchive).where(
                CcrFacilityArchive.building_id == building_id,
                CcrFacilityArchive.log_date == log_date,
            )
        )
    ).scalar_one_or_none()
    fname = f"중앙관제실설비_{log_date.isoformat()}.xlsx"
    if arch:
        arch.file_data = xbytes
        arch.file_size = len(xbytes)
        arch.original_name = fname
    else:
        session.add(
            CcrFacilityArchive(
                building_id=building_id,
                log_date=log_date,
                original_name=fname,
                file_data=xbytes,
                file_size=len(xbytes),
            )
        )


def ccr_facility_daily_qr_url(building_code: str, request: Any | None = None) -> str:
    import os

    env = (os.environ.get("PUBLIC_BASE_URL") or "").strip().rstrip("/")
    if env:
        base = env
    elif request is not None:
        base = str(request.base_url).rstrip("/")
    else:
        base = "http://127.0.0.1:8000"
    return f"{base}/ccrf/{building_code}/daily"


def qr_png_bytes(url: str) -> bytes:
    import qrcode

    img = qrcode.make(url)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


async def get_building_for_qr(session: AsyncSession, code: str) -> Building | None:
    from models import InspectionLogBuilding2

    building = (
        await session.execute(
            select(Building)
            .join(InspectionLogBuilding2, InspectionLogBuilding2.building_id == Building.id)
            .where(
                Building.code == (code or "").strip(),
                Building.is_active == True,  # noqa: E712
            )
        )
    ).scalar_one_or_none()
    if not building or not is_ccr_facility_building(building):
        return None
    return building


async def ensure_tables(engine) -> None:
    from sqlalchemy import text

    url = str(engine.url).lower()
    is_pg = "postgresql" in url or "postgres" in url
    async with engine.begin() as conn:
        if is_pg:
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS ccr_facility_daily (
                        id SERIAL PRIMARY KEY,
                        building_id INTEGER NOT NULL REFERENCES buildings(id),
                        log_date DATE NOT NULL,
                        data JSONB NOT NULL DEFAULT '{}',
                        created_at TIMESTAMP WITHOUT TIME ZONE DEFAULT NOW(),
                        updated_at TIMESTAMP WITHOUT TIME ZONE DEFAULT NOW(),
                        UNIQUE (building_id, log_date)
                    )
                    """
                )
            )
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS ccr_facility_archives (
                        id SERIAL PRIMARY KEY,
                        building_id INTEGER NOT NULL REFERENCES buildings(id),
                        log_date DATE NOT NULL,
                        original_name VARCHAR(300),
                        file_data BYTEA,
                        file_size INTEGER,
                        created_at TIMESTAMP WITHOUT TIME ZONE DEFAULT NOW(),
                        UNIQUE (building_id, log_date)
                    )
                    """
                )
            )
        else:
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS ccr_facility_daily (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        building_id INTEGER NOT NULL,
                        log_date DATE NOT NULL,
                        data TEXT NOT NULL DEFAULT '{}',
                        created_at DATETIME,
                        updated_at DATETIME,
                        UNIQUE (building_id, log_date)
                    )
                    """
                )
            )
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS ccr_facility_archives (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        building_id INTEGER NOT NULL,
                        log_date DATE NOT NULL,
                        original_name VARCHAR(300),
                        file_data BLOB,
                        file_size INTEGER,
                        created_at DATETIME,
                        UNIQUE (building_id, log_date)
                    )
                    """
                )
            )


def register_scheduler(scheduler, session_factory, kst) -> None:
    async def _close_all() -> None:
        from datetime import datetime as dt

        today = dt.now(kst).date()
        async with session_factory() as session:
            from models import InspectionLogBuilding2

            rows = (
                await session.execute(
                    select(InspectionLogBuilding2.building_id, Building)
                    .join(Building, Building.id == InspectionLogBuilding2.building_id)
                )
            ).all()
            for bid, bld in rows:
                if not is_ccr_facility_building(bld):
                    continue
                try:
                    await archive_daily_excel(session, bid, today)
                    await propagate_to_next_day(session, bid, today, {})
                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    print(f"[ccrf] close {bid}: {e}", flush=True)

    scheduler.add_job(_close_all, "cron", hour=23, minute=59, timezone=kst, id="ccrf_daily_close")
