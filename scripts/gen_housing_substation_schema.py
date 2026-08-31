# -*- coding: utf-8 -*-
"""주택변전소.xlsx → resources/housing_substation_schema.json"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "resources" / "housing_substation_template.xlsx"
if not XLSX.is_file():
    XLSX = ROOT / "data" / "주택변전소.xlsx"
OUT = ROOT / "resources" / "housing_substation_schema.json"


def parse_mult(s) -> int | None:
    if not s:
        return None
    m = re.search(r"(\d+)", str(s))
    return int(m.group(1)) if m else None


def main() -> None:
    wb = load_workbook(XLSX)
    ws = wb["1일"]
    blocks_cfg = []
    meta = [
        ("no3", 7, 12, 3, "No.3 6.6KV INCOMMING"),
        ("no2", 20, 25, 2, "No.2 6.6KV INCOMMING"),
        ("no1", 32, 37, 1, "No.1 6.6KV INCOMMING"),
    ]
    for bid, rs, _re, monthly_sec, title in meta:
        header = None
        for hr in range(rs - 5, rs):
            v6 = ws.cell(hr, 6).value
            v8 = ws.cell(hr, 8).value
            if v6 in ("V", "kV") or v8 in ("V", "kV"):
                header = hr
        cols = []
        groups: dict[int, str] = {}
        for hr in range(rs - 6, rs):
            for c in range(6, 100):
                v = ws.cell(hr, c).value
                if not v or not isinstance(v, str):
                    continue
                t = v.strip()
                if len(t) < 3 or "주택" in t or t == "구분":
                    continue
                if any(
                    k in t
                    for k in (
                        "INCOMMING",
                        "S/S",
                        "MAIN",
                        "백운",
                        "몰오",
                        "학교",
                        "마트",
                        "센타",
                        "쇼핑",
                        "웰빙",
                        "동백",
                        "제철",
                    )
                ):
                    groups[c] = t
        if not header:
            continue
        for c in range(6, 100):
            unit = ws.cell(header, c).value
            metric = ws.cell(header - 1, c).value if header > 1 else None
            rng = ws.cell(header + 1, c).value
            if not unit and not metric:
                continue
            g = ""
            for gc, gv in sorted(groups.items()):
                if gc <= c:
                    g = gv
            mult = parse_mult(metric) or parse_mult(ws.cell(header - 2, c).value if header > 2 else None)
            if not mult:
                # look left for 적산량×NNNN in same breaker group
                for lc in range(c, 5, -1):
                    lm = ws.cell(header - 1, lc).value
                    if lm and "적산" in str(lm):
                        mult = parse_mult(lm)
                        break
            cols.append(
                {
                    "col": get_column_letter(c),
                    "group": g,
                    "metric": str(metric or ""),
                    "unit": str(unit or ""),
                    "range": str(rng or ""),
                    "multiplier": mult,
                    "is_cumulative": "전일" in str(unit or "") or "전일" in str(metric or ""),
                }
            )
        input_cols = [x for x in cols if not x["is_cumulative"]]
        meters = []
        for i, col in enumerate(cols):
            if not col["is_cumulative"]:
                continue
            cc = None
            for j in range(i - 1, -1, -1):
                if cols[j]["unit"] == "A":
                    cc = cols[j]["col"]
                    break
            meters.append(
                {
                    "id": f"{bid}_{col['col']}",
                    "name": col["group"] or title,
                    "reading_col": col["col"],
                    "current_col": cc,
                    "multiplier": col["multiplier"] or 4800,
                }
            )
        blocks_cfg.append(
            {
                "id": bid,
                "title": title,
                "monthly_section": monthly_sec,
                "times": ["06:00", "10:00", "14:00", "17:00", "20:00", "22:00"],
                "input_columns": input_cols,
                "meters": meters,
            }
        )

    ws2 = wb["월보"]
    monthly = []
    for start, sec_id in [(5, "no1"), (44, "no2"), (83, "no3")]:
        br = []
        for c in range(2, 40):
            name = ws2.cell(start, c).value
            if not name or not str(name).strip():
                continue
            name_s = str(name).strip()
            if name_s.lower() == "spare":
                continue
            mult = parse_mult(ws2.cell(start + 1, c).value)
            if mult is None and c > 2:
                mult = parse_mult(ws2.cell(start + 1, c - 1).value)
            br.append(
                {
                    "name": name_s,
                    "multiplier": mult,
                    "cols": {
                        "reading": get_column_letter(c),
                        "usage": get_column_letter(c + 1),
                        "max_a": get_column_letter(c + 2),
                    },
                }
            )
        monthly.append({"id": sec_id, "breakers": br})

    cfg = {
        "building_name": "주택변전소",
        "daily_blocks": blocks_cfg,
        "monthly_sections": monthly,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT} blocks={len(blocks_cfg)}")


if __name__ == "__main__":
    main()
