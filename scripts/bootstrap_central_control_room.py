# -*- coding: utf-8 -*-
"""주택변전소 모듈 기반 중앙관제실 모듈 생성."""
from __future__ import annotations

import json
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]


def build_schema() -> dict:
    sections = [
        (
            "line",
            "한전 / 수전 LINE INCOMING",
            [
                ("B", "전압", "V", "6300~6600"),
                ("C", "전류", "A", "0~800"),
                ("D", "유효", "KW", "0~4800"),
                ("E", "무효", "KVAR", "0~4800"),
                ("F", "역률", "COS", "0.9±0.1"),
                ("G", "주파수", "HZ", "59.5~60.5"),
            ],
            "H",
            "C",
            4800,
        ),
        (
            "baegun_life",
            "백운생활관",
            [("I", "유효", "KW", "0~2400"), ("J", "전류", "A", "0~200")],
            "K",
            "J",
            2400,
        ),
        (
            "sc1",
            "SC I TR",
            [("L", "유효", "KW", "0~2400"), ("M", "전류", "A", "0~200")],
            "N",
            "M",
            2400,
        ),
        (
            "plaza",
            "백운플라자",
            [("O", "유효", "KW", "0~2400"), ("P", "전류", "A", "0~200")],
            "Q",
            "P",
            2400,
        ),
        (
            "no2",
            "NO.2 TR",
            [("R", "유효", "KW", "0~2400"), ("S", "전류", "A", "0~200")],
            "T",
            "S",
            2400,
        ),
        (
            "no3",
            "NO.3 TR",
            [("U", "유효", "KW", "0~2400"), ("V", "전류", "A", "0~200")],
            "W",
            "V",
            2400,
        ),
        (
            "argpol",
            "아르폴, 그린, 교육센터",
            [("X", "유효", "KW", "0~2400"), ("Y", "전류", "A", "0~200")],
            "Z",
            "Y",
            2400,
        ),
        (
            "laundry",
            "휴먼스 세탁소",
            [("AA", "유효", "KW", "0~2400"), ("AB", "전류", "A", "0~200")],
            "AC",
            "AB",
            2400,
        ),
    ]
    input_cols: list[dict] = []
    meters: list[dict] = []
    breakers: list[dict] = []
    for _sid, gname, cols, rc, cc, mult in sections:
        for col, metric, unit, rng in cols:
            input_cols.append(
                {
                    "col": col,
                    "group": gname,
                    "metric": metric,
                    "unit": unit,
                    "range": rng,
                    "multiplier": mult,
                    "is_cumulative": False,
                }
            )
        meters.append(
            {
                "id": f"main_{rc}",
                "name": gname,
                "reading_col": rc,
                "current_col": cc,
                "multiplier": mult,
            }
        )
        breakers.append({"name": gname, "multiplier": mult, "cols": {}})

    from openpyxl.utils import get_column_letter

    for i, br in enumerate(breakers):
        c = 2 + i * 3
        br["cols"] = {
            "reading": get_column_letter(c),
            "usage": get_column_letter(c + 1),
            "max_a": get_column_letter(c + 2),
        }

    checklist_items = [
        ("vcb", "VCB 상태 및 계기 이상유무"),
        ("ct_pt", "CT, PT 및 각 계기 정상동작 유무"),
        ("ct_pt_bus", "CT, PT 단자조임 및 BUS-BAR 열화상태"),
        ("relay", "보호계전기 상태"),
        ("meter_relay", "계기지시 및 보호계전기 상태"),
        ("cable", "CABLE 단말부 및 접속부 열화상태"),
        ("acb", "ACB, MCCB 및 CABLE 접속부 열화상태"),
        ("battery", "BATTERY CHARGER PANEL 정상동작 유무"),
        ("plc", "PLC PANEL 및 UPS 이상유무"),
    ]

    return {
        "building_name": "중앙관제실",
        "daily_blocks": [
            {
                "id": "main",
                "title": "중앙관제실 전기운전 및 설비점검 일지",
                "monthly_section": 1,
                "times": ["06:00", "10:00", "14:00", "16:00", "20:00", "22:00"],
                "input_columns": input_cols,
                "meters": meters,
            }
        ],
        "monthly_sections": [{"id": "main", "breakers": breakers}],
        "daily_footer": {
            "times": ["06:00", "10:00", "14:00", "16:00", "20:00", "22:00"],
            "transformer": {
                "title": "변압기온도(105℃이하)",
                "subtitle": "",
                "fields": [
                    {"id": "tr1", "name": "#1", "col": "N", "range": "105℃이하"},
                    {"id": "tr2", "name": "#2", "col": "O", "range": "105℃이하"},
                    {"id": "tr3", "name": "#3", "col": "P", "range": "105℃이하"},
                ],
            },
            "checklist": {
                "title": "점검사항",
                "shifts": ["주간", "야간"],
                "items": [{"id": i, "name": n} for i, n in checklist_items],
            },
            "notes": {"title": "특이사항"},
        },
    }


def copy_python_module() -> None:
    src = (ROOT / "housing_substation.py").read_text(encoding="utf-8")
    repl = [
        ("HousingSubstation", "CentralControlRoom"),
        ("housing_substation", "central_control_room"),
        ("주택변전소", "중앙관제실"),
        ("housing_substation_", "central_control_room_"),
        ("uq_housing_sub_daily", "uq_ccr_daily"),
        ("uq_housing_sub_archive", "uq_ccr_archive"),
        ("/hs/", "/ccr/"),
        ("X-HS-", "X-CCR-"),
        ("is_housing_substation_building", "is_central_control_room_building"),
        ("housing_daily_qr_url", "ccr_daily_qr_url"),
        ("get_housing_building_for_qr", "get_ccr_building_for_qr"),
        ("[housing_sub]", "[ccr]"),
        ("housing_sub", "ccr"),
        ("housing_substation_schema.json", "central_control_room_schema.json"),
        ("housing_substation_template.xlsx", "central_control_room_template.xlsx"),
    ]
    for a, b in repl:
        src = src.replace(a, b)
    src = src.replace(
        'block_starts = {"no3": 7, "no2": 20, "no1": 32}',
        'block_starts = {"main": 7}',
    )
    src = src.replace('roman = {"no1": "Ⅰ", "no2": "Ⅱ", "no3": "Ⅲ"}', 'roman = {"main": ""}')
    src = src.replace('order = {"no1": 0, "no2": 1, "no3": 2}', 'order = {"main": 0}')
    src = src.replace(
        '"no1": {"header": 5, "prev": 8, "day1": 9, "total": 40},\n'
        '        "no2": {"header": 44, "prev": 47, "day1": 48, "total": 79},\n'
        '        "no3": {"header": 83, "prev": 86, "day1": 87, "total": 118},',
        '"main": {"header": 5, "prev": 8, "day1": 9, "total": 40},',
    )
    (ROOT / "central_control_room.py").write_text(src, encoding="utf-8")


def copy_front() -> None:
    pairs = [
        (ROOT / "templates" / "housing_substation.html", ROOT / "templates" / "central_control_room.html"),
        (ROOT / "static" / "css" / "housing_substation.css", ROOT / "static" / "css" / "central_control_room.css"),
        (ROOT / "static" / "js" / "housing_substation.js", ROOT / "static" / "js" / "central_control_room.js"),
    ]
    for src_path, dst_path in pairs:
        text = src_path.read_text(encoding="utf-8")
        for a, b in [
            ("housing_substation", "central_control_room"),
            ("주택변전소", "중앙관제실"),
            ("/hs/", "/ccr/"),
            ("X-HS-", "X-CCR-"),
            ("hs-daily-form", "ccr-daily-form"),
            ("hs-save-status", "ccr-save-status"),
        ]:
            text = text.replace(a, b)
        dst_path.write_text(text, encoding="utf-8")


def main() -> None:
    schema = build_schema()
    out = ROOT / "resources" / "central_control_room_schema.json"
    out.write_text(json.dumps(schema, ensure_ascii=False, indent=2), encoding="utf-8")
    copy_python_module()
    copy_front()
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
