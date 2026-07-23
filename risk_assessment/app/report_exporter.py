"""위험성평가서 양식 — Excel/텍스트 출력"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.local_engine import RiskRow
from app.risk_form import (
    FormRow,
    convert_to_form_rows,
    excel_r_formula,
    excel_result_formula,
    merge_sequence_cells,
)
from app.field_content import format_improvement_summary
from app.risk_grade_matrix import insert_approval_image, insert_risk_matrix_image

TRADE_FILL = PatternFill("solid", fgColor="FCE4D6")
GREY_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUB_FILL = PatternFill("solid", fgColor="B4C6E7")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")
ROW_FILL = PatternFill("solid", fgColor="FFFFFF")
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
HDR_FONT = Font(bold=True, size=10)
NORMAL = Font(size=10)
LINK_FONT = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")


@dataclass
class ReportMeta:
    seq: int = 1
    group: str = ""
    section: str = ""
    job_name: str = ""
    evaluator: str = ""
    department: str = ""
    assessment_no: str = ""
    apply_type: str = "정기평가"
    ai_name: str = "P-WIDE V1 Local"
    ai_link: str = ""
    work_trade: str = ""       # 공종
    work_content: str = ""     # 작업내용
    progress_rate: str = ""    # 공정율
    person_in_charge: str = "" # 담당자


@dataclass
class AssessmentBundle:
    meta: ReportMeta
    rows: list[RiskRow] = field(default_factory=list)
    created_at: datetime = field(default_factory=datetime.now)
    mode: str = "local"

    @property
    def form_rows(self) -> list[FormRow]:
        return convert_to_form_rows(self.rows, self.meta.job_name)


def _cell(ws, row, col, value, fill=ROW_FILL, bold=False, wrap=True, center=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill
    c.font = Font(bold=bold, size=10)
    c.border = BORDER
    c.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center" if center else "top",
        wrap_text=wrap,
    )
    return c


def _formula_cell(ws, row, col, formula, fill=ROW_FILL, bold=False, wrap=True, center=False):
    c = ws.cell(row=row, column=col)
    c.value = formula
    c.fill = fill
    c.font = Font(bold=bold, size=10)
    c.border = BORDER
    c.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center" if center else "top",
        wrap_text=wrap,
    )
    return c


def _apply_fs_data_validation(ws, data_start: int, last_row: int) -> None:
    """F(1~5), S(1~4) 입력 제한 — 수정 시 R·결과 수식 자동 반영."""
    if last_row < data_start:
        return
    dv_f = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=5,
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="빈도(F)",
        error="1~5 사이 정수를 입력하세요.",
    )
    dv_s = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=4,
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="강도(S)",
        error="1~4 사이 정수를 입력하세요.",
    )
    ws.add_data_validation(dv_f)
    ws.add_data_validation(dv_s)
    for col in (5, 10):
        dv_f.add(f"{get_column_letter(col)}{data_start}:{get_column_letter(col)}{last_row}")
    for col in (6, 11):
        dv_s.add(f"{get_column_letter(col)}{data_start}:{get_column_letter(col)}{last_row}")


def format_report_text(bundle: AssessmentBundle) -> str:
    m = bundle.meta
    form = bundle.form_rows
    lines = [
        "■ 위험성평가표 (Risk Assessment)",
        "",
        f"  부서명: {m.department or m.group or '-'}",
        f"  평가서NO: {m.assessment_no or '-'}",
        f"  작업명: {m.job_name}",
        f"  평가일: {bundle.created_at:%Y. %m. %d}",
        f"  평가담당: {m.evaluator or '-'}",
        "",
    ]
    before_sum, after_sum = format_improvement_summary(bundle.rows)
    if before_sum:
        lines.extend([
            "■ 개선전 (유해·위험요인)",
            before_sum,
            "",
            "■ 개선후 (안전보건 대책)",
            after_sum,
            "",
        ])
    lines.extend([
        "| 작업순서 | 작업공정 | 재해형태 | 유해위험요인 | "
        "F | S | R | 현재안전조치(전) | 결과(전) | "
        "F | S | R | 현재안전조치(후) | 결과(후) | 법적근거 |",
        "|" + "|".join([" :--- "] * 15) + "|",
    ])
    prev_seq = ""
    for r in form:
        seq = r.work_sequence if r.work_sequence != prev_seq else "〃"
        prev_seq = r.work_sequence
        law_cell = f"[{r.law or '-'}]({r.law_url})" if r.law_url else (r.law or "-")
        lines.append(
            f"| {seq} | {r.work_process} | {r.disaster_type} | {r.hazard_factor} | "
            f"{r.f_before} | {r.s_before} | {r.r_before} | {r.measures_before} | {r.result_before()} | "
            f"{r.f_after} | {r.s_after} | {r.r_after} | {r.measures_after} | {r.result_after()} | {law_cell} |"
        )
    lines.extend([
        "",
        f"※ R = F × S  |  생성: {bundle.created_at:%Y-%m-%d %H:%M}  |  {m.ai_name}",
    ])
    return "\n".join(lines)


def _style_cell(
    ws,
    row: int,
    col: int,
    value="",
    *,
    fill=ROW_FILL,
    bold: bool = False,
    size: int = 11,
    horizontal: str = "left",
    vertical: str = "center",
    wrap: bool = False,
):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill
    c.font = Font(name="맑은 고딕", bold=bold, size=size)
    c.border = BORDER
    c.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)
    return c


def _merge_style(
    ws,
    r1: int,
    c1: int,
    r2: int,
    c2: int,
    value="",
    **kwargs,
):
    if r2 > r1 or c2 > c1:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return _style_cell(ws, r1, c1, value, **kwargs)


def _border_block(ws, r1: int, c1: int, r2: int, c2: int):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            if cell.border.left.style is None:
                cell.border = BORDER


def _write_summary_sheet(wb: Workbook, bundle: AssessmentBundle) -> None:
    """일일작업일보 Sheet2 양식 — 공종·작업사항 + 잠재위험/안전작업대책"""
    ws = wb.create_sheet("위험성평가 요약")
    m = bundle.meta
    before_sum, after_sum = format_improvement_summary(bundle.rows)

    col_widths = {
        1: 9.625, 2: 7.5, 3: 11.0, 4: 46.375, 5: 53.875,
        6: 44.125, 7: 13.5, 8: 10.625, 9: 13.0, 10: 13.0,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[2].height = 17.25
    ws.row_dimensions[3].height = 17.25
    line_count = max(before_sum.count("\n"), after_sum.count("\n"), 1)
    ws.row_dimensions[7].height = max(74.25, 15 * line_count + 24)

    hdr = dict(fill=GREY_FILL, bold=True, size=12, horizontal="center", vertical="center")

    # ── 2~3행: 헤더 ──
    _merge_style(ws, 2, 1, 2, 2, "구분", **hdr)
    _merge_style(ws, 2, 3, 3, 4, "작업 사항", **hdr)
    _merge_style(ws, 2, 5, 3, 5, "잠재위험 내용", **hdr)
    _merge_style(ws, 2, 6, 3, 7, " 안전작업 대책", **hdr)
    _merge_style(ws, 2, 8, 3, 8, "관리등급", **hdr, wrap=True)
    _merge_style(ws, 2, 9, 3, 9, "정비\n감독자", **hdr, wrap=True)
    _merge_style(ws, 2, 10, 3, 10, "관리주체", **hdr, wrap=True)
    _style_cell(ws, 3, 1, "공종", **hdr)
    _style_cell(ws, 3, 2, "NO.", **hdr)

    # ── 4~7행: 데이터 ──
    trade = dict(fill=TRADE_FILL, bold=True, size=12, horizontal="center", vertical="center")
    _merge_style(ws, 4, 1, 7, 1, m.work_trade, **trade, wrap=True)
    no_val = m.seq if m.seq else ""
    _merge_style(ws, 4, 2, 7, 2, no_val, **trade)

    label = dict(fill=ROW_FILL, bold=True, size=11, horizontal="center", vertical="center")
    _style_cell(ws, 4, 3, "작 업 명", **label)
    _style_cell(ws, 5, 3, "작업내용", **label)
    _style_cell(ws, 6, 3, "공정율", **label)
    _style_cell(ws, 7, 3, "담당자", **label)

    _merge_style(
        ws, 4, 4, 4, 7, m.job_name,
        fill=ROW_FILL, bold=True, size=11, horizontal="left", vertical="center",
    )
    _style_cell(ws, 5, 4, m.work_content, wrap=True)
    _style_cell(ws, 6, 4, m.progress_rate)
    _style_cell(ws, 7, 4, m.person_in_charge)

    _merge_style(
        ws, 5, 5, 7, 5, before_sum,
        fill=ROW_FILL, bold=False, size=12, horizontal="left", vertical="top", wrap=True,
    )
    _merge_style(
        ws, 5, 6, 7, 7, after_sum,
        fill=ROW_FILL, bold=False, size=12, horizontal="left", vertical="top", wrap=True,
    )

    for col in (8, 9, 10):
        _merge_style(
            ws, 4, col, 7, col, "",
            fill=ROW_FILL, bold=True, size=12, horizontal="center", vertical="center", wrap=True,
        )

    _border_block(ws, 2, 1, 7, 10)


def _write_law_cell(ws, row: int, col: int, law: str, law_url: str):
    """법적근거 셀 — law.go.kr 하이퍼링크"""
    c = ws.cell(row=row, column=col, value=law or "")
    c.fill = ROW_FILL
    c.border = BORDER
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    if law_url:
        c.hyperlink = law_url
        c.font = LINK_FONT
    else:
        c.font = NORMAL
    return c


def export_to_excel(bundle: AssessmentBundle, path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "위험성평가서"
    m = bundle.meta
    form = bundle.form_rows

    col_widths = [12, 18, 10, 36, 4, 4, 4, 28, 14, 4, 4, 4, 28, 14, 32]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 제목 ──
    ws.merge_cells("A1:O1")
    t = ws["A1"]
    t.value = "위험성평가표 (Risk Assessment)"
    t.fill = TITLE_FILL
    t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    for r in range(2, 5):
        ws.row_dimensions[r].height = 33
    for r in range(5, 7):
        ws.row_dimensions[r].height = 18

    # ── 기본정보 ──
    info = [
        ("A2", "B2", "부서명", m.department or m.group),
        ("A3", "B3", "평가서NO", m.assessment_no),
        ("A4", "B4", "작 업 명", m.job_name),
        ("C2", "D2", "평가일", f"{bundle.created_at:%Y. %m. %d}"),
        ("C3", "D3", "평가담당", m.evaluator),
        ("C4", "D4", "섹션", m.section),
    ]
    for c1, c2, label, val in info:
        ws.merge_cells(f"{c1}:{c2}")
        cell = ws[c1]
        cell.value = f"{label} : {val or ''}"
        cell.fill = HEADER_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # ── 표 헤더 (5행~6행) ──
    header_row = 6
    sub_row = 7
    headers = [
        (1, 1, "작업순서", 2),
        (2, 2, "작업공정", 2),
        (3, 3, "재해형태", 2),
        (4, 4, "유해위험요인", 2),
        (5, 9, "위험도(개선전)", 1),
        (10, 14, "위험도(개선후)", 1),
        (15, 15, "법적근거", 2),
    ]
    for c1, c2, title, rowspan in headers:
        if rowspan == 2:
            ws.merge_cells(start_row=header_row, start_column=c1, end_row=sub_row, end_column=c2)
        else:
            ws.merge_cells(start_row=header_row, start_column=c1, end_row=header_row, end_column=c2)
        cell = ws.cell(row=header_row, column=c1, value=title)
        cell.fill = SUB_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sub_headers = ["F", "S", "R", "현재안전조치", "결과"]
    for i, h in enumerate(sub_headers, 5):
        _cell(ws, sub_row, i, h, SUB_FILL, bold=True, center=True)
    for i, h in enumerate(sub_headers, 10):
        _cell(ws, sub_row, i, h, SUB_FILL, bold=True, center=True)

    ws.row_dimensions[header_row].height = 22
    ws.row_dimensions[sub_row].height = 22

    # ── 데이터 (F·S 값 입력 → R·결과는 수식) ──
    data_start = 8
    col_f_before, col_s_before, col_r_before = 5, 6, 7
    col_result_before = 9
    col_f_after, col_s_after, col_r_after = 10, 11, 12
    col_result_after = 14

    if form:
        for i, r in enumerate(form):
            row = data_start + i
            _cell(ws, row, 2, r.work_process)
            _cell(ws, row, 3, r.disaster_type, center=True)
            _cell(ws, row, 4, r.hazard_factor)
            _cell(ws, row, col_f_before, r.f_before, center=True)
            _cell(ws, row, col_s_before, r.s_before, center=True)
            _formula_cell(
                ws, row, col_r_before,
                excel_r_formula(row, col_f_before, col_s_before),
                center=True,
            )
            _cell(ws, row, 8, r.measures_before)
            _formula_cell(
                ws, row, col_result_before,
                excel_result_formula(row, col_r_before),
                center=True,
            )
            _cell(ws, row, col_f_after, r.f_after, center=True)
            _cell(ws, row, col_s_after, r.s_after, center=True)
            _formula_cell(
                ws, row, col_r_after,
                excel_r_formula(row, col_f_after, col_s_after),
                center=True,
            )
            _cell(ws, row, 13, r.measures_after)
            _formula_cell(
                ws, row, col_result_after,
                excel_result_formula(row, col_r_after),
                center=True,
            )
            _write_law_cell(ws, row, 15, r.law, r.law_url)
            ws.row_dimensions[row].height = 48

        last_data_row = data_start + len(form) - 1
        _apply_fs_data_validation(ws, data_start, last_data_row)

        for start, end in merge_sequence_cells(form):
            r1 = data_start + start
            r2 = data_start + end
            if r2 > r1:
                ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
            cell = ws.cell(row=r1, column=1, value=form[start].work_sequence)
            cell.fill = ROW_FILL
            cell.font = NORMAL
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A8"

    # 양식 이미지 (표 작성 후 삽입)
    insert_risk_matrix_image(ws)
    insert_approval_image(ws)

    _write_summary_sheet(wb, bundle)
    wb.save(path)
    return path
