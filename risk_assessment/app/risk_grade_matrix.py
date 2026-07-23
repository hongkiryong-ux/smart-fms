"""위험성평가서 양식 이미지 — 매트릭스(E2:I6), 결재란(M2:N4)"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

from app.runtime_paths import APP_ROOT

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

MATRIX_IMAGE_NAME = "risk_grade_matrix.png"
MATRIX_ANCHOR_ROW, MATRIX_ANCHOR_COL = 2, 5  # E2 (좌측 상단)
MATRIX_WIDTH_CM = 9.87
MATRIX_HEIGHT_CM = 3.55
MATRIX_ROW2, MATRIX_COL2 = 6, 9  # 참고 영역 (E2:I6)

APPROVAL_IMAGE_NAME = "approval_box.png"
APPROVAL_ROW1, APPROVAL_COL1 = 2, 13  # M2
APPROVAL_ROW2, APPROVAL_COL2 = 4, 14  # N4


def _asset_path(name: str) -> Path | None:
    p = APP_ROOT / "assets" / name
    return p if p.is_file() else None


def _col_width_px(ws: Worksheet, col: int) -> float:
    letter = get_column_letter(col)
    w = ws.column_dimensions[letter].width
    if w is None:
        w = 8.43
    return w * 7 + 5


def _row_height_px(ws: Worksheet, row: int) -> float:
    h = ws.row_dimensions[row].height
    if h is None:
        h = 15.0
    return h * 96.0 / 72.0


def _range_pixels(ws: Worksheet, row1: int, col1: int, row2: int, col2: int) -> tuple[int, int]:
    width = sum(_col_width_px(ws, c) for c in range(col1, col2 + 1))
    height = sum(_row_height_px(ws, r) for r in range(row1, row2 + 1))
    return int(width), int(height)


def cm_to_pixels(cm: float, dpi: int = 96) -> int:
    """Excel 이미지 크기 — cm → px (96 DPI)."""
    return int(round(cm / 2.54 * dpi))


def _insert_image(
    ws: Worksheet,
    image_name: str,
    row1: int,
    col1: int,
    row2: int,
    col2: int,
) -> bool:
    path = _asset_path(image_name)
    if not path:
        return False
    img = Image(str(path))
    w_px, h_px = _range_pixels(ws, row1, col1, row2, col2)
    if w_px > 0 and h_px > 0:
        img.width = w_px
        img.height = h_px
    ws.add_image(img, f"{get_column_letter(col1)}{row1}")
    return True


def _insert_image_fixed_cm(
    ws: Worksheet,
    image_name: str,
    anchor_row: int,
    anchor_col: int,
    width_cm: float,
    height_cm: float,
) -> bool:
    """좌측 상단 셀 기준 고정 크기(cm)로 이미지 삽입."""
    path = _asset_path(image_name)
    if not path:
        return False
    img = Image(str(path))
    img.width = cm_to_pixels(width_cm)
    img.height = cm_to_pixels(height_cm)
    ws.add_image(img, f"{get_column_letter(anchor_col)}{anchor_row}")
    return True


def insert_risk_matrix_image(
    ws: Worksheet,
    anchor_row: int = MATRIX_ANCHOR_ROW,
    anchor_col: int = MATRIX_ANCHOR_COL,
    width_cm: float = MATRIX_WIDTH_CM,
    height_cm: float = MATRIX_HEIGHT_CM,
) -> bool:
    """위험등급 판단기준 — E2 좌상단, 9.87cm × 3.55cm."""
    return _insert_image_fixed_cm(
        ws, MATRIX_IMAGE_NAME, anchor_row, anchor_col, width_cm, height_cm
    )


def insert_approval_image(
    ws: Worksheet,
    row1: int = APPROVAL_ROW1,
    col1: int = APPROVAL_COL1,
    row2: int = APPROVAL_ROW2,
    col2: int = APPROVAL_COL2,
) -> bool:
    """결재란 PNG → M2:N4."""
    return _insert_image(ws, APPROVAL_IMAGE_NAME, row1, col1, row2, col2)


def format_matrix_text() -> str:
    return (
        "■ 위험등급 판단기준\n"
        "  가능성(빈도) × 중대성(강도) 매트릭스 — Excel 양식 E2:I6 참조 (6등급 A~F)"
    )
