"""위험성평가서 양식 변환 (작업순서·공정·F×S·등급)"""

from __future__ import annotations

from dataclasses import dataclass

from app.field_content import clean_improvement_text, polish_hazard, polish_improvement
from app.local_engine import RiskRow, risk_grade

# 작업 전/중/후 → 작업순서(대분류)
PHASE_SEQUENCE = {
    "작업 전": "작업준비",
    "작업 중": "작업중",
    "작업 후": "작업 후",
}

SEQUENCE_ORDER = {
    "작업준비": 0,
    "점검": 1,
    "측정": 2,
    "보수": 3,
    "작업": 4,
    "작업중": 4,
    "운전": 4,
    "비상대응": 5,
    "작업 후": 6,
}

GRADE_LABEL = {
    "A급": "허용불가위험",
    "B급": "중대위험",
    "C급": "상당위험",
    "D급": "경미위험",
    "E급": "미미위험",
    "F급": "무시위험",
}


@dataclass
class FormRow:
    work_sequence: str
    work_process: str
    disaster_type: str
    hazard_factor: str
    f_before: int
    s_before: int
    measures_before: str
    f_after: int
    s_after: int
    measures_after: str
    law: str = ""
    law_url: str = ""

    @property
    def r_before(self) -> int:
        return self.f_before * self.s_before

    @property
    def r_after(self) -> int:
        return self.f_after * self.s_after

    def result_before(self) -> str:
        return format_result(self.r_before)

    def result_after(self) -> str:
        return format_result(self.r_after)


def format_result(score: int) -> str:
    grade, level, _ = risk_grade(score)
    label = GRADE_LABEL.get(grade, level)
    return f"{label} ({grade})"


def excel_r_formula(row: int, f_col: int, s_col: int) -> str:
    """Excel R = F × S (행·열 번호)."""
    from openpyxl.utils import get_column_letter

    return f"={get_column_letter(f_col)}{row}*{get_column_letter(s_col)}{row}"


def excel_result_formula(row: int, r_col: int) -> str:
    """Excel 결과(등급) — R 셀 참조, format_result와 동일 구간."""
    from openpyxl.utils import get_column_letter

    r = f"{get_column_letter(r_col)}{row}"
    return (
        f'=IF({r}>=16,"허용불가위험 (A급)",'
        f'IF({r}=15,"중대위험 (B급)",'
        f'IF(AND({r}>=9,{r}<=12),"상당위험 (C급)",'
        f'IF({r}=8,"경미위험 (D급)",'
        f'IF(AND({r}>=4,{r}<=7),"미미위험 (E급)",'
        f'"무시위험 (F급)")))))'
    )


def _primary_disaster(injury: str) -> str:
    for sep in (",", "，", "·", "/"):
        if sep in injury:
            return injury.split(sep)[0].strip()
    return injury.strip() or "기타"


def _resolve_sequence(work_class: str, phase: str, unit_task: str) -> str:
    if phase in PHASE_SEQUENCE:
        return PHASE_SEQUENCE[phase]

    task = unit_task.lower()
    if work_class == "돌발대응":
        return "비상대응"
    if work_class == "정비작업":
        if any(k in unit_task for k in ("점검", "측정", "검사")):
            return "점검"
        return "보수"
    if any(k in unit_task for k in ("측정", "열화상", "절연")):
        return "측정"
    if any(k in unit_task for k in ("점검", "순회", "확인")):
        return "점검"
    if any(k in unit_task for k in ("운전", "운행", "가동")):
        return "작업"
    return "작업중"


def _resolve_process(work_class: str, phase: str, unit_task: str, job_name: str) -> str:
    if phase == "작업 전":
        if any(k in unit_task for k in ("이동", "준비", "TBM", "교육")):
            return "현장이동·작업준비"
        if any(k in unit_task for k in ("점검", "시동", "LOTO", "에너지")):
            return "작업 전 점검·LOTO"
        if any(k in unit_task for k in ("설치", "사다리", "비계")):
            return "사다리·비계 설치"
        return unit_task

    if phase == "작업 후":
        if any(k in unit_task for k in ("철거", "해체")):
            return "사다리·비계 철거"
        if any(k in unit_task for k in ("정리", "복구", "5S")):
            return "작업장 정리·복구"
        return unit_task

    if unit_task and unit_task != f"{job_name} 본작업":
        return unit_task
    if work_class == "정비작업":
        return f"{job_name} 보수·정비"
    if work_class == "돌발대응":
        return f"{job_name} 비상대응"
    return f"{job_name} 본작업"


def _measures_after(row: RiskRow) -> str:
    imp = polish_improvement(row.improvements).strip()
    if imp:
        return imp
    if row.score_after >= row.score_before and row.freq_after == row.freq_before:
        if row.sev_after == row.sev_before:
            return "좌동"
    return "좌동"


def risk_row_to_form(row: RiskRow, job_name: str = "") -> FormRow:
    src = (row.source or "").upper()
    if src.startswith("DOC"):
        seq = PHASE_SEQUENCE.get(row.phase, row.phase or "작업중")
        proc = (row.unit_task or "").strip() or row.phase or job_name
        imp = (row.improvements or "").strip()
        cur = (row.current or "").strip() or imp
        return FormRow(
            work_sequence=seq,
            work_process=proc,
            disaster_type=_primary_disaster(row.injury) if row.injury else "기타",
            hazard_factor=row.hazard,
            f_before=row.freq_before,
            s_before=row.sev_before,
            measures_before=cur,
            f_after=row.freq_after,
            s_after=row.sev_after,
            measures_after=imp or cur,
            law=row.law,
            law_url=row.law_url,
        )

    seq = _resolve_sequence(row.work_class, row.phase, row.unit_task)
    proc = _resolve_process(row.work_class, row.phase, row.unit_task, job_name)
    return FormRow(
        work_sequence=seq,
        work_process=proc,
        disaster_type=_primary_disaster(row.injury),
        hazard_factor=polish_hazard(row.hazard),
        f_before=row.freq_before,
        s_before=row.sev_before,
        measures_before=row.current,
        f_after=row.freq_after,
        s_after=row.sev_after,
        measures_after=_measures_after(row),
        law=row.law,
        law_url=row.law_url,
    )


def convert_to_form_rows(rows: list[RiskRow], job_name: str = "") -> list[FormRow]:
    indexed = [(i, risk_row_to_form(r, job_name)) for i, r in enumerate(rows)]
    indexed.sort(
        key=lambda x: (SEQUENCE_ORDER.get(x[1].work_sequence, 99), x[0]),
    )
    return [f for _, f in indexed]


def merge_sequence_cells(form_rows: list[FormRow]) -> list[tuple[int, int]]:
    """(start_index, end_index) per 작업순서 group for Excel merge."""
    if not form_rows:
        return []
    groups: list[tuple[int, int]] = []
    start = 0
    current = form_rows[0].work_sequence
    for i, row in enumerate(form_rows[1:], 1):
        if row.work_sequence != current:
            groups.append((start, i - 1))
            start = i
            current = row.work_sequence
    groups.append((start, len(form_rows) - 1))
    return groups
