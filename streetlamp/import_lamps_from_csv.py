# import_lamps_from_csv.py — 업로드/CSV/엑셀 → DB lamps
from __future__ import annotations

import asyncio
import csv
import io
import os
import tempfile
from pathlib import Path

from sqlalchemy import func, select, text

from database import AsyncSessionLocal, engine
import streetlamp.models  # noqa: F401
from streetlamp.models import Lamp

CSV_PATH = Path(__file__).resolve().parent.parent / "data" / "streetlamp" / "lamp_codes.csv"
XLSX_PATH = Path(__file__).resolve().parent.parent / "data" / "streetlamp" / "가로등 adress.xlsx"
UPLOAD_MAX_BYTES = 15 * 1024 * 1024

# 백그라운드 작업 상태 (프로세스 단위)
_import_lock = asyncio.Lock()
_import_status: dict[str, object] = {
    "running": False,
    "message": "",
    "added": 0,
    "updated": 0,
    "removed": 0,
    "total": 0,
    "source": "",
    "error": "",
}


def get_import_status() -> dict[str, object]:
    return dict(_import_status)


def _row(code: str, location: str | None = None, prefix: str | None = None) -> dict[str, str]:
    code = (code or "").strip()
    if not code:
        raise ValueError("empty code")
    pfx = (prefix or code.rsplit("-", 1)[0] if "-" in code else code).strip()
    loc = (location or code).strip() or code
    return {"code": code, "group_prefix": pfx, "location": loc}


def _dedupe_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in rows:
        code = row["code"]
        if code in seen:
            continue
        seen.add(code)
        out.append(row)
    return out


def parse_rows_from_csv_bytes(raw: bytes) -> list[dict[str, str]]:
    text_data = raw.decode("utf-8-sig", errors="replace")
    if not text_data.strip():
        return []
    sample = text_data[:4096]
    try:
        dialect = csv.Sniffer().has_header(sample)
    except csv.Error:
        dialect = False
    reader = csv.reader(io.StringIO(text_data))
    rows_iter = list(reader)
    if not rows_iter:
        return []
    header = [str(c or "").strip().lower() for c in rows_iter[0]]
    code_keys = {"code", "코드", "가로등코드", "lamp_code"}
    loc_keys = {"location", "address", "addr", "위치", "주소", "어드레스"}
    prefix_keys = {"group_prefix", "prefix", "구역", "그룹"}
    has_code_header = any(h in code_keys for h in header)
    if (dialect or has_code_header) and has_code_header:
        idx = {h: i for i, h in enumerate(header)}
        code_i = next((idx[k] for k in code_keys if k in idx), None)
        loc_i = next((idx[k] for k in loc_keys if k in idx), None)
        pfx_i = next((idx[k] for k in prefix_keys if k in idx), None)
        if code_i is None:
            raise ValueError("CSV 헤더에 code(코드) 열이 필요합니다.")
        out: list[dict[str, str]] = []
        for line in rows_iter[1:]:
            if code_i >= len(line):
                continue
            code = str(line[code_i] or "").strip()
            if not code:
                continue
            loc = str(line[loc_i] or "").strip() if loc_i is not None and loc_i < len(line) else ""
            pfx = str(line[pfx_i] or "").strip() if pfx_i is not None and pfx_i < len(line) else ""
            out.append(_row(code, loc or None, pfx or None))
        return _dedupe_rows(out)
    # 헤더 없음: 모든 셀 값을 코드로
    out = []
    for line in rows_iter:
        for cell in line:
            s = str(cell or "").strip()
            if s and s.lower() not in code_keys:
                out.append(_row(s))
    return _dedupe_rows(out)


def parse_rows_from_xlsx_bytes(raw: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb.active
    out: list[dict[str, str]] = []
    for row in ws.iter_rows(values_only=True):
        for value in row:
            if value is None:
                continue
            s = str(value).strip()
            if s:
                out.append(_row(s))
    wb.close()
    return _dedupe_rows(out)


def parse_rows_from_xls_bytes(raw: bytes) -> list[dict[str, str]]:
    import xlrd

    book = xlrd.open_workbook(file_contents=raw)
    sheet = book.sheet_by_index(0)
    out: list[dict[str, str]] = []
    for r in range(sheet.nrows):
        for c in range(sheet.ncols):
            val = sheet.cell_value(r, c)
            if val is None or val == "":
                continue
            s = str(val).strip()
            if s.endswith(".0") and s[:-2].isdigit():
                s = s[:-2]
            if s:
                out.append(_row(s))
    return _dedupe_rows(out)


def parse_upload_to_rows(content: bytes, filename: str) -> list[dict[str, str]]:
    if len(content) > UPLOAD_MAX_BYTES:
        raise ValueError(f"파일 크기는 {UPLOAD_MAX_BYTES // (1024 * 1024)}MB 이하여야 합니다.")
    name = (filename or "").lower()
    if name.endswith(".csv"):
        rows = parse_rows_from_csv_bytes(content)
    elif name.endswith(".xlsx"):
        rows = parse_rows_from_xlsx_bytes(content)
    elif name.endswith(".xls"):
        rows = parse_rows_from_xls_bytes(content)
    else:
        raise ValueError("xlsx, xls, csv 파일만 업로드할 수 있습니다.")
    if not rows:
        raise ValueError("파일에서 가로등 코드를 찾지 못했습니다.")
    return rows


def save_rows_as_csv(rows: list[dict[str, str]], path: Path | None = None) -> Path:
    out = path or CSV_PATH
    try:
        out.parent.mkdir(parents=True, exist_ok=True)
        _write_csv(out, rows)
        return out
    except OSError:
        tmp = Path(tempfile.gettempdir()) / "smart_fms_lamp_codes.csv"
        _write_csv(tmp, rows)
        return tmp


def _is_postgres() -> bool:
    url = (
        os.environ.get("DATABASE_INTERNAL_URL", "")
        or os.environ.get("DATABASE_URL", "")
        or ""
    ).lower()
    return "postgres" in url


async def _ensure_lamps_table() -> None:
    """가로등 테이블만 빠르게 보장 (전체 스키마 마이그레이션은 하지 않음)."""
    ddl = """
    CREATE TABLE IF NOT EXISTS lamps (
      id SERIAL PRIMARY KEY,
      code VARCHAR(64) UNIQUE,
      location VARCHAR(255) NOT NULL,
      description TEXT
    )
    """
    if not _is_postgres():
        ddl = """
        CREATE TABLE IF NOT EXISTS lamps (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          code VARCHAR(64) UNIQUE,
          location VARCHAR(255) NOT NULL,
          description TEXT
        )
        """
    async with engine.begin() as conn:
        await conn.execute(text(ddl))
        if _is_postgres():
            await conn.execute(text("CREATE INDEX IF NOT EXISTS ix_lamps_code ON lamps (code)"))


async def _sync_lamps_id_sequence(session) -> None:
    if not _is_postgres():
        return
    await session.execute(
        text(
            "SELECT setval("
            "pg_get_serial_sequence('lamps', 'id'), "
            "COALESCE((SELECT MAX(id) FROM lamps), 1))"
        )
    )


def load_rows_from_csv(path: Path | None = None) -> list[dict[str, str]]:
    csv_path = path or CSV_PATH
    if not csv_path.is_file():
        raise FileNotFoundError(f"{csv_path} 없음")
    rows: list[dict[str, str]] = []
    with csv_path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            code = (row.get("code") or "").strip()
            if not code:
                continue
            prefix = (row.get("group_prefix") or code.rsplit("-", 1)[0]).strip()
            location = (row.get("location") or code).strip() or code
            rows.append({"code": code, "group_prefix": prefix, "location": location})
    return rows


def rebuild_csv_from_xlsx(xlsx_path: Path | None = None) -> tuple[int, Path]:
    """엑셀 → CSV. Render 등 읽기전용 배포본은 /tmp 에 기록.

    반환: (건수, 사용한 CSV 경로)
    """
    from openpyxl import load_workbook

    path = xlsx_path or XLSX_PATH
    if not path.is_file():
        raise FileNotFoundError(f"엑셀 없음: {path}")

    # 배포본 CSV가 이미 있으면 재생성 생략 (버튼 응답 지연·디스크 쓰기 방지)
    if CSV_PATH.is_file() and os.environ.get("STREETLAMP_FORCE_XLSX_REBUILD", "").strip() not in (
        "1",
        "true",
        "yes",
    ):
        rows = load_rows_from_csv(CSV_PATH)
        return len(rows), CSV_PATH

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[dict[str, str]] = []
    seen: set[str] = set()
    for line in ws.iter_rows(values_only=True):
        for value in line:
            if value is None:
                continue
            code = str(value).strip()
            if not code or code in seen:
                continue
            seen.add(code)
            rows.append(_row(code))

    out = CSV_PATH
    try:
        out.parent.mkdir(parents=True, exist_ok=True)
        _write_csv(out, rows)
    except OSError:
        out = Path(tempfile.gettempdir()) / "smart_fms_lamp_codes.csv"
        _write_csv(out, rows)
    return len(rows), out


def _write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["code", "group_prefix", "location"])
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "code": row["code"],
                    "group_prefix": row["group_prefix"],
                    "location": row["location"],
                }
            )


async def import_lamps(
    *,
    replace_all: bool = False,
    csv_path: Path | None = None,
    rows: list[dict[str, str]] | None = None,
) -> int:
    """가로등 목록 등록/갱신. 반환: 신규 추가 건수."""
    if rows is None:
        path = csv_path or CSV_PATH
        if not path.is_file() and XLSX_PATH.is_file():
            _, path = rebuild_csv_from_xlsx()
        rows = load_rows_from_csv(path)

    await _ensure_lamps_table()

    added = 0
    updated = 0
    removed = 0
    async with AsyncSessionLocal() as session:
        await _sync_lamps_id_sequence(session)

        if replace_all:
            from streetlamp.models import MaintenanceRequest

            await session.execute(MaintenanceRequest.__table__.delete())
            await session.execute(Lamp.__table__.delete())
            await session.commit()

        existing_by_code: dict[str, Lamp] = {
            lamp.code: lamp
            for lamp in (
                await session.scalars(select(Lamp).where(Lamp.code.isnot(None)))
            ).all()
            if lamp.code
        }

        incoming_codes = {r["code"] for r in rows}

        if not replace_all:
            # 파일에 없는 코드는 삭제 (업로드 파일 = 최신 목록)
            stale = [lamp for code, lamp in existing_by_code.items() if code not in incoming_codes]
            if stale:
                from streetlamp.models import MaintenanceRequest

                stale_ids = [l.id for l in stale]
                if stale_ids:
                    await session.execute(
                        MaintenanceRequest.__table__.delete().where(
                            MaintenanceRequest.lamp_id.in_(stale_ids)
                        )
                    )
                for lamp in stale:
                    await session.delete(lamp)
                    removed += 1
                await session.commit()
                existing_by_code = {
                    code: lamp
                    for code, lamp in existing_by_code.items()
                    if code in incoming_codes
                }

        batch_new: list[Lamp] = []
        for row in rows:
            code = row["code"]
            location = row["location"]
            prefix = row["group_prefix"]
            description = f"구역 {prefix}" if prefix else None
            existing = existing_by_code.get(code)
            if existing:
                changed = False
                if existing.location != location:
                    existing.location = location
                    changed = True
                if (existing.description or "") != (description or ""):
                    existing.description = description
                    changed = True
                if changed:
                    updated += 1
                continue

            lamp = Lamp(code=code, location=location, description=description)
            batch_new.append(lamp)
            existing_by_code[code] = lamp
            added += 1
            if len(batch_new) >= 200:
                session.add_all(batch_new)
                await session.commit()
                batch_new.clear()

        if batch_new:
            session.add_all(batch_new)
        await _sync_lamps_id_sequence(session)
        await session.commit()

    mode = "전체 교체" if replace_all else "파일 기준 동기화"
    _import_status.update(
        {
            "added": added,
            "updated": updated,
            "removed": removed,
            "total": len(rows),
            "message": f"완료({mode}): 신규 {added} · 갱신 {updated} · 삭제 {removed} · 파일 {len(rows)}건",
        }
    )
    print(
        f"[lamp-import] rows={len(rows)} added={added} updated={updated} removed={removed}",
        flush=True,
    )
    return added


async def run_import_job(
    *,
    replace_all: bool = False,
    upload_bytes: bytes | None = None,
    upload_name: str = "",
    use_bundled: bool = False,
) -> None:
    """웹 버튼용 백그라운드 임포트."""
    if _import_lock.locked():
        return
    async with _import_lock:
        _import_status.update(
            {
                "running": True,
                "message": "가로등 어드레스 등록 중…",
                "added": 0,
                "updated": 0,
                "removed": 0,
                "total": 0,
                "source": upload_name or ("bundled" if use_bundled else ""),
                "error": "",
            }
        )
        try:
            rows: list[dict[str, str]] | None = None
            if upload_bytes:
                _import_status["message"] = f"파일 분석 중… ({upload_name})"
                rows = await asyncio.to_thread(parse_upload_to_rows, upload_bytes, upload_name)
                _import_status["message"] = f"{len(rows)}건 → DB 등록 중…"
                try:
                    save_rows_as_csv(rows)
                except Exception:
                    pass
            elif use_bundled:
                csv_path = CSV_PATH
                try:
                    _, csv_path = rebuild_csv_from_xlsx()
                except FileNotFoundError:
                    if not CSV_PATH.is_file():
                        raise
                rows = load_rows_from_csv(csv_path)
                _import_status["message"] = f"배포본 {len(rows)}건 → DB 등록 중…"
            else:
                raise ValueError("업로드 파일을 선택해 주세요.")

            await import_lamps(replace_all=replace_all, rows=rows)
        except Exception as exc:
            _import_status.update(
                {
                    "running": False,
                    "error": str(exc)[:300],
                    "message": f"실패: {exc}",
                }
            )
            print(f"[lamp-import] failed: {exc}", flush=True)
            return
        _import_status["running"] = False


async def import_lamps_if_needed() -> int:
    """CSV 대비 누락 코드가 있으면 전체 upsert (기동 시)."""
    if not CSV_PATH.is_file():
        if XLSX_PATH.is_file():
            n, _ = rebuild_csv_from_xlsx()
            print(f"[lamp-import] rebuilt CSV from xlsx: {n}", flush=True)
        else:
            print(f"[lamp-import] skip: no CSV at {CSV_PATH}", flush=True)
            return 0

    rows = load_rows_from_csv()
    if not rows:
        return 0
    expected = {r["code"] for r in rows}

    async with AsyncSessionLocal() as session:
        existing = set(
            await session.scalars(select(Lamp.code).where(Lamp.code.isnot(None)))
        )
        missing = expected - existing
        db_n = await session.scalar(
            select(func.count()).select_from(Lamp).where(Lamp.code.isnot(None))
        )

    force = os.environ.get("STREETLAMP_SYNC_LAMPS", "").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
    )
    if not force and not missing and (db_n or 0) >= len(expected):
        print(
            f"[lamp-import] skip: DB has {db_n} codes, csv={len(expected)}",
            flush=True,
        )
        return 0

    if missing:
        print(f"[lamp-import] missing {len(missing)} codes → sync", flush=True)
    return await import_lamps(replace_all=False)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--from-xlsx", action="store_true", help="엑셀에서 CSV 재생성")
    parser.add_argument("--replace", action="store_true", help="기존 lamps/의뢰 삭제 후 재등록")
    args = parser.parse_args()
    if args.from_xlsx:
        count, path = rebuild_csv_from_xlsx()
        print(f"엑셀→CSV {count}건: {path}")
    n = asyncio.run(import_lamps(replace_all=args.replace))
    print(f"완료. 신규 {n}건 (CSV: {CSV_PATH})")
