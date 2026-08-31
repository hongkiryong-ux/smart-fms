# -*- coding: utf-8 -*-
"""점검일지2 엑셀 양식 분석 — 추후 건물 추가 시 자동 스키마 생성 기반."""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def analyze_workbook(path: str | Path) -> dict[str, Any]:
    """엑셀 일지 파일을 분석해 시트·헤더·미터 후보를 반환."""
    path = Path(path)
    wb = load_workbook(path, data_only=False)
    result: dict[str, Any] = {
        "source_file": path.name,
        "sheets": [],
        "suggested_building_name": _guess_building_name(wb),
        "schema_skeleton": {"building_name": "", "electrical": {"blocks": []}, "facility": {}},
        "notes": [
            "자동 생성 스키마는 초안입니다. 배율·전일지침·월보 집계 규칙은 수동 보완이 필요할 수 있습니다.",
            "전기+설비 2시트는 electrical/facility로 분리해 등록하세요.",
        ],
    }
    for sn in wb.sheetnames:
        ws = wb[sn]
        result["sheets"].append(_analyze_sheet(ws, sn))
    bn = result["suggested_building_name"]
    if bn:
        result["schema_skeleton"]["building_name"] = bn
    return result


def _guess_building_name(wb) -> str:
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(1, 6):
            for c in range(1, 8):
                v = ws.cell(r, c).value
                if not v or not isinstance(v, str):
                    continue
                m = re.search(r"[●◆]?\s*(.+?)\s*(전기|운영|점검)", v)
                if m:
                    return m.group(1).strip()
    return ""


def _analyze_sheet(ws, name: str) -> dict[str, Any]:
    headers: list[dict[str, Any]] = []
    time_rows: list[int] = []
    for r in range(1, min(ws.max_row + 1, 60)):
        b = ws.cell(r, 2).value
        if isinstance(b, str) and re.match(r"^\d{1,2}:\d{2}", b.strip()):
            time_rows.append(r)
        for c in range(1, min(ws.max_column + 1, 30)):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and any(k in v for k in ("지침", "유효", "적산", "전압", "전류")):
                headers.append({"row": r, "col": get_column_letter(c), "text": v.strip()})
    kind = "electrical" if "전기" in name else "facility" if "설비" in name or "운영" in name else "unknown"
    return {
        "name": name,
        "kind": kind,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "time_rows": time_rows,
        "header_samples": headers[:40],
    }
