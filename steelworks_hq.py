# -*- coding: utf-8 -*-
"""제철소본부 점검일지 — 전기·설비 1일·월보·엑셀 아카이브."""
from __future__ import annotations

import io
import json
import re
from copy import deepcopy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models import Building, SteelworksHqArchive, SteelworksHqDaily

ROOT = Path(__file__).resolve().parent
SCHEMA_PATH = ROOT / "resources" / "steelworks_hq_schema.json"
TEMPLATE_XLSX = ROOT / "resources" / "steelworks_hq_template.xlsx"

_schema_cache: dict | None = None

_ELEC_BLOCK_ROWS = {
    "incoming": {"prev": 6, "start": 7},
    "turbo": {"prev": 14, "start": 15},
    "room": {"prev": 22, "start": 23},
    "hv": {"start": 31},
}
_ELEC_TIME_INDEX = {
    "06:00": 0,
    "10:00": 1,
    "14:00": 2,
    "18:00": 3,
    "22:00": 4,
}
_HV_TIME_INDEX = {"10:00": 0, "14:00": 1}
_FAC_UTILITY_ROWS = {
    "pwr1": 6,
    "pwr2": 7,
    "heat": 8,
    "flow": 9,
    "water1": 10,
    "water_main": 11,
}
_FAC_HEATING_ROWS = {"b1": 17, "b3": 18}
_FAC_FIRE_ROWS = {"b1": 20, "b3": 21}
_FAC_AHU_ROWS = {"09:50": 23, "13:30": 24}
_FAC_AHU_COLS = {
    "AHU-1": ("E", "F"),
    "AHU-2": ("G", "H"),
    "AHU-3": ("I", "J"),
    "AHU-4": ("K", "L"),
    "AHU-5": ("M", "N"),
    "AHU-6": ("O", "P"),
}


def public_base_url(request: Any | None = None) -> str:
    import os

    env = (os.environ.get("PUBLIC_BASE_URL") or "").strip().rstrip("/")
    if env:
        return env
    if request is not None:
        return str(request.base_url).rstrip("/")
    return "http://127.0.0.1:8000"


def sw_hq_daily_qr_url(building_code: str, request: Any | None = None) -> str:
    code = (building_code or "").strip()
    return f"{public_base_url(request)}/swhq/{code}/daily"


def qr_png_bytes(url: str) -> bytes:
    import qrcode

    img = qrcode.make(url)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


async def get_building_for_qr(session: AsyncSession, code: str) -> Building | None:
    from models import InspectionLogBuilding2

    building = (
        await session.execute(
            select(Building)
            .join(InspectionLogBuilding2, InspectionLogBuilding2.building_id == Building.id)
            .where(
                Building.code == (code or "").strip(),
                Building.is_active == True,  # noqa: E712
            )
        )
    ).scalar_one_or_none()
    if not building or not is_steelworks_hq_building(building):
        return None
    return building


def load_schema() -> dict:
    global _schema_cache
    if _schema_cache is None:
        _schema_cache = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))
    return _schema_cache


def is_steelworks_hq_building(building: Building | None) -> bool:
    if not building:
        return False
    name = (building.name or "").strip()
    return name == load_schema().get("building_name", "제철소본부")


def _electrical_blocks(schema: dict | None = None) -> list[dict]:
    schema = schema or load_schema()
    return (schema.get("electrical") or {}).get("blocks") or []


def _facility_schema(schema: dict | None = None) -> dict:
    schema = schema or load_schema()
    return schema.get("facility") or {}


def _block_meters(block: dict) -> list[dict]:
    if block.get("meters"):
        return list(block["meters"])
    meters: list[dict] = []
    for grp in block.get("groups") or []:
        meter = grp.get("meter")
        if meter:
            meters.append(meter)
    return meters


def _chiller_fields(unit_id: str, fac: dict | None = None) -> list[str]:
    fac = fac or _facility_schema()
    ch = fac.get("chiller") or {}
    if unit_id.startswith("c260") or unit_id == "c80":
        return list(ch.get("fields_260") or [])
    return list(ch.get("fields_300") or [])


def _empty_electrical_block(block: dict) -> dict[str, Any]:
    bid = block["id"]
    if bid == "hv":
        times = block.get("times") or []
        return {
            "times": {t: {} for t in times},
            "notes": {t: "" for t in times},
        }
    meters = _block_meters(block)
    return {
        "prev": {m["id"]: "" for m in meters},
        "prev_manual": {},
        "times": {t: {} for t in (block.get("times") or [])},
    }


def _empty_facility_payload(fac: dict | None = None) -> dict[str, Any]:
    fac = fac or _facility_schema()
    utility = {
        row["id"]: {
            "prev": "",
            "today": "",
            "daily": "",
            "monthly": "",
            "prev_manual": False,
        }
        for row in (fac.get("utility") or {}).get("rows") or []
    }
    heating = {
        row["id"]: {
            "supply_temp": "",
            "return_temp": "",
            "supply_pressure": "",
            "return_pressure": "",
            "location": "",
            "location_temp": "",
            "location_pressure": "",
        }
        for row in (fac.get("heating") or {}).get("rows") or []
    }
    heating["_tank"] = {"tank": "", "level": ""}
    fire = {
        row["id"]: {"time": "", "pressure": ""}
        for row in (fac.get("fire") or {}).get("rows") or []
    }
    ahu_times = (fac.get("ahu") or {}).get("times") or []
    ahu_units = (fac.get("ahu") or {}).get("units") or []
    ahu = {
        t: {u: {"supply": "", "return": ""} for u in ahu_units} for t in ahu_times
    }
    outdoor_times = (fac.get("outdoor") or {}).get("times") or []
    outdoor_groups = (fac.get("outdoor") or {}).get("groups") or []
    outdoor = {
        t: {g["id"]: {f: "" for f in g.get("fields") or []} for g in outdoor_groups}
        for t in outdoor_times
    }
    chiller: dict[str, Any] = {}
    for unit in (fac.get("chiller") or {}).get("units") or []:
        uid = unit["id"]
        times = unit.get("times")
        fields = _chiller_fields(uid, fac)
        if isinstance(times, int):
            chiller[uid] = {f"t{i}": {fld: "" for fld in fields} for i in range(1, times + 1)}
        else:
            chiller[uid] = {t: {fld: "" for fld in fields} for t in (times or [])}
    return {
        "utility": utility,
        "heating": heating,
        "fire": fire,
        "ahu": ahu,
        "outdoor": outdoor,
        "chiller": chiller,
        "notes": "",
    }


def empty_daily_payload() -> dict[str, Any]:
    schema = load_schema()
    electrical = {b["id"]: _empty_electrical_block(b) for b in _electrical_blocks(schema)}
    return {"electrical": electrical, "facility": _empty_facility_payload(_facility_schema(schema))}


def build_electrical_block_layout(block: dict) -> dict[str, Any]:
    """엑셀 전기 시트와 동일한 다단 헤더·열 순서."""
    from openpyxl.utils import column_index_from_string as col_idx

    bid = block.get("id", "")
    sheet_title = block.get("title") or "제철소본부 전기설비 점검일지"
    all_cols: list[dict[str, Any]] = []

    for grp in block.get("groups") or []:
        gname = (grp.get("name") or "").strip()
        for c in grp.get("columns") or []:
            mult = c.get("multiplier")
            is_cum = bool(c.get("is_cumulative"))
            all_cols.append(
                {
                    "sort": col_idx(c["col"]),
                    "col": c["col"],
                    "kind": "cumulative" if is_cum else "input",
                    "group": gname,
                    "meter_id": (grp.get("meter") or {}).get("id") if is_cum else None,
                    "metric": c.get("metric") or "",
                    "unit": c.get("unit") or "",
                    "range": c.get("range") or "",
                    "multiplier": mult,
                }
            )
    for tr in block.get("tr_temps") or []:
        all_cols.append(
            {
                "sort": col_idx(tr["col"]),
                "col": tr["col"],
                "kind": "tr_temp",
                "group": "TR 온도 측정(℃)",
                "metric": tr.get("name") or "",
                "unit": "℃",
                "range": tr.get("range") or "",
            }
        )
    notes_col = block.get("notes_col")
    if notes_col:
        all_cols.append(
            {
                "sort": col_idx(notes_col),
                "col": notes_col,
                "kind": "notes",
                "group": "특이사항",
                "metric": "특이사항",
                "unit": "",
                "range": "",
            }
        )
    all_cols.sort(key=lambda x: x["sort"])

    sections: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for col in all_cols:
        g = col.get("group") or ""
        if current is None or current["name"] != g:
            if current:
                sections.append(current)
            current = {"name": g, "columns": []}
        current["columns"].append(col)
    if current:
        sections.append(current)

    meters = _block_meters(block)
    return {
        "id": bid,
        "title": sheet_title,
        "sections": sections,
        "times": block.get("times") or [],
        "meters": meters,
    }


def build_electrical_layouts(schema: dict | None = None) -> dict[str, dict]:
    schema = schema or load_schema()
    return {b["id"]: build_electrical_block_layout(b) for b in _electrical_blocks(schema)}


def _parse_num(raw: Any) -> float | None:
    if raw is None or raw == "":
        return None
    try:
        return float(str(raw).replace(",", "").strip())
    except ValueError:
        return None


def parse_time_range_hours(raw: str) -> float | None:
    """08:00~20:00 → 12.0, 단일 숫자도 허용."""
    s = (raw or "").strip()
    if not s:
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})\s*[~\-–—]\s*(\d{1,2}):(\d{2})$", s)
    if m:
        h1, m1, h2, m2 = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
        start = h1 + m1 / 60
        end = h2 + m2 / 60
        if end < start:
            end += 24
        return round(end - start, 1)
    n = _parse_num(s)
    return round(n, 1) if n is not None else None


def _fmt_num(n: float | None) -> str:
    if n is None:
        return ""
    if abs(n - round(n)) < 0.05:
        return str(int(round(n)))
    return str(round(n, 1))


def _calc_meter(prev: str, today: str, prev_monthly: str = "") -> tuple[str, str]:
    p, t = _parse_num(prev), _parse_num(today)
    daily = ""
    if p is not None and t is not None:
        daily = _fmt_num(t - p)
    pm = _parse_num(prev_monthly)
    monthly = ""
    d = _parse_num(daily)
    if pm is not None and d is not None:
        monthly = _fmt_num(pm + d)
    elif d is not None:
        monthly = daily
    return daily, monthly


def _incoming_meter_today(data: dict, meter_id: str) -> str:
    """수전 블록 22:00 적산 → 설비 utility 금일지침."""
    incoming = (data.get("electrical") or {}).get("incoming") or {}
    times = incoming.get("times") or {}
    t2200 = times.get("22:00") or {}
    for block in _electrical_blocks():
        if block["id"] != "incoming":
            continue
        for grp in block.get("groups") or []:
            meter = grp.get("meter") or {}
            if meter.get("id") != meter_id:
                continue
            return str(t2200.get(meter.get("reading_col"), "") or "")
    return ""


def recompute_daily(data: dict) -> dict:
    data = deepcopy(data or empty_daily_payload())
    fac = data.setdefault("facility", _empty_facility_payload())
    utility = fac.setdefault("utility", {})

    for uid, row in utility.items():
        prev_m = ""
        if uid in ("pwr1", "pwr2"):
            mid = "m_J" if uid == "pwr1" else "m_S"
            auto = _incoming_meter_today(data, mid)
            if auto and not row.get("today"):
                row["today"] = auto
        daily, monthly = _calc_meter(row.get("prev", ""), row.get("today", ""), prev_m)
        row["daily"] = daily
        row["monthly"] = monthly

    return data


def prev_electrical_from_daily_data(source_data: dict) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for block in _electrical_blocks():
        bid = block["id"]
        if bid == "hv":
            continue
        block_data = ((source_data or {}).get("electrical") or {}).get(bid) or {}
        times = block_data.get("times") or {}
        t2200 = times.get("22:00") or {}
        prev: dict[str, Any] = {}
        for m in _block_meters(block):
            prev[m["id"]] = t2200.get(m["reading_col"], "")
        out[bid] = prev
    return out


def _block_by_id(bid: str) -> dict:
    for b in _electrical_blocks():
        if b["id"] == bid:
            return b
    return {"id": bid, "times": []}


def apply_electrical_prev(data: dict, prev_vals: dict[str, dict[str, Any]]) -> bool:
    if not prev_vals:
        return False
    changed = False
    elec = data.setdefault("electrical", {})
    for bid, meters in prev_vals.items():
        block = elec.setdefault(bid, _empty_electrical_block(_block_by_id(bid)))
        block.setdefault("prev", {})
        block.setdefault("prev_manual", {})
        manual = block["prev_manual"]
        for mid, val in meters.items():
            if manual.get(mid):
                continue
            if block["prev"].get(mid) != val:
                block["prev"][mid] = val
                changed = True
    return changed


async def get_daily_row(
    session: AsyncSession, building_id: int, log_date: date
) -> SteelworksHqDaily | None:
    return (
        await session.execute(
            select(SteelworksHqDaily).where(
                SteelworksHqDaily.building_id == building_id,
                SteelworksHqDaily.log_date == log_date,
            )
        )
    ).scalar_one_or_none()


async def _fetch_day_data(session: AsyncSession, building_id: int, log_date: date) -> dict:
    row = await get_daily_row(session, building_id, log_date)
    return (row.data or {}) if row else {}


async def sync_prev_values(
    session: AsyncSession, building_id: int, log_date: date, data: dict
) -> tuple[dict, bool]:
    """전일 22:00 전기 지침 + 설비 utility prev/monthly 동기화."""
    merged = merge_daily_save(empty_daily_payload(), data)
    changed = False

    prev_date = log_date - timedelta(days=1)
    prev_row = await get_daily_row(session, building_id, prev_date)
    if prev_row and prev_row.data:
        prev_elec = prev_electrical_from_daily_data(prev_row.data)
        if apply_electrical_prev(merged, prev_elec):
            changed = True

        prev_data = recompute_daily(prev_row.data)
        prev_fac = (prev_data.get("facility") or {}).get("utility") or {}
        fac_util = merged.setdefault("facility", {}).setdefault("utility", {})
        for uid in fac_util:
            row = fac_util.setdefault(uid, {})
            prev_u = prev_fac.get(uid) or {}
            if not row.get("prev_manual"):
                pt = prev_u.get("today", "")
                if row.get("prev") != pt:
                    row["prev"] = pt
                    changed = True
            daily, monthly = _calc_meter(row.get("prev", ""), row.get("today", ""), "")
            if row.get("daily") != daily:
                row["daily"] = daily
                changed = True
            pm = prev_u.get("monthly", "")
            _, monthly2 = _calc_meter(row.get("prev", ""), row.get("today", ""), pm)
            if row.get("monthly") != monthly2:
                row["monthly"] = monthly2
                changed = True

    recomputed = recompute_daily(merged)
    if recomputed != merged:
        changed = True
    return recomputed, changed


def merge_daily_save(existing: dict, posted: dict) -> dict:
    base = deepcopy(existing) if existing else empty_daily_payload()
    posted = posted or {}

    elec_post = posted.get("electrical") or {}
    elec = base.setdefault("electrical", {})
    for bid, block_post in elec_post.items():
        block_schema = next((b for b in _electrical_blocks() if b["id"] == bid), None)
        if not block_schema:
            continue
        block = elec.setdefault(bid, _empty_electrical_block(block_schema))
        if bid == "hv":
            for t, cells in (block_post.get("times") or {}).items():
                block["times"].setdefault(t, {})
                block["times"][t].update(cells or {})
            notes = block_post.get("notes") or {}
            block.setdefault("notes", {})
            block["notes"].update(notes)
            continue
        if "prev" in block_post:
            block.setdefault("prev", {})
            block["prev"].update(block_post.get("prev") or {})
        if "prev_manual" in block_post:
            block["prev_manual"] = {
                k: v for k, v in (block_post.get("prev_manual") or {}).items() if v
            }
        for t, cells in (block_post.get("times") or {}).items():
            block["times"].setdefault(t, {})
            block["times"][t].update(cells or {})

    fac_post = posted.get("facility") or {}
    if fac_post:
        fac = base.setdefault("facility", _empty_facility_payload())
        util_post = fac_post.get("utility") or {}
        for uid, cells in util_post.items():
            row = fac.setdefault("utility", {}).setdefault(uid, {})
            for k, v in (cells or {}).items():
                if k == "prev_manual":
                    row[k] = bool(v) if isinstance(v, bool) else v == "1"
                else:
                    row[k] = v
        for sec in ("heating", "fire", "ahu", "outdoor", "chiller"):
            sec_post = fac_post.get(sec)
            if not sec_post:
                continue
            fac.setdefault(sec, {})
            if isinstance(sec_post, dict):
                _deep_merge(fac[sec], sec_post)
        if "notes" in fac_post:
            fac["notes"] = fac_post["notes"]

    return recompute_daily(base)


def _deep_merge(base: dict, patch: dict) -> None:
    for k, v in patch.items():
        if isinstance(v, dict) and isinstance(base.get(k), dict):
            _deep_merge(base[k], v)
        else:
            base[k] = v


async def get_or_create_daily(
    session: AsyncSession, building_id: int, log_date: date
) -> SteelworksHqDaily:
    row = await get_daily_row(session, building_id, log_date)
    if row:
        synced, changed = await sync_prev_values(session, building_id, log_date, row.data or {})
        if changed:
            row.data = synced
            await session.flush()
        return row
    synced, _ = await sync_prev_values(session, building_id, log_date, empty_daily_payload())
    row = SteelworksHqDaily(building_id=building_id, log_date=log_date, data=synced)
    session.add(row)
    await session.flush()
    return row


async def propagate_to_next_day(
    session: AsyncSession, building_id: int, log_date: date, saved_data: dict
) -> None:
    next_date = log_date + timedelta(days=1)
    next_row = await get_daily_row(session, building_id, next_date)
    if not next_row:
        return
    prev_elec = prev_electrical_from_daily_data(saved_data)
    next_data = merge_daily_save(empty_daily_payload(), next_row.data or {})
    apply_electrical_prev(next_data, prev_elec)

    prev_data = recompute_daily(saved_data)
    prev_fac = (prev_data.get("facility") or {}).get("utility") or {}
    next_fac = next_data.setdefault("facility", {}).setdefault("utility", {})
    for uid, prev_u in prev_fac.items():
        row = next_fac.setdefault(uid, {})
        if not row.get("prev_manual"):
            row["prev"] = prev_u.get("today", "")

    synced, _ = await sync_prev_values(session, building_id, next_date, next_data)
    next_row.data = synced
    next_row.updated_at = datetime.utcnow()


def parse_daily_form(form) -> dict:
    data = empty_daily_payload()
    items = form.multi_items() if hasattr(form, "multi_items") else form.items()
    for key, val in items:
        if not isinstance(key, str):
            continue
        raw = (val or "").strip() if isinstance(val, str) else str(val or "")
        if key.startswith("eprev__"):
            parts = key.split("__")
            if len(parts) < 3:
                continue
            bid, rest = parts[1], parts[2:]
            if rest[-1] == "manual" and len(rest) >= 2:
                mid = rest[0]
                data["electrical"].setdefault(bid, {}).setdefault("prev_manual", {})[mid] = raw == "1"
            elif len(rest) == 1:
                data["electrical"].setdefault(bid, {}).setdefault("prev", {})[rest[0]] = raw
        elif key.startswith("e__"):
            parts = key.split("__")
            if len(parts) < 4:
                continue
            bid, t, col = parts[1], parts[2], parts[3]
            if bid == "hv" and col == "notes":
                data["electrical"]["hv"].setdefault("notes", {})[t] = raw
            else:
                data["electrical"].setdefault(bid, {}).setdefault("times", {}).setdefault(t, {})[col] = raw
        elif key.startswith("f__"):
            parts = key.split("__")
            if len(parts) < 3:
                continue
            sec = parts[1]
            if sec == "utility" and len(parts) >= 4:
                uid, field = parts[2], parts[3]
                row = data["facility"]["utility"].setdefault(uid, {})
                if field == "prev_manual":
                    row[field] = raw == "1"
                else:
                    row[field] = raw
            elif sec == "heating" and len(parts) >= 4:
                rid, field = parts[2], parts[3]
                data["facility"]["heating"].setdefault(rid, {})[field] = raw
            elif sec == "fire" and len(parts) >= 4:
                rid, field = parts[2], parts[3]
                data["facility"]["fire"].setdefault(rid, {})[field] = raw
            elif sec == "ahu" and len(parts) >= 5:
                t, unit, field = parts[2], parts[3], parts[4]
                data["facility"]["ahu"].setdefault(t, {}).setdefault(unit, {})[field] = raw
            elif sec == "outdoor" and len(parts) >= 5:
                t, gid = parts[2], parts[3]
                if len(parts) == 5:
                    data["facility"]["outdoor"].setdefault(t, {}).setdefault(gid, {})[parts[4]] = raw
                elif len(parts) >= 6:
                    fld = parts[5]
                    data["facility"]["outdoor"].setdefault(t, {}).setdefault(gid, {})[fld] = raw
            elif sec == "chiller" and len(parts) >= 5:
                uid, slot, field = parts[2], parts[3], parts[4]
                data["facility"]["chiller"].setdefault(uid, {}).setdefault(slot, {})[field] = raw
            elif sec == "notes":
                data["facility"]["notes"] = raw
    return data


def _meter_reading_at_2200(block_data: dict, meter: dict) -> Any:
    times = block_data.get("times") or {}
    return (times.get("22:00") or {}).get(meter["reading_col"], "")


def _meter_max_current(block_data: dict, meter: dict, block: dict) -> float | None:
    rc = meter.get("reading_col")
    current_cols: list[str] = []
    for grp in block.get("groups") or []:
        gm = grp.get("meter") or {}
        if gm.get("reading_col") == rc:
            for c in grp.get("columns") or []:
                if c.get("metric") == "전류":
                    current_cols.append(c["col"])
    amps: list[float] = []
    for cells in (block_data.get("times") or {}).values():
        for cc in current_cols:
            av = _parse_num((cells or {}).get(cc))
            if av is not None:
                amps.append(av)
    return max(amps) if amps else None


def _breaker_row_from_daily(
    block_data: dict,
    meter: dict,
    block: dict,
    prev_reading: float | None,
) -> tuple[dict[str, Any], float | None]:
    reading_raw = _meter_reading_at_2200(block_data, meter)
    read_v = _parse_num(reading_raw)
    mult = meter.get("multiplier") or 1
    usage: Any = ""
    if prev_reading is not None and read_v is not None:
        usage = round((read_v - prev_reading) * mult, 2)
    max_a_val = _meter_max_current(block_data, meter, block)
    max_a: Any = max_a_val if max_a_val is not None else ""
    next_prev = read_v if read_v is not None else prev_reading
    return (
        {
            "meter_id": meter.get("id"),
            "name": meter["name"],
            "multiplier": mult,
            "reading": reading_raw if reading_raw not in (None, "") else "",
            "usage": usage,
            "max_a": max_a,
        },
        next_prev,
    )


def build_monthly_electrical_sections(schema: dict | None = None) -> list[dict[str, Any]]:
    schema = schema or load_schema()
    sections: list[dict[str, Any]] = []
    for block in _electrical_blocks(schema):
        bid = block["id"]
        if bid == "hv":
            continue
        meters = _block_meters(block)
        if not meters:
            continue
        breakers = [
            {
                "meter_id": m["id"],
                "name": m["name"],
                "multiplier": m.get("multiplier") or 1,
                "reading_col": m["reading_col"],
            }
            for m in meters
        ]
        sections.append(
            {
                "id": bid,
                "title": f"제철소본부 월보[{block.get('title', bid)}]",
                "breakers": breakers,
                "block": block,
            }
        )
    return sections


def _max_numeric(values: list[Any]) -> Any:
    nums = [_parse_num(v) for v in values]
    nums = [n for n in nums if n is not None]
    return round(max(nums), 1) if nums else ""


def compute_tr_temp_monthly(
    year: int,
    month: int,
    daily_rows: list[SteelworksHqDaily],
) -> dict[str, Any]:
    hv_block = next((b for b in _electrical_blocks() if b["id"] == "hv"), {})
    tr_fields = hv_block.get("tr_temps") or []
    hv_groups = hv_block.get("groups") or []
    tr_cols = {f["col"] for f in tr_fields}
    for grp in hv_groups:
        for c in grp.get("columns") or []:
            if c.get("metric") == "TR온도":
                tr_cols.add(c["col"])

    by_date = {r.log_date: r.data or {} for r in daily_rows}
    days = []
    peaks: dict[str, float | None] = {f["col"]: None for f in tr_fields}
    times = hv_block.get("times") or ["10:00", "14:00"]

    for day in range(1, 32):
        try:
            d = date(year, month, day)
        except ValueError:
            break
        hv_data = ((by_date.get(d) or {}).get("electrical") or {}).get("hv") or {}
        hv_times = hv_data.get("times") or {}
        values: dict[str, Any] = {}
        for col in tr_cols:
            vals = [(hv_times.get(t) or {}).get(col, "") for t in times]
            values[col] = _max_numeric(vals)
        days.append({"day": day, "date": d.isoformat(), "values": values})
        for field in tr_fields:
            col = field["col"]
            val = _parse_num(values.get(col))
            if val is None:
                continue
            cur = peaks[col]
            peaks[col] = val if cur is None else max(cur, val)

    totals = {col: (round(peaks[col], 1) if peaks[col] is not None else "") for col in peaks}
    return {
        "title": "TR온도",
        "fields": tr_fields,
        "days": days,
        "totals": totals,
    }


def compute_monthly_report(
    building_id: int,
    year: int,
    month: int,
    daily_rows: list[SteelworksHqDaily],
    prev_month_last_row: SteelworksHqDaily | None = None,
) -> dict[str, Any]:
    del building_id
    by_date = {r.log_date: r.data or {} for r in daily_rows}
    sections_out = []

    for sec in build_monthly_electrical_sections():
        sec_id = sec["id"]
        block = sec["block"]
        breakers = sec["breakers"]
        prev_day_cells: list[dict[str, Any]] = []
        prev_readings: dict[str, float | None] = {}
        prev_month_data = (
            ((prev_month_last_row.data or {}).get("electrical") or {}).get(sec_id, {})
            if prev_month_last_row
            else {}
        )
        for br in breakers:
            mid = br["meter_id"]
            meter = next(m for m in _block_meters(block) if m["id"] == mid)
            val = _parse_num(_meter_reading_at_2200(prev_month_data, meter))
            if val is None:
                first_of_month = date(year, month, 1)
                first_data = ((by_date.get(first_of_month) or {}).get("electrical") or {}).get(sec_id, {})
                prev_map = first_data.get("prev") or {}
                val = _parse_num(prev_map.get(mid))
            prev_readings[mid] = val
            prev_day_cells.append(
                {
                    "meter_id": mid,
                    "name": br["name"],
                    "multiplier": br["multiplier"],
                    "reading": val if val is not None else "",
                    "usage": "",
                    "max_a": "",
                }
            )

        days = []
        usage_sums: dict[str, float] = {b["meter_id"]: 0.0 for b in breakers}
        max_a_peaks: dict[str, float | None] = {b["meter_id"]: None for b in breakers}
        rolling_prev = dict(prev_readings)

        for day in range(1, 32):
            try:
                d = date(year, month, day)
            except ValueError:
                break
            block_data = ((by_date.get(d) or {}).get("electrical") or {}).get(sec_id) or {}
            prev_map = block_data.get("prev") or {} if block_data else {}
            breaker_rows = []
            for br in breakers:
                mid = br["meter_id"]
                meter = next(m for m in _block_meters(block) if m["id"] == mid)
                prev_v = rolling_prev.get(mid)
                if block_data and mid in prev_map and prev_map.get(mid) not in (None, ""):
                    pv = _parse_num(prev_map.get(mid))
                    if pv is not None:
                        prev_v = pv
                row, next_prev = _breaker_row_from_daily(block_data, meter, block, prev_v)
                breaker_rows.append(row)
                if row["usage"] != "":
                    usage_sums[mid] += float(row["usage"])
                if row["max_a"] != "":
                    mv = float(row["max_a"])
                    cur = max_a_peaks[mid]
                    max_a_peaks[mid] = mv if cur is None else max(cur, mv)
                if next_prev is not None:
                    rolling_prev[mid] = next_prev
            days.append({"day": day, "date": d.isoformat(), "breakers": breaker_rows})

        totals = []
        for br in breakers:
            mid = br["meter_id"]
            u_sum = usage_sums[mid]
            peak = max_a_peaks[mid]
            totals.append(
                {
                    "meter_id": mid,
                    "usage_sum": round(u_sum, 2) if u_sum else "",
                    "max_a_max": peak if peak is not None else "",
                }
            )
        sections_out.append(
            {
                "id": sec_id,
                "title": sec["title"],
                "breakers": breakers,
                "prev_day": prev_day_cells,
                "days": days,
                "totals": totals,
            }
        )

    import calendar

    last = calendar.monthrange(year, month)[1]
    fac_days = []
    for d in range(1, last + 1):
        ld = date(year, month, d)
        dd = recompute_daily(by_date.get(ld, {}))
        util = (dd.get("facility") or {}).get("utility") or {}
        fac_days.append(
            {
                "day": d,
                "date": ld.isoformat(),
                "utility": {
                    uid: {
                        "daily": (util.get(uid) or {}).get("daily", ""),
                        "monthly": (util.get(uid) or {}).get("monthly", ""),
                    }
                    for uid in util
                },
            }
        )

    return {
        "year": year,
        "month": month,
        "electrical_sections": sections_out,
        "facility_utility": {"days": fac_days},
        "tr_temps": compute_tr_temp_monthly(year, month, daily_rows),
    }


async def fetch_monthly_report_data(
    session: AsyncSession,
    building_id: int,
    year: int,
    month: int,
) -> dict[str, Any]:
    import calendar

    last_day = calendar.monthrange(year, month)[1]
    d_from = date(year, month, 1)
    d_to = date(year, month, last_day)
    daily_rows = (
        await session.execute(
            select(SteelworksHqDaily).where(
                SteelworksHqDaily.building_id == building_id,
                SteelworksHqDaily.log_date >= d_from,
                SteelworksHqDaily.log_date <= d_to,
            )
        )
    ).scalars().all()
    prev_month_last = d_from - timedelta(days=1)
    prev_month_row = (
        await session.execute(
            select(SteelworksHqDaily).where(
                SteelworksHqDaily.building_id == building_id,
                SteelworksHqDaily.log_date == prev_month_last,
            )
        )
    ).scalar_one_or_none()
    return compute_monthly_report(
        building_id, year, month, list(daily_rows), prev_month_row
    )


async def fetch_yearly_report_data(
    session: AsyncSession,
    building_id: int,
    year: int,
) -> dict[str, Any]:
    d_from = date(year, 1, 1)
    d_to = date(year, 12, 31)
    daily_rows = (
        await session.execute(
            select(SteelworksHqDaily).where(
                SteelworksHqDaily.building_id == building_id,
                SteelworksHqDaily.log_date >= d_from,
                SteelworksHqDaily.log_date <= d_to,
            )
        )
    ).scalars().all()
    rows_list = list(daily_rows)
    months = []
    for m in range(1, 13):
        month_rows = [r for r in rows_list if r.log_date.year == year and r.log_date.month == m]
        prev_month_last = date(year, m, 1) - timedelta(days=1)
        prev_row = (
            await session.execute(
                select(SteelworksHqDaily).where(
                    SteelworksHqDaily.building_id == building_id,
                    SteelworksHqDaily.log_date == prev_month_last,
                )
            )
        ).scalar_one_or_none()
        rep = compute_monthly_report(building_id, year, m, month_rows, prev_row)
        elec_usage: dict[str, float] = {}
        for sec in rep.get("electrical_sections") or []:
            for tot in sec.get("totals") or []:
                u = tot.get("usage_sum")
                if u != "":
                    elec_usage[tot["meter_id"]] = elec_usage.get(tot["meter_id"], 0) + float(u)
        fac_usage: dict[str, float] = {}
        for day in (rep.get("facility_utility") or {}).get("days") or []:
            for uid, vals in (day.get("utility") or {}).items():
                dv = _parse_num(vals.get("daily"))
                if dv is not None:
                    fac_usage[uid] = fac_usage.get(uid, 0) + dv
        months.append(
            {
                "month": m,
                "electrical_usage": {k: round(v, 2) for k, v in elec_usage.items()},
                "facility_usage": {k: _fmt_num(v) for k, v in fac_usage.items()},
                "tr_temp_peaks": (rep.get("tr_temps") or {}).get("totals") or {},
            }
        )
    return {"year": year, "months": months}


async def fetch_notes_list(
    session: AsyncSession,
    building_id: int,
    year: int,
    month: int,
) -> dict[str, Any]:
    import calendar

    last_day = calendar.monthrange(year, month)[1]
    d_from = date(year, month, 1)
    d_to = date(year, month, last_day)
    rows = (
        await session.execute(
            select(SteelworksHqDaily).where(
                SteelworksHqDaily.building_id == building_id,
                SteelworksHqDaily.log_date >= d_from,
                SteelworksHqDaily.log_date <= d_to,
            )
            .order_by(SteelworksHqDaily.log_date.asc())
        )
    ).scalars().all()

    hv_block = next((b for b in _electrical_blocks() if b["id"] == "hv"), {})
    hv_times = hv_block.get("times") or ["10:00", "14:00"]
    entries: list[dict[str, Any]] = []

    for row in rows:
        data = row.data or {}
        items: list[dict[str, Any]] = []
        hv_notes = ((data.get("electrical") or {}).get("hv") or {}).get("notes") or {}
        for t in hv_times:
            text = (hv_notes.get(t) or "").strip()
            if text:
                items.append({"time": t, "text": text, "section": "전기 HV"})
        fac_note = ((data.get("facility") or {}).get("notes") or "").strip()
        if fac_note:
            items.append({"time": "", "text": fac_note, "section": "설비"})
        if items:
            entries.append(
                {
                    "date": row.log_date.isoformat(),
                    "day": row.log_date.day,
                    "items": items,
                }
            )
    return {"year": year, "month": month, "entries": entries}


def _safe_cell(ws, addr: str):
    """병합 셀일 때 좌상단 셀 반환."""
    from openpyxl.utils import coordinate_to_tuple

    cell = ws[addr]
    if type(cell).__name__ == "MergedCell":
        row, col = coordinate_to_tuple(addr)
        for merged in ws.merged_cells.ranges:
            if row >= merged.min_row and row <= merged.max_row and col >= merged.min_col and col <= merged.max_col:
                from openpyxl.utils import get_column_letter

                return ws[f"{get_column_letter(merged.min_col)}{merged.min_row}"]
    return cell


def _write_cell(ws, addr: str, value: Any) -> None:
    if value in (None, ""):
        return
    _safe_cell(ws, addr).value = value


def _sheet_by_title(wb, preferred: str):
    if preferred in wb.sheetnames:
        return wb[preferred]
    for name in wb.sheetnames:
        if name == preferred:
            return wb[name]
    if len(wb.sheetnames) >= 2:
        idx = 0 if preferred == "전기" else 1
        if idx < len(wb.sheetnames):
            return wb[wb.sheetnames[idx]]
    return wb.active


def _write_electrical_sheet(ws, data: dict, log_date: date) -> None:
    elec = data.get("electrical") or {}
    for block in _electrical_blocks():
        bid = block["id"]
        layout = _ELEC_BLOCK_ROWS.get(bid)
        if not layout:
            continue
        block_data = elec.get(bid) or {}
        if bid == "hv":
            start = layout["start"]
            times = block.get("times") or []
            hv_times = block_data.get("times") or {}
            notes = block_data.get("notes") or {}
            all_cols: set[str] = set()
            for grp in block.get("groups") or []:
                for c in grp.get("columns") or []:
                    all_cols.add(c["col"])
            for tr in block.get("tr_temps") or []:
                all_cols.add(tr["col"])
            notes_col = block.get("notes_col")
            for i, t in enumerate(times):
                row = start + i
                cells = hv_times.get(t) or {}
                for col in all_cols:
                    _write_cell(ws, f"{col}{row}", cells.get(col))
                if notes_col:
                    _write_cell(ws, f"{notes_col}{row}", notes.get(t))
            continue

        start = layout["start"]
        prev_row = layout["prev"]
        times_map = block_data.get("times") or {}
        prev_map = block_data.get("prev") or {}
        all_cols: set[str] = set()
        for grp in block.get("groups") or []:
            for c in grp.get("columns") or []:
                all_cols.add(c["col"])
        for m in _block_meters(block):
            _write_cell(ws, f"{m['reading_col']}{prev_row}", prev_map.get(m["id"]))
        time_index = _ELEC_TIME_INDEX if bid != "hv" else _HV_TIME_INDEX
        for t, idx in time_index.items():
            row = start + idx
            cells = times_map.get(t) or {}
            for col in all_cols:
                _write_cell(ws, f"{col}{row}", cells.get(col))

    _write_cell(ws, "A1", log_date.day)


def _write_facility_sheet(ws, data: dict) -> None:
    data = recompute_daily(data)
    fac = data.get("facility") or {}
    utility = fac.get("utility") or {}
    for uid, row_idx in _FAC_UTILITY_ROWS.items():
        row = utility.get(uid) or {}
        _write_cell(ws, f"F{row_idx}", row.get("prev"))
        _write_cell(ws, f"H{row_idx}", row.get("today"))
        _write_cell(ws, f"J{row_idx}", row.get("daily"))
        _write_cell(ws, f"L{row_idx}", row.get("monthly"))

    heating = fac.get("heating") or {}
    tank = heating.get("_tank") or {}
    _write_cell(ws, "E19", tank.get("tank"))
    _write_cell(ws, "H19", tank.get("level"))
    for rid, row_idx in _FAC_HEATING_ROWS.items():
        h = heating.get(rid) or {}
        _write_cell(ws, f"E{row_idx}", h.get("supply_temp"))
        _write_cell(ws, f"F{row_idx}", h.get("return_temp"))
        _write_cell(ws, f"H{row_idx}", h.get("supply_pressure"))
        _write_cell(ws, f"J{row_idx}", h.get("return_pressure"))
        _write_cell(ws, f"L{row_idx}", h.get("location"))
        _write_cell(ws, f"N{row_idx}", h.get("location_temp"))
        _write_cell(ws, f"O{row_idx}", h.get("location_pressure"))

    fire = fac.get("fire") or {}
    for rid, row_idx in _FAC_FIRE_ROWS.items():
        f = fire.get(rid) or {}
        _write_cell(ws, f"D{row_idx}", f.get("time"))
        _write_cell(ws, f"H{row_idx}", f.get("pressure"))

    ahu = fac.get("ahu") or {}
    for t, row_idx in _FAC_AHU_ROWS.items():
        slot = ahu.get(t) or {}
        for unit, (sc, rc) in _FAC_AHU_COLS.items():
            u = slot.get(unit) or {}
            _write_cell(ws, f"{sc}{row_idx}", u.get("supply"))
            _write_cell(ws, f"{rc}{row_idx}", u.get("return"))

    outdoor = fac.get("outdoor") or {}
    _OUTDOOR_CELLS = {
        ("09:50", "g1", "제철소장실"): "E26",
        ("09:50", "g1", "행정부소장실"): "H26",
        ("09:50", "g1", "안전환경부소장실"): "K26",
        ("09:50", "g1", "환경기획실장실"): "N26",
        ("13:30", "g1", "제철소장실"): "E27",
        ("13:30", "g1", "행정부소장실"): "H27",
        ("13:30", "g1", "안전환경부소장실"): "K27",
        ("13:30", "g1", "환경기획실장실"): "N27",
        ("09:50", "g2", "회장실"): "E29",
        ("09:50", "g2", "본부장실"): "H29",
        ("09:50", "g2", "대응접실"): "K29",
        ("09:50", "g2", "영상회의실"): "N29",
        ("13:30", "g2", "회장실"): "E30",
        ("13:30", "g2", "본부장실"): "H30",
        ("13:30", "g2", "대응접실"): "K30",
        ("13:30", "g2", "영상회의실"): "N30",
    }
    for (t, gid, fld), addr in _OUTDOOR_CELLS.items():
        val = ((outdoor.get(t) or {}).get(gid) or {}).get(fld)
        _write_cell(ws, addr, val)

    chiller = fac.get("chiller") or {}
    fields_260 = _chiller_fields("c260_1")
    _CHILLER_260_ROWS = {"c260_1": (38, 39), "c260_2": (40, 41)}
    col_letters = ["E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]
    for uid, (r1, r2) in _CHILLER_260_ROWS.items():
        unit = chiller.get(uid) or {}
        for slot, row in (("t1", r1), ("t2", r2)):
            cells = unit.get(slot) or {}
            for i, fld in enumerate(fields_260):
                if i < len(col_letters):
                    _write_cell(ws, f"{col_letters[i]}{row}", cells.get(fld))
    c80 = chiller.get("c80") or {}
    for t, row in (("09:50", 42), ("13:30", 43)):
        cells = c80.get(t) or {}
        for i, fld in enumerate(fields_260):
            if i < len(col_letters):
                _write_cell(ws, f"{col_letters[i]}{row}", cells.get(fld))
    fields_300 = _chiller_fields("c300")
    c300 = chiller.get("c300") or {}
    for t, row in (("09:00", 45), ("13:10", 46)):
        cells = c300.get(t) or {}
        for i, fld in enumerate(fields_300):
            if i < len(col_letters):
                _write_cell(ws, f"{col_letters[i]}{row}", cells.get(fld))

    notes = fac.get("notes")
    _write_cell(ws, "D48", notes)


def export_daily_to_excel(data: dict, log_date: date) -> bytes:
    if TEMPLATE_XLSX.is_file():
        wb = load_workbook(TEMPLATE_XLSX)
    else:
        from openpyxl import Workbook

        wb = Workbook()
        wb.active.title = "전기"
        wb.create_sheet("설비")

    ws_elec = _sheet_by_title(wb, "전기")
    ws_fac = _sheet_by_title(wb, "설비")
    _write_electrical_sheet(ws_elec, data, log_date)
    _write_facility_sheet(ws_fac, data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def archive_daily_excel(
    session: AsyncSession,
    building_id: int,
    log_date: date,
) -> SteelworksHqArchive | None:
    row = await get_daily_row(session, building_id, log_date)
    if not row:
        return None
    existing = (
        await session.execute(
            select(SteelworksHqArchive).where(
                SteelworksHqArchive.building_id == building_id,
                SteelworksHqArchive.log_date == log_date,
            )
        )
    ).scalar_one_or_none()
    if existing:
        return existing

    xbytes = export_daily_to_excel(row.data or {}, log_date)
    fname = f"제철소본부_1일_{log_date.isoformat()}.xlsx"
    arch = SteelworksHqArchive(
        building_id=building_id,
        log_date=log_date,
        original_name=fname,
        file_data=xbytes,
        file_size=len(xbytes),
    )
    session.add(arch)
    await session.flush()
    return arch


async def rollover_at_midnight(session: AsyncSession, building_id: int, closing_date: date) -> None:
    row = await get_daily_row(session, building_id, closing_date)
    if row and row.data:
        await propagate_to_next_day(session, building_id, closing_date, row.data)
    await archive_daily_excel(session, building_id, closing_date)
    await get_or_create_daily(session, building_id, closing_date + timedelta(days=1))
    await session.commit()


async def ensure_tables(engine) -> None:
    from sqlalchemy import text

    url = str(engine.url).lower()
    is_pg = "postgresql" in url or "postgres" in url
    async with engine.begin() as conn:
        if is_pg:
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS steelworks_hq_daily (
                        id SERIAL PRIMARY KEY,
                        building_id INTEGER NOT NULL REFERENCES buildings(id),
                        log_date DATE NOT NULL,
                        data JSONB NOT NULL DEFAULT '{}',
                        created_at TIMESTAMP WITHOUT TIME ZONE DEFAULT NOW(),
                        updated_at TIMESTAMP WITHOUT TIME ZONE DEFAULT NOW(),
                        UNIQUE (building_id, log_date)
                    )
                    """
                )
            )
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS steelworks_hq_archives (
                        id SERIAL PRIMARY KEY,
                        building_id INTEGER NOT NULL REFERENCES buildings(id),
                        log_date DATE NOT NULL,
                        original_name VARCHAR(300),
                        file_data BYTEA,
                        file_size INTEGER,
                        created_at TIMESTAMP WITHOUT TIME ZONE DEFAULT NOW(),
                        UNIQUE (building_id, log_date)
                    )
                    """
                )
            )
        else:
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS steelworks_hq_daily (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        building_id INTEGER NOT NULL,
                        log_date DATE NOT NULL,
                        data TEXT NOT NULL DEFAULT '{}',
                        created_at DATETIME,
                        updated_at DATETIME,
                        UNIQUE (building_id, log_date)
                    )
                    """
                )
            )
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS steelworks_hq_archives (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        building_id INTEGER NOT NULL,
                        log_date DATE NOT NULL,
                        original_name VARCHAR(300),
                        file_data BLOB,
                        file_size INTEGER,
                        created_at DATETIME,
                        UNIQUE (building_id, log_date)
                    )
                    """
                )
            )


def register_scheduler(scheduler, session_factory, kst) -> None:
    """매일 23:59 KST — 제철소본부 1일 마감."""

    async def _daily_job():
        async with session_factory() as session:
            try:
                from models import InspectionLogBuilding2

                rows = (
                    await session.execute(
                        select(InspectionLogBuilding2.building_id, Building)
                        .join(Building, Building.id == InspectionLogBuilding2.building_id)
                    )
                ).all()
                today = datetime.now(kst).date()
                for bid, bld in rows:
                    if not is_steelworks_hq_building(bld):
                        continue
                    try:
                        await rollover_at_midnight(session, bid, today)
                    except Exception as e:
                        await session.rollback()
                        print(f"[swhq] rollover building={bid}: {e}", flush=True)
            except Exception as e:
                print(f"[swhq] daily scheduler: {e}", flush=True)

    scheduler.add_job(
        _daily_job,
        trigger="cron",
        hour=23,
        minute=59,
        timezone=kst,
        id="steelworks_hq_rollover",
        replace_existing=True,
    )
