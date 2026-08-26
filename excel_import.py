# excel_import.py
"""설비현황 엑셀 파싱 · 등록 · 출력."""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import delete, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from models import (
    Building,
    Equipment,
    Floor,
    MaintenanceRecord,
    PMFrequency,
    PMInspection,
    PMResult,
    PMSchedule,
    Site,
    WorkOrder,
    WorkOrderStatus,
    Zone,
)

SKIP_SHEETS = {
    "총괄",
    "총괄표",
    "Sheet1",
    "TOTAL",
    "개요",
    "표지",
    "정비이력",
    "점검이력",
    "점검주기",
    "정비의뢰",
    "사양",
}
SITE_NAME = "광양운영그룹"
SITE_CODE = "GY-OP"

# 사용자 제공 건물 목록 (파일명 기준)
BUILDING_NAMES = [
    "러닝센타",
    "기술연구원",
    "기술교육센터",
    "금호빗물펌프장",
    "금당어린이집",
    "60서브",
    "57서브",
    "56서브",
    "55서브",
    "54서브",
    "53서브",
    "52서브",
    "51서브",
    "18서브",
    "16서브",
    "12서브",
    "8서브,백운그린랜드",
    "7서브",
    "6서브",
    "5서브",
    "3서브",
    "2서브",
    "휴먼센터",
    "축구전용구장",
    "중앙관제실",
    "주택변전소",
    "제철회관",
    "제철소본부",
    "임원숙소 1,2,3,5,금호어버이집",
    "어울림체육관",
    "복지센터",
    "백운플라자",
    "백운아트홀",
    "백운쇼핑센터",
    "백운생활관5,6동",
    "백운생활관3,4동",
    "백운생활관1,2동",
    "백운대",
]


def _safe_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _building_code(name: str) -> str:
    code = re.sub(r"[^\w가-힣]", "", name)[:20]
    return code or "BLD"


def _open_workbook(path: str | Path):
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        import openpyxl

        return ("openpyxl", openpyxl.load_workbook(path, read_only=True, data_only=True))
    return ("xlrd", xlrd.open_workbook(str(path)))


def _sheet_names(wb_kind: str, wb) -> list[str]:
    if wb_kind == "xlrd":
        return wb.sheet_names()
    return wb.sheetnames


def _get_sheet(wb_kind: str, wb, name: str):
    if wb_kind == "xlrd":
        return wb.sheet_by_name(name)
    return wb[name]


def _cell(sheet, wb_kind: str, r: int, c: int) -> str:
    if wb_kind == "xlrd":
        return _safe_str(sheet.cell_value(r, c))
    row = list(sheet.iter_rows(min_row=r + 1, max_row=r + 1, values_only=True))
    if not row:
        return ""
    vals = row[0]
    return _safe_str(vals[c] if c < len(vals) else "")


def _row_values(sheet, wb_kind: str, r: int, max_col: int) -> list[str]:
    return [_cell(sheet, wb_kind, r, c) for c in range(max_col)]


def _find_header_row(sheet, wb_kind: str, nrows: int, ncols: int) -> int | None:
    keywords = ("구분", "구 분", "명칭", "TYPE", "형식", "PUMP", "FAN")
    for r in range(min(10, nrows)):
        row = _row_values(sheet, wb_kind, r, ncols)
        joined = " ".join(row)
        if any(k in joined for k in keywords):
            return r
    return None


def _is_data_row(cells: list[str]) -> bool:
    text = "".join(cells).strip()
    if not text:
        return False
    if cells[0] in ("계", "합계", "소계"):
        return False
    if "합계" in text or "소계" in text:
        return False
    return True


def parse_excel_file(path: str | Path) -> dict[str, list[dict[str, Any]]]:
    """엑셀 파일 → {시트명: [행 dict]} (총괄 제외)."""
    wb_kind, wb = _open_workbook(path)
    result: dict[str, list[dict]] = {}

    try:
        for sheet_name in _sheet_names(wb_kind, wb):
            if sheet_name in SKIP_SHEETS:
                continue
            sheet = _get_sheet(wb_kind, wb, sheet_name)
            if wb_kind == "xlrd":
                nrows, ncols = sheet.nrows, sheet.ncols
            else:
                nrows = sheet.max_row or 0
                ncols = sheet.max_column or 0

            if nrows < 2:
                continue

            header_row = _find_header_row(sheet, wb_kind, nrows, ncols)
            if header_row is None:
                continue

            headers = [_safe_str(h) or f"col{i}" for i, h in enumerate(_row_values(sheet, wb_kind, header_row, ncols))]
            rows: list[dict] = []

            for r in range(header_row + 1, nrows):
                cells = _row_values(sheet, wb_kind, r, ncols)
                if not _is_data_row(cells):
                    continue
                row_dict = {headers[i]: cells[i] for i in range(len(headers)) if headers[i] and cells[i]}
                if not row_dict:
                    continue
                # 설비명 추출
                name = (
                    row_dict.get("구분")
                    or row_dict.get("구 분")
                    or row_dict.get("명칭")
                    or cells[0]
                    or cells[2] if len(cells) > 2 else ""
                )
                if not name or name in ("PUMP", "FAN", "MOTOR"):
                    continue
                row_dict["_name"] = name
                rows.append(row_dict)

            if rows:
                result[sheet_name] = rows
    finally:
        if wb_kind == "openpyxl":
            wb.close()

    return result


def _equipment_code(building_code: str, sheet: str, idx: int, name: str) -> str:
    base = re.sub(r"[^\w]", "", sheet)[:6].upper()
    nm = re.sub(r"[^\w가-힣]", "", name)[:10]
    return f"{building_code}-{base}-{idx:03d}"[:64]


# 건물 등록 시 기본으로 만드는 대분류(설비 탭) — 각 1코드
DEFAULT_BUILDING_CATEGORIES = ("위생기기", "조명기기", "기타 설비")


async def _lookup_building(
    session: AsyncSession, site_id: int, building_name: str, bcode: str
) -> Building | None:
    """동일 코드/이름 건물이 여러 개(삭제 후 재등록)여도 활성·최신 1건만 반환."""
    return (
        await session.execute(
            select(Building)
            .where(
                Building.site_id == site_id,
                or_(Building.code == bcode, Building.name == building_name),
            )
            .order_by(Building.is_active.desc(), Building.id.desc())
            .limit(1)
        )
    ).scalars().first()


async def _lookup_equipment_by_code(session: AsyncSession, code: str) -> Equipment | None:
    return (
        await session.execute(
            select(Equipment)
            .where(Equipment.code == code)
            .order_by(Equipment.is_active.desc(), Equipment.id.desc())
            .limit(1)
        )
    ).scalars().first()


async def ensure_building_floor_zone(
    session: AsyncSession, building: Building
) -> Zone:
    """건물에 기본 1층/전체 구역이 없으면 생성. 중복 행이 있어도 첫 행을 사용."""
    floor = (
        await session.execute(
            select(Floor)
            .where(
                Floor.building_id == building.id,
                Floor.name == "1층",
                Floor.is_active == True,
            )
            .order_by(Floor.id)
            .limit(1)
        )
    ).scalars().first()
    if not floor:
        floor = (
            await session.execute(
                select(Floor)
                .where(Floor.building_id == building.id, Floor.is_active == True)
                .order_by(Floor.id)
                .limit(1)
            )
        ).scalars().first()
    if not floor:
        floor = Floor(building_id=building.id, name="1층", level=1, is_active=True)
        session.add(floor)
        await session.flush()

    zone = (
        await session.execute(
            select(Zone)
            .where(
                Zone.floor_id == floor.id,
                Zone.name == "전체",
                Zone.is_active == True,
            )
            .order_by(Zone.id)
            .limit(1)
        )
    ).scalars().first()
    if not zone:
        zone = (
            await session.execute(
                select(Zone)
                .where(Zone.floor_id == floor.id, Zone.is_active == True)
                .order_by(Zone.id)
                .limit(1)
            )
        ).scalars().first()
    if not zone:
        zone = Zone(floor_id=floor.id, name="전체", code="ALL", is_active=True)
        session.add(zone)
        await session.flush()
    return zone


async def ensure_building_default_categories(
    session: AsyncSession, building: Building
) -> int:
    """건물마다 위생기기·조명기기·기타 설비 대분류와 코드 1개씩 보장. 신규 생성 건수 반환."""
    if not building or not building.id:
        return 0
    zone = await ensure_building_floor_zone(session, building)
    created = 0
    for cat in DEFAULT_BUILDING_CATEGORIES:
        exists = (
            await session.execute(
                select(Equipment.id)
                .join(Zone)
                .join(Floor)
                .where(
                    Floor.building_id == building.id,
                    Equipment.category == cat,
                    Equipment.is_active == True,
                )
                .limit(1)
            )
        ).scalars().first()
        if exists:
            continue

        code_val = None
        for idx in range(1, 100):
            candidate = _equipment_code(building.code, cat, idx, cat)
            clash = (
                await session.execute(
                    select(Equipment)
                    .where(Equipment.code == candidate)
                    .order_by(Equipment.id)
                    .limit(1)
                )
            ).scalars().first()
            if clash is None:
                code_val = candidate
                break
            if not clash.is_active:
                clash.is_active = True
                clash.zone_id = zone.id
                clash.name = cat
                clash.category = cat
                clash.status = "normal"
                created += 1
                code_val = None
                break
        if code_val is None:
            continue

        session.add(
            Equipment(
                zone_id=zone.id,
                code=code_val,
                name=cat,
                category=cat,
                status="normal",
                extra_data={},
            )
        )
        created += 1

    if created:
        await session.flush()
    return created


async def ensure_site_and_building(
    session: AsyncSession, building_name: str
) -> tuple[Site, Building, Zone]:
    site = (
        await session.execute(select(Site).where(Site.code == SITE_CODE))
    ).scalar_one_or_none()
    if not site:
        site = Site(name=SITE_NAME, code=SITE_CODE, address="전라남도 광양시")
        session.add(site)
        await session.flush()

    bcode = _building_code(building_name)
    building = await _lookup_building(session, site.id, building_name, bcode)
    if building:
        if not building.is_active:
            building.is_active = True
        if building.name != building_name:
            building.name = building_name
    else:
        building = Building(site_id=site.id, name=building_name, code=bcode)
        session.add(building)
        await session.flush()

    zone = await ensure_building_floor_zone(session, building)
    await ensure_building_default_categories(session, building)
    return site, building, zone


async def import_excel_to_building(
    session: AsyncSession,
    building_name: str,
    file_path: str | Path,
    replace: bool = False,
    building_id: int | None = None,
) -> dict[str, int]:
    """엑셀 파일을 건물에 import. replace=True면 기존 설비 비활성화 후 재등록."""
    if building_id is not None:
        building = await session.get(Building, building_id)
        if not building:
            raise ValueError("건물을 찾을 수 없습니다.")
        if not building.is_active:
            raise ValueError("비활성화된 건물입니다.")
        zone = await ensure_building_floor_zone(session, building)
        await ensure_building_default_categories(session, building)
    else:
        _, building, zone = await ensure_site_and_building(session, building_name)
    parsed = parse_excel_file(file_path)

    if replace:
        existing = (
            await session.execute(
                select(Equipment)
                .join(Zone)
                .join(Floor)
                .where(Floor.building_id == building.id, Equipment.is_active == True)
            )
        ).scalars().all()
        for eq in existing:
            eq.is_active = False

    stats = {"sheets": 0, "created": 0, "updated": 0, "history_added": 0, "pm_added": 0}
    bcode = building.code

    for sheet_name, rows in parsed.items():
        stats["sheets"] += 1
        for idx, row in enumerate(rows, start=1):
            name = row.pop("_name", f"항목{idx}")
            code = _equipment_code(bcode, sheet_name, idx, name)

            existing = await _lookup_equipment_by_code(session, code)

            manufacturer = row.get("제조사") or row.get("제조사/년") or ""
            model = row.get("TYPE") or row.get("Type or Model명") or row.get("MODEL NO.") or row.get("형식") or ""
            serial_no = row.get("Serial No") or row.get("Serial No.") or ""

            if existing:
                existing.is_active = True
                existing.name = name
                existing.category = sheet_name
                existing.zone_id = zone.id
                existing.manufacturer = manufacturer or existing.manufacturer
                existing.model = model or existing.model
                existing.serial_no = serial_no or existing.serial_no
                existing.extra_data = row
                stats["updated"] += 1
            else:
                session.add(
                    Equipment(
                        zone_id=zone.id,
                        code=code,
                        name=name,
                        category=sheet_name,
                        manufacturer=manufacturer,
                        model=model,
                        serial_no=serial_no or None,
                        extra_data=row,
                        status="normal",
                    )
                )
                stats["created"] += 1

    # replace 시 기본 대분류가 비활성화됐을 수 있으므로 재보장
    await ensure_building_default_categories(session, building)

    # 정비이력·점검이력 시트도 함께 대체/반영
    hist_stats = await _import_building_history_sheets(
        session, building.id, file_path, replace=replace
    )
    stats["history_added"] = hist_stats.get("history_added", 0)
    stats["pm_added"] = hist_stats.get("pm_added", 0)

    await session.commit()
    return stats


async def ensure_all_buildings(session: AsyncSession) -> int:
    """건물 목록만 등록 (엑셀 없이) + 기본 대분류/코드 보장."""
    count = 0
    for name in BUILDING_NAMES:
        await ensure_site_and_building(session, name)
        count += 1
    await backfill_all_building_default_categories(session)
    return count


async def backfill_all_building_default_categories(session: AsyncSession) -> int:
    """기존 활성 건물 전체에 기본 대분류(위생기기·조명기기·기타 설비)와 코드 1개씩 보강."""
    buildings = (
        await session.execute(select(Building).where(Building.is_active == True))
    ).scalars().all()
    total = 0
    for b in buildings:
        try:
            total += await ensure_building_default_categories(session, b)
        except Exception as e:
            print(f"[seed] default categories skip building={b.id} {b.name}: {e}", flush=True)
            await session.rollback()
    await session.commit()
    print(
        f"[seed] default categories backfill: buildings={len(buildings)} created={total}",
        flush=True,
    )
    return total


async def import_from_directory(
    session: AsyncSession,
    directory: str | Path,
    replace: bool = True,
) -> dict[str, Any]:
    """디렉터리 내 xls/xlsx 파일 일괄 import."""
    directory = Path(directory)
    results: dict[str, Any] = {"buildings": 0, "total_created": 0, "total_updated": 0, "errors": []}

    # 먼저 모든 건물 등록
    results["buildings"] = await ensure_all_buildings(session)

    for name in BUILDING_NAMES:
        matched = None
        for ext in (".xls", ".xlsx", ".XLS", ".XLSX"):
            p = directory / f"{name}{ext}"
            if p.exists():
                matched = p
                break
        if not matched:
            # fuzzy: 파일명에 건물명 포함
            for f in directory.iterdir():
                if f.suffix.lower() in (".xls", ".xlsx") and name in f.stem:
                    matched = f
                    break
        if not matched:
            results["errors"].append(f"파일 없음: {name}")
            continue
        try:
            stats = await import_excel_to_building(session, name, matched, replace=replace)
            results["total_created"] += stats["created"]
            results["total_updated"] += stats["updated"]
        except Exception as e:
            results["errors"].append(f"{name}: {e}")

    return results


def export_equipment_excel(eq: Equipment) -> bytes:
    """단일 설비에 등록된 사양·정비이력·점검이력·점검주기·작업 전체를 엑셀로 출력."""
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="003876", end_color="003876", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    def _style_header(ws) -> None:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

    # ── 사양 시트 ──
    info_ws = wb.create_sheet(title="사양", index=0)
    info_ws.append(["항목", "내용"])
    _style_header(info_ws)

    zone = getattr(eq, "zone", None)
    floor = getattr(zone, "floor", None) if zone else None
    building = getattr(floor, "building", None) if floor else None
    location = ""
    if zone:
        parts = []
        if building:
            parts.append(building.name or "")
        if floor:
            parts.append(floor.name or "")
        parts.append(zone.name or "")
        location = " / ".join(p for p in parts if p)

    info_rows = [
        ("코드", eq.code or ""),
        ("명칭", eq.name or ""),
        ("분류", eq.category or ""),
        ("상태", eq.status or ""),
        ("위치", location),
        ("제조사", eq.manufacturer or ""),
        ("모델", eq.model or ""),
        ("Serial", eq.serial_no or ""),
    ]
    for label, value in info_rows:
        info_ws.append([label, value])

    extra = eq.extra_data or {}
    for k, v in extra.items():
        if str(k).startswith("_"):
            continue
        info_ws.append([str(k), "" if v is None else str(v)])

    info_ws.column_dimensions["A"].width = 22
    info_ws.column_dimensions["B"].width = 48

    # ── 정비이력 시트 ──
    hist_ws = wb.create_sheet(title="정비이력")
    hist_ws.append(
        ["작업일", "제목", "작업자", "원인", "작업내용(조치)", "사용부품", "비고", "구분", "작업번호"]
    )
    _style_header(hist_ws)
    history = sorted(
        eq.maintenance_records or [],
        key=lambda h: (str(h.work_date or ""), h.id or 0),
        reverse=True,
    )
    for h in history:
        hist_ws.append(
            [
                str(h.work_date) if h.work_date else "",
                h.title or "",
                h.worker_name or "",
                h.cause or "",
                h.action or "",
                h.parts_used or "",
                h.note or "",
                "수동" if h.is_manual else "자동",
                h.work_order_id or "",
            ]
        )
    if not history:
        hist_ws.append(["정비이력 없음"])

    # ── 점검이력 시트 ──
    pm_ws = wb.create_sheet(title="점검이력")
    pm_ws.append(["점검일시", "점검명", "결과", "점검자", "점검내용", "정비의뢰번호"])
    _style_header(pm_ws)
    pm_result_labels = {"normal": "정상", "caution": "주의", "fault": "고장"}
    inspections = sorted(
        eq.pm_inspections or [],
        key=lambda i: (i.inspected_at or i.id or 0),
        reverse=True,
    )
    for insp in inspections:
        result_key = (
            insp.result.value
            if hasattr(insp.result, "value")
            else str(insp.result or "")
        )
        schedule = getattr(insp, "schedule", None)
        inspected = insp.inspected_at.strftime("%Y-%m-%d %H:%M") if insp.inspected_at else ""
        pm_ws.append(
            [
                inspected,
                (schedule.title if schedule else "예방점검"),
                pm_result_labels.get(result_key, result_key),
                insp.inspector_name or "",
                insp.note or "",
                insp.work_order_id or "",
            ]
        )
    if not inspections:
        pm_ws.append(["점검이력 없음"])

    # ── 점검주기 시트 ──
    sched_ws = wb.create_sheet(title="점검주기")
    sched_ws.append(["점검명", "주기", "담당자", "다음예정일", "최근완료일", "활성"])
    _style_header(sched_ws)
    freq_labels = {
        "daily": "매일",
        "weekly": "매주",
        "monthly": "매월",
        "quarterly": "분기",
        "semi_annual": "반기",
        "annual": "연간",
        "custom": "사용자정의",
    }
    schedules = list(eq.pm_schedules or [])
    for pm in schedules:
        freq = pm.frequency.value if hasattr(pm.frequency, "value") else str(pm.frequency or "")
        if freq == "custom" and pm.custom_days:
            freq_label = f"{pm.custom_days}일"
        else:
            freq_label = freq_labels.get(freq, freq)
        sched_ws.append(
            [
                pm.title or "",
                freq_label,
                pm.assignee_name or "",
                str(pm.next_due) if pm.next_due else "",
                str(pm.last_done) if pm.last_done else "",
                "Y" if pm.is_active else "N",
            ]
        )
    if not schedules:
        sched_ws.append(["점검주기 없음"])

    # ── 정비의뢰 시트 ──
    wo_ws = wb.create_sheet(title="정비의뢰")
    wo_ws.append(["번호", "제목", "상태", "우선순위", "담당자", "접수일", "설명"])
    _style_header(wo_ws)
    status_labels = {
        "received": "접수",
        "assigned": "배정",
        "in_progress": "진행",
        "completed": "완료",
        "verified": "확인",
        "closed": "종료",
        "cancelled": "취소",
    }
    orders = sorted(
        [wo for wo in (eq.work_orders or []) if getattr(wo, "is_active", True)],
        key=lambda w: (w.created_at or w.id or 0),
        reverse=True,
    )
    for wo in orders:
        st = wo.status.value if hasattr(wo.status, "value") else str(wo.status or "")
        created = wo.created_at.strftime("%Y-%m-%d %H:%M") if wo.created_at else ""
        wo_ws.append(
            [
                wo.id,
                wo.title or "",
                status_labels.get(st, st),
                wo.priority or "",
                wo.assignee_name or "",
                created,
                wo.description or "",
            ]
        )
    if not orders:
        wo_ws.append(["정비의뢰 없음"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _parse_excel_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _cell_str(v)
    if not s or s in ("정비이력 없음", "점검이력 없음", "점검주기 없음", "정비의뢰 없음"):
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s[:19], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace("/", "-")[:10]).date()
    except ValueError:
        return None


def _parse_excel_datetime(v: Any) -> datetime | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date) and not isinstance(v, datetime):
        return datetime.combine(v, datetime.min.time())
    s = _cell_str(v)
    if not s or s in ("정비이력 없음", "점검이력 없음"):
        return None
    for fmt in (
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(s[:19], fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace("/", "-"))
    except ValueError:
        return None


def _pm_result_from_label(raw: str) -> PMResult:
    key = (raw or "").strip().lower()
    mapping = {
        "정상": PMResult.normal,
        "normal": PMResult.normal,
        "주의": PMResult.caution,
        "caution": PMResult.caution,
        "고장": PMResult.fault,
        "fault": PMResult.fault,
    }
    # 한글 라벨은 lower()가 그대로이므로 원문도 조회
    return mapping.get(key) or mapping.get((raw or "").strip()) or PMResult.normal


def _pm_freq_from_label(raw: str) -> tuple[PMFrequency, int | None]:
    s = (raw or "").strip()
    mapping = {
        "매일": PMFrequency.daily,
        "daily": PMFrequency.daily,
        "매주": PMFrequency.weekly,
        "weekly": PMFrequency.weekly,
        "매월": PMFrequency.monthly,
        "monthly": PMFrequency.monthly,
        "분기": PMFrequency.quarterly,
        "quarterly": PMFrequency.quarterly,
        "반기": PMFrequency.semi_annual,
        "semi_annual": PMFrequency.semi_annual,
        "연간": PMFrequency.annual,
        "annual": PMFrequency.annual,
    }
    if s in mapping:
        return mapping[s], None
    m = re.match(r"^(\d+)\s*일$", s)
    if m:
        return PMFrequency.custom, int(m.group(1))
    return PMFrequency.monthly, None


def _wo_status_from_label(raw: str) -> WorkOrderStatus:
    s = (raw or "").strip()
    mapping = {
        "접수": WorkOrderStatus.received,
        "received": WorkOrderStatus.received,
        "배정": WorkOrderStatus.assigned,
        "assigned": WorkOrderStatus.assigned,
        "진행": WorkOrderStatus.in_progress,
        "in_progress": WorkOrderStatus.in_progress,
        "완료": WorkOrderStatus.completed,
        "completed": WorkOrderStatus.completed,
        "확인": WorkOrderStatus.verified,
        "verified": WorkOrderStatus.verified,
        "종료": WorkOrderStatus.closed,
        "closed": WorkOrderStatus.closed,
    }
    return mapping.get(s) or mapping.get(s.lower()) or WorkOrderStatus.received


async def _clear_equipment_history(session: AsyncSession, eq: Equipment) -> None:
    """설비의 정비이력·점검이력·점검주기를 삭제 (대체 import용)."""
    await session.execute(
        delete(PMInspection).where(PMInspection.equipment_id == eq.id)
    )
    await session.execute(
        delete(PMSchedule).where(PMSchedule.equipment_id == eq.id)
    )
    await session.execute(
        delete(MaintenanceRecord).where(MaintenanceRecord.equipment_id == eq.id)
    )
    # 관계 캐시 비움
    if eq.pm_inspections is not None:
        eq.pm_inspections.clear()
    if eq.pm_schedules is not None:
        eq.pm_schedules.clear()
    if eq.maintenance_records is not None:
        eq.maintenance_records.clear()
    await session.flush()


async def import_equipment_excel(
    session: AsyncSession,
    eq: Equipment,
    file_bytes: bytes,
    *,
    replace: bool = True,
) -> dict[str, Any]:
    """단일 설비 Excel(사양·정비이력·점검이력·점검주기·정비의뢰)을 읽어 반영.

    replace=True(기본): 기존 이력/주기를 지우고 엑셀 내용으로 대체.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    stats = {
        "spec_updated": 0,
        "history_added": 0,
        "history_cleared": 0,
        "pm_added": 0,
        "schedule_added": 0,
        "wo_added": 0,
        "skipped": 0,
        "warnings": [],
        "replaced": bool(replace),
    }

    if replace:
        # 기존 건수 (메시지용)
        hist_n = len(eq.maintenance_records or [])
        pm_n = len(eq.pm_inspections or [])
        sch_n = len(eq.pm_schedules or [])
        stats["history_cleared"] = hist_n + pm_n + sch_n
        await _clear_equipment_history(session, eq)
        # 정비의뢰는 활성 건만 비활성 처리 (외래키·D-1 보존)
        for wo in list(eq.work_orders or []):
            if getattr(wo, "is_active", True):
                wo.is_active = False

    # ── 사양 ──
    if "사양" in wb.sheetnames:
        ws = wb["사양"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        field_map = {
            "명칭": "name",
            "분류": "category",
            "상태": "status",
            "제조사": "manufacturer",
            "모델": "model",
            "Serial": "serial_no",
            "시리얼": "serial_no",
            "시리얼번호": "serial_no",
        }
        extra = dict(eq.extra_data or {}) if not replace else {}
        for row in rows:
            if not row or len(row) < 1:
                continue
            label = _cell_str(row[0])
            value = _cell_str(row[1]) if len(row) > 1 else ""
            if not label:
                continue
            if label == "코드":
                if value and value != (eq.code or ""):
                    stats["warnings"].append(
                        f"엑셀 코드({value})가 현재 설비({eq.code})와 달라 코드는 변경하지 않았습니다."
                    )
                continue
            if label == "위치":
                continue
            attr = field_map.get(label)
            if attr:
                current = getattr(eq, attr, None)
                if attr == "name" and not value:
                    continue
                new_val = value or None
                if attr == "name":
                    new_val = value
                if (current or "") != (new_val or ""):
                    setattr(eq, attr, new_val)
                    stats["spec_updated"] += 1
            else:
                if extra.get(label) != value:
                    extra[label] = value
                    stats["spec_updated"] += 1
        eq.extra_data = extra

    # ── 정비이력 ──
    if "정비이력" in wb.sheetnames:
        ws = wb["정비이력"]
        existing: set[tuple] = set()
        if not replace:
            existing = {
                (
                    str(h.work_date or ""),
                    (h.title or "").strip(),
                    (h.worker_name or "").strip(),
                    (h.action or "").strip(),
                )
                for h in (eq.maintenance_records or [])
            }
        header_cells = list(next(ws.iter_rows(min_row=1, max_row=1)))
        headers = [_cell_str(c.value) for c in header_cells]
        col = {name: idx for idx, name in enumerate(headers) if name}

        def hist_val(row, *names):
            for n in names:
                if n in col and col[n] < len(row):
                    return row[col[n]]
            return None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            first = _cell_str(row[0])
            if first in ("정비이력 없음", ""):
                continue
            work_date = _parse_excel_date(hist_val(row, "작업일") if col else row[0])
            title = _cell_str(hist_val(row, "제목") if col else (row[1] if len(row) > 1 else ""))
            if not work_date or not title:
                stats["skipped"] += 1
                continue
            worker = _cell_str(hist_val(row, "작업자") if col else (row[2] if len(row) > 2 else ""))
            cause = _cell_str(hist_val(row, "원인") if col else (row[3] if len(row) > 3 else ""))
            action = _cell_str(
                hist_val(row, "작업내용(조치)", "작업내용") if col else (row[4] if len(row) > 4 else "")
            )
            parts = _cell_str(hist_val(row, "사용부품") if col else (row[5] if len(row) > 5 else ""))
            note = _cell_str(hist_val(row, "비고") if col else (row[6] if len(row) > 6 else ""))
            kind = _cell_str(hist_val(row, "구분") if col else (row[7] if len(row) > 7 else ""))
            key = (str(work_date), title, worker, action)
            if key in existing:
                stats["skipped"] += 1
                continue
            existing.add(key)
            session.add(
                MaintenanceRecord(
                    equipment_id=eq.id,
                    title=title[:300],
                    work_date=work_date,
                    worker_name=worker or None,
                    cause=cause or None,
                    action=action or None,
                    parts_used=parts or None,
                    note=note or None,
                    is_manual=(kind != "자동"),
                )
            )
            stats["history_added"] += 1

    # ── 점검주기 (점검이력보다 먼저 등록) ──
    schedules_by_title: dict[str, PMSchedule] = {}
    if "점검주기" in wb.sheetnames:
        ws = wb["점검주기"]
        header_cells = list(next(ws.iter_rows(min_row=1, max_row=1)))
        headers = [_cell_str(c.value) for c in header_cells]
        col = {name: idx for idx, name in enumerate(headers) if name}

        def sch_val(row, *names):
            for n in names:
                if n in col and col[n] < len(row):
                    return row[col[n]]
            return None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            first = _cell_str(row[0])
            if first in ("점검주기 없음", ""):
                continue
            title = _cell_str(sch_val(row, "점검명") if col else row[0]) or "예방점검"
            freq_raw = _cell_str(sch_val(row, "주기") if col else (row[1] if len(row) > 1 else ""))
            assignee = _cell_str(
                sch_val(row, "담당자") if col else (row[2] if len(row) > 2 else "")
            )
            next_due = _parse_excel_date(
                sch_val(row, "다음예정일") if col else (row[3] if len(row) > 3 else None)
            )
            last_done = _parse_excel_date(
                sch_val(row, "최근완료일") if col else (row[4] if len(row) > 4 else None)
            )
            active_raw = _cell_str(
                sch_val(row, "활성") if col else (row[5] if len(row) > 5 else "Y")
            )
            freq, custom_days = _pm_freq_from_label(freq_raw)
            schedule = PMSchedule(
                equipment_id=eq.id,
                title=title[:200],
                frequency=freq,
                custom_days=custom_days,
                assignee_name=assignee or None,
                next_due=next_due,
                last_done=last_done,
                is_active=active_raw.upper() not in ("N", "NO", "FALSE", "0", "비활성"),
            )
            session.add(schedule)
            await session.flush()
            schedules_by_title[title] = schedule
            stats["schedule_added"] += 1

    # replace가 아니면 기존 스케줄도 맵에 포함
    if not replace:
        for s in eq.pm_schedules or []:
            schedules_by_title.setdefault((s.title or "").strip(), s)

    # ── 점검이력 ──
    if "점검이력" in wb.sheetnames:
        ws = wb["점검이력"]
        existing_insp: set[tuple] = set()
        if not replace:
            existing_insp = {
                (
                    (i.inspected_at.strftime("%Y-%m-%d %H:%M") if i.inspected_at else ""),
                    (i.result.value if hasattr(i.result, "value") else str(i.result or "")),
                    (i.inspector_name or "").strip(),
                    (i.note or "").strip(),
                )
                for i in (eq.pm_inspections or [])
            }
        header_cells = list(next(ws.iter_rows(min_row=1, max_row=1)))
        headers = [_cell_str(c.value) for c in header_cells]
        col = {name: idx for idx, name in enumerate(headers) if name}

        def pm_val(row, *names):
            for n in names:
                if n in col and col[n] < len(row):
                    return row[col[n]]
            return None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            first = _cell_str(row[0])
            if first in ("점검이력 없음", ""):
                continue
            inspected_at = _parse_excel_datetime(pm_val(row, "점검일시") if col else row[0])
            title = _cell_str(
                pm_val(row, "점검명") if col else (row[1] if len(row) > 1 else "")
            ) or "예방점검"
            result_label = _cell_str(
                pm_val(row, "결과") if col else (row[2] if len(row) > 2 else "")
            )
            inspector = _cell_str(
                pm_val(row, "점검자") if col else (row[3] if len(row) > 3 else "")
            )
            note = _cell_str(
                pm_val(row, "점검내용") if col else (row[4] if len(row) > 4 else "")
            )
            if not inspected_at:
                stats["skipped"] += 1
                continue
            result = _pm_result_from_label(result_label)
            key = (
                inspected_at.strftime("%Y-%m-%d %H:%M"),
                result.value,
                inspector,
                note,
            )
            if key in existing_insp:
                stats["skipped"] += 1
                continue
            existing_insp.add(key)

            schedule = schedules_by_title.get(title)
            if not schedule:
                schedule = next(
                    (s for s in schedules_by_title.values() if getattr(s, "is_active", True)),
                    None,
                )
            if not schedule:
                schedule = PMSchedule(
                    equipment_id=eq.id,
                    title=title[:200],
                    frequency=PMFrequency.monthly,
                    is_active=True,
                )
                session.add(schedule)
                await session.flush()
                schedules_by_title[title] = schedule
                stats["schedule_added"] += 1

            session.add(
                PMInspection(
                    schedule_id=schedule.id,
                    equipment_id=eq.id,
                    result=result,
                    note=note or None,
                    inspector_name=inspector or None,
                    inspected_at=inspected_at,
                )
            )
            stats["pm_added"] += 1

    # ── 정비의뢰 ──
    if "정비의뢰" in wb.sheetnames:
        ws = wb["정비의뢰"]
        header_cells = list(next(ws.iter_rows(min_row=1, max_row=1)))
        headers = [_cell_str(c.value) for c in header_cells]
        col = {name: idx for idx, name in enumerate(headers) if name}

        def wo_val(row, *names):
            for n in names:
                if n in col and col[n] < len(row):
                    return row[col[n]]
            return None

        site_id = None
        try:
            zone = getattr(eq, "zone", None)
            floor = getattr(zone, "floor", None) if zone else None
            building = getattr(floor, "building", None) if floor else None
            site_id = getattr(building, "site_id", None)
        except Exception:
            site_id = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            first = _cell_str(row[0])
            if first in ("정비의뢰 없음", ""):
                continue
            title = _cell_str(wo_val(row, "제목") if col else (row[1] if len(row) > 1 else ""))
            if not title:
                # 번호만 있고 제목 없는 경우 skip
                if col and "제목" in col:
                    stats["skipped"] += 1
                    continue
                title = first or f"[가져오기] {eq.code}"
            status = _wo_status_from_label(
                _cell_str(wo_val(row, "상태") if col else (row[2] if len(row) > 2 else ""))
            )
            priority = _cell_str(
                wo_val(row, "우선순위") if col else (row[3] if len(row) > 3 else "")
            ) or "normal"
            if priority in ("긴급", "high"):
                priority = "high"
            elif priority in ("일반", "normal", ""):
                priority = "normal"
            assignee = _cell_str(
                wo_val(row, "담당자") if col else (row[4] if len(row) > 4 else "")
            )
            created = _parse_excel_datetime(
                wo_val(row, "접수일") if col else (row[5] if len(row) > 5 else None)
            )
            desc = _cell_str(
                wo_val(row, "설명") if col else (row[6] if len(row) > 6 else "")
            )
            wo = WorkOrder(
                equipment_id=eq.id,
                site_id=site_id,
                title=title[:300],
                description=desc or None,
                status=status,
                priority=priority[:20],
                assignee_name=assignee or None,
                is_active=True,
            )
            if created:
                wo.created_at = created
            session.add(wo)
            stats["wo_added"] += 1

    await session.flush()
    return stats


async def _import_building_history_sheets(
    session: AsyncSession,
    building_id: int,
    file_path: str | Path,
    *,
    replace: bool,
) -> dict[str, int]:
    """건물 설비현황 엑셀의 정비이력·점검이력 시트를 반영."""
    from openpyxl import load_workbook

    stats = {"history_added": 0, "pm_added": 0, "cleared_equipment": 0}
    path = Path(file_path)

    eq_rows = (
        await session.execute(
            select(Equipment)
            .join(Zone)
            .join(Floor)
            .where(Floor.building_id == building_id, Equipment.is_active == True)
            .options(
                selectinload(Equipment.maintenance_records),
                selectinload(Equipment.pm_inspections),
                selectinload(Equipment.pm_schedules),
            )
        )
    ).scalars().unique().all()
    by_code = {(eq.code or "").strip(): eq for eq in eq_rows if eq.code}

    # 대체 import: 엑셀에 이력이 없어도 기존 정비·점검 이력을 비운 뒤 엑셀 내용만 남긴다
    if replace:
        for eq in eq_rows:
            await _clear_equipment_history(session, eq)
            stats["cleared_equipment"] += 1

    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        # .xls 는 openpyxl 미지원 → 설비 시트만 반영되고 이력 시트는 스킵
        return stats

    wb = load_workbook(filename=str(path), data_only=True)
    try:
        if "정비이력" in wb.sheetnames:
            ws = wb["정비이력"]
            header_cells = list(next(ws.iter_rows(min_row=1, max_row=1)))
            headers = [_cell_str(c.value) for c in header_cells]
            col = {name: idx for idx, name in enumerate(headers) if name}

            def hv(row, *names):
                for n in names:
                    if n in col and col[n] < len(row):
                        return row[col[n]]
                return None

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                first = _cell_str(row[0])
                if first in ("정비이력 없음", ""):
                    continue
                code = _cell_str(hv(row, "설비코드") if col else row[0])
                eq = by_code.get(code)
                if not eq:
                    continue
                work_date = _parse_excel_date(hv(row, "작업일") if col else None)
                title = _cell_str(hv(row, "제목") if col else "")
                if not work_date or not title:
                    continue
                session.add(
                    MaintenanceRecord(
                        equipment_id=eq.id,
                        title=title[:300],
                        work_date=work_date,
                        worker_name=_cell_str(hv(row, "작업자") if col else "") or None,
                        cause=_cell_str(hv(row, "원인") if col else "") or None,
                        action=_cell_str(hv(row, "작업내용(조치)", "작업내용") if col else "") or None,
                        parts_used=_cell_str(hv(row, "사용부품") if col else "") or None,
                        note=_cell_str(hv(row, "비고") if col else "") or None,
                        is_manual=_cell_str(hv(row, "구분") if col else "") != "자동",
                    )
                )
                stats["history_added"] += 1

        if "점검이력" in wb.sheetnames:
            ws = wb["점검이력"]
            header_cells = list(next(ws.iter_rows(min_row=1, max_row=1)))
            headers = [_cell_str(c.value) for c in header_cells]
            col = {name: idx for idx, name in enumerate(headers) if name}

            def pv(row, *names):
                for n in names:
                    if n in col and col[n] < len(row):
                        return row[col[n]]
                return None

            schedule_cache: dict[tuple[int, str], PMSchedule] = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                first = _cell_str(row[0])
                if first in ("점검이력 없음", ""):
                    continue
                code = _cell_str(pv(row, "설비코드") if col else row[0])
                eq = by_code.get(code)
                if not eq:
                    continue
                inspected_at = _parse_excel_datetime(pv(row, "점검일시") if col else None)
                title = _cell_str(pv(row, "점검명") if col else "") or "예방점검"
                if not inspected_at:
                    continue
                result = _pm_result_from_label(_cell_str(pv(row, "결과") if col else ""))
                cache_key = (eq.id, title)
                schedule = schedule_cache.get(cache_key)
                if not schedule:
                    for s in eq.pm_schedules or []:
                        if (s.title or "").strip() == title:
                            schedule = s
                            break
                if not schedule:
                    schedule = PMSchedule(
                        equipment_id=eq.id,
                        title=title[:200],
                        frequency=PMFrequency.monthly,
                        is_active=True,
                    )
                    session.add(schedule)
                    await session.flush()
                    if eq.pm_schedules is None:
                        eq.pm_schedules = []
                    eq.pm_schedules.append(schedule)
                schedule_cache[cache_key] = schedule
                session.add(
                    PMInspection(
                        schedule_id=schedule.id,
                        equipment_id=eq.id,
                        result=result,
                        note=_cell_str(pv(row, "점검내용") if col else "") or None,
                        inspector_name=_cell_str(pv(row, "점검자") if col else "") or None,
                        inspected_at=inspected_at,
                    )
                )
                stats["pm_added"] += 1
    finally:
        wb.close()
    return stats


def export_building_excel(
    building_name: str,
    equipment_by_sheet: dict[str, list[Equipment]],
) -> bytes:
    """건물 설비를 엑셀 파일(bytes)로 출력 (시트별 설비 + 정비이력 + 점검이력)."""
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="003876", end_color="003876", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    all_equipment: list[Equipment] = []

    for sheet_name, items in equipment_by_sheet.items():
        safe_name = sheet_name[:31]
        ws = wb.create_sheet(title=safe_name)
        all_equipment.extend(items)

        if not items:
            ws.append(["데이터 없음"])
            continue

        # extra_data 키 수집
        all_keys: list[str] = []
        for eq in items:
            for k in (eq.extra_data or {}):
                if k not in all_keys and not k.startswith("_"):
                    all_keys.append(k)

        headers = ["코드", "명칭", "제조사", "모델", "Serial"] + all_keys
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for eq in items:
            extra = eq.extra_data or {}
            row = [
                eq.code,
                eq.name,
                eq.manufacturer or "",
                eq.model or "",
                eq.serial_no or "",
            ] + [extra.get(k, "") for k in all_keys]
            ws.append(row)

    # 정비이력 시트
    hist_ws = wb.create_sheet(title="정비이력")
    hist_headers = [
        "설비코드",
        "설비명",
        "시트",
        "작업일",
        "제목",
        "작업자",
        "원인",
        "작업내용(조치)",
        "사용부품",
        "비고",
        "구분",
        "작업번호",
    ]
    hist_ws.append(hist_headers)
    for cell in hist_ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    history_rows: list[tuple] = []
    for eq in all_equipment:
        for h in eq.maintenance_records or []:
            history_rows.append((eq.code or "", eq.name or "", eq.category or "", h))
    history_rows.sort(
        key=lambda x: (str(x[3].work_date or ""), x[0]),
        reverse=True,
    )

    for code, name, category, h in history_rows:
        hist_ws.append(
            [
                code,
                name,
                category,
                str(h.work_date) if h.work_date else "",
                h.title or "",
                h.worker_name or "",
                h.cause or "",
                h.action or "",
                h.parts_used or "",
                h.note or "",
                "수동" if h.is_manual else "자동",
                h.work_order_id or "",
            ]
        )
    if not history_rows:
        hist_ws.append(["정비이력 없음"])

    # 점검이력 시트
    pm_ws = wb.create_sheet(title="점검이력")
    pm_headers = [
        "설비코드",
        "설비명",
        "시트",
        "점검일시",
        "점검명",
        "결과",
        "점검자",
        "점검내용",
        "정비의뢰번호",
    ]
    pm_ws.append(pm_headers)
    for cell in pm_ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    pm_result_labels = {"normal": "정상", "caution": "주의", "fault": "고장"}
    pm_rows: list[tuple] = []
    for eq in all_equipment:
        for insp in eq.pm_inspections or []:
            pm_rows.append((eq, insp))
    pm_rows.sort(
        key=lambda x: (x[1].inspected_at or x[1].id or 0),
        reverse=True,
    )

    for eq, insp in pm_rows:
        result_key = (
            insp.result.value
            if hasattr(insp.result, "value")
            else str(insp.result or "")
        )
        schedule = getattr(insp, "schedule", None)
        inspected = insp.inspected_at.strftime("%Y-%m-%d %H:%M") if insp.inspected_at else ""
        pm_ws.append(
            [
                eq.code or "",
                eq.name or "",
                eq.category or "",
                inspected,
                (schedule.title if schedule else "예방점검"),
                pm_result_labels.get(result_key, result_key),
                insp.inspector_name or "",
                insp.note or "",
                insp.work_order_id or "",
            ]
        )
    if not pm_rows:
        pm_ws.append(["점검이력 없음"])

    # 총괄 시트
    summary = wb.create_sheet(title="총괄", index=0)
    summary.append([f"{building_name} 설비현황"])
    summary.append(["시트", "건수"])
    for sheet_name, items in equipment_by_sheet.items():
        summary.append([sheet_name, len(items)])
    summary.append(["정비이력", len(history_rows)])
    summary.append(["점검이력", len(pm_rows)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def match_building_filename(stem: str) -> str | None:
    """파일명(stem)에서 건물명 매칭."""
    stem = stem.strip()
    if stem in BUILDING_NAMES:
        return stem
    for name in BUILDING_NAMES:
        if name in stem or stem in name:
            return name
    return stem
