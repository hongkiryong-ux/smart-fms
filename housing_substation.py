# -*- coding: utf-8 -*-
"""주택변전소 점검일지2 — 1일·월보 웹 입력·자동 집계·엑셀 아카이브."""
from __future__ import annotations

import io
import json
import re
from copy import deepcopy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models import Building, HousingSubstationArchive, HousingSubstationDaily

ROOT = Path(__file__).resolve().parent
SCHEMA_PATH = ROOT / "resources" / "housing_substation_schema.json"
TEMPLATE_XLSX = ROOT / "resources" / "housing_substation_template.xlsx"

_schema_cache: dict | None = None


def public_base_url(request: Any | None = None) -> str:
    import os

    env = (os.environ.get("PUBLIC_BASE_URL") or "").strip().rstrip("/")
    if env:
        return env
    if request is not None:
        return str(request.base_url).rstrip("/")
    return "http://127.0.0.1:8000"


def housing_daily_qr_url(building_code: str, request: Any | None = None) -> str:
    code = (building_code or "").strip()
    return f"{public_base_url(request)}/hs/{code}/daily"


def qr_png_bytes(url: str) -> bytes:
    import qrcode

    img = qrcode.make(url)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


async def get_housing_building_for_qr(
    session: AsyncSession, code: str
) -> Building | None:
    from models import InspectionLogBuilding2

    building = (
        await session.execute(
            select(Building)
            .join(InspectionLogBuilding2, InspectionLogBuilding2.building_id == Building.id)
            .where(
                Building.code == (code or "").strip(),
                Building.is_active == True,  # noqa: E712
            )
        )
    ).scalar_one_or_none()
    if not building or not is_housing_substation_building(building):
        return None
    return building


def load_schema() -> dict:
    global _schema_cache
    if _schema_cache is None:
        _schema_cache = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))
    return _schema_cache


def build_daily_block_layout(block: dict) -> dict:
    """엑셀 1일 시트와 동일한 다단 헤더·열 순서."""
    from openpyxl.utils import column_index_from_string as col_idx

    roman = {"no1": "Ⅰ", "no2": "Ⅱ", "no3": "Ⅲ"}
    bid = block.get("id", "")
    sheet_title = f"주택변전소 운전일지[{roman.get(bid, bid)}]"

    all_cols: list[dict[str, Any]] = []
    for c in block.get("input_columns") or []:
        all_cols.append(
            {
                "sort": col_idx(c["col"]),
                "col": c["col"],
                "kind": "input",
                "group": (c.get("group") or "").strip(),
                "metric": c.get("metric") or "",
                "unit": c.get("unit") or "",
                "range": c.get("range") or "",
            }
        )
    for m in block.get("meters") or []:
        mult = m.get("multiplier") or 4800
        all_cols.append(
            {
                "sort": col_idx(m["reading_col"]),
                "col": m["reading_col"],
                "kind": "cumulative",
                "group": (m.get("name") or "").strip(),
                "meter_id": m["id"],
                "metric": f"적산량×{mult}",
                "unit": "전일 지침",
                "range": "",
            }
        )
    all_cols.sort(key=lambda x: x["sort"])

    sections: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for col in all_cols:
        g = col.get("group") or ""
        if current is None or current["name"] != g:
            if current:
                sections.append(current)
            current = {"name": g, "columns": []}
        current["columns"].append(col)
    if current:
        sections.append(current)

    return {"title": sheet_title, "sections": sections, "times": block.get("times") or []}


def build_all_daily_layouts(schema: dict | None = None) -> dict[str, dict]:
    schema = schema or load_schema()
    return {b["id"]: build_daily_block_layout(b) for b in schema.get("daily_blocks") or []}


def build_monthly_sections(schema: dict | None = None) -> list[dict[str, Any]]:
    """1일 시트 meters 순서·명칭과 동일한 월보 열 구성."""
    schema = schema or load_schema()
    roman = {"no1": "Ⅰ", "no2": "Ⅱ", "no3": "Ⅲ"}
    order = {"no1": 0, "no2": 1, "no3": 2}
    blocks = sorted(schema.get("daily_blocks") or [], key=lambda b: order.get(b["id"], 99))
    sections: list[dict[str, Any]] = []
    for block in blocks:
        bid = block["id"]
        breakers = []
        for m in block.get("meters") or []:
            breakers.append(
                {
                    "meter_id": m["id"],
                    "name": m["name"],
                    "multiplier": m.get("multiplier") or 4800,
                    "reading_col": m["reading_col"],
                    "current_col": m.get("current_col"),
                }
            )
        sections.append(
            {
                "id": bid,
                "title": f"주택변전소 월보[{roman.get(bid, bid)}]",
                "breakers": breakers,
            }
        )
    return sections


def is_housing_substation_building(building: Building | None) -> bool:
    if not building:
        return False
    name = (building.name or "").strip()
    target = load_schema().get("building_name", "주택변전소")
    return name == target or target in name


def empty_daily_payload() -> dict:
    schema = load_schema()
    out: dict[str, Any] = {}
    for block in schema["daily_blocks"]:
        bid = block["id"]
        prev: dict[str, Any] = {}
        for m in block["meters"]:
            prev[m["id"]] = ""
        out[bid] = {"prev": prev, "times": {t: {} for t in block["times"]}}
    return out


def _parse_num(raw: Any) -> float | None:
    if raw is None or raw == "":
        return None
    try:
        return float(str(raw).replace(",", "").strip())
    except ValueError:
        return None


async def get_daily_row(
    session: AsyncSession, building_id: int, log_date: date
) -> HousingSubstationDaily | None:
    return (
        await session.execute(
            select(HousingSubstationDaily).where(
                HousingSubstationDaily.building_id == building_id,
                HousingSubstationDaily.log_date == log_date,
            )
        )
    ).scalar_one_or_none()


async def _prev_day_readings(
    session: AsyncSession, building_id: int, log_date: date
) -> dict[str, dict[str, Any]]:
    """전일 22:00 적산(지침) → 오늘 prev."""
    prev_date = log_date - timedelta(days=1)
    row = await get_daily_row(session, building_id, prev_date)
    if not row or not row.data:
        return {}
    return prev_readings_from_daily_data(row.data)


def prev_readings_from_daily_data(source_data: dict) -> dict[str, dict[str, Any]]:
    """저장된 1일 데이터에서 블록별 전일(22:00) 지침 추출."""
    schema = load_schema()
    out: dict[str, dict[str, Any]] = {}
    for block in schema["daily_blocks"]:
        bid = block["id"]
        block_data = (source_data or {}).get(bid) or {}
        times = block_data.get("times") or {}
        t2200 = times.get("22:00") or {}
        prev: dict[str, Any] = {}
        for m in block["meters"]:
            rc = m["reading_col"]
            val = t2200.get(rc, "")
            prev[m["id"]] = val
        out[bid] = prev
    return out


def apply_prev_readings(data: dict, prev_vals: dict[str, dict[str, Any]]) -> bool:
    """전일 22:00 지침을 prev에 반영. 변경 여부 반환."""
    if not prev_vals:
        return False
    changed = False
    for bid, meters in prev_vals.items():
        if bid not in data:
            continue
        data[bid].setdefault("prev", {})
        for mid, val in meters.items():
            if data[bid]["prev"].get(mid) != val:
                data[bid]["prev"][mid] = val
                changed = True
    return changed


async def sync_daily_prev(
    session: AsyncSession, building_id: int, log_date: date, data: dict
) -> tuple[dict, bool]:
    """스키마 키 보정 후 전일 22:00 지침으로 prev 동기화."""
    merged = merge_daily_save(empty_daily_payload(), data)
    prev_vals = await _prev_day_readings(session, building_id, log_date)
    changed = apply_prev_readings(merged, prev_vals)
    return merged, changed


async def propagate_prev_to_next_day(
    session: AsyncSession, building_id: int, log_date: date, saved_data: dict
) -> None:
    """당일 22:00 저장 내용을 다음 날 prev에 반영."""
    next_date = log_date + timedelta(days=1)
    next_row = await get_daily_row(session, building_id, next_date)
    if not next_row:
        return
    prev_vals = prev_readings_from_daily_data(saved_data)
    next_data = merge_daily_save(empty_daily_payload(), next_row.data or {})
    if apply_prev_readings(next_data, prev_vals):
        next_row.data = next_data


async def get_or_create_daily(
    session: AsyncSession, building_id: int, log_date: date
) -> HousingSubstationDaily:
    row = await get_daily_row(session, building_id, log_date)
    if row:
        synced, changed = await sync_daily_prev(session, building_id, log_date, row.data or {})
        if changed:
            row.data = synced
            await session.flush()
        return row
    synced, _ = await sync_daily_prev(session, building_id, log_date, empty_daily_payload())
    row = HousingSubstationDaily(
        building_id=building_id,
        log_date=log_date,
        data=synced,
    )
    session.add(row)
    await session.flush()
    return row


def merge_daily_save(existing: dict, posted: dict) -> dict:
    data = deepcopy(existing) if existing else empty_daily_payload()
    for bid, block_post in (posted or {}).items():
        if bid not in data:
            data[bid] = {"prev": {}, "times": {}}
        if "prev" in block_post:
            data[bid]["prev"].update(block_post.get("prev") or {})
        for t, cells in (block_post.get("times") or {}).items():
            data[bid]["times"].setdefault(t, {})
            data[bid]["times"][t].update(cells or {})
    return data


def _meter_reading_at_2200(block_data: dict, meter: dict) -> Any:
    times = block_data.get("times") or {}
    return (times.get("22:00") or {}).get(meter["reading_col"], "")


def _meter_max_current(block_data: dict, meter: dict) -> float | None:
    cc = meter.get("current_col")
    if not cc:
        return None
    times = block_data.get("times") or {}
    amps: list[float] = []
    for cells in times.values():
        av = _parse_num((cells or {}).get(cc))
        if av is not None:
            amps.append(av)
    return max(amps) if amps else None


def _breaker_row_from_daily(
    block_data: dict,
    meter: dict,
    prev_reading: float | None,
) -> tuple[dict[str, Any], float | None]:
    """지침·사용량·최대(A). 사용량 = (당일 22:00 지침 − 전일 지침) × 배율."""
    reading_raw = _meter_reading_at_2200(block_data, meter)
    read_v = _parse_num(reading_raw)
    mult = meter.get("multiplier") or 4800
    usage: Any = ""
    if prev_reading is not None and read_v is not None:
        usage = round((read_v - prev_reading) * mult, 2)
    max_a_val = _meter_max_current(block_data, meter)
    max_a: Any = max_a_val if max_a_val is not None else ""
    next_prev = read_v if read_v is not None else prev_reading
    return (
        {
            "meter_id": meter.get("id") or meter.get("meter_id"),
            "name": meter["name"],
            "multiplier": mult,
            "reading": reading_raw if reading_raw not in (None, "") else "",
            "usage": usage,
            "max_a": max_a,
        },
        next_prev,
    )


def compute_monthly_report(
    building_id: int,
    year: int,
    month: int,
    daily_rows: list[HousingSubstationDaily],
    prev_month_last_row: HousingSubstationDaily | None = None,
) -> dict[str, Any]:
    """월보: 전일(노란 칸) · 1~31일 · 계(사용량 합·최대A)."""
    del building_id
    by_date = {r.log_date: r.data or {} for r in daily_rows}
    sections_out = []
    for sec in build_monthly_sections():
        sec_id = sec["id"]
        breakers = sec["breakers"]
        prev_day_cells: list[dict[str, Any]] = []
        prev_readings: dict[str, float | None] = {}
        prev_month_data = (prev_month_last_row.data or {}).get(sec_id, {}) if prev_month_last_row else {}
        for br in breakers:
            mid = br["meter_id"]
            val = _parse_num(_meter_reading_at_2200(prev_month_data, br))
            if val is None:
                first_of_month = date(year, month, 1)
                first_data = by_date.get(first_of_month, {}).get(sec_id, {})
                prev_map = first_data.get("prev") or {}
                val = _parse_num(prev_map.get(mid))
            prev_readings[mid] = val
            prev_day_cells.append(
                {
                    "meter_id": mid,
                    "name": br["name"],
                    "multiplier": br["multiplier"],
                    "reading": val if val is not None else "",
                    "usage": "",
                    "max_a": "",
                }
            )
        days = []
        usage_sums: dict[str, float] = {b["meter_id"]: 0.0 for b in breakers}
        max_a_peaks: dict[str, float | None] = {b["meter_id"]: None for b in breakers}
        rolling_prev = dict(prev_readings)
        for day in range(1, 32):
            try:
                d = date(year, month, day)
            except ValueError:
                break
            block_data = by_date.get(d, {}).get(sec_id, {})
            if block_data:
                prev_map = block_data.get("prev") or {}
            else:
                prev_map = {}
            breaker_rows = []
            for br in breakers:
                mid = br["meter_id"]
                prev_v = rolling_prev.get(mid)
                if block_data and mid in prev_map and prev_map.get(mid) not in (None, ""):
                    pv = _parse_num(prev_map.get(mid))
                    if pv is not None:
                        prev_v = pv
                row, next_prev = _breaker_row_from_daily(block_data, br, prev_v)
                breaker_rows.append(row)
                if row["usage"] != "":
                    usage_sums[mid] += float(row["usage"])
                if row["max_a"] != "":
                    mv = float(row["max_a"])
                    cur = max_a_peaks[mid]
                    max_a_peaks[mid] = mv if cur is None else max(cur, mv)
                if next_prev is not None:
                    rolling_prev[mid] = next_prev
            days.append({"day": day, "date": d.isoformat(), "breakers": breaker_rows})
        totals = []
        for br in breakers:
            mid = br["meter_id"]
            u_sum = usage_sums[mid]
            peak = max_a_peaks[mid]
            totals.append(
                {
                    "meter_id": mid,
                    "usage_sum": round(u_sum, 2) if u_sum else "",
                    "max_a_max": peak if peak is not None else "",
                }
            )
        sections_out.append(
            {
                "id": sec_id,
                "title": sec["title"],
                "breakers": breakers,
                "prev_day": prev_day_cells,
                "days": days,
                "totals": totals,
            }
        )
    return {"year": year, "month": month, "sections": sections_out}


def build_yearly_sections(schema: dict | None = None) -> list[dict[str, Any]]:
    """월보와 동일한 설비·열 구성."""
    schema = schema or load_schema()
    roman = {"no1": "Ⅰ", "no2": "Ⅱ", "no3": "Ⅲ"}
    order = {"no1": 0, "no2": 1, "no3": 2}
    blocks = sorted(schema.get("daily_blocks") or [], key=lambda b: order.get(b["id"], 99))
    sections: list[dict[str, Any]] = []
    for block in blocks:
        bid = block["id"]
        breakers = []
        for m in block.get("meters") or []:
            breakers.append(
                {
                    "meter_id": m["id"],
                    "name": m["name"],
                    "multiplier": m.get("multiplier") or 4800,
                    "reading_col": m["reading_col"],
                    "current_col": m.get("current_col"),
                }
            )
        sections.append(
            {
                "id": bid,
                "title": f"주택변전소 년보[{roman.get(bid, bid)}]",
                "breakers": breakers,
            }
        )
    return sections


def _end_of_month_readings(
    year: int,
    month: int,
    sec_id: str,
    breakers: list[dict],
    rows_by_date: dict[date, HousingSubstationDaily],
) -> dict[str, Any]:
    """해당 월 말일(또는 마지막 입력일) 22:00 지침."""
    import calendar

    out = {br["meter_id"]: "" for br in breakers}
    last_day = calendar.monthrange(year, month)[1]
    for day in range(last_day, 0, -1):
        dt = date(year, month, day)
        row = rows_by_date.get(dt)
        if not row:
            continue
        block_data = (row.data or {}).get(sec_id) or {}
        t2200 = (block_data.get("times") or {}).get("22:00") or {}
        found = False
        for br in breakers:
            val = t2200.get(br["reading_col"], "")
            if val not in (None, ""):
                out[br["meter_id"]] = val
                found = True
        if found:
            break
    return out


def compute_yearly_report(
    building_id: int,
    year: int,
    daily_rows: list[HousingSubstationDaily],
    prev_year_last_row: HousingSubstationDaily | None = None,
) -> dict[str, Any]:
    """년보: 전년(12/31) · 1~12월 · 계(사용량 합·최대A)."""
    del building_id
    rows_by_date = {r.log_date: r for r in daily_rows}
    if prev_year_last_row:
        rows_by_date[prev_year_last_row.log_date] = prev_year_last_row
    sections_out = []
    for sec in build_yearly_sections():
        sec_id = sec["id"]
        breakers = sec["breakers"]
        prev_year_data = (
            (prev_year_last_row.data or {}).get(sec_id, {}) if prev_year_last_row else {}
        )
        prev_year_cells: list[dict[str, Any]] = []
        for br in breakers:
            val = _parse_num(_meter_reading_at_2200(prev_year_data, br))
            prev_year_cells.append(
                {
                    "meter_id": br["meter_id"],
                    "name": br["name"],
                    "multiplier": br["multiplier"],
                    "reading": val if val is not None else "",
                    "usage": "",
                    "max_a": "",
                }
            )
        months: list[dict[str, Any]] = []
        usage_sums_year: dict[str, float] = {b["meter_id"]: 0.0 for b in breakers}
        max_a_year: dict[str, float | None] = {b["meter_id"]: None for b in breakers}
        for month in range(1, 13):
            month_rows = [r for r in daily_rows if r.log_date.year == year and r.log_date.month == month]
            prev_month_last = date(year, month, 1) - timedelta(days=1)
            prev_month_row = rows_by_date.get(prev_month_last)
            monthly = compute_monthly_report(
                0, year, month, month_rows, prev_month_row
            )
            msec = next((s for s in monthly["sections"] if s["id"] == sec_id), None)
            end_readings = _end_of_month_readings(year, month, sec_id, breakers, rows_by_date)
            breaker_rows = []
            if msec:
                for i, br in enumerate(breakers):
                    mid = br["meter_id"]
                    tot = msec["totals"][i]
                    usage = tot.get("usage_sum", "")
                    max_a = tot.get("max_a_max", "")
                    breaker_rows.append(
                        {
                            "meter_id": mid,
                            "name": br["name"],
                            "multiplier": br["multiplier"],
                            "reading": end_readings.get(mid, ""),
                            "usage": usage,
                            "max_a": max_a,
                        }
                    )
                    if usage != "":
                        usage_sums_year[mid] += float(usage)
                    if max_a != "":
                        mv = float(max_a)
                        cur = max_a_year[mid]
                        max_a_year[mid] = mv if cur is None else max(cur, mv)
            else:
                for br in breakers:
                    breaker_rows.append(
                        {
                            "meter_id": br["meter_id"],
                            "name": br["name"],
                            "multiplier": br["multiplier"],
                            "reading": end_readings.get(br["meter_id"], ""),
                            "usage": "",
                            "max_a": "",
                        }
                    )
            months.append({"month": month, "breakers": breaker_rows})
        totals = []
        for br in breakers:
            mid = br["meter_id"]
            u_sum = usage_sums_year[mid]
            peak = max_a_year[mid]
            totals.append(
                {
                    "meter_id": mid,
                    "usage_sum": round(u_sum, 2) if u_sum else "",
                    "max_a_max": peak if peak is not None else "",
                }
            )
        sections_out.append(
            {
                "id": sec_id,
                "title": sec["title"],
                "breakers": breakers,
                "prev_year": prev_year_cells,
                "months": months,
                "totals": totals,
            }
        )
    return {"year": year, "sections": sections_out}


def _norm_name(s: str) -> str:
    return re.sub(r"\s+", "", (s or "").upper().replace(".", ""))


def _breaker_match(daily_name: str, monthly_name: str) -> bool:
    a, b = _norm_name(daily_name), _norm_name(monthly_name)
    if a == b:
        return True
    if a in b or b in a:
        return True
    ma = re.search(r"NO\s*(\d+)", a)
    mb = re.search(r"NO\s*(\d+)", b)
    if ma and mb and ma.group(1) == mb.group(1) and "INCOMMING" in a and "INCOMMING" in b:
        return True
    return False


def _monthly_excel_block_rows() -> dict[str, dict[str, int]]:
    return {
        "no1": {"header": 5, "prev": 8, "day1": 9, "total": 40},
        "no2": {"header": 44, "prev": 47, "day1": 48, "total": 79},
        "no3": {"header": 83, "prev": 86, "day1": 87, "total": 118},
    }


def _excel_cols_for_meter(sec_id: str, meter: dict, br_idx: int) -> dict[str, str] | None:
    schema = load_schema()
    schema_sec = next((s for s in schema.get("monthly_sections") or [] if s["id"] == sec_id), None)
    if not schema_sec:
        return None
    for sb in schema_sec.get("breakers") or []:
        if _breaker_match(meter.get("name") or "", sb.get("name") or ""):
            return sb.get("cols") or {}
    breakers = schema_sec.get("breakers") or []
    if br_idx < len(breakers):
        return breakers[br_idx].get("cols") or {}
    return None


def export_monthly_to_excel(monthly_report: dict) -> bytes:
    """월보 집계 결과를 템플릿 월보 시트에 채워 반환."""
    if not TEMPLATE_XLSX.is_file():
        wb = load_workbook()
        ws = wb.active
        ws.title = "월보"
    else:
        wb = load_workbook(TEMPLATE_XLSX)
        ws = wb["월보"] if "월보" in wb.sheetnames else wb.active

    schema = load_schema()
    block_rows = _monthly_excel_block_rows()
    for sec in monthly_report.get("sections") or []:
        sec_id = sec["id"]
        rows = block_rows.get(sec_id)
        if not rows:
            continue
        daily_block = next((b for b in schema["daily_blocks"] if b["id"] == sec_id), None)
        if not daily_block:
            continue
        prev_row = rows["prev"]
        day1_row = rows["day1"]
        total_row = rows["total"]
        for br_idx, br in enumerate(sec.get("breakers") or []):
            mid = br.get("meter_id")
            daily_meter = next((m for m in daily_block["meters"] if m["id"] == mid), br)
            cols = _excel_cols_for_meter(sec_id, daily_meter, br_idx)
            if not cols:
                continue
            rc, uc, mc = cols.get("reading"), cols.get("usage"), cols.get("max_a")
            if not rc or not uc or not mc:
                continue
            prev_cell = (sec.get("prev_day") or [])[br_idx] if br_idx < len(sec.get("prev_day") or []) else {}
            if prev_cell.get("reading") not in (None, ""):
                ws[f"{rc}{prev_row}"] = prev_cell["reading"]
            for day_row in sec.get("days") or []:
                excel_row = day1_row + int(day_row["day"]) - 1
                dr = (day_row.get("breakers") or [])[br_idx] if br_idx < len(day_row.get("breakers") or []) else {}
                if dr.get("reading") not in (None, ""):
                    ws[f"{rc}{excel_row}"] = dr["reading"]
                if dr.get("usage") not in (None, ""):
                    ws[f"{uc}{excel_row}"] = dr["usage"]
                if dr.get("max_a") not in (None, ""):
                    ws[f"{mc}{excel_row}"] = dr["max_a"]
            tot = (sec.get("totals") or [])[br_idx] if br_idx < len(sec.get("totals") or []) else {}
            if tot.get("usage_sum") not in (None, ""):
                ws[f"{uc}{total_row}"] = tot["usage_sum"]
            if tot.get("max_a_max") not in (None, ""):
                ws[f"{mc}{total_row}"] = tot["max_a_max"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_yearly_to_excel(yearly_report: dict) -> bytes:
    """년보 집계 결과를 년보 시트로 생성."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if TEMPLATE_XLSX.is_file():
        wb = load_workbook(TEMPLATE_XLSX)
    else:
        wb = load_workbook()
    if "년보" in wb.sheetnames:
        del wb["년보"]
    ws = wb.create_sheet("년보")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="DCE6F1")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    green_fill = PatternFill("solid", fgColor="C6EFCE")

    row_ptr = 1
    for sec in yearly_report.get("sections") or []:
        ws.cell(row_ptr, 1, sec.get("title", "")).font = Font(bold=True)
        row_ptr += 1
        breakers = sec.get("breakers") or []
        col = 2
        for br in breakers:
            c = ws.cell(row_ptr, col, br.get("name", ""))
            c.font = Font(bold=True)
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=row_ptr, start_column=col, end_row=row_ptr, end_column=col + 2)
            col += 3
        row_ptr += 1
        col = 2
        for br in breakers:
            ws.cell(row_ptr, col, f"적산량×{br.get('multiplier', '')}")
            ws.merge_cells(start_row=row_ptr, start_column=col, end_row=row_ptr, end_column=col + 2)
            col += 3
        row_ptr += 1
        ws.cell(row_ptr, 1, "구분").font = Font(bold=True)
        col = 2
        for _br in breakers:
            for label in ("지침(KW)", "사용량(KWH)", "최대(A)"):
                ws.cell(row_ptr, col, label).font = Font(bold=True)
                col += 1
        row_ptr += 1
        ws.cell(row_ptr, 1, "전년")
        col = 2
        for cell in sec.get("prev_year") or []:
            c = ws.cell(row_ptr, col, cell.get("reading", ""))
            c.fill = yellow_fill
            col += 3
        row_ptr += 1
        for mon in sec.get("months") or []:
            ws.cell(row_ptr, 1, f"{mon['month']}월")
            col = 2
            for br in mon.get("breakers") or []:
                ws.cell(row_ptr, col, br.get("reading", ""))
                ws.cell(row_ptr, col + 1, br.get("usage", ""))
                ws.cell(row_ptr, col + 2, br.get("max_a", ""))
                col += 3
            row_ptr += 1
        ws.cell(row_ptr, 1, "계").font = Font(bold=True)
        col = 2
        for tot in sec.get("totals") or []:
            ws.cell(row_ptr, col + 1, tot.get("usage_sum", "")).fill = green_fill
            ws.cell(row_ptr, col + 2, tot.get("max_a_max", "")).fill = green_fill
            col += 3
        row_ptr += 2

    max_col = ws.max_column or 2
    for r in ws.iter_rows(min_row=1, max_row=max(1, row_ptr - 1), min_col=1, max_col=max_col):
        for cell in r:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def fetch_monthly_report_data(
    session: AsyncSession,
    building_id: int,
    year: int,
    month: int,
) -> dict[str, Any]:
    import calendar

    last_day = calendar.monthrange(year, month)[1]
    d_from = date(year, month, 1)
    d_to = date(year, month, last_day)
    daily_rows = (
        await session.execute(
            select(HousingSubstationDaily).where(
                HousingSubstationDaily.building_id == building_id,
                HousingSubstationDaily.log_date >= d_from,
                HousingSubstationDaily.log_date <= d_to,
            )
        )
    ).scalars().all()
    prev_month_last = d_from - timedelta(days=1)
    prev_month_row = (
        await session.execute(
            select(HousingSubstationDaily).where(
                HousingSubstationDaily.building_id == building_id,
                HousingSubstationDaily.log_date == prev_month_last,
            )
        )
    ).scalar_one_or_none()
    return compute_monthly_report(
        building_id, year, month, list(daily_rows), prev_month_row
    )


async def fetch_yearly_report_data(
    session: AsyncSession,
    building_id: int,
    year: int,
) -> dict[str, Any]:
    d_from = date(year, 1, 1)
    d_to = date(year, 12, 31)
    daily_rows = (
        await session.execute(
            select(HousingSubstationDaily).where(
                HousingSubstationDaily.building_id == building_id,
                HousingSubstationDaily.log_date >= d_from,
                HousingSubstationDaily.log_date <= d_to,
            )
        )
    ).scalars().all()
    prev_year_last = date(year - 1, 12, 31)
    prev_year_row = (
        await session.execute(
            select(HousingSubstationDaily).where(
                HousingSubstationDaily.building_id == building_id,
                HousingSubstationDaily.log_date == prev_year_last,
            )
        )
    ).scalar_one_or_none()
    return compute_yearly_report(
        building_id, year, list(daily_rows), prev_year_row
    )


async def list_archives(
    session: AsyncSession, building_id: int, limit: int = 60
) -> list[HousingSubstationArchive]:
    rows = (
        await session.execute(
            select(HousingSubstationArchive)
            .where(HousingSubstationArchive.building_id == building_id)
            .order_by(HousingSubstationArchive.log_date.desc())
            .limit(limit)
        )
    ).scalars().all()
    return list(rows)


async def archive_daily_excel(
    session: AsyncSession,
    building_id: int,
    log_date: date,
) -> HousingSubstationArchive | None:
    """해당 일자 1일 시트를 엑셀으로 저장."""
    row = await get_daily_row(session, building_id, log_date)
    if not row:
        return None
    existing = (
        await session.execute(
            select(HousingSubstationArchive).where(
                HousingSubstationArchive.building_id == building_id,
                HousingSubstationArchive.log_date == log_date,
            )
        )
    ).scalar_one_or_none()
    if existing:
        return existing

    xbytes = export_daily_to_excel(row.data or {}, log_date)
    fname = f"주택변전소_1일_{log_date.isoformat()}.xlsx"
    arch = HousingSubstationArchive(
        building_id=building_id,
        log_date=log_date,
        original_name=fname,
        file_data=xbytes,
        file_size=len(xbytes),
    )
    session.add(arch)
    await session.flush()
    return arch


def export_daily_to_excel(data: dict, log_date: date) -> bytes:
    """템플릿 xlsx에 당일 입력값을 채워 반환."""
    if not TEMPLATE_XLSX.is_file():
        wb = load_workbook()
        ws = wb.active
        ws.title = "1일"
        ws["A1"] = log_date.isoformat()
    else:
        wb = load_workbook(TEMPLATE_XLSX)
        ws = wb["1일"] if "1일" in wb.sheetnames else wb.active
        ws["A1"] = log_date.day

    schema = load_schema()
    block_starts = {"no3": 7, "no2": 20, "no1": 32}
    time_index = {"06:00": 0, "10:00": 1, "14:00": 2, "17:00": 3, "20:00": 4, "22:00": 5}

    for block in schema["daily_blocks"]:
        bid = block["id"]
        start = block_starts.get(bid, 7)
        block_data = data.get(bid) or {}
        times = block_data.get("times") or {}
        prev_map = block_data.get("prev") or {}
        all_cols = {c["col"] for c in block.get("input_columns") or []}
        for m in block.get("meters") or []:
            all_cols.add(m["reading_col"])
        prev_row = start - 1
        for m in block.get("meters") or []:
            val = prev_map.get(m["id"])
            if val not in (None, ""):
                ws[f"{m['reading_col']}{prev_row}"] = val
        for t, idx in time_index.items():
            row = start + idx
            cells = times.get(t) or {}
            for col_letter in all_cols:
                val = cells.get(col_letter)
                if val not in (None, ""):
                    ws[f"{col_letter}{row}"] = val

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def rollover_at_midnight(session: AsyncSession, building_id: int, closing_date: date) -> None:
    """당일 23:59 마감: 해당 일자 엑셀 아카이브 후 다음 날 일지 생성."""
    await archive_daily_excel(session, building_id, closing_date)
    await get_or_create_daily(session, building_id, closing_date + timedelta(days=1))
    await session.commit()


async def ensure_tables(engine) -> None:
    from sqlalchemy import text

    url = str(engine.url).lower()
    is_pg = "postgresql" in url or "postgres" in url
    async with engine.begin() as conn:
        if is_pg:
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS housing_substation_daily (
                        id SERIAL PRIMARY KEY,
                        building_id INTEGER NOT NULL REFERENCES buildings(id),
                        log_date DATE NOT NULL,
                        data JSONB NOT NULL DEFAULT '{}',
                        created_at TIMESTAMP WITHOUT TIME ZONE DEFAULT NOW(),
                        updated_at TIMESTAMP WITHOUT TIME ZONE DEFAULT NOW(),
                        UNIQUE (building_id, log_date)
                    )
                    """
                )
            )
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS housing_substation_archives (
                        id SERIAL PRIMARY KEY,
                        building_id INTEGER NOT NULL REFERENCES buildings(id),
                        log_date DATE NOT NULL,
                        original_name VARCHAR(300),
                        file_data BYTEA,
                        file_size INTEGER,
                        created_at TIMESTAMP WITHOUT TIME ZONE DEFAULT NOW(),
                        UNIQUE (building_id, log_date)
                    )
                    """
                )
            )
        else:
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS housing_substation_daily (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        building_id INTEGER NOT NULL,
                        log_date DATE NOT NULL,
                        data TEXT NOT NULL DEFAULT '{}',
                        created_at DATETIME,
                        updated_at DATETIME,
                        UNIQUE (building_id, log_date)
                    )
                    """
                )
            )
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS housing_substation_archives (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        building_id INTEGER NOT NULL,
                        log_date DATE NOT NULL,
                        original_name VARCHAR(300),
                        file_data BLOB,
                        file_size INTEGER,
                        created_at DATETIME,
                        UNIQUE (building_id, log_date)
                    )
                    """
                )
            )


def register_scheduler(scheduler, session_factory, kst) -> None:
    """매일 23:59 KST — 주택변전소 1일 마감."""

    async def _daily_job():
        from sqlalchemy import select

        async with session_factory() as session:
            try:
                buildings = (
                    await session.execute(select(Building).where(Building.is_active == True))  # noqa: E712
                ).scalars().all()
                target = load_schema().get("building_name", "주택변전소")
                today = datetime.now(kst).date()
                for b in buildings:
                    if not b.name or target not in b.name:
                        continue
                    try:
                        await rollover_at_midnight(session, b.id, today)
                    except Exception as e:
                        print(f"[housing_sub] rollover building={b.id}: {e}", flush=True)
            except Exception as e:
                print(f"[housing_sub] daily scheduler: {e}", flush=True)

    scheduler.add_job(
        _daily_job,
        trigger="cron",
        hour=23,
        minute=59,
        id="housing_substation_rollover",
        replace_existing=True,
    )
