# main.py
from __future__ import annotations

import json
import os
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.exception_handlers import http_exception_handler
from fastapi.responses import JSONResponse, RedirectResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.middleware.sessions import SessionMiddleware

from auth import (
    MENU_ITEMS,
    ROLE_LABELS,
    SIGNUP_ROLES,
    apply_role_permissions,
    can_access_equipment_pm,
    can_access_menu,
    can_create,
    can_delete,
    can_edit,
    default_menu_access,
    default_permissions,
    effective_menu_access,
    get_current_user,
    group_buildings_by_site,
    hash_password,
    home_path_for_user,
    menu_key_for_path,
    normalize_menu_access,
    require_can_create,
    require_can_delete,
    require_can_edit,
    require_login,
    require_user_manager,
    verify_password,
)
from database import AsyncSessionLocal, Base, engine, get_db, ensure_schema_updates
from init_data import seed_if_empty
import onlyoffice as oo
from models import (
    Building,
    BuildingDrawing,
    BuildingStandard,
    Consumable,
    D1Plan,
    D1Status,
    Equipment,
    EquipmentChangeLog,
    EquipmentTemplate,
    EquipmentType,
    Floor,
    InspectionLogBuilding,
    InspectionLogFile,
    MaintenanceRecord,
    MaterialItem,
    MaterialGroup,
    MaterialLog,
    Partner,
    PMFrequency,
    PMInspection,
    PMResult,
    PMSchedule,
    Site,
    User,
    UserRole,
    WorkOrder,
    WorkOrderStatus,
    Zone,
)

KST = ZoneInfo("Asia/Seoul")

# 도면·표준서 공통 업로드 제한
UPLOAD_MAX_FILE_MB = 20
UPLOAD_MAX_FILE_BYTES = UPLOAD_MAX_FILE_MB * 1024 * 1024
UPLOAD_MAX_FILES_PER_REQUEST = 10


def _today_kst() -> date:
    return datetime.now(KST).date()


def _building_sort_key(name: str | None) -> tuple:
    """건물명 가나다 → ABC → 기타 순."""
    from auth import nav_building_sort_key

    return nav_building_sort_key(name)


def _sort_buildings(buildings: list) -> list:
    return sorted(buildings, key=lambda b: _building_sort_key(getattr(b, "name", None)))


def _fmt_kst(dt: datetime | None) -> str:
    if dt is None:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(KST).strftime("%Y-%m-%d %H:%M:%S")


def _fmt_kst_date(dt: datetime | date | None) -> str:
    """등록일 등 연-월-일만 표시."""
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(KST).strftime("%Y-%m-%d")
    return dt.strftime("%Y-%m-%d")


def _fmt_file_size(n: int | None) -> str:
    """바이트 수를 B / KB / MB로 표시."""
    if n is None:
        return "-"
    try:
        n = int(n)
    except (TypeError, ValueError):
        return "-"
    if n < 0:
        return "-"
    if n < 1024:
        return f"{n} B"
    if n < 1024 * 1024:
        kb = n / 1024
        return f"{kb:.1f} KB" if kb < 10 else f"{int(round(kb))} KB"
    mb = n / (1024 * 1024)
    return f"{mb:.2f} MB" if mb < 10 else f"{mb:.1f} MB"


def _sync_attachment_file_size(obj) -> bool:
    """file_size가 비어 있으면 file_data 길이로 채운다. 변경 시 True."""
    if getattr(obj, "file_size", None) is not None:
        return False
    data = getattr(obj, "file_data", None)
    if data is None:
        return False
    obj.file_size = len(data)
    return True


async def _backfill_attachment_sizes(db: AsyncSession, items) -> None:
    dirty = False
    for obj in items or []:
        if _sync_attachment_file_size(obj):
            dirty = True
    if dirty:
        try:
            await db.commit()
        except Exception:
            await db.rollback()


async def _load_change_logs_by_eq(
    db: AsyncSession, eq_ids: list[int], per_eq: int = 20
) -> dict[int, list]:
    """설비별 최근 변경 로그."""
    if not eq_ids:
        return {}
    rows = (
        await db.execute(
            select(EquipmentChangeLog)
            .where(EquipmentChangeLog.equipment_id.in_(eq_ids))
            .order_by(EquipmentChangeLog.changed_at.desc())
        )
    ).scalars().all()
    by_eq: dict[int, list] = {}
    for row in rows:
        bucket = by_eq.setdefault(row.equipment_id, [])
        if len(bucket) < per_eq:
            bucket.append(row)
    return by_eq


async def _record_equipment_change(
    db: AsyncSession,
    eq: Equipment,
    before: dict[str, str],
    user: User | None = None,
    source: str = "수정",
) -> EquipmentChangeLog | None:
    """사양 스냅샷 비교 후 변경이 있으면 로그 저장."""
    from equipment_schema import (
        diff_equipment_snapshots,
        equipment_snapshot,
        summarize_equipment_changes,
    )

    after = equipment_snapshot(eq)
    changes = diff_equipment_snapshots(before, after)
    if not changes:
        return None
    summary = f"[{source}] {summarize_equipment_changes(changes)}"
    changer = None
    if user is not None:
        changer = (getattr(user, "name", None) or getattr(user, "username", None) or "").strip() or None
    log = EquipmentChangeLog(
        equipment_id=eq.id,
        changed_by=changer,
        summary=summary[:300],
        changes=changes,
    )
    db.add(log)
    return log


def _status_label(status: WorkOrderStatus | str) -> str:
    labels = {
        "received": "정비의뢰",
        "assigned": "정비의뢰",
        "in_progress": "정비진행",
        "completed": "정비완료",
        "verified": "정비완료",
        "closed": "정비완료",
    }
    key = status.value if isinstance(status, WorkOrderStatus) else str(status)
    return labels.get(key, key)


def _pm_freq_label(freq: PMFrequency | str | None) -> str:
    labels = {
        "daily": "매일",
        "weekly": "매주",
        "monthly": "매월",
        "quarterly": "분기",
        "semi_annual": "반기",
        "annual": "연간",
        "custom": "사용자지정",
    }
    if freq is None:
        return "-"
    key = freq.value if isinstance(freq, PMFrequency) else str(freq)
    return labels.get(key, key)


def _pm_result_label(result: PMResult | str | None) -> str:
    labels = {"normal": "정상", "caution": "주의", "fault": "고장"}
    if result is None:
        return "-"
    key = result.value if isinstance(result, PMResult) else str(result)
    return labels.get(key, key)


def _pm_cycle_days(freq: PMFrequency | str | None, custom_days: int | None) -> int:
    key = freq.value if isinstance(freq, PMFrequency) else str(freq or "monthly")
    if key == "custom":
        return max(1, int(custom_days or 30))
    mapping = {
        "daily": 1,
        "weekly": 7,
        "monthly": 30,
        "quarterly": 90,
        "semi_annual": 180,
        "annual": 365,
    }
    return mapping.get(key, 30)


def _pm_advance_schedule(schedule: PMSchedule, done: date | None = None) -> None:
    done = done or _today_kst()
    days = _pm_cycle_days(schedule.frequency, schedule.custom_days)
    schedule.last_done = done
    schedule.next_due = done + timedelta(days=days)


def _parse_pm_frequency(raw: str) -> PMFrequency:
    try:
        return PMFrequency(raw)
    except ValueError:
        return PMFrequency.monthly


def _equipment_building(eq: Equipment | None) -> Building | None:
    if not eq or not eq.zone or not eq.zone.floor:
        return None
    return eq.zone.floor.building


def _equipment_pm_inspections_map(equipment_list: list) -> dict[int, list[dict]]:
    """설비별 점검 이력(팝업·기간 필터용)."""
    out: dict[int, list[dict]] = {}
    for eq in equipment_list:
        insps = sorted(
            eq.pm_inspections or [],
            key=lambda x: x.inspected_at or datetime.min,
            reverse=True,
        )
        out[eq.id] = [
            {
                "id": i.id,
                "at": _fmt_kst(i.inspected_at) if i.inspected_at else "",
                "at_iso": i.inspected_at.isoformat() if i.inspected_at else "",
                "result": i.result.value if i.result else "normal",
                "result_label": _pm_result_label(i.result),
                "note": i.note or "",
                "inspector": i.inspector_name or "",
                "title": (i.schedule.title if i.schedule else "예방점검"),
                "work_order_id": i.work_order_id,
            }
            for i in insps
        ]
    return out


def _wo_process_step(status: WorkOrderStatus | str) -> int:
    """1=정비의뢰, 2=정비중, 3=정비완료."""
    key = status.value if isinstance(status, WorkOrderStatus) else str(status)
    if key in ("completed", "verified", "closed"):
        return 3
    if key == "in_progress":
        return 2
    return 1


def _parse_wo_status_filters(status: list[str] | None) -> list[str]:
    allowed = {"received", "in_progress", "completed"}
    out: list[str] = []
    seen: set[str] = set()
    for raw in status or []:
        for part in str(raw).split(","):
            key = part.strip()
            if key in allowed and key not in seen:
                seen.add(key)
                out.append(key)
    return out


LIST_PAGE_SIZE = 20
D1_BOARD_KEYS = ("today", "tomorrow", "scheduled", "completed")
D1_ALL_BOARD_KEYS = ("receipt",) + D1_BOARD_KEYS


def _parse_d1_boards(raw: list[str] | None, *, allow_receipt: bool = False) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    allowed = set(D1_ALL_BOARD_KEYS if allow_receipt else D1_BOARD_KEYS)
    for item in raw or []:
        for part in str(item).split(","):
            key = part.strip().lower()
            if key == "all":
                return list(D1_BOARD_KEYS)
            if key in allowed and key not in seen:
                seen.add(key)
                out.append(key)
    return out


def _d1_resolve_boards(
    request: Request,
    board_param: list[str] | str | None,
    *,
    partner_id: int = 0,
) -> tuple[list[str], str]:
    """단일 선택: board 미지정 시 일반=오늘작업, 협력사 선택=오늘접수. board=all 이면 전체항목."""
    allow_receipt = bool(partner_id)
    if "board" not in request.query_params:
        if allow_receipt:
            return ["receipt"], "receipt"
        return ["today"], "today"
    if isinstance(board_param, str):
        raw = [board_param]
    else:
        raw = list(board_param or [])
    flat: list[str] = []
    for item in raw:
        for part in str(item).split(","):
            key = part.strip().lower()
            if key:
                flat.append(key)
    if "all" in flat:
        return list(D1_BOARD_KEYS), "all"
    parsed = _parse_d1_boards(raw, allow_receipt=allow_receipt)
    if not parsed:
        if allow_receipt:
            return ["receipt"], "receipt"
        return ["today"], "today"
    # 단일 선택 — 첫 번째만 사용
    mode = parsed[0]
    return [mode], mode


def _d1_board_select_urls(
    status_vals: list[str],
    partner_id: int = 0,
) -> dict[str, str]:
    """작업 구분 단일 선택 URL (전체/오늘접수/오늘/내일/예정/완료)."""
    from urllib.parse import urlencode

    keys = ("all",) + (D1_ALL_BOARD_KEYS if partner_id else D1_BOARD_KEYS)
    links: dict[str, str] = {}
    for key in keys:
        params: list[tuple[str, str]] = [("board", key)]
        params.extend(("status", s) for s in status_vals)
        if partner_id:
            params.append(("partner_id", str(partner_id)))
        links[key] = "/admin/d1?" + urlencode(params)
    return links


def _paginate(items: list, page: int | str | None, per_page: int = LIST_PAGE_SIZE) -> dict:
    """목록 페이지네이션 정보 생성."""
    total = len(items)
    total_pages = max(1, (total + per_page - 1) // per_page) if total else 1
    try:
        page_n = int(page or 1)
    except (TypeError, ValueError):
        page_n = 1
    page_n = max(1, min(page_n, total_pages))
    start = (page_n - 1) * per_page
    window = 5
    start_p = max(1, page_n - window // 2)
    end_p = min(total_pages, start_p + window - 1)
    start_p = max(1, end_p - window + 1)
    return {
        "items": items[start : start + per_page],
        "page": page_n,
        "per_page": per_page,
        "total": total,
        "total_pages": total_pages,
        "page_numbers": list(range(start_p, end_p + 1)),
        "has_prev": page_n > 1,
        "has_next": page_n < total_pages,
    }


def _wo_status_sql_filter(status_vals: list[str]):
    if not status_vals:
        return None
    status_conds = []
    if "received" in status_vals:
        status_conds.append(
            WorkOrder.status.in_([WorkOrderStatus.received, WorkOrderStatus.assigned])
        )
    if "in_progress" in status_vals:
        status_conds.append(WorkOrder.status == WorkOrderStatus.in_progress)
    if "completed" in status_vals:
        status_conds.append(
            WorkOrder.status.in_(
                [
                    WorkOrderStatus.completed,
                    WorkOrderStatus.verified,
                    WorkOrderStatus.closed,
                ]
            )
        )
    if not status_conds:
        return None
    return or_(*status_conds)


def _wo_matches_status_filter(wo: WorkOrder, status_vals: list[str]) -> bool:
    if not status_vals:
        return True
    key = wo.status.value if isinstance(wo.status, WorkOrderStatus) else str(wo.status)
    if "received" in status_vals and key in ("received", "assigned"):
        return True
    if "in_progress" in status_vals and key == "in_progress":
        return True
    if "completed" in status_vals and key in ("completed", "verified", "closed"):
        return True
    return False


async def _apply_user_company(
    db: AsyncSession,
    target: User,
    *,
    company_partner_id: str = "",
    company_name: str = "",
) -> None:
    """협력사 선택 또는 직접입력 회사명을 User에 반영."""
    raw = (company_partner_id or "").strip()
    custom = (company_name or "").strip()[:200]
    if raw in ("", "0"):
        target.partner_id = None
        target.company_name = None
        return
    if raw == "__custom__":
        target.partner_id = None
        target.company_name = custom or None
        return
    if raw.isdigit():
        pid = int(raw)
        partner = await db.get(Partner, pid)
        if partner and getattr(partner, "is_active", True):
            target.partner_id = partner.id
            target.company_name = (partner.name or custom or "")[:200] or None
            return
    target.partner_id = None
    target.company_name = custom or None


def _wo_created_date_kst(wo: WorkOrder) -> date | None:
    dt = getattr(wo, "created_at", None)
    if dt is None:
        return None
    if isinstance(dt, datetime):
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(KST).date()
    if isinstance(dt, date):
        return dt
    return None


def _wo_is_today_partner_receipt(wo: WorkOrder, today: date | None = None) -> bool:
    """당일 접수 + 협력사 지정된 정비의뢰."""
    if not getattr(wo, "partner_id", None):
        return False
    key = wo.status.value if isinstance(wo.status, WorkOrderStatus) else str(wo.status)
    if key not in ("received", "assigned"):
        return False
    day = today or _today_kst()
    return _wo_created_date_kst(wo) == day


def _wo_is_d1_eligible(wo: WorkOrder) -> bool:
    """업체 지정 + D-1 승인된 정비만 D-1 보드에 표시."""
    if not getattr(wo, "partner_id", None):
        return False
    return bool(getattr(wo, "d1_approved", False))


def _wo_d1_sql_gate():
    """D-1 노출 SQL 조건."""
    return (
        WorkOrder.partner_id.is_not(None),
        WorkOrder.d1_approved.is_(True),
    )


def _wo_approver_label(user: User) -> str:
    name = (getattr(user, "name", None) or "").strip()
    uname = (getattr(user, "username", None) or "").strip()
    if name and uname and name != uname:
        return f"{name} ({uname})"[:100]
    return (name or uname or f"user-{user.id}")[:100]


def _wo_apply_d1_approve(wo: WorkOrder, approver: str, now: datetime | None = None) -> None:
    """정비섹션 D-1 승인 (협력사 작업허가 승인요청은 별도)."""
    ts = now or datetime.utcnow()
    wo.d1_approved = True
    wo.approved_by = approver
    wo.approved_at = ts


def _wo_apply_d1_unapprove(wo: WorkOrder) -> None:
    """D-1 승인 취소 시 하위 승인요청·작업허가도 함께 해제."""
    wo.d1_approved = False
    wo.approved_by = None
    wo.approved_at = None
    wo.approval_requested = False
    wo.approval_requested_by = None
    wo.approval_requested_at = None
    wo.work_permitted = False
    wo.work_permitted_by = None
    wo.work_permitted_at = None


def _wo_clear_auto_approval_request(wo: WorkOrder) -> bool:
    """정비섹션 승인 시 잘못 자동 반영된 작업허가 승인요청을 되돌린다. 변경 시 True."""
    if not getattr(wo, "d1_approved", False):
        return False
    if not getattr(wo, "approval_requested", False):
        return False
    if getattr(wo, "work_permitted", False):
        return False
    req_at = getattr(wo, "approval_requested_at", None)
    appr_at = getattr(wo, "approved_at", None)
    # 동일 시각으로 함께 세팅된 자동 반영분만 해제
    if req_at is None or appr_at is None or req_at != appr_at:
        return False
    wo.approval_requested = False
    wo.approval_requested_by = None
    wo.approval_requested_at = None
    return True


def _wo_list_redirect_params(
    *,
    q: str = "",
    filter_status: str = "",
    filter_priority: str = "",
    date_from: str = "",
    date_to: str = "",
    page: str = "",
    filter_partner_id: str = "",
    message: str = "",
    error: str = "",
) -> str:
    from urllib.parse import urlencode

    params = {
        k: v
        for k, v in {
            "q": (q or "").strip(),
            "status": (filter_status or "").strip(),
            "priority": (filter_priority or "").strip(),
            "date_from": (date_from or "").strip(),
            "date_to": (date_to or "").strip(),
            "partner_id": (filter_partner_id or "").strip()
            if (filter_partner_id or "").strip() not in ("", "0")
            else "",
            "page": (page or "").strip() if str(page or "").strip() not in ("", "1") else "",
            "message": (message or "").strip(),
            "error": (error or "").strip(),
        }.items()
        if v
    }
    return f"?{urlencode(params)}" if params else ""


def _sort_orders_today_receipt_first(orders: list, today: date | None = None) -> list:
    day = today or _today_kst()

    def _key(w: WorkOrder):
        pin = 0 if _wo_is_today_partner_receipt(w, day) else 1
        created = getattr(w, "created_at", None) or datetime.min
        ts = created.timestamp() if isinstance(created, datetime) else 0.0
        return (pin, -ts, -(getattr(w, "id", 0) or 0))

    return sorted(orders, key=_key)


def _partner_groups_for_orders(orders: list, today: date | None = None) -> list[dict]:
    day = today or _today_kst()
    by_partner: dict[str, list] = {}
    for w in orders:
        name = w.partner.name if getattr(w, "partner", None) else "미지정"
        by_partner.setdefault(name, []).append(w)
    groups: list[dict] = []
    for name in sorted(by_partner.keys(), key=lambda n: (n == "미지정", n)):
        items = _sort_orders_today_receipt_first(by_partner[name], day)
        groups.append(
            {
                "name": name,
                "orders": items,
                "today_count": sum(1 for w in items if _wo_is_today_partner_receipt(w, day)),
            }
        )
    return groups


def _work_orders_excel_response(orders: list, filename_prefix: str = "정비목록"):
    from io import BytesIO
    from urllib.parse import quote

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "정비목록"
    ws.append(
        [
            "설비코드",
            "설비명",
            "정비의뢰내용",
            "등록일",
            "예정일",
            "우선순위",
            "협력사",
            "진행",
            "조치내용",
            "작업 승인자",
            "D-1승인",
            "담당자",
        ]
    )
    for wo in orders:
        eq = wo.equipment
        ws.append(
            [
                eq.code if eq else "",
                eq.name if eq else "",
                (wo.description or wo.title or "").strip(),
                _fmt_kst_date(wo.created_at),
                wo.scheduled_date.isoformat() if wo.scheduled_date else "",
                "긴급" if wo.priority == "high" else "일반",
                wo.partner.name if wo.partner else "미지정",
                _status_label(wo.status),
                wo.action or "",
                getattr(wo, "approved_by", None) or "",
                "Y" if getattr(wo, "d1_approved", False) else "N",
                wo.assignee_name or "",
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now(KST).strftime("%Y%m%d_%H%M")
    filename = quote(f"{filename_prefix}_{stamp}.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


async def _ensure_maintenance_history(db: AsyncSession, wo: WorkOrder) -> None:
    """정비완료 시 설비 정비이력 자동 등록 (중복 방지)."""
    if not wo.equipment_id:
        return
    existing = (
        await db.execute(
            select(MaintenanceRecord).where(MaintenanceRecord.work_order_id == wo.id)
        )
    ).scalar_one_or_none()
    if existing:
        return
    db.add(
        MaintenanceRecord(
            equipment_id=wo.equipment_id,
            work_order_id=wo.id,
            title=wo.title,
            work_date=(wo.completed_at or datetime.utcnow()).date(),
            worker_name=wo.assignee_name,
            cause=wo.cause,
            action=wo.action,
            parts_used=wo.parts_used,
            work_hours=wo.work_hours,
            cost=wo.cost,
            is_manual=False,
        )
    )


def _d1_status_label(status: D1Status) -> str:
    return {
        D1Status.draft: "작성중",
        D1Status.review: "검토",
        D1Status.approved: "승인",
        D1Status.jsa_pending: "JSA 대기",
        D1Status.tbm_pending: "TBM 대기",
        D1Status.permit_pending: "작업허가 대기",
        D1Status.in_progress: "작업중",
        D1Status.completed: "완료",
    }.get(status, status.value)


async def _startup_db_init() -> None:
    """기동 후 백그라운드에서 DB 스키마/시드 준비."""
    import os as _os

    _os.environ.setdefault("LAW_WEB_SEARCH", "0")
    last_err: Exception | None = None
    for attempt in range(1, 6):
        try:
            async with engine.begin() as conn:
                await conn.run_sync(Base.metadata.create_all)
            await ensure_schema_updates()
            async with AsyncSessionLocal() as session:
                await seed_if_empty(session)
                from excel_import import backfill_all_building_default_categories

                await backfill_all_building_default_categories(session)
            print(f"[startup] DB ready (attempt {attempt})", flush=True)
            return
        except Exception as e:
            last_err = e
            print(f"[startup] DB init failed ({attempt}/5): {e}", flush=True)
            if attempt < 5:
                import asyncio

                await asyncio.sleep(3)
    print(
        f"[startup] WARNING: continuing without full DB init: {last_err}",
        flush=True,
    )


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Render health check가 막히지 않도록 DB 초기화는 백그라운드로 수행."""
    import asyncio
    import os as _os

    _os.environ.setdefault("LAW_WEB_SEARCH", "0")
    init_task = asyncio.create_task(_startup_db_init())
    try:
        yield
    finally:
        if not init_task.done():
            init_task.cancel()
            try:
                await init_task
            except asyncio.CancelledError:
                pass
            except Exception as e:
                print(f"[startup] background init cancel: {e}", flush=True)


app = FastAPI(title="POSCO WIDE Smart FMS", lifespan=lifespan)


class _MenuAccessMiddleware(BaseHTTPMiddleware):
    """메뉴 권한이 없는 /admin/* 경로는 내 계정으로 리다이렉트."""

    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        menu_key = menu_key_for_path(path)
        if menu_key is None:
            return await call_next(request)
        if "session" not in request.scope:
            return await call_next(request)
        user_id = request.session.get("user_id")
        if not user_id:
            return await call_next(request)
        try:
            async with AsyncSessionLocal() as session:
                u = (
                    await session.execute(
                        select(User).where(
                            User.id == user_id,
                            User.is_active == True,  # noqa: E712
                            User.is_approved == True,  # noqa: E712
                        )
                    )
                ).scalar_one_or_none()
                if u is not None and not can_access_menu(u, menu_key):
                    return RedirectResponse("/admin/account?error=no_menu", status_code=303)
        except Exception as e:
            print(f"[menu] access check skip: {e}", flush=True)
        return await call_next(request)


SECRET_KEY = os.environ.get("APP_SECRET_KEY", "change_this_secret_in_prod")
_session_kw: dict = {"secret_key": SECRET_KEY, "same_site": "lax"}
if os.environ.get("RENDER", "").lower() in ("true", "1", "yes") or os.environ.get(
    "COOKIE_HTTPS_ONLY", ""
).lower() in ("1", "true", "yes"):
    _session_kw["https_only"] = True
# 안쪽(메뉴) → Session → Proxy 순으로 add (마지막이 가장 바깥)
app.add_middleware(_MenuAccessMiddleware)
app.add_middleware(SessionMiddleware, **_session_kw)

try:
    from starlette.middleware.proxy_headers import ProxyHeadersMiddleware

    app.add_middleware(ProxyHeadersMiddleware, trusted_hosts="*")
except ImportError:
    pass

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

from equipment_schema import field_value, get_category_fields, list_display_fields, NAME_KEYS

templates.env.globals["field_value"] = field_value
templates.env.globals["name_fields"] = set(NAME_KEYS)
templates.env.globals["user_can_create"] = can_create
templates.env.globals["user_can_edit"] = can_edit
templates.env.globals["user_can_delete"] = can_delete
templates.env.globals["user_can_access_equipment_pm"] = can_access_equipment_pm
templates.env.globals["user_can_access_menu"] = can_access_menu
templates.env.globals["user_menu_access"] = effective_menu_access
templates.env.globals["menu_items"] = MENU_ITEMS
templates.env.globals["upload_max_mb"] = UPLOAD_MAX_FILE_MB
templates.env.globals["upload_max_files"] = UPLOAD_MAX_FILES_PER_REQUEST
templates.env.globals.update(
    fmt_kst=_fmt_kst,
    fmt_kst_date=_fmt_kst_date,
    fmt_file_size=_fmt_file_size,
    role_labels=ROLE_LABELS,
    wo_status_label=_status_label,
    wo_process_step=_wo_process_step,
    d1_status_label=_d1_status_label,
    pm_freq_label=_pm_freq_label,
    pm_result_label=_pm_result_label,
)


@app.exception_handler(HTTPException)
async def _http_exception_handler(request: Request, exc: HTTPException):
    # 미로그인 → 로그인 페이지로 이동
    if exc.status_code in (401, 303) and (
        exc.detail == "login_required"
        or (exc.headers or {}).get("Location") == "/admin/login"
        or (exc.headers or {}).get("X-Redirect") == "/admin/login"
    ):
        return RedirectResponse("/admin/login", status_code=303)
    return await http_exception_handler(request, exc)


@app.exception_handler(Exception)
async def _unhandled_exception_handler(request: Request, exc: Exception):
    import traceback

    # HTTPException은 전용 핸들러로
    if isinstance(exc, HTTPException):
        return await _http_exception_handler(request, exc)

    print(f"[error] {request.method} {request.url.path}: {exc}", flush=True)
    traceback.print_exc()
    accept = (request.headers.get("accept") or "").lower()
    if "text/html" in accept or str(request.url.path).startswith("/admin"):
        try:
            return templates.TemplateResponse(
                request,
                "error.html",
                {
                    "user": None,
                    "status_code": 500,
                    "message": "일시적인 오류가 발생했습니다. 잠시 후 다시 시도해 주세요.",
                    "detail": str(exc)[:500],
                },
                status_code=500,
            )
        except Exception:
            pass
    return JSONResponse(
        {"detail": "internal_error", "message": str(exc)[:300]},
        status_code=500,
    )


@app.get("/health")
async def health():
    return {"status": "ok", "service": "smart-fms"}


# ── Auth ──────────────────────────────────────────────────────────────


@app.get("/admin/login")
async def admin_login_page(request: Request, user: User | None = Depends(get_current_user)):
    if user:
        return RedirectResponse(home_path_for_user(user), status_code=303)
    return templates.TemplateResponse(
        request,
        "login.html",
        {
            "error": request.query_params.get("error"),
            "message": request.query_params.get("message"),
        },
    )


@app.post("/admin/login")
async def admin_login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(User).where(User.username == username.strip()))
    user = result.scalar_one_or_none()
    if not user or not user.is_active or not verify_password(password, user.password_hash):
        return RedirectResponse("/admin/login?error=1", status_code=303)
    if not getattr(user, "is_approved", True):
        return RedirectResponse("/admin/login?error=pending", status_code=303)
    request.session["user_id"] = user.id
    return RedirectResponse(home_path_for_user(user), status_code=303)


@app.get("/admin/logout")
async def admin_logout(request: Request):
    request.session.clear()
    return RedirectResponse("/admin/login", status_code=303)


@app.get("/admin/signup")
async def admin_signup_page(
    request: Request,
    user: User | None = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if user:
        return RedirectResponse("/admin/dashboard", status_code=303)
    partners = (
        await db.execute(
            select(Partner).where(Partner.is_active == True).order_by(Partner.name)
        )
    ).scalars().all()
    return templates.TemplateResponse(
        request,
        "signup.html",
        {
            "error": request.query_params.get("error"),
            "signup_roles": SIGNUP_ROLES,
            "partners": partners,
        },
    )


@app.post("/admin/signup")
async def admin_signup(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    password2: str = Form(...),
    name: str = Form(...),
    role: str = Form("facility_manager"),
    phone: str = Form(""),
    email: str = Form(""),
    company_partner_id: str = Form(""),
    company_name: str = Form(""),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    uname = username.strip()
    nm = name.strip()
    pw = password.strip()
    if not uname or not nm or not pw:
        return RedirectResponse("/admin/signup?error=required", status_code=303)
    if len(pw) < 6:
        return RedirectResponse("/admin/signup?error=short", status_code=303)
    if pw != password2.strip():
        return RedirectResponse("/admin/signup?error=mismatch", status_code=303)
    if not all(c.isalnum() or c in "._-" for c in uname):
        return RedirectResponse("/admin/signup?error=username", status_code=303)
    exists = (
        await db.execute(select(User).where(User.username == uname))
    ).scalar_one_or_none()
    if exists:
        return RedirectResponse("/admin/signup?error=exists", status_code=303)

    allowed = {r.value for r in SIGNUP_ROLES}
    try:
        role_val = UserRole(role) if role in allowed else UserRole.facility_manager
    except ValueError:
        role_val = UserRole.facility_manager

    new_user = User(
        username=uname,
        password_hash=hash_password(pw),
        name=nm,
        phone=phone.strip() or None,
        email=email.strip() or None,
        role=role_val,
        is_active=True,
        is_approved=False,
    )
    await _apply_user_company(
        db,
        new_user,
        company_partner_id=company_partner_id,
        company_name=company_name,
    )
    apply_role_permissions(new_user)
    db.add(new_user)
    await db.commit()
    return RedirectResponse(
        "/admin/login?message=" + quote("가입 신청이 접수되었습니다. 관리자 승인 후 로그인할 수 있습니다."),
        status_code=303,
    )


@app.get("/admin/account")
async def account_page(
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    db_user = (
        await db.execute(
            select(User)
            .where(User.id == user.id)
            .options(selectinload(User.partner))
        )
    ).scalar_one_or_none() or user
    partners = (
        await db.execute(
            select(Partner).where(Partner.is_active == True).order_by(Partner.name)
        )
    ).scalars().all()
    return templates.TemplateResponse(
        request,
        "account.html",
        {
            "user": db_user,
            "partners": partners,
            "error": request.query_params.get("error"),
            "message": request.query_params.get("message"),
        },
    )


@app.post("/admin/account/profile")
async def account_update_profile(
    name: str = Form(...),
    phone: str = Form(""),
    email: str = Form(""),
    company_partner_id: str = Form(""),
    company_name: str = Form(""),
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    db_user = await db.get(User, user.id)
    if not db_user:
        raise HTTPException(404)
    nm = name.strip()
    if not nm:
        return RedirectResponse("/admin/account?error=required", status_code=303)
    db_user.name = nm
    db_user.phone = phone.strip() or None
    db_user.email = email.strip() or None
    await _apply_user_company(
        db,
        db_user,
        company_partner_id=company_partner_id,
        company_name=company_name,
    )
    await db.commit()
    return RedirectResponse(
        "/admin/account?message=" + quote("개인정보가 저장되었습니다."),
        status_code=303,
    )


@app.post("/admin/account/password")
async def account_change_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    new_password2: str = Form(...),
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    db_user = await db.get(User, user.id)
    if not db_user:
        raise HTTPException(404)
    if not verify_password(current_password, db_user.password_hash):
        return RedirectResponse("/admin/account?error=current", status_code=303)
    if len(new_password.strip()) < 6:
        return RedirectResponse("/admin/account?error=short", status_code=303)
    if new_password.strip() != new_password2.strip():
        return RedirectResponse("/admin/account?error=mismatch", status_code=303)
    db_user.password_hash = hash_password(new_password.strip())
    await db.commit()
    return RedirectResponse(
        "/admin/account?message=" + quote("비밀번호가 변경되었습니다."),
        status_code=303,
    )


# ── Users / Account Admin ─────────────────────────────────────────────


def _menu_keys_from_form(form, role: UserRole | None = None) -> list[str]:
    """multipart/form에서 menu_key 다중 체크박스 값 수집."""
    try:
        raw = form.getlist("menu_key")
    except Exception:
        v = form.get("menu_key")
        raw = [v] if v else []
    keys = normalize_menu_access(list(raw))
    # 계정관리 메뉴는 시스템관리자만 (라우트도 system_admin 전용)
    if role != UserRole.system_admin:
        keys = [k for k in keys if k != "users"]
    return keys


def _force_admin_menus(user_obj: User) -> None:
    if user_obj.role == UserRole.system_admin:
        user_obj.can_create = user_obj.can_edit = user_obj.can_delete = True
        user_obj.menu_access = list(default_menu_access(UserRole.system_admin))
    else:
        # 비관리자 계정에 users 키가 남지 않도록 정리
        keys = normalize_menu_access(getattr(user_obj, "menu_access", None) or [])
        user_obj.menu_access = [k for k in keys if k != "users"]


@app.get("/admin/users")
async def users_manage_page(
    request: Request,
    user: User = Depends(require_user_manager),
    db: AsyncSession = Depends(get_db),
):
    rows = (
        await db.execute(
            select(User)
            .options(selectinload(User.partner))
            .order_by(User.is_approved.asc(), User.created_at.desc())
        )
    ).scalars().unique().all()
    pending = [u for u in rows if not u.is_approved and u.is_active]
    active = [u for u in rows if u.is_approved and u.is_active]
    inactive = [u for u in rows if not u.is_active]
    partners = (
        await db.execute(
            select(Partner).where(Partner.is_active == True).order_by(Partner.name)
        )
    ).scalars().all()
    return templates.TemplateResponse(
        request,
        "users.html",
        {
            "user": user,
            "pending_users": pending,
            "active_users": active,
            "inactive_users": inactive,
            "partners": partners,
            "roles": list(UserRole),
            "menu_items": MENU_ITEMS,
            "error": request.query_params.get("error"),
            "message": request.query_params.get("message"),
        },
    )


@app.post("/admin/users/create")
async def users_create(
    username: str = Form(...),
    password: str = Form(...),
    name: str = Form(...),
    role: str = Form("viewer"),
    phone: str = Form(""),
    email: str = Form(""),
    company_partner_id: str = Form(""),
    company_name: str = Form(""),
    can_create_flag: str = Form(""),
    can_edit_flag: str = Form(""),
    can_delete_flag: str = Form(""),
    user: User = Depends(require_user_manager),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    uname = username.strip()
    nm = name.strip()
    pw = password.strip()
    if not uname or not nm or not pw:
        return RedirectResponse("/admin/users?error=required", status_code=303)
    if len(pw) < 6:
        return RedirectResponse("/admin/users?error=short", status_code=303)
    if (await db.execute(select(User).where(User.username == uname))).scalar_one_or_none():
        return RedirectResponse("/admin/users?error=exists", status_code=303)
    try:
        role_val = UserRole(role)
    except ValueError:
        role_val = UserRole.viewer
    new_user = User(
        username=uname,
        password_hash=hash_password(pw),
        name=nm,
        role=role_val,
        phone=phone.strip() or None,
        email=email.strip() or None,
        is_active=True,
        is_approved=True,
        can_create=can_create_flag == "1",
        can_edit=can_edit_flag == "1",
        can_delete=can_delete_flag == "1",
        menu_access=default_menu_access(role_val),
    )
    await _apply_user_company(
        db,
        new_user,
        company_partner_id=company_partner_id,
        company_name=company_name,
    )
    _force_admin_menus(new_user)
    db.add(new_user)
    await db.commit()
    return RedirectResponse(
        "/admin/users?message=" + quote(f"계정 {uname} 이(가) 추가되었습니다."),
        status_code=303,
    )


@app.post("/admin/users/{uid}/approve")
async def users_approve(
    uid: int,
    role: str = Form("facility_manager"),
    company_partner_id: str = Form(""),
    company_name: str = Form(""),
    can_create_flag: str = Form(""),
    can_edit_flag: str = Form(""),
    can_delete_flag: str = Form(""),
    user: User = Depends(require_user_manager),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    target = await db.get(User, uid)
    if not target:
        raise HTTPException(404)
    try:
        role_val = UserRole(role)
    except ValueError:
        role_val = UserRole.facility_manager
    target.role = role_val
    target.is_approved = True
    target.is_active = True
    target.can_create = can_create_flag == "1"
    target.can_edit = can_edit_flag == "1"
    target.can_delete = can_delete_flag == "1"
    target.menu_access = default_menu_access(role_val)
    await _apply_user_company(
        db,
        target,
        company_partner_id=company_partner_id,
        company_name=company_name,
    )
    _force_admin_menus(target)
    await db.commit()
    return RedirectResponse(
        "/admin/users?message=" + quote(f"{target.username} 승인이 완료되었습니다."),
        status_code=303,
    )


@app.post("/admin/users/{uid}/reject")
async def users_reject(
    uid: int,
    user: User = Depends(require_user_manager),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    target = await db.get(User, uid)
    if not target:
        raise HTTPException(404)
    if target.id == user.id:
        return RedirectResponse("/admin/users?error=self", status_code=303)
    target.is_active = False
    target.is_approved = False
    await db.commit()
    return RedirectResponse(
        "/admin/users?message=" + quote(f"{target.username} 가입이 거절되었습니다."),
        status_code=303,
    )


@app.post("/admin/users/{uid}/update")
async def users_update(
    request: Request,
    uid: int,
    name: str = Form(...),
    role: str = Form(...),
    phone: str = Form(""),
    email: str = Form(""),
    company_partner_id: str = Form(""),
    company_name: str = Form(""),
    can_create_flag: str = Form(""),
    can_edit_flag: str = Form(""),
    can_delete_flag: str = Form(""),
    is_active_flag: str = Form(""),
    user: User = Depends(require_user_manager),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    target = await db.get(User, uid)
    if not target:
        raise HTTPException(404)
    try:
        role_val = UserRole(role)
    except ValueError:
        role_val = target.role
    # 자기 자신의 시스템관리자 역할/활성은 유지
    if target.id == user.id and role_val != UserRole.system_admin:
        return RedirectResponse("/admin/users?error=self_role", status_code=303)
    target.name = name.strip() or target.name
    target.role = role_val
    target.phone = phone.strip() or None
    target.email = email.strip() or None
    await _apply_user_company(
        db,
        target,
        company_partner_id=company_partner_id,
        company_name=company_name,
    )
    target.can_create = can_create_flag == "1"
    target.can_edit = can_edit_flag == "1"
    target.can_delete = can_delete_flag == "1"
    if role_val == UserRole.system_admin:
        target.can_create = target.can_edit = target.can_delete = True
        target.menu_access = default_menu_access(UserRole.system_admin)
    if target.id != user.id:
        target.is_active = is_active_flag == "1"
        if target.is_active:
            target.is_approved = True
    await db.commit()
    return RedirectResponse(
        "/admin/users?message=" + quote(f"{target.username} 정보가 저장되었습니다."),
        status_code=303,
    )


@app.post("/admin/users/{uid}/menu-access")
async def users_menu_access(
    request: Request,
    uid: int,
    user: User = Depends(require_user_manager),
    db: AsyncSession = Depends(get_db),
):
    """하단 메뉴 접근 설정 전용 저장."""
    from urllib.parse import quote

    form = await request.form()
    target = await db.get(User, uid)
    if not target:
        raise HTTPException(404)
    target.menu_access = _menu_keys_from_form(form, target.role)
    _force_admin_menus(target)
    await db.commit()
    return RedirectResponse(
        "/admin/users?message="
        + quote(f"{target.username} 메뉴 접근이 저장되었습니다.")
        + f"#menu-access-{target.id}",
        status_code=303,
    )


@app.post("/admin/users/{uid}/password")
async def users_reset_password(
    uid: int,
    new_password: str = Form(...),
    new_password2: str = Form(...),
    user: User = Depends(require_user_manager),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    target = await db.get(User, uid)
    if not target:
        raise HTTPException(404)
    if len(new_password.strip()) < 6:
        return RedirectResponse("/admin/users?error=short", status_code=303)
    if new_password.strip() != new_password2.strip():
        return RedirectResponse("/admin/users?error=mismatch", status_code=303)
    target.password_hash = hash_password(new_password.strip())
    await db.commit()
    return RedirectResponse(
        "/admin/users?message=" + quote(f"{target.username} 비밀번호가 변경되었습니다."),
        status_code=303,
    )


@app.post("/admin/users/{uid}/delete")
async def users_delete(
    uid: int,
    user: User = Depends(require_user_manager),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    target = await db.get(User, uid)
    if not target:
        raise HTTPException(404)
    if target.id == user.id:
        return RedirectResponse("/admin/users?error=self", status_code=303)
    target.is_active = False
    await db.commit()
    return RedirectResponse(
        "/admin/users?message=" + quote(f"{target.username} 계정이 비활성화되었습니다."),
        status_code=303,
    )


# ── Dashboard ─────────────────────────────────────────────────────────

_WO_OPEN = (
    WorkOrderStatus.received,
    WorkOrderStatus.assigned,
    WorkOrderStatus.in_progress,
)
_WO_DONE = (
    WorkOrderStatus.completed,
    WorkOrderStatus.verified,
    WorkOrderStatus.closed,
)
_WO_REQUEST = (WorkOrderStatus.received, WorkOrderStatus.assigned)


def _kst_day_naive_utc_range(day: date) -> tuple[datetime, datetime]:
    """KST 하루를 naive UTC datetime 구간으로 변환 (completed_at 비교용)."""
    start_kst = datetime(day.year, day.month, day.day, tzinfo=KST)
    start = start_kst.astimezone(timezone.utc).replace(tzinfo=None)
    end = start + timedelta(days=1)
    return start, end


def _demo_energy_payload(today: date | None = None) -> dict:
    """에너지 사용현황 예시 데이터 (점검일지 연동 전 데모)."""
    import random

    day = today or _today_kst()
    rng = random.Random(int(day.strftime("%Y%m%d")))
    labels = [(day - timedelta(days=i)).strftime("%m/%d") for i in range(6, -1, -1)]

    def series(base: float, variance: float) -> list[float]:
        return [round(base + rng.uniform(-variance, variance), 1) for _ in labels]

    kepco = series(8200, 450)
    receive = series(6100, 380)
    rolling = series(2400, 220)
    medium = series(185, 18)  # 중온 (t/h 예시)
    water = series(920, 70)  # 급수 (m³ 예시)
    return {
        "is_demo": True,
        "note": "예시 데이터입니다. 추후 점검일지와 연동하여 자동 집계·그래프화합니다.",
        "as_of": day.isoformat(),
        "labels": labels,
        "power": {
            "unit": "kWh",
            "kepco": kepco,
            "receive": receive,
            "rolling": rolling,
            "today": {
                "kepco": kepco[-1],
                "receive": receive[-1],
                "rolling": rolling[-1],
            },
        },
        "medium_temp": {
            "unit": "t/h",
            "values": medium,
            "today": medium[-1],
        },
        "water": {
            "unit": "m³",
            "values": water,
            "today": water[-1],
        },
    }


async def _compute_dashboard_kpi(db: AsyncSession) -> dict:
    """대시보드 전 구역 KPI (KST 기준, 활성 데이터 연동)."""
    today = _today_kst()
    yesterday = today - timedelta(days=1)
    today_start, today_end = _kst_day_naive_utc_range(today)

    sites = (
        await db.execute(select(func.count(Site.id)).where(Site.is_active == True))  # noqa: E712
    ).scalar() or 0
    buildings = (
        await db.execute(
            select(func.count(Building.id))
            .join(Site, Building.site_id == Site.id)
            .where(Building.is_active == True, Site.is_active == True)  # noqa: E712
        )
    ).scalar() or 0
    equipment = (
        await db.execute(
            select(func.count(Equipment.id))
            .join(Zone, Equipment.zone_id == Zone.id)
            .join(Floor, Zone.floor_id == Floor.id)
            .join(Building, Floor.building_id == Building.id)
            .join(Site, Building.site_id == Site.id)
            .where(
                Equipment.is_active == True,  # noqa: E712
                Building.is_active == True,  # noqa: E712
                Site.is_active == True,  # noqa: E712
            )
        )
    ).scalar() or 0

    # 점검: 금일 계획(next_due==today) / 금일 완료(last_done==today) / 전일 실적
    pm_today_plan = (
        await db.execute(
            select(func.count(PMSchedule.id)).where(
                PMSchedule.is_active == True,  # noqa: E712
                PMSchedule.next_due == today,
            )
        )
    ).scalar() or 0
    pm_today_done = (
        await db.execute(
            select(func.count(PMSchedule.id)).where(
                PMSchedule.is_active == True,  # noqa: E712
                PMSchedule.last_done == today,
            )
        )
    ).scalar() or 0
    pm_yesterday_done = (
        await db.execute(
            select(func.count(PMSchedule.id)).where(
                PMSchedule.is_active == True,  # noqa: E712
                PMSchedule.last_done == yesterday,
            )
        )
    ).scalar() or 0
    pm_overdue = (
        await db.execute(
            select(func.count(PMSchedule.id)).where(
                PMSchedule.is_active == True,  # noqa: E712
                PMSchedule.next_due.is_not(None),
                PMSchedule.next_due < today,
            )
        )
    ).scalar() or 0

    # 정비관리: 의뢰 / 진행 / 완료 (누적 활성)
    wo_request = (
        await db.execute(
            select(func.count(WorkOrder.id)).where(
                WorkOrder.is_active == True,  # noqa: E712
                WorkOrder.status.in_(_WO_REQUEST),
            )
        )
    ).scalar() or 0
    wo_progress = (
        await db.execute(
            select(func.count(WorkOrder.id)).where(
                WorkOrder.is_active == True,  # noqa: E712
                WorkOrder.status == WorkOrderStatus.in_progress,
            )
        )
    ).scalar() or 0
    wo_done = (
        await db.execute(
            select(func.count(WorkOrder.id)).where(
                WorkOrder.is_active == True,  # noqa: E712
                WorkOrder.status.in_(_WO_DONE),
            )
        )
    ).scalar() or 0
    wo_urgent = (
        await db.execute(
            select(func.count(WorkOrder.id)).where(
                WorkOrder.is_active == True,  # noqa: E712
                WorkOrder.priority == "high",
                WorkOrder.status.in_(_WO_OPEN),
            )
        )
    ).scalar() or 0

    # D-1 (업체 지정 + 승인된 WorkOrder만)
    d1_gate = _wo_d1_sql_gate()
    d1_today_incomplete = (
        await db.execute(
            select(func.count(WorkOrder.id)).where(
                WorkOrder.is_active == True,  # noqa: E712
                WorkOrder.status.in_(_WO_OPEN),
                WorkOrder.scheduled_date == today,
                *d1_gate,
            )
        )
    ).scalar() or 0
    d1_today_done = (
        await db.execute(
            select(func.count(WorkOrder.id)).where(
                WorkOrder.is_active == True,  # noqa: E712
                WorkOrder.status.in_(_WO_DONE),
                *d1_gate,
                (
                    (WorkOrder.scheduled_date == today)
                    | (
                        WorkOrder.completed_at.is_not(None)
                        & (WorkOrder.completed_at >= today_start)
                        & (WorkOrder.completed_at < today_end)
                    )
                ),
            )
        )
    ).scalar() or 0
    d1_today_plan = int(d1_today_incomplete) + int(d1_today_done)

    energy = _demo_energy_payload(today)

    return {
        "as_of": today.isoformat(),
        "sites": int(sites),
        "buildings": int(buildings),
        "equipment": int(equipment),
        "pm_today_plan": int(pm_today_plan),
        "pm_today_done": int(pm_today_done),
        "pm_yesterday_done": int(pm_yesterday_done),
        "pm_overdue": int(pm_overdue),
        "wo_request": int(wo_request),
        "wo_progress": int(wo_progress),
        "wo_done": int(wo_done),
        "wo_urgent": int(wo_urgent),
        "d1_today_plan": int(d1_today_plan),
        "d1_today_done": int(d1_today_done),
        "d1_today_incomplete": int(d1_today_incomplete),
        "energy": energy,
    }


@app.get("/admin/dashboard")
async def dashboard(
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    kpi = await _compute_dashboard_kpi(db)

    return templates.TemplateResponse(
        request,
        "dashboard.html",
        {
            "user": user,
            "kpi": kpi,
        },
    )


@app.get("/admin/dashboard/kpi")
async def dashboard_kpi(
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    """대시보드 KPI JSON (자동 갱신)."""
    from fastapi.responses import JSONResponse

    return JSONResponse(await _compute_dashboard_kpi(db))


# ── Sites & Hierarchy ─────────────────────────────────────────────────


@app.get("/admin/sites")
async def sites_list(
    request: Request,
    error: str | None = None,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Site)
        .where(Site.is_active == True)
        .options(selectinload(Site.buildings))
        .order_by(Site.name)
    )
    sites = result.scalars().all()
    return templates.TemplateResponse(
        request, "sites.html", {"user": user, "sites": sites, "error": error or ""}
    )


@app.post("/admin/sites")
async def site_create(
    name: str = Form(...),
    code: str = Form(...),
    address: str = Form(""),
    manager_name: str = Form(""),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    site = Site(name=name.strip(), code=code.strip(), address=address, manager_name=manager_name)
    db.add(site)
    await db.commit()
    return RedirectResponse("/admin/sites", status_code=303)


@app.get("/admin/sites/{site_id}/edit")
async def site_edit_page(
    site_id: int,
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    site = await db.get(Site, site_id)
    if not site or not site.is_active:
        raise HTTPException(404)
    return templates.TemplateResponse(
        request, "site_edit.html", {"user": user, "site": site}
    )


@app.post("/admin/sites/{site_id}/edit")
async def site_edit(
    site_id: int,
    name: str = Form(...),
    code: str = Form(...),
    address: str = Form(""),
    manager_name: str = Form(""),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    site = await db.get(Site, site_id)
    if not site or not site.is_active:
        raise HTTPException(404)
    site.name = name.strip()
    site.code = code.strip()
    site.address = address
    site.manager_name = manager_name
    await db.commit()
    return RedirectResponse("/admin/sites", status_code=303)


@app.post("/admin/sites/{site_id}/delete")
async def site_delete(
    site_id: int,
    user: User = Depends(require_can_delete),
    db: AsyncSession = Depends(get_db),
):
    site = await db.get(Site, site_id)
    if not site:
        raise HTTPException(404)
    site.is_active = False
    result = await db.execute(select(Building).where(Building.site_id == site_id))
    for b in result.scalars().all():
        b.is_active = False
    await db.commit()
    return RedirectResponse("/admin/sites", status_code=303)


@app.get("/admin/buildings/{building_id}")
async def building_detail(
    building_id: int,
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
    message: str = "",
    error: str = "",
):
    result = await db.execute(
        select(Building)
        .where(Building.id == building_id)
        .options(
            selectinload(Building.site),
            selectinload(Building.floors).selectinload(Floor.zones),
            selectinload(Building.drawings).selectinload(BuildingDrawing.floor),
        )
    )
    building = result.scalar_one_or_none()
    if not building or not building.is_active:
        raise HTTPException(404)
    drawings = sorted(
        building.drawings or [],
        key=_attachment_sort_key,
        reverse=True,
    )
    try:
        standards = list(
            (
                await db.execute(
                    select(BuildingStandard)
                    .where(BuildingStandard.building_id == building_id)
                    .order_by(BuildingStandard.id.desc())
                )
            ).scalars().all()
        )
        standards = sorted(standards, key=_attachment_sort_key, reverse=True)
    except Exception as e:
        print(f"[standards] load skip building={building_id}: {e}", flush=True)
        try:
            await db.rollback()
        except Exception:
            pass
        standards = []
    await _backfill_attachment_sizes(db, drawings)
    await _backfill_attachment_sizes(db, standards)
    active_floors = sorted(
        [f for f in (building.floors or []) if getattr(f, "is_active", True)],
        key=lambda f: (f.level or 0, f.name or "", f.id),
    )
    return templates.TemplateResponse(
        request,
        "building_detail.html",
        {
            "user": user,
            "building": building,
            "drawings": drawings,
            "standards": standards,
            "active_floors": active_floors,
            "message": message,
            "error": error,
        },
    )


DRAWING_ALLOWED_EXT = {
    ".png",
    ".jpg",
    ".jpeg",
    ".gif",
    ".webp",
    ".bmp",
    ".pdf",
    ".dwg",
    ".dxf",
}

# 도면·표준서 공통 업로드 제한 (상단 UPLOAD_MAX_* 상수 사용)


def _attachment_sort_key(item) -> tuple:
    """created_at 정렬용 — aware/naive 혼용 방지."""
    created = getattr(item, "created_at", None)
    if isinstance(created, datetime):
        if created.tzinfo is not None:
            created = created.replace(tzinfo=None)
        ts = created.timestamp()
    else:
        ts = 0.0
    return (ts, int(getattr(item, "id", 0) or 0))


def _building_upload_dir(building_id: int) -> Path | None:
    path = Path("static") / "uploads" / "buildings" / str(building_id)
    try:
        path.mkdir(parents=True, exist_ok=True)
        return path
    except OSError as e:
        print(f"[upload] mkdir skip: {e}", flush=True)
        return None


async def _ensure_building_standards_table() -> None:
    """표준서 테이블만 빠르게 보장 (전체 스키마 재실행 금지 — 서비스 지연/잠금 방지)."""
    from sqlalchemy import text
    from sqlalchemy.exc import DBAPIError, OperationalError, ProgrammingError

    url = (os.environ.get("DATABASE_URL") or os.environ.get("DATABASE_INTERNAL_URL") or "").lower()
    is_pg = "postgres" in url
    if is_pg:
        stmts = [
            """
            CREATE TABLE IF NOT EXISTS building_standards (
                id SERIAL PRIMARY KEY,
                building_id INTEGER NOT NULL REFERENCES buildings(id),
                title VARCHAR(200) NOT NULL,
                original_name VARCHAR(300),
                stored_name VARCHAR(300) NOT NULL,
                content_type VARCHAR(100),
                file_data BYTEA,
                created_at TIMESTAMP WITHOUT TIME ZONE
            )
            """,
            "CREATE INDEX IF NOT EXISTS ix_building_standards_building_id ON building_standards (building_id)",
            "ALTER TABLE building_standards ADD COLUMN IF NOT EXISTS file_data BYTEA",
            "ALTER TABLE building_standards ADD COLUMN IF NOT EXISTS file_size INTEGER",
        ]
    else:
        stmts = [
            """
            CREATE TABLE IF NOT EXISTS building_standards (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                building_id INTEGER NOT NULL,
                title VARCHAR(200) NOT NULL,
                original_name VARCHAR(300),
                stored_name VARCHAR(300) NOT NULL,
                content_type VARCHAR(100),
                file_data BLOB,
                created_at DATETIME
            )
            """,
            "ALTER TABLE building_standards ADD COLUMN file_size INTEGER",
        ]
    for stmt in stmts:
        try:
            async with engine.begin() as conn:
                await conn.execute(text(stmt))
        except (OperationalError, ProgrammingError, DBAPIError) as e:
            print(f"[standards] schema skip: {e}", flush=True)


def _drawing_media_type(drawing: BuildingDrawing) -> str:
    ct = (drawing.content_type or "").strip()
    name = (drawing.original_name or drawing.stored_name or "").lower()
    if ct and ct.lower() not in {"application/octet-stream", "binary/octet-stream"}:
        return ct
    if name.endswith(".pdf"):
        return "application/pdf"
    if name.endswith((".png",)):
        return "image/png"
    if name.endswith((".jpg", ".jpeg")):
        return "image/jpeg"
    if name.endswith(".gif"):
        return "image/gif"
    if name.endswith(".webp"):
        return "image/webp"
    if name.endswith(".bmp"):
        return "image/bmp"
    if name.endswith(".dwg"):
        return "image/vnd.dwg"
    if name.endswith(".dxf"):
        return "image/vnd.dxf"
    return "application/octet-stream"


def _drawing_content_disposition(drawing: BuildingDrawing, *, as_attachment: bool = False) -> str:
    from urllib.parse import quote

    filename = drawing.original_name or drawing.stored_name or "drawing"
    # DWG/DXF는 브라우저에서 미리보기 불가 → 다운로드
    force_attach = as_attachment or not (drawing.is_image or drawing.is_pdf)
    kind = "attachment" if force_attach else "inline"
    ascii_name = filename.encode("ascii", "ignore").decode("ascii") or "drawing"
    return f"{kind}; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(filename)}"


async def _load_drawing_bytes(
    drawing: BuildingDrawing,
    building_id: int,
    db: AsyncSession,
) -> bytes | None:
    """DB 우선, 없으면 디스크에서 읽어 DB에 백필."""
    data = drawing.file_data
    if data:
        return bytes(data)

    file_path = Path("static") / "uploads" / "buildings" / str(building_id) / drawing.stored_name
    if file_path.is_file():
        raw = file_path.read_bytes()
        if raw:
            drawing.file_data = raw
            drawing.file_size = len(raw)
            await db.commit()
            return raw
    return None


@app.get("/admin/buildings/{building_id}/drawings/{drawing_id}/file")
async def building_drawing_file(
    building_id: int,
    drawing_id: int,
    download: int = Query(0),
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    drawing = await db.get(BuildingDrawing, drawing_id)
    if not drawing or drawing.building_id != building_id:
        raise HTTPException(404, detail="도면을 찾을 수 없습니다.")

    data = await _load_drawing_bytes(drawing, building_id, db)
    if not data:
        raise HTTPException(
            404,
            detail="도면 파일이 없습니다. 배포 환경에서는 파일이 유지되지 않을 수 있으니 다시 업로드해 주세요.",
        )

    return Response(
        content=data,
        media_type=_drawing_media_type(drawing),
        headers={
            "Content-Disposition": _drawing_content_disposition(
                drawing, as_attachment=bool(download)
            ),
            "Cache-Control": "private, max-age=3600",
        },
    )


@app.post("/admin/buildings/{building_id}/drawings")
async def building_drawing_upload(
    building_id: int,
    title: str = Form(""),
    floor_id: str = Form(""),
    files: list[UploadFile] = File(default=[]),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote
    import uuid

    building = await db.get(Building, building_id)
    if not building or not building.is_active:
        raise HTTPException(404)

    raw_files = files if isinstance(files, list) else ([files] if files else [])
    uploads = [f for f in raw_files if f and getattr(f, "filename", None)]
    if not uploads:
        return RedirectResponse(
            f"/admin/buildings/{building_id}?error={quote('도면 파일을 선택하세요.')}",
            status_code=303,
        )
    if len(uploads) > UPLOAD_MAX_FILES_PER_REQUEST:
        return RedirectResponse(
            f"/admin/buildings/{building_id}?error="
            + quote(f"한 번에 최대 {UPLOAD_MAX_FILES_PER_REQUEST}개까지 업로드할 수 있습니다."),
            status_code=303,
        )

    linked_floor_id = None
    if str(floor_id).strip().isdigit():
        fid = int(str(floor_id).strip())
        floor = await db.get(Floor, fid)
        if floor and floor.building_id == building_id and getattr(floor, "is_active", True):
            linked_floor_id = fid

    upload_dir = _building_upload_dir(building_id)
    saved = 0
    skipped = 0
    common_title = (title or "").strip()

    try:
        for f in uploads:
            original = Path(f.filename or "drawing").name
            suffix = Path(original).suffix.lower()
            if suffix not in DRAWING_ALLOWED_EXT:
                skipped += 1
                continue
            stored = f"{uuid.uuid4().hex}{suffix}"
            data = await f.read()
            if not data:
                skipped += 1
                continue
            if len(data) > UPLOAD_MAX_FILE_BYTES:
                skipped += 1
                continue
            if upload_dir is not None:
                try:
                    (upload_dir / stored).write_bytes(data)
                except OSError:
                    pass
            draw_title = common_title or Path(original).stem or "도면"
            db.add(
                BuildingDrawing(
                    building_id=building_id,
                    floor_id=linked_floor_id,
                    title=draw_title[:200],
                    original_name=original[:300],
                    stored_name=stored,
                    content_type=f.content_type or None,
                    file_data=bytes(data),
                    file_size=len(data),
                )
            )
            saved += 1

        if not saved:
            return RedirectResponse(
                f"/admin/buildings/{building_id}?error="
                + quote(
                    f"업로드 가능한 파일이 없습니다. (이미지/PDF/DWG/DXF, 파일당 최대 {UPLOAD_MAX_FILE_MB}MB, "
                    f"한 번에 {UPLOAD_MAX_FILES_PER_REQUEST}개)"
                ),
                status_code=303,
            )

        await db.commit()
    except Exception as e:
        await db.rollback()
        print(f"[drawings] upload failed building={building_id}: {e}", flush=True)
        return RedirectResponse(
            f"/admin/buildings/{building_id}?error="
            + quote(f"도면 업로드 중 오류: {str(e)[:180]}"),
            status_code=303,
        )

    msg = f"도면 {saved}건 첨부 완료"
    if skipped:
        msg += f" (제외 {skipped}건)"
    return RedirectResponse(
        f"/admin/buildings/{building_id}?message={quote(msg)}",
        status_code=303,
    )


@app.post("/admin/buildings/{building_id}/drawings/{drawing_id}/delete")
async def building_drawing_delete(
    building_id: int,
    drawing_id: int,
    user: User = Depends(require_can_delete),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    drawing = await db.get(BuildingDrawing, drawing_id)
    if not drawing or drawing.building_id != building_id:
        raise HTTPException(404)

    file_path = Path("static") / "uploads" / "buildings" / str(building_id) / drawing.stored_name
    await db.delete(drawing)
    await db.commit()
    try:
        if file_path.is_file():
            file_path.unlink()
    except OSError:
        pass

    return RedirectResponse(
        f"/admin/buildings/{building_id}?message={quote('도면이 삭제되었습니다.')}",
        status_code=303,
    )


STANDARD_ALLOWED_EXT = {
    ".pdf",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
    ".ppt",
    ".pptx",
    ".hwp",
    ".hwpx",
    ".txt",
    ".png",
    ".jpg",
    ".jpeg",
}


def _building_standards_dir(building_id: int) -> Path | None:
    path = Path("static") / "uploads" / "buildings" / str(building_id) / "standards"
    try:
        path.mkdir(parents=True, exist_ok=True)
        return path
    except OSError as e:
        print(f"[standards] mkdir skip: {e}", flush=True)
        return None


def _standard_media_type(doc: BuildingStandard) -> str:
    ct = (doc.content_type or "").strip()
    name = (doc.original_name or doc.stored_name or "").lower()
    if ct and ct.lower() not in {"application/octet-stream", "binary/octet-stream"}:
        return ct
    if name.endswith(".pdf"):
        return "application/pdf"
    if name.endswith(".doc"):
        return "application/msword"
    if name.endswith(".docx"):
        return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    if name.endswith(".xls"):
        return "application/vnd.ms-excel"
    if name.endswith(".xlsx"):
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if name.endswith(".ppt"):
        return "application/vnd.ms-powerpoint"
    if name.endswith(".pptx"):
        return "application/vnd.openxmlformats-officedocument.presentationml.presentation"
    if name.endswith(".txt"):
        return "text/plain"
    if name.endswith(".png"):
        return "image/png"
    if name.endswith((".jpg", ".jpeg")):
        return "image/jpeg"
    return "application/octet-stream"


def _standard_content_disposition(doc: BuildingStandard) -> str:
    from urllib.parse import quote

    filename = doc.original_name or doc.stored_name or "standard"
    ascii_name = filename.encode("ascii", "ignore").decode("ascii") or "standard"
    return f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{quote(filename)}'


async def _load_standard_bytes(
    doc: BuildingStandard,
    building_id: int,
    db: AsyncSession,
) -> bytes | None:
    data = doc.file_data
    if data:
        return bytes(data)
    file_path = (
        Path("static")
        / "uploads"
        / "buildings"
        / str(building_id)
        / "standards"
        / doc.stored_name
    )
    if file_path.is_file():
        raw = file_path.read_bytes()
        if raw:
            doc.file_data = raw
            doc.file_size = len(raw)
            await db.commit()
            return raw
    return None


@app.get("/admin/buildings/{building_id}/standards/{standard_id}/file")
async def building_standard_file(
    building_id: int,
    standard_id: int,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    doc = await db.get(BuildingStandard, standard_id)
    if not doc or doc.building_id != building_id:
        raise HTTPException(404, detail="표준서를 찾을 수 없습니다.")

    data = await _load_standard_bytes(doc, building_id, db)
    if not data:
        raise HTTPException(
            404,
            detail="표준서 파일이 없습니다. 다시 업로드해 주세요.",
        )

    return Response(
        content=data,
        media_type=_standard_media_type(doc),
        headers={
            "Content-Disposition": _standard_content_disposition(doc),
            "Cache-Control": "private, max-age=3600",
        },
    )


@app.post("/admin/buildings/{building_id}/standards")
async def building_standard_upload(
    building_id: int,
    request: Request,
    title: str = Form(""),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote
    import uuid

    building = await db.get(Building, building_id)
    if not building or not building.is_active:
        raise HTTPException(404)

    # 배포 DB에 테이블이 아직 없으면 표준서 테이블만 빠르게 생성
    try:
        await _ensure_building_standards_table()
    except Exception as e:
        print(f"[standards] schema ensure: {e}", flush=True)

    try:
        form = await request.form()
        raw_files = form.getlist("files")
    except Exception as e:
        print(f"[standards] form parse error: {e}", flush=True)
        return RedirectResponse(
            f"/admin/buildings/{building_id}?error="
            + quote(f"파일 전송 오류: {e}"),
            status_code=303,
        )

    uploads = [
        f
        for f in raw_files
        if f is not None and getattr(f, "filename", None) and hasattr(f, "read")
    ]
    if not uploads:
        return RedirectResponse(
            f"/admin/buildings/{building_id}?error={quote('표준서 파일을 선택하세요.')}",
            status_code=303,
        )
    if len(uploads) > UPLOAD_MAX_FILES_PER_REQUEST:
        return RedirectResponse(
            f"/admin/buildings/{building_id}?error="
            + quote(f"한 번에 최대 {UPLOAD_MAX_FILES_PER_REQUEST}개까지 업로드할 수 있습니다."),
            status_code=303,
        )

    upload_dir = _building_standards_dir(building_id)
    saved = 0
    skipped = 0
    common_title = (title or "").strip() or str(form.get("title") or "").strip()

    try:
        for f in uploads:
            original = Path(str(getattr(f, "filename", None) or "standard")).name
            suffix = Path(original).suffix.lower()
            if suffix not in STANDARD_ALLOWED_EXT:
                skipped += 1
                continue
            stored = f"{uuid.uuid4().hex}{suffix}"
            data = await f.read()
            if not data:
                skipped += 1
                continue
            if len(data) > UPLOAD_MAX_FILE_BYTES:
                skipped += 1
                continue
            if upload_dir is not None:
                try:
                    (upload_dir / stored).write_bytes(data)
                except OSError:
                    pass
            doc_title = common_title or Path(original).stem or "표준서"
            db.add(
                BuildingStandard(
                    building_id=building_id,
                    title=doc_title[:200],
                    original_name=original[:300],
                    stored_name=stored,
                    content_type=getattr(f, "content_type", None) or None,
                    file_data=bytes(data),
                    file_size=len(data),
                )
            )
            saved += 1

        if not saved:
            return RedirectResponse(
                f"/admin/buildings/{building_id}?error="
                + quote(
                    f"업로드 가능한 파일이 없습니다. (PDF/DOC/XLS/PPT/HWP/이미지, "
                    f"파일당 최대 {UPLOAD_MAX_FILE_MB}MB, 한 번에 {UPLOAD_MAX_FILES_PER_REQUEST}개)"
                ),
                status_code=303,
            )

        await db.commit()
    except Exception as e:
        await db.rollback()
        print(f"[standards] upload failed building={building_id}: {e}", flush=True)
        try:
            await _ensure_building_standards_table()
        except Exception:
            pass
        detail = str(e)
        if "building_standards" in detail.lower() or "does not exist" in detail.lower():
            msg = "표준서 저장 테이블 준비 중입니다. 잠시 후 다시 시도해 주세요."
        elif "memory" in detail.lower() or "too large" in detail.lower():
            msg = f"파일이 너무 큽니다. 파일당 {UPLOAD_MAX_FILE_MB}MB 이하로 올려 주세요."
        else:
            msg = f"표준서 업로드 중 오류: {detail[:180]}"
        return RedirectResponse(
            f"/admin/buildings/{building_id}?error={quote(msg)}",
            status_code=303,
        )

    msg = f"표준서 {saved}건 등록 완료"
    if skipped:
        msg += f" (제외 {skipped}건)"
    return RedirectResponse(
        f"/admin/buildings/{building_id}?message={quote(msg)}",
        status_code=303,
    )


@app.post("/admin/buildings/{building_id}/standards/{standard_id}/delete")
async def building_standard_delete(
    building_id: int,
    standard_id: int,
    user: User = Depends(require_can_delete),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    doc = await db.get(BuildingStandard, standard_id)
    if not doc or doc.building_id != building_id:
        raise HTTPException(404)

    file_path = (
        Path("static")
        / "uploads"
        / "buildings"
        / str(building_id)
        / "standards"
        / doc.stored_name
    )
    await db.delete(doc)
    await db.commit()
    try:
        if file_path.is_file():
            file_path.unlink()
    except OSError:
        pass

    return RedirectResponse(
        f"/admin/buildings/{building_id}?message={quote('표준서가 삭제되었습니다.')}",
        status_code=303,
    )


@app.post("/admin/buildings")
async def building_create(
    site_id: int = Form(...),
    name: str = Form(...),
    code: str = Form(...),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote
    from excel_import import ensure_building_default_categories

    name = name.strip()
    code = code.strip()

    # 같은 사업장 내 동일 건물명 중복 방지
    dup = (
        await db.execute(
            select(Building)
            .where(Building.site_id == site_id, Building.name == name, Building.is_active == True)
            .limit(1)
        )
    ).scalars().first()
    if dup:
        return RedirectResponse(
            f"/admin/sites?error={quote(f'이미 같은 이름의 건물이 있습니다: {name}')}",
            status_code=303,
        )

    # 코드 중복 방지 (같은 사업장)
    dup_code = (
        await db.execute(
            select(Building)
            .where(Building.site_id == site_id, Building.code == code, Building.is_active == True)
            .limit(1)
        )
    ).scalars().first()
    if dup_code:
        return RedirectResponse(
            f"/admin/sites?error={quote(f'이미 같은 코드의 건물이 있습니다: {code}')}",
            status_code=303,
        )

    building = Building(site_id=site_id, name=name, code=code)
    db.add(building)
    await db.flush()
    await ensure_building_default_categories(db, building)
    await db.commit()
    return RedirectResponse("/admin/sites", status_code=303)


@app.get("/admin/buildings/{building_id}/edit")
async def building_edit_page(
    building_id: int,
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
    message: str = "",
    error: str = "",
):
    result = await db.execute(
        select(Building)
        .where(Building.id == building_id)
        .options(
            selectinload(Building.site),
            selectinload(Building.floors).selectinload(Floor.zones),
        )
    )
    building = result.scalar_one_or_none()
    if not building or not building.is_active:
        raise HTTPException(404)
    sites = (
        await db.execute(select(Site).where(Site.is_active == True).order_by(Site.name))
    ).scalars().all()

    floors = sorted(
        [f for f in (building.floors or []) if getattr(f, "is_active", True)],
        key=lambda f: (f.level or 0, f.name or "", f.id),
    )
    zone_eq_counts: dict[int, int] = {}
    if floors:
        count_rows = (
            await db.execute(
                select(Equipment.zone_id, func.count(Equipment.id))
                .join(Zone, Equipment.zone_id == Zone.id)
                .join(Floor, Zone.floor_id == Floor.id)
                .where(
                    Floor.building_id == building_id,
                    Equipment.is_active == True,
                    Zone.is_active == True,
                    Floor.is_active == True,
                )
                .group_by(Equipment.zone_id)
            )
        ).all()
        zone_eq_counts = {int(zid): int(cnt) for zid, cnt in count_rows if zid is not None}

    return templates.TemplateResponse(
        request,
        "building_edit.html",
        {
            "user": user,
            "building": building,
            "sites": sites,
            "floors": floors,
            "zone_eq_counts": zone_eq_counts,
            "message": message,
            "error": error,
        },
    )


@app.post("/admin/buildings/{building_id}/edit")
async def building_edit(
    building_id: int,
    site_id: int = Form(...),
    name: str = Form(...),
    code: str = Form(...),
    manager_name: str = Form(""),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    building = await db.get(Building, building_id)
    if not building or not building.is_active:
        raise HTTPException(404)
    building.site_id = site_id
    building.name = name.strip()
    building.code = code.strip()
    building.manager_name = manager_name
    await db.commit()
    return RedirectResponse(
        f"/admin/buildings/{building_id}/edit?message={quote('건물 정보가 저장되었습니다.')}",
        status_code=303,
    )


@app.post("/admin/buildings/{building_id}/delete")
async def building_delete(
    building_id: int,
    user: User = Depends(require_can_delete),
    db: AsyncSession = Depends(get_db),
):
    building = await db.get(Building, building_id)
    if not building:
        raise HTTPException(404)
    building.is_active = False
    await db.commit()
    return RedirectResponse("/admin/sites", status_code=303)


@app.post("/admin/floors")
async def floor_create(
    building_id: int = Form(...),
    name: str = Form(...),
    level: int = Form(1),
    return_to: str = Form("detail"),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    building = await db.get(Building, building_id)
    if not building or not building.is_active:
        raise HTTPException(404)
    db.add(Floor(building_id=building_id, name=name.strip(), level=level, is_active=True))
    await db.commit()
    if return_to == "edit":
        return RedirectResponse(
            f"/admin/buildings/{building_id}/edit?message={quote('층이 추가되었습니다.')}",
            status_code=303,
        )
    return RedirectResponse(f"/admin/buildings/{building_id}", status_code=303)


@app.post("/admin/floors/{floor_id}/edit")
async def floor_edit(
    floor_id: int,
    name: str = Form(...),
    level: int = Form(1),
    building_id: int = Form(...),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    floor = await db.get(Floor, floor_id)
    if not floor or floor.building_id != building_id or not getattr(floor, "is_active", True):
        raise HTTPException(404)
    floor.name = name.strip()
    floor.level = level
    await db.commit()
    return RedirectResponse(
        f"/admin/buildings/{building_id}/edit?message={quote('층 정보가 수정되었습니다.')}",
        status_code=303,
    )


@app.post("/admin/floors/{floor_id}/delete")
async def floor_delete(
    floor_id: int,
    building_id: int = Form(...),
    user: User = Depends(require_can_delete),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    result = await db.execute(
        select(Floor)
        .where(Floor.id == floor_id, Floor.building_id == building_id)
        .options(selectinload(Floor.zones).selectinload(Zone.equipment))
    )
    floor = result.scalar_one_or_none()
    if not floor or not getattr(floor, "is_active", True):
        raise HTTPException(404)

    eq_count = 0
    for zone in floor.zones or []:
        if not getattr(zone, "is_active", True):
            continue
        for eq in zone.equipment or []:
            if getattr(eq, "is_active", True):
                eq.is_active = False
                eq_count += 1
        zone.is_active = False
    floor.is_active = False
    await db.commit()
    msg = f"층 «{floor.name}»이(가) 삭제되었습니다."
    if eq_count:
        msg += f" (하위 설비 {eq_count}대 비활성화)"
    return RedirectResponse(
        f"/admin/buildings/{building_id}/edit?message={quote(msg)}",
        status_code=303,
    )


@app.post("/admin/zones")
async def zone_create(
    floor_id: int = Form(...),
    building_id: int = Form(...),
    name: str = Form(...),
    code: str = Form(""),
    return_to: str = Form("detail"),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    floor = await db.get(Floor, floor_id)
    if not floor or floor.building_id != building_id or not getattr(floor, "is_active", True):
        raise HTTPException(404)
    db.add(
        Zone(
            floor_id=floor_id,
            name=name.strip(),
            code=(code or "").strip() or None,
            is_active=True,
        )
    )
    await db.commit()
    if return_to == "edit":
        return RedirectResponse(
            f"/admin/buildings/{building_id}/edit?message={quote('구역이 추가되었습니다.')}",
            status_code=303,
        )
    return RedirectResponse(f"/admin/buildings/{building_id}", status_code=303)


@app.post("/admin/zones/{zone_id}/edit")
async def zone_edit(
    zone_id: int,
    building_id: int = Form(...),
    floor_id: int = Form(...),
    name: str = Form(...),
    code: str = Form(""),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    result = await db.execute(
        select(Zone)
        .where(Zone.id == zone_id)
        .options(selectinload(Zone.floor))
    )
    zone = result.scalar_one_or_none()
    if not zone or not getattr(zone, "is_active", True):
        raise HTTPException(404)
    if not zone.floor or zone.floor.building_id != building_id:
        raise HTTPException(404)

    new_floor = await db.get(Floor, floor_id)
    if (
        not new_floor
        or new_floor.building_id != building_id
        or not getattr(new_floor, "is_active", True)
    ):
        return RedirectResponse(
            f"/admin/buildings/{building_id}/edit?error={quote('선택한 층이 없습니다.')}",
            status_code=303,
        )

    zone.floor_id = floor_id
    zone.name = name.strip()
    zone.code = (code or "").strip() or None
    await db.commit()
    return RedirectResponse(
        f"/admin/buildings/{building_id}/edit?message={quote('구역 정보가 수정되었습니다.')}",
        status_code=303,
    )


@app.post("/admin/zones/{zone_id}/delete")
async def zone_delete(
    zone_id: int,
    building_id: int = Form(...),
    user: User = Depends(require_can_delete),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    result = await db.execute(
        select(Zone)
        .where(Zone.id == zone_id)
        .options(selectinload(Zone.floor), selectinload(Zone.equipment))
    )
    zone = result.scalar_one_or_none()
    if not zone or not getattr(zone, "is_active", True):
        raise HTTPException(404)
    if not zone.floor or zone.floor.building_id != building_id:
        raise HTTPException(404)

    eq_count = 0
    for eq in zone.equipment or []:
        if getattr(eq, "is_active", True):
            eq.is_active = False
            eq_count += 1
    zone.is_active = False
    await db.commit()
    msg = f"구역 «{zone.name}»이(가) 삭제되었습니다."
    if eq_count:
        msg += f" (하위 설비 {eq_count}대 비활성화)"
    return RedirectResponse(
        f"/admin/buildings/{building_id}/edit?message={quote(msg)}",
        status_code=303,
    )


# ── Equipment ─────────────────────────────────────────────────────────

async def _building_categories(db: AsyncSession, building_id: int) -> list[str]:
    """건물별 엑셀 시트(카테고리) 목록."""
    rows = await db.execute(
        select(Equipment.category)
        .join(Zone)
        .join(Floor)
        .where(Floor.building_id == building_id, Equipment.is_active == True)
        .distinct()
        .order_by(Equipment.category)
    )
    return [r[0] for r in rows.all() if r[0]]


async def _building_category_counts(
    db: AsyncSession, building_id: int
) -> dict[str, int]:
    count_q = await db.execute(
        select(Equipment.category, func.count(Equipment.id))
        .join(Zone)
        .join(Floor)
        .where(Floor.building_id == building_id, Equipment.is_active == True)
        .group_by(Equipment.category)
    )
    return {cat: cnt for cat, cnt in count_q.all() if cat}


@app.get("/admin/equipment")
async def equipment_list(
    request: Request,
    building_id: int | None = None,
    category: str | None = None,
    error: str | None = None,
    message: str | None = None,
    open_eq: int | None = None,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    buildings = _sort_buildings(
        list(
            (
                await db.execute(
                    select(Building)
                    .where(Building.is_active == True)
                    .options(selectinload(Building.site))
                )
            ).scalars().all()
        )
    )
    building_groups = group_buildings_by_site(buildings)

    selected_building = None
    categories: list[str] = []
    category_counts: dict[str, int] = {}
    equipment: list = []
    zones = []
    sheet_fields: list[str] = []
    list_columns: list[str] = []
    show_all_categories = False

    if building_id:
        selected_building = await db.get(Building, building_id)
        if selected_building and not selected_building.is_active:
            selected_building = None
        if selected_building:
            from excel_import import ensure_building_default_categories

            # 기존 건물 진입 시에도 기본 대분류·코드 보장 (중복 층/구역이 있어도 실패하지 않음)
            try:
                added = await ensure_building_default_categories(db, selected_building)
                if added:
                    await db.commit()
            except Exception as e:
                print(f"[equipment_list] default categories skip: {e}", flush=True)
                await db.rollback()

            category_counts = await _building_category_counts(db, building_id)
            categories = sorted(category_counts.keys(), key=_building_sort_key)

            zones = (
                await db.execute(
                    select(Zone)
                    .join(Floor)
                    .where(
                        Floor.building_id == building_id,
                        Floor.is_active == True,
                        Zone.is_active == True,
                    )
                    .order_by(Zone.name)
                )
            ).scalars().all()

            active_category = category if category and category in categories else None
            show_all_categories = active_category is None

            try:
                q = (
                    select(Equipment)
                    .join(Zone)
                    .join(Floor)
                    .where(
                        Floor.building_id == building_id,
                        Equipment.is_active == True,
                    )
                    .options(
                        selectinload(Equipment.zone)
                        .selectinload(Zone.floor)
                        .selectinload(Floor.building),
                        selectinload(Equipment.equipment_type),
                        selectinload(Equipment.work_orders),
                        selectinload(Equipment.maintenance_records),
                        selectinload(Equipment.pm_inspections).selectinload(
                            PMInspection.schedule
                        ),
                    )
                    .order_by(Equipment.category, Equipment.code)
                )
                if active_category:
                    q = q.where(Equipment.category == active_category)
                equipment = (await db.execute(q)).scalars().all()
            except Exception as e:
                print(f"[equipment_list] fallback load: {e}", flush=True)
                await db.rollback()
                q = (
                    select(Equipment)
                    .join(Zone)
                    .join(Floor)
                    .where(
                        Floor.building_id == building_id,
                        Equipment.is_active == True,
                    )
                    .options(
                        selectinload(Equipment.zone)
                        .selectinload(Zone.floor)
                        .selectinload(Floor.building),
                        selectinload(Equipment.work_orders),
                        selectinload(Equipment.pm_inspections).selectinload(
                            PMInspection.schedule
                        ),
                    )
                    .order_by(Equipment.category, Equipment.code)
                )
                if active_category:
                    q = q.where(Equipment.category == active_category)
                equipment = (await db.execute(q)).scalars().all()

            if active_category:
                sheet_fields = get_category_fields(active_category, equipment)
                list_columns = list_display_fields(active_category, equipment)
            else:
                list_columns = ["명칭", "제조사", "모델"]

            category = active_category

    pm_inspections_by_eq = _equipment_pm_inspections_map(equipment) if equipment else {}
    change_logs_by_eq = await _load_change_logs_by_eq(
        db, [e.id for e in equipment]
    ) if equipment else {}

    return templates.TemplateResponse(
        request,
        "equipment.html",
        {
            "user": user,
            "buildings": buildings,
            "building_groups": building_groups,
            "selected_building": selected_building,
            "building_id": building_id,
            "category": category,
            "show_all_categories": show_all_categories,
            "categories": categories,
            "category_counts": category_counts,
            "equipment": equipment,
            "zones": zones,
            "sheet_fields": sheet_fields,
            "list_columns": list_columns,
            "error": error,
            "message": message,
            "open_eq": open_eq,
            "change_logs_by_eq": change_logs_by_eq,
            "pm_inspections_by_eq": pm_inspections_by_eq,
            "pm_inspections_json": json.dumps(pm_inspections_by_eq, ensure_ascii=False),
        },
    )


@app.post("/admin/equipment")
async def equipment_create(
    request: Request,
    zone_id: int = Form(...),
    code: str = Form(""),
    category: str = Form(""),
    building_id: int = Form(0),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote
    from sqlalchemy.exc import IntegrityError
    from equipment_schema import (
        merge_extra_for_save,
        parse_extra_form,
        resolve_core_fields,
    )
    from excel_import import _equipment_code

    form = await request.form()
    extra = parse_extra_form(form)
    cat = category.strip() if category else "기타"
    code_val = code.strip()
    name_val, manufacturer, model, serial_no = resolve_core_fields(extra)
    bld_id = building_id if building_id > 0 else 0

    def _list_url(error: str | None = None) -> str:
        if bld_id:
            url = f"/admin/equipment?building_id={bld_id}&category={quote(cat)}"
        else:
            url = "/admin/equipment"
        if error:
            sep = "&" if "?" in url else "?"
            url = f"{url}{sep}error={quote(error)}"
        return url

    if not name_val:
        return RedirectResponse(_list_url("구분/명칭을 입력하세요."), status_code=303)
    if zone_id <= 0:
        return RedirectResponse(_list_url("구역을 선택하세요."), status_code=303)

    zone = await db.get(Zone, zone_id)
    if not zone:
        return RedirectResponse(_list_url("선택한 구역이 없습니다."), status_code=303)

    if not code_val and bld_id:
        building = await db.get(Building, bld_id)
        if building:
            count = (
                await db.execute(
                    select(func.count(Equipment.id))
                    .join(Zone)
                    .join(Floor)
                    .where(Floor.building_id == bld_id, Equipment.category == cat)
                )
            ).scalar() or 0
            code_val = _equipment_code(building.code, cat, count + 1, name_val)

    if not code_val:
        return RedirectResponse(_list_url("코드를 입력하세요."), status_code=303)

    extra = merge_extra_for_save(extra, name_val, manufacturer, model, serial_no)

    existing = (
        await db.execute(
            select(Equipment)
            .where(Equipment.code == code_val)
            .order_by(Equipment.is_active.desc(), Equipment.id.desc())
            .limit(1)
        )
    ).scalars().first()

    try:
        if existing and existing.is_active:
            return RedirectResponse(
                _list_url(f"이미 사용 중인 코드입니다: {code_val}"),
                status_code=303,
            )

        if existing and not existing.is_active:
            eq = existing
            eq.is_active = True
            eq.zone_id = zone_id
            eq.name = name_val
            eq.category = cat
            eq.manufacturer = manufacturer or None
            eq.model = model or None
            eq.serial_no = serial_no or None
            eq.extra_data = extra
            eq.status = "normal"
        else:
            eq = Equipment(
                zone_id=zone_id,
                code=code_val,
                name=name_val,
                category=cat,
                manufacturer=manufacturer or None,
                model=model or None,
                serial_no=serial_no or None,
                extra_data=extra,
            )
            db.add(eq)

        await db.commit()
    except IntegrityError:
        await db.rollback()
        return RedirectResponse(
            _list_url(f"등록 실패: 코드 중복 또는 DB 제약 오류 ({code_val})"),
            status_code=303,
        )
    except Exception as e:
        await db.rollback()
        print(f"[equipment_create] error: {e}", flush=True)
        return RedirectResponse(
            _list_url(f"등록 실패: {e}"),
            status_code=303,
        )

    return RedirectResponse(_list_url(), status_code=303)


@app.get("/admin/equipment/import")
async def equipment_import_page(
    request: Request,
    building_id: int | None = None,
    message: str | None = None,
    error: str | None = None,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    buildings = (
        await db.execute(
            select(Building)
            .where(Building.is_active == True)
            .options(selectinload(Building.site))
            .order_by(Building.name)
        )
    ).scalars().all()
    selected = await db.get(Building, building_id) if building_id else None
    return templates.TemplateResponse(
        request,
        "equipment_import.html",
        {
            "user": user,
            "buildings": buildings,
            "selected_building": selected,
            "message": message,
            "error": error,
        },
    )


@app.post("/admin/equipment/import")
async def equipment_import_upload(
    request: Request,
    file: UploadFile = File(...),
    building_id: int = Form(...),
    replace: str = Form("0"),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote
    import tempfile
    from excel_import import import_excel_to_building

    building = await db.get(Building, building_id)
    if not building or not building.is_active:
        return RedirectResponse(
            "/admin/equipment/import?error=" + quote("건물을 찾을 수 없습니다."),
            status_code=303,
        )

    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in (".xls", ".xlsx"):
        return RedirectResponse(
            f"/admin/equipment/import?building_id={building_id}&error="
            + quote("xls 또는 xlsx 파일만 업로드할 수 있습니다."),
            status_code=303,
        )

    content = await file.read()
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        stats = await import_excel_to_building(
            db,
            building.name,
            tmp_path,
            replace=replace == "1",
            building_id=building_id,
        )
        msg = f"시트 {stats['sheets']}개 · 신규 {stats['created']}건 · 갱신 {stats['updated']}건"
        return RedirectResponse(
            f"/admin/equipment/import?building_id={building_id}&message={quote(msg)}",
            status_code=303,
        )
    except Exception as e:
        print(f"[equipment_import] error: {e}", flush=True)
        return RedirectResponse(
            f"/admin/equipment/import?building_id={building_id}&error={quote(str(e))}",
            status_code=303,
        )
    finally:
        Path(tmp_path).unlink(missing_ok=True)


@app.get("/admin/equipment/export/{building_id}")
async def equipment_export(
    building_id: int,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote
    from excel_import import export_building_excel

    building = await db.get(Building, building_id)
    if not building:
        raise HTTPException(404)

    result = await db.execute(
        select(Equipment)
        .join(Zone)
        .join(Floor)
        .where(Floor.building_id == building_id, Equipment.is_active == True)
        .options(
            selectinload(Equipment.maintenance_records),
            selectinload(Equipment.pm_inspections).selectinload(PMInspection.schedule),
        )
        .order_by(Equipment.category, Equipment.code)
    )
    items = result.scalars().unique().all()

    by_sheet: dict[str, list] = {}
    for eq in items:
        by_sheet.setdefault(eq.category or "기타", []).append(eq)

    data = export_building_excel(building.name, by_sheet)
    fname = quote(f"{building.name}_설비현황.xlsx")
    return StreamingResponse(
        BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"},
    )


@app.get("/admin/equipment/{eq_id}/export")
async def equipment_one_export(
    eq_id: int,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    """단일 설비에 등록된 사양·이력·점검·작업 전체를 엑셀로 다운로드."""
    from urllib.parse import quote
    from excel_import import export_equipment_excel

    result = await db.execute(
        select(Equipment)
        .where(Equipment.id == eq_id, Equipment.is_active == True)
        .options(
            selectinload(Equipment.zone).selectinload(Zone.floor).selectinload(Floor.building),
            selectinload(Equipment.maintenance_records),
            selectinload(Equipment.pm_inspections).selectinload(PMInspection.schedule),
            selectinload(Equipment.pm_schedules),
            selectinload(Equipment.work_orders),
        )
    )
    eq = result.scalar_one_or_none()
    if not eq:
        raise HTTPException(404)

    data = export_equipment_excel(eq)
    safe_code = (eq.code or f"eq{eq.id}").replace("/", "-").replace("\\", "-")
    fname = quote(f"{safe_code}_설비상세.xlsx")
    return StreamingResponse(
        BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"},
    )


@app.post("/admin/equipment/{eq_id}/import")
async def equipment_one_import(
    eq_id: int,
    request: Request,
    file: UploadFile = File(...),
    redirect_building_id: int = Form(0),
    redirect_category: str = Form(""),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    """단일 설비 Excel(사양·정비이력·점검이력) 가져오기."""
    from urllib.parse import quote
    from excel_import import import_equipment_excel

    result = await db.execute(
        select(Equipment)
        .where(Equipment.id == eq_id, Equipment.is_active == True)
        .options(
            selectinload(Equipment.zone).selectinload(Zone.floor),
            selectinload(Equipment.maintenance_records),
            selectinload(Equipment.pm_inspections),
            selectinload(Equipment.pm_schedules),
        )
    )
    eq = result.scalar_one_or_none()
    if not eq:
        raise HTTPException(404)

    building_id = redirect_building_id or (
        eq.zone.floor.building_id if eq.zone and eq.zone.floor else 0
    )

    def _back(msg: str = "", err: str = ""):
        qs = []
        if building_id:
            qs.append(f"building_id={building_id}")
        if redirect_category:
            qs.append(f"category={quote(redirect_category)}")
        if msg:
            qs.append(f"message={quote(msg)}")
        if err:
            qs.append(f"error={quote(err)}")
        return RedirectResponse(
            "/admin/equipment" + (("?" + "&".join(qs)) if qs else ""),
            status_code=303,
        )

    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in (".xlsx", ".xlsm"):
        return _back(err="xlsx 파일만 가져올 수 있습니다. (내보내기 Excel과 동일 형식)")

    content = await file.read()
    if not content:
        return _back(err="빈 파일입니다.")

    try:
        from equipment_schema import equipment_snapshot

        before = equipment_snapshot(eq)
        stats = await import_equipment_excel(db, eq, content)
        await _record_equipment_change(db, eq, before, user=user, source="Excel")
        await db.commit()
    except Exception as e:
        await db.rollback()
        return _back(err=f"가져오기 실패: {e}")

    msg = (
        f"{eq.code} 가져오기 완료 "
        f"(사양 {stats['spec_updated']} · 정비이력 +{stats['history_added']} · "
        f"점검이력 +{stats['pm_added']})"
    )
    if stats.get("warnings"):
        msg += " · " + " / ".join(stats["warnings"][:2])
    if building_id:
        from urllib.parse import quote as _q

        qs = [f"building_id={building_id}"]
        if redirect_category:
            qs.append(f"category={_q(redirect_category)}")
        qs.append(f"message={_q(msg)}")
        qs.append(f"open_eq={eq_id}")
        return RedirectResponse("/admin/equipment?" + "&".join(qs), status_code=303)
    return _back(msg=msg)


@app.post("/admin/equipment/bulk-import")
async def equipment_bulk_import(
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    """네트워크 공유 또는 data 폴더에서 일괄 import."""
    from urllib.parse import quote
    from excel_import import import_from_directory

    candidates = [
        Path(r"\\poscowide1\홍기룡\202010 설비현황"),
        Path("data/excel"),
        Path("data"),
    ]
    directory = next((p for p in candidates if p.is_dir()), None)
    if not directory:
        return RedirectResponse(
            "/admin/equipment/import?error=" + quote("import 대상 폴더를 찾을 수 없습니다."),
            status_code=303,
        )

    try:
        results = await import_from_directory(db, directory, replace=True)
        msg = (
            f"건물 {results['buildings']}개 · 신규 {results['total_created']}건 · "
            f"갱신 {results['total_updated']}건"
        )
        if results["errors"]:
            msg += f" · 오류 {len(results['errors'])}건"
        return RedirectResponse(
            f"/admin/equipment/import?message={quote(msg)}",
            status_code=303,
        )
    except Exception as e:
        return RedirectResponse(
            f"/admin/equipment/import?error={quote(str(e))}",
            status_code=303,
        )


@app.get("/admin/equipment/{eq_id}")
async def equipment_detail(
    eq_id: int,
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Equipment)
        .where(Equipment.id == eq_id, Equipment.is_active == True)
        .options(
            selectinload(Equipment.zone).selectinload(Zone.floor).selectinload(Floor.building),
            selectinload(Equipment.consumables),
            selectinload(Equipment.pm_schedules),
            selectinload(Equipment.pm_inspections).selectinload(PMInspection.schedule),
            selectinload(Equipment.work_orders),
            selectinload(Equipment.maintenance_records),
            selectinload(Equipment.equipment_type),
            selectinload(Equipment.template),
            selectinload(Equipment.change_logs),
        )
    )
    eq = result.scalar_one_or_none()
    if not eq:
        raise HTTPException(404)
    base_url = _public_base_url(request)
    sheet_fields = get_category_fields(eq.category, [eq])
    history = sorted(
        eq.maintenance_records or [],
        key=lambda r: (r.work_date or date.min, r.id),
        reverse=True,
    )
    open_orders = [
        wo
        for wo in (eq.work_orders or [])
        if getattr(wo, "is_active", True)
        and wo.status
        not in (WorkOrderStatus.completed, WorkOrderStatus.verified, WorkOrderStatus.closed)
    ]
    pm_inspections_by_eq = _equipment_pm_inspections_map([eq])
    change_logs = list(eq.change_logs or [])[:30]
    return templates.TemplateResponse(
        request,
        "equipment_detail.html",
        {
            "user": user,
            "eq": eq,
            "qr_url": f"{base_url}/eq/{eq.code}",
            "sheet_fields": sheet_fields,
            "history": history,
            "open_orders": open_orders,
            "change_logs": change_logs,
            "pm_inspections_json": json.dumps(pm_inspections_by_eq, ensure_ascii=False),
        },
    )


@app.get("/admin/equipment/{eq_id}/popup")
async def equipment_popup(
    eq_id: int,
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Equipment)
        .where(Equipment.id == eq_id, Equipment.is_active == True)
        .options(
            selectinload(Equipment.zone).selectinload(Zone.floor).selectinload(Floor.building),
            selectinload(Equipment.work_orders),
            selectinload(Equipment.maintenance_records),
            selectinload(Equipment.change_logs),
        )
    )
    eq = result.scalar_one_or_none()
    if not eq:
        raise HTTPException(404)
    sheet_fields = get_category_fields(eq.category, [eq])
    history = sorted(
        eq.maintenance_records or [],
        key=lambda r: (r.work_date or date.min, r.id),
        reverse=True,
    )[:10]
    open_orders = [
        wo
        for wo in (eq.work_orders or [])
        if getattr(wo, "is_active", True)
        and wo.status
        not in (WorkOrderStatus.completed, WorkOrderStatus.verified, WorkOrderStatus.closed)
    ]
    change_logs = list(eq.change_logs or [])[:15]
    return templates.TemplateResponse(
        request,
        "partials/equipment_popup.html",
        {
            "user": user,
            "eq": eq,
            "sheet_fields": sheet_fields,
            "history": history,
            "open_orders": open_orders,
            "change_logs": change_logs,
        },
    )


@app.post("/admin/equipment/{eq_id}/maintenance-request")
async def equipment_maintenance_request(
    eq_id: int,
    title: str = Form(""),
    description: str = Form(""),
    priority: str = Form("normal"),
    assignee_name: str = Form(""),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    eq = (
        await db.execute(
            select(Equipment)
            .where(Equipment.id == eq_id, Equipment.is_active == True)
            .options(
                selectinload(Equipment.zone).selectinload(Zone.floor).selectinload(Floor.building)
            )
        )
    ).scalar_one_or_none()
    if not eq:
        raise HTTPException(404)

    site_id = None
    if eq.zone and eq.zone.floor and eq.zone.floor.building:
        site_id = eq.zone.floor.building.site_id

    wo_title = title.strip() or f"[정비의뢰] {eq.code} {eq.name}"
    wo = WorkOrder(
        title=wo_title,
        description=description.strip() or f"{eq.category} 설비 정비의뢰",
        priority=priority,
        assignee_name=assignee_name.strip() or None,
        equipment_id=eq.id,
        site_id=site_id,
        status=WorkOrderStatus.received,
        work_type="정비",
    )
    db.add(wo)
    await db.commit()
    await db.refresh(wo)
    return RedirectResponse(f"/admin/work-orders/{wo.id}", status_code=303)


@app.post("/admin/equipment/{eq_id}/history")
async def equipment_history_create(
    eq_id: int,
    title: str = Form(...),
    work_date: str = Form(...),
    worker_name: str = Form(""),
    cause: str = Form(""),
    action: str = Form(""),
    parts_used: str = Form(""),
    note: str = Form(""),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    eq = await db.get(Equipment, eq_id)
    if not eq or not eq.is_active:
        raise HTTPException(404)
    try:
        wd = date.fromisoformat(work_date)
    except ValueError:
        wd = date.today()
    db.add(
        MaintenanceRecord(
            equipment_id=eq_id,
            title=title.strip(),
            work_date=wd,
            worker_name=worker_name.strip() or None,
            cause=cause.strip() or None,
            action=action.strip() or None,
            parts_used=parts_used.strip() or None,
            note=note.strip() or None,
            is_manual=True,
        )
    )
    await db.commit()
    return RedirectResponse(f"/admin/equipment/{eq_id}", status_code=303)


@app.get("/admin/equipment/{eq_id}/edit")
async def equipment_edit_page(
    eq_id: int,
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Equipment)
        .where(Equipment.id == eq_id, Equipment.is_active == True)
        .options(
            selectinload(Equipment.zone).selectinload(Zone.floor).selectinload(Floor.building),
            selectinload(Equipment.equipment_type),
        )
    )
    eq = result.scalar_one_or_none()
    if not eq:
        raise HTTPException(404)

    building_id = eq.zone.floor.building_id if eq.zone and eq.zone.floor else 0
    zones = []
    if building_id:
        zones = (
            await db.execute(
                select(Zone)
                .join(Floor)
                .where(
                    Floor.building_id == building_id,
                    Floor.is_active == True,
                    Zone.is_active == True,
                )
                .order_by(Zone.name)
            )
        ).scalars().all()
    bld_cats = await _building_categories(db, building_id) if building_id else []
    if eq.category and eq.category not in bld_cats:
        bld_cats = [eq.category] + bld_cats
    sheet_fields = get_category_fields(eq.category, [eq])
    field_values = {f: field_value(eq, f) for f in sheet_fields}
    return templates.TemplateResponse(
        request,
        "equipment_edit.html",
        {
            "user": user,
            "eq": eq,
            "zones": zones,
            "categories": bld_cats or [eq.category],
            "building_id": building_id,
            "sheet_fields": sheet_fields,
            "field_values": field_values,
            "category": eq.category,
        },
    )


@app.post("/admin/equipment/{eq_id}/edit")
async def equipment_edit(
    eq_id: int,
    request: Request,
    zone_id: int = Form(...),
    code: str = Form(...),
    category: str = Form(""),
    status: str = Form("normal"),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    from equipment_schema import (
        equipment_snapshot,
        merge_extra_for_save,
        parse_extra_form,
        resolve_core_fields,
    )

    result = await db.execute(
        select(Equipment)
        .where(Equipment.id == eq_id, Equipment.is_active == True)
        .options(selectinload(Equipment.zone).selectinload(Zone.floor))
    )
    eq = result.scalar_one_or_none()
    if not eq:
        raise HTTPException(404)

    before = equipment_snapshot(eq)

    form = await request.form()
    extra = parse_extra_form(form)
    cat = category.strip() if category else eq.category
    name_val, manufacturer, model, serial_no = resolve_core_fields(extra, eq.name)

    if not name_val:
        name_val = eq.name

    extra = merge_extra_for_save(extra, name_val, manufacturer, model, serial_no)

    eq.zone_id = zone_id
    eq.code = code.strip()
    eq.name = name_val
    eq.category = cat
    eq.manufacturer = manufacturer or None
    eq.model = model or None
    eq.serial_no = serial_no or None
    eq.extra_data = extra
    eq.status = status

    # after 스냅샷용 위치 라벨 갱신
    new_zone = (
        await db.execute(
            select(Zone)
            .where(Zone.id == zone_id)
            .options(selectinload(Zone.floor))
        )
    ).scalar_one_or_none()
    if new_zone is not None:
        eq.zone = new_zone

    await _record_equipment_change(db, eq, before, user=user, source="수정")
    await db.commit()

    building_id = 0
    if new_zone and new_zone.floor:
        building_id = new_zone.floor.building_id
    if building_id:
        return RedirectResponse(
            f"/admin/equipment?building_id={building_id}&category={cat}&open_eq={eq_id}",
            status_code=303,
        )
    return RedirectResponse(f"/admin/equipment/{eq_id}", status_code=303)


@app.post("/admin/equipment/{eq_id}/delete")
async def equipment_delete(
    eq_id: int,
    redirect_building_id: int = Form(0),
    redirect_category: str = Form(""),
    user: User = Depends(require_can_delete),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    result = await db.execute(
        select(Equipment)
        .where(Equipment.id == eq_id)
        .options(
            selectinload(Equipment.zone).selectinload(Zone.floor),
        )
    )
    eq = result.scalar_one_or_none()
    if not eq:
        raise HTTPException(404)
    building_id = (
        redirect_building_id
        if redirect_building_id > 0
        else (eq.zone.floor.building_id if eq.zone and eq.zone.floor else 0)
    )
    cat = (redirect_category or "").strip() or eq.category
    eq.is_active = False
    await db.commit()
    if building_id:
        if cat:
            return RedirectResponse(
                f"/admin/equipment?building_id={building_id}&category={quote(cat)}",
                status_code=303,
            )
        return RedirectResponse(
            f"/admin/equipment?building_id={building_id}",
            status_code=303,
        )
    return RedirectResponse("/admin/equipment", status_code=303)


# ── Equipment Templates ───────────────────────────────────────────────


@app.get("/admin/templates")
async def templates_page(
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    tpls = (
        await db.execute(
            select(EquipmentTemplate)
            .options(selectinload(EquipmentTemplate.equipment_type))
            .order_by(EquipmentTemplate.name)
        )
    ).scalars().all()
    types = (await db.execute(select(EquipmentType).order_by(EquipmentType.name))).scalars().all()
    return templates.TemplateResponse(
        request, "templates.html", {"user": user, "templates": tpls, "types": types}
    )


@app.post("/admin/templates")
async def template_create(
    equipment_type_id: int = Form(...),
    name: str = Form(...),
    manufacturer: str = Form(""),
    model: str = Form(""),
    pm_items: str = Form(""),
    consumables: str = Form(""),
    pm_cycle_days: int = Form(30),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    pm_list = [x.strip() for x in pm_items.split("\n") if x.strip()]
    cons_list = []
    for line in consumables.split("\n"):
        parts = [p.strip() for p in line.split(",") if p.strip()]
        if parts:
            item = {"name": parts[0]}
            if len(parts) > 1:
                try:
                    item["interval_days"] = int(parts[1])
                except ValueError:
                    pass
            cons_list.append(item)

    db.add(
        EquipmentTemplate(
            equipment_type_id=equipment_type_id,
            name=name.strip(),
            manufacturer=manufacturer,
            model=model,
            pm_items=pm_list,
            consumables=cons_list,
            pm_cycle_days=pm_cycle_days,
        )
    )
    await db.commit()
    return RedirectResponse("/admin/templates", status_code=303)


@app.post("/admin/equipment-types")
async def equipment_type_create(
    name: str = Form(...),
    category: str = Form("설비"),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    cat = category.strip() if category else "기타"
    db.add(EquipmentType(name=name.strip(), category=cat))
    await db.commit()
    return RedirectResponse("/admin/templates", status_code=303)


# ── Work Orders (CMMS) ─────────────────────────────────────────────────


@app.get("/admin/work-orders")
async def work_orders_list(
    request: Request,
    q: str = "",
    status: list[str] = Query(default=[]),
    priority: str = "",
    date_from: str = "",
    date_to: str = "",
    page: int = Query(1),
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    q_val = (q or "").strip()
    status_vals = _parse_wo_status_filters(status)
    # 처음 진입(상태 미지정) 시 정비의뢰만 표시
    if not status_vals:
        status_vals = ["received"]
    priority_val = (priority or "").strip()
    date_from_val = (date_from or "").strip()
    date_to_val = (date_to or "").strip()

    stmt = (
        select(WorkOrder)
        .outerjoin(Equipment, WorkOrder.equipment_id == Equipment.id)
        .options(selectinload(WorkOrder.equipment), selectinload(WorkOrder.partner))
        .where(WorkOrder.is_active == True)
    )
    filters = []

    if q_val:
        like = f"%{q_val}%"
        filters.append(
            or_(
                WorkOrder.title.ilike(like),
                WorkOrder.description.ilike(like),
                WorkOrder.action.ilike(like),
                WorkOrder.assignee_name.ilike(like),
                Equipment.code.ilike(like),
                Equipment.name.ilike(like),
            )
        )

    status_f = _wo_status_sql_filter(status_vals)
    if status_f is not None:
        filters.append(status_f)

    if priority_val in ("normal", "high"):
        filters.append(WorkOrder.priority == priority_val)

    if date_from_val:
        try:
            filters.append(
                WorkOrder.created_at
                >= datetime.fromisoformat(date_from_val).replace(tzinfo=None)
            )
        except ValueError:
            pass
    if date_to_val:
        try:
            # 종료일 포함 (다음날 0시 미만)
            end = datetime.fromisoformat(date_to_val).replace(tzinfo=None)
            filters.append(WorkOrder.created_at < end.replace(hour=23, minute=59, second=59, microsecond=999999))
        except ValueError:
            pass

    if filters:
        stmt = stmt.where(and_(*filters))

    all_orders = list(
        (
            await db.execute(stmt.order_by(WorkOrder.created_at.desc()))
        ).scalars().unique().all()
    )
    pager = _paginate(all_orders, page)
    orders = pager["items"]

    partners = (
        await db.execute(
            select(Partner).where(Partner.is_active == True).order_by(Partner.name)
        )
    ).scalars().all()

    buildings = (
        await db.execute(
            select(Building)
            .join(Site, Building.site_id == Site.id)
            .where(Building.is_active == True, Site.is_active == True)
            .order_by(Building.name)
        )
    ).scalars().all()
    equipment = (
        await db.execute(
            select(Equipment)
            .where(Equipment.is_active == True)
            .options(
                selectinload(Equipment.zone)
                .selectinload(Zone.floor)
                .selectinload(Floor.building)
            )
            .order_by(Equipment.code)
        )
    ).scalars().all()
    # 건물별 설비 옵션용 building_id 매핑
    equipment_opts = []
    for eq in equipment:
        bld_id = 0
        if eq.zone and eq.zone.floor and eq.zone.floor.building:
            bld_id = eq.zone.floor.building.id
        equipment_opts.append(
            {
                "id": eq.id,
                "code": eq.code,
                "name": eq.name,
                "category": eq.category or "",
                "building_id": bld_id,
            }
        )
    return templates.TemplateResponse(
        request,
        "work_orders.html",
        {
            "user": user,
            "orders": orders,
            "pager": pager,
            "partners": partners,
            "buildings": buildings,
            "equipment_opts": equipment_opts,
            "equipment_opts_json": json.dumps(equipment_opts, ensure_ascii=False),
            "filters": {
                "q": q_val,
                "status": ",".join(status_vals),
                "statuses": status_vals,
                "priority": priority_val,
                "date_from": date_from_val,
                "date_to": date_to_val,
                "page": pager["page"],
            },
            "result_count": pager["total"],
            "flash_message": request.query_params.get("message") or "",
            "flash_error": request.query_params.get("error") or "",
        },
    )


@app.get("/admin/work-orders/export")
async def work_orders_export(
    q: str = "",
    status: list[str] = Query(default=[]),
    priority: str = "",
    date_from: str = "",
    date_to: str = "",
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    """정비관리 목록(현재 필터) Excel 내보내기."""
    q_val = (q or "").strip()
    status_vals = _parse_wo_status_filters(status)
    if not status_vals:
        status_vals = ["received"]
    priority_val = (priority or "").strip()
    date_from_val = (date_from or "").strip()
    date_to_val = (date_to or "").strip()

    stmt = (
        select(WorkOrder)
        .outerjoin(Equipment, WorkOrder.equipment_id == Equipment.id)
        .options(selectinload(WorkOrder.equipment), selectinload(WorkOrder.partner))
        .where(WorkOrder.is_active == True)
    )
    filters = []
    if q_val:
        like = f"%{q_val}%"
        filters.append(
            or_(
                WorkOrder.title.ilike(like),
                WorkOrder.description.ilike(like),
                WorkOrder.action.ilike(like),
                WorkOrder.assignee_name.ilike(like),
                Equipment.code.ilike(like),
                Equipment.name.ilike(like),
            )
        )
    status_f = _wo_status_sql_filter(status_vals)
    if status_f is not None:
        filters.append(status_f)
    if priority_val in ("normal", "high"):
        filters.append(WorkOrder.priority == priority_val)
    if date_from_val:
        try:
            filters.append(
                WorkOrder.created_at
                >= datetime.fromisoformat(date_from_val).replace(tzinfo=None)
            )
        except ValueError:
            pass
    if date_to_val:
        try:
            end = datetime.fromisoformat(date_to_val).replace(tzinfo=None)
            filters.append(
                WorkOrder.created_at
                < end.replace(hour=23, minute=59, second=59, microsecond=999999)
            )
        except ValueError:
            pass
    if filters:
        stmt = stmt.where(and_(*filters))

    orders = (
        await db.execute(stmt.order_by(WorkOrder.created_at.desc()))
    ).scalars().unique().all()
    return _work_orders_excel_response(orders, "정비관리")


@app.post("/admin/work-orders")
async def work_order_create(
    equipment_id: int = Form(...),
    description: str = Form(...),
    priority: str = Form("normal"),
    assignee_name: str = Form(""),
    scheduled_date: str = Form(""),
    partner_id: int = Form(0),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    if equipment_id <= 0:
        return RedirectResponse("/admin/work-orders?error=eq&open=1", status_code=303)

    desc = description.strip()
    if not desc:
        return RedirectResponse("/admin/work-orders?error=desc&open=1", status_code=303)

    eq = (
        await db.execute(
            select(Equipment)
            .where(Equipment.id == equipment_id, Equipment.is_active == True)
            .options(
                selectinload(Equipment.zone)
                .selectinload(Zone.floor)
                .selectinload(Floor.building)
            )
        )
    ).scalar_one_or_none()
    if not eq:
        return RedirectResponse("/admin/work-orders?error=eq&open=1", status_code=303)

    site_id = None
    if eq.zone and eq.zone.floor and eq.zone.floor.building:
        site_id = eq.zone.floor.building.site_id

    sched = None
    if scheduled_date.strip():
        try:
            sched = date.fromisoformat(scheduled_date.strip())
        except ValueError:
            sched = None

    partner_fk = None
    if partner_id and partner_id > 0:
        partner = await db.get(Partner, partner_id)
        if partner and partner.is_active:
            partner_fk = partner.id

    title = f"[정비의뢰] {eq.code} {eq.name}"
    wo = WorkOrder(
        title=title,
        description=desc,
        priority=priority if priority in ("normal", "high") else "normal",
        assignee_name=assignee_name.strip() or None,
        equipment_id=eq.id,
        site_id=site_id,
        partner_id=partner_fk,
        scheduled_date=sched,
        status=WorkOrderStatus.received,
        work_type="정비",
    )
    db.add(wo)
    await db.commit()
    return RedirectResponse("/admin/work-orders", status_code=303)


@app.get("/admin/work-orders/{wo_id}")
async def work_order_detail(
    wo_id: int,
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    wo = (
        await db.execute(
            select(WorkOrder)
            .where(WorkOrder.id == wo_id, WorkOrder.is_active == True)
            .options(selectinload(WorkOrder.equipment), selectinload(WorkOrder.partner))
        )
    ).scalar_one_or_none()
    if not wo:
        raise HTTPException(404)
    partners = (
        await db.execute(
            select(Partner).where(Partner.is_active == True).order_by(Partner.name)
        )
    ).scalars().all()
    return templates.TemplateResponse(
        request,
        "work_order_detail.html",
        {
            "user": user,
            "wo": wo,
            "partners": partners,
            "process_step": _wo_process_step(wo.status),
            "flash_message": request.query_params.get("message") or "",
            "flash_error": request.query_params.get("error") or "",
        },
    )


@app.post("/admin/work-orders/{wo_id}/status")
async def work_order_status(
    wo_id: int,
    status: str = Form(...),
    action: str = Form(""),
    cause: str = Form(""),
    assignee_name: str = Form(""),
    partner_id: int = Form(0),
    scheduled_date: str = Form(""),
    redirect: str = Form(""),
    q: str = Form(""),
    filter_status: str = Form(""),
    filter_priority: str = Form(""),
    date_from: str = Form(""),
    date_to: str = Form(""),
    page: str = Form(""),
    filter_partner_id: str = Form(""),
    d1_board: str = Form(""),
    d1_partner_id: str = Form(""),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import urlencode

    wo = await db.get(WorkOrder, wo_id)
    if not wo or not wo.is_active:
        raise HTTPException(404)

    # 3단계 프로세스만 허용
    allowed = {"received", "in_progress", "completed"}
    if status not in allowed:
        status = "received"

    wo.status = WorkOrderStatus(status)
    wo.action = action.strip() or None
    if cause.strip():
        wo.cause = cause.strip()
    if assignee_name.strip():
        wo.assignee_name = assignee_name.strip()

    # 협력사 지정/해제
    if partner_id and partner_id > 0:
        partner = await db.get(Partner, partner_id)
        wo.partner_id = partner.id if partner and partner.is_active else None
    else:
        wo.partner_id = None

    # 업체 미지정 시 D-1 승인·하위 단계 해제
    if not wo.partner_id and getattr(wo, "d1_approved", False):
        _wo_apply_d1_unapprove(wo)
    elif not wo.partner_id:
        if getattr(wo, "approval_requested", False):
            wo.approval_requested = False
            wo.approval_requested_by = None
            wo.approval_requested_at = None
        if getattr(wo, "work_permitted", False):
            wo.work_permitted = False
            wo.work_permitted_by = None
            wo.work_permitted_at = None

    # 정비 예정일
    if scheduled_date.strip():
        try:
            wo.scheduled_date = date.fromisoformat(scheduled_date.strip())
        except ValueError:
            pass
    else:
        wo.scheduled_date = None
        # 예정일 제거 시 승인요청/작업허가 초기화
        if getattr(wo, "approval_requested", False):
            wo.approval_requested = False
            wo.approval_requested_by = None
            wo.approval_requested_at = None
        if getattr(wo, "work_permitted", False):
            wo.work_permitted = False
            wo.work_permitted_by = None
            wo.work_permitted_at = None

    if status == "completed":
        wo.completed_at = datetime.utcnow()
        await _ensure_maintenance_history(db, wo)
    elif status != "completed":
        # 완료가 아니면 완료시각 유지/해제 — 재진행 시 완료시각 비움
        if status in ("received", "in_progress"):
            wo.completed_at = None

    await db.commit()
    if redirect == "d1":
        board_raw = (d1_board or "").strip().lower()
        has_partner = (d1_partner_id or "").strip() not in ("", "0")
        if board_raw == "all":
            params: list[tuple[str, str]] = [("board", "all")]
        else:
            board_vals = _parse_d1_boards(
                [board_raw] if board_raw else [],
                allow_receipt=has_partner,
            )
            default_board = "receipt" if has_partner else "today"
            params = (
                [("board", board_vals[0])]
                if board_vals
                else [("board", default_board)]
            )
        if filter_status.strip():
            for part in filter_status.split(","):
                p = part.strip()
                if p:
                    params.append(("status", p))
        if has_partner:
            params.append(("partner_id", (d1_partner_id or "").strip()))
        if page.strip() and page.strip() not in ("", "1"):
            params.append(("page", page.strip()))
        qs = f"?{urlencode(params)}" if params else ""
        return RedirectResponse(f"/admin/d1{qs}", status_code=303)
    if redirect == "facility":
        fac_board = (d1_board or "day_before").strip() or "day_before"
        return _facility_redirect(board=fac_board)
    if redirect == "list":
        qs = _wo_list_redirect_params(
            q=q,
            filter_status=filter_status,
            filter_priority=filter_priority,
            date_from=date_from,
            date_to=date_to,
            page=page,
            filter_partner_id=filter_partner_id,
        )
        return RedirectResponse(f"/admin/work-orders{qs}", status_code=303)
    return RedirectResponse(f"/admin/work-orders/{wo_id}", status_code=303)


@app.post("/admin/work-orders/{wo_id}/approve-d1")
async def work_order_approve_d1(
    wo_id: int,
    partner_id: int = Form(0),
    redirect: str = Form("list"),
    q: str = Form(""),
    filter_status: str = Form(""),
    filter_priority: str = Form(""),
    date_from: str = Form(""),
    date_to: str = Form(""),
    page: str = Form(""),
    filter_partner_id: str = Form(""),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    """업체 지정된 정비를 D-1 작업으로 승인 전송."""
    from urllib.parse import quote, urlencode

    wo = await db.get(WorkOrder, wo_id)
    if not wo or not wo.is_active:
        raise HTTPException(404)

    # 폼의 협력사 선택을 먼저 반영 (저장 전 승인해도 적용)
    if partner_id and partner_id > 0:
        partner = await db.get(Partner, partner_id)
        wo.partner_id = partner.id if partner and partner.is_active else None

    def _back(error: str = "", message: str = ""):
        if redirect == "detail":
            qs = []
            if error:
                qs.append("error=" + quote(error))
            if message:
                qs.append("message=" + quote(message))
            suffix = ("?" + "&".join(qs)) if qs else ""
            return RedirectResponse(f"/admin/work-orders/{wo_id}{suffix}", status_code=303)
        qs = _wo_list_redirect_params(
            q=q,
            filter_status=filter_status,
            filter_priority=filter_priority,
            date_from=date_from,
            date_to=date_to,
            page=page,
            filter_partner_id=filter_partner_id,
            error=error,
            message=message,
        )
        return RedirectResponse(f"/admin/work-orders{qs}", status_code=303)

    if not wo.partner_id:
        return _back(error="업체를 선택한 뒤 승인하세요.")

    _wo_apply_d1_approve(wo, _wo_approver_label(user))
    await db.commit()
    return _back(message="D-1 작업으로 승인되었습니다.")


@app.post("/admin/work-orders/approve-d1-bulk")
async def work_order_approve_d1_bulk(
    request: Request,
    q: str = Form(""),
    filter_status: str = Form(""),
    filter_priority: str = Form(""),
    date_from: str = Form(""),
    date_to: str = Form(""),
    page: str = Form(""),
    filter_partner_id: str = Form(""),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    """선택한 정비를 일괄 D-1 승인 (업체는 항목별 지정값 사용)."""
    form = await request.form()
    raw_ids = form.getlist("wo_ids")
    ids: list[int] = []
    for raw in raw_ids:
        try:
            ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    # 중복 제거, 순서 유지
    seen: set[int] = set()
    uniq_ids: list[int] = []
    for i in ids:
        if i in seen:
            continue
        seen.add(i)
        uniq_ids.append(i)

    if not uniq_ids:
        qs = _wo_list_redirect_params(
            q=q,
            filter_status=filter_status,
            filter_priority=filter_priority,
            date_from=date_from,
            date_to=date_to,
            page=page,
            filter_partner_id=filter_partner_id,
            error="승인할 항목을 선택하세요.",
        )
        return RedirectResponse(f"/admin/work-orders{qs}", status_code=303)

    approver = _wo_approver_label(user)
    now = datetime.utcnow()
    approved = 0
    already = 0
    no_partner = 0
    missing = 0

    for wo_id in uniq_ids:
        wo = await db.get(WorkOrder, wo_id)
        if not wo or not wo.is_active:
            missing += 1
            continue

        partner_raw = form.get(f"partner_{wo_id}") or form.get(f"partner_id_{wo_id}") or "0"
        try:
            partner_id = int(partner_raw)
        except (TypeError, ValueError):
            partner_id = 0
        if partner_id > 0:
            partner = await db.get(Partner, partner_id)
            wo.partner_id = partner.id if partner and partner.is_active else None
        # partner_id==0이면 기존 DB 값 유지 (행에서 미지정으로 바꾼 경우는 0으로 옴)

        if getattr(wo, "d1_approved", False):
            already += 1
            continue
        if not wo.partner_id:
            no_partner += 1
            continue

        _wo_apply_d1_approve(wo, approver, now)
        approved += 1

    await db.commit()

    parts: list[str] = []
    if approved:
        parts.append(f"{approved}건 승인")
    if no_partner:
        parts.append(f"{no_partner}건 업체 미지정으로 제외")
    if already:
        parts.append(f"{already}건 이미 승인됨")
    if missing:
        parts.append(f"{missing}건 없음")

    if approved and not no_partner and not missing:
        msg = " · ".join(parts) if parts else "처리되었습니다."
        err = ""
    elif approved:
        msg = " · ".join(parts)
        err = ""
    elif no_partner and not already:
        msg = ""
        err = "선택한 항목에 업체가 지정되지 않았습니다. 항목별로 업체를 선택한 뒤 다시 승인하세요."
    else:
        msg = ""
        err = " · ".join(parts) if parts else "승인된 항목이 없습니다."

    qs = _wo_list_redirect_params(
        q=q,
        filter_status=filter_status,
        filter_priority=filter_priority,
        date_from=date_from,
        date_to=date_to,
        page=page,
        filter_partner_id=filter_partner_id,
        message=msg,
        error=err,
    )
    return RedirectResponse(f"/admin/work-orders{qs}", status_code=303)


@app.post("/admin/work-orders/{wo_id}/unapprove-d1")
async def work_order_unapprove_d1(
    wo_id: int,
    redirect: str = Form("list"),
    q: str = Form(""),
    filter_status: str = Form(""),
    filter_priority: str = Form(""),
    date_from: str = Form(""),
    date_to: str = Form(""),
    page: str = Form(""),
    filter_partner_id: str = Form(""),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    """D-1 승인 취소 (보드에서 제거)."""
    from urllib.parse import quote

    wo = await db.get(WorkOrder, wo_id)
    if not wo or not wo.is_active:
        raise HTTPException(404)
    _wo_apply_d1_unapprove(wo)
    await db.commit()

    if redirect == "detail":
        return RedirectResponse(
            f"/admin/work-orders/{wo_id}?message="
            + quote("D-1 승인이 취소되었습니다. 작업허가 승인요청·허가도 해제되었습니다."),
            status_code=303,
        )
    qs = _wo_list_redirect_params(
        q=q,
        filter_status=filter_status,
        filter_priority=filter_priority,
        date_from=date_from,
        date_to=date_to,
        page=page,
        filter_partner_id=filter_partner_id,
        message="D-1 승인이 취소되었습니다. 작업허가 승인요청·허가도 해제되었습니다.",
    )
    return RedirectResponse(f"/admin/work-orders{qs}", status_code=303)


@app.post("/admin/work-orders/{wo_id}/delete")
async def work_order_delete(
    wo_id: int,
    redirect: str = Form("list"),
    q: str = Form(""),
    filter_status: str = Form(""),
    filter_priority: str = Form(""),
    date_from: str = Form(""),
    date_to: str = Form(""),
    page: str = Form(""),
    filter_partner_id: str = Form(""),
    d1_board: str = Form(""),
    d1_partner_id: str = Form(""),
    user: User = Depends(require_can_delete),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import urlencode

    wo = await db.get(WorkOrder, wo_id)
    if not wo or not wo.is_active:
        raise HTTPException(404)
    wo.is_active = False
    await db.commit()
    if redirect == "d1":
        board_raw = (d1_board or "").strip().lower()
        has_partner = (d1_partner_id or "").strip() not in ("", "0")
        if board_raw == "all":
            params: list[tuple[str, str]] = [("board", "all")]
        else:
            board_vals = _parse_d1_boards(
                [board_raw] if board_raw else [],
                allow_receipt=has_partner,
            )
            default_board = "receipt" if has_partner else "today"
            params = (
                [("board", board_vals[0])]
                if board_vals
                else [("board", default_board)]
            )
        if filter_status.strip():
            for part in filter_status.split(","):
                p = part.strip()
                if p:
                    params.append(("status", p))
        if has_partner:
            params.append(("partner_id", (d1_partner_id or "").strip()))
        if page.strip() and page.strip() not in ("", "1"):
            params.append(("page", page.strip()))
        qs = f"?{urlencode(params)}" if params else ""
        return RedirectResponse(f"/admin/d1{qs}", status_code=303)
    params = {
        k: v
        for k, v in {
            "q": q.strip(),
            "status": filter_status.strip(),
            "priority": filter_priority.strip(),
            "date_from": date_from.strip(),
            "date_to": date_to.strip(),
            "partner_id": (filter_partner_id or "").strip()
            if (filter_partner_id or "").strip() not in ("", "0")
            else "",
            "page": page.strip() if str(page).strip() not in ("", "1") else "",
        }.items()
        if v
    }
    qs = f"?{urlencode(params)}" if params else ""
    return RedirectResponse(f"/admin/work-orders{qs}", status_code=303)


def _d1_redirect(
    *,
    d1_board: str = "",
    d1_partner_id: str = "",
    filter_status: str = "",
    page: str = "",
    message: str = "",
    error: str = "",
) -> RedirectResponse:
    from urllib.parse import urlencode

    board_raw = (d1_board or "").strip().lower()
    has_partner = (d1_partner_id or "").strip() not in ("", "0")
    if board_raw == "all":
        params: list[tuple[str, str]] = [("board", "all")]
    else:
        board_vals = _parse_d1_boards(
            [board_raw] if board_raw else [],
            allow_receipt=has_partner,
        )
        default_board = "receipt" if has_partner else "today"
        params = (
            [("board", board_vals[0])] if board_vals else [("board", default_board)]
        )
    if filter_status.strip():
        for part in filter_status.split(","):
            p = part.strip()
            if p:
                params.append(("status", p))
    if has_partner:
        params.append(("partner_id", (d1_partner_id or "").strip()))
    if page.strip() and page.strip() not in ("", "1"):
        params.append(("page", page.strip()))
    if message.strip():
        params.append(("message", message.strip()))
    if error.strip():
        params.append(("error", error.strip()))
    qs = f"?{urlencode(params)}" if params else ""
    return RedirectResponse(f"/admin/d1{qs}", status_code=303)


async def _apply_partner_and_schedule_from_form(
    db: AsyncSession,
    wo: WorkOrder,
    *,
    partner_id: int = 0,
    scheduled_date: str = "",
) -> None:
    if partner_id and partner_id > 0:
        partner = await db.get(Partner, partner_id)
        wo.partner_id = partner.id if partner and partner.is_active else None
    if scheduled_date and scheduled_date.strip():
        try:
            wo.scheduled_date = date.fromisoformat(scheduled_date.strip())
        except ValueError:
            pass


@app.post("/admin/work-orders/{wo_id}/request-approval")
async def work_order_request_approval(
    wo_id: int,
    partner_id: int = Form(0),
    scheduled_date: str = Form(""),
    d1_board: str = Form(""),
    d1_partner_id: str = Form(""),
    filter_status: str = Form(""),
    page: str = Form(""),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    """D-1(협력사): 예정일 지정 후 시설섹션으로 승인요청."""
    wo = await db.get(WorkOrder, wo_id)
    if not wo or not wo.is_active:
        raise HTTPException(404)
    await _apply_partner_and_schedule_from_form(
        db, wo, partner_id=partner_id, scheduled_date=scheduled_date
    )
    if not wo.partner_id:
        return _d1_redirect(
            d1_board=d1_board,
            d1_partner_id=d1_partner_id,
            filter_status=filter_status,
            page=page,
            error="업체가 지정되지 않았습니다.",
        )
    if not wo.scheduled_date:
        return _d1_redirect(
            d1_board=d1_board,
            d1_partner_id=d1_partner_id,
            filter_status=filter_status,
            page=page,
            error="예정일을 지정한 뒤 작업허가 승인요청하세요.",
        )
    if getattr(wo, "approval_requested", False):
        return _d1_redirect(
            d1_board=d1_board,
            d1_partner_id=d1_partner_id,
            filter_status=filter_status,
            page=page,
            message="이미 작업허가 승인요청된 항목입니다.",
        )
    wo.approval_requested = True
    wo.approval_requested_by = _wo_approver_label(user)
    wo.approval_requested_at = datetime.utcnow()
    await db.commit()
    return _d1_redirect(
        d1_board=d1_board,
        d1_partner_id=d1_partner_id,
        filter_status=filter_status,
        page=page,
        message="작업허가 승인요청되었습니다. 시설섹션에서 확인됩니다.",
    )


@app.post("/admin/d1/request-approval-bulk")
async def d1_request_approval_bulk(
    request: Request,
    d1_board: str = Form(""),
    d1_partner_id: str = Form(""),
    filter_status: str = Form(""),
    page: str = Form(""),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    """D-1(협력사): 선택 항목 일괄 승인요청."""
    form = await request.form()
    raw_ids = form.getlist("wo_ids")
    ids: list[int] = []
    for raw in raw_ids:
        try:
            ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    seen: set[int] = set()
    uniq: list[int] = []
    for i in ids:
        if i in seen:
            continue
        seen.add(i)
        uniq.append(i)
    if not uniq:
        return _d1_redirect(
            d1_board=d1_board,
            d1_partner_id=d1_partner_id,
            filter_status=filter_status,
            page=page,
            error="작업허가 승인요청할 항목을 선택하세요.",
        )

    label = _wo_approver_label(user)
    now = datetime.utcnow()
    ok = 0
    no_date = 0
    no_partner = 0
    already = 0

    for wo_id in uniq:
        wo = await db.get(WorkOrder, wo_id)
        if not wo or not wo.is_active:
            continue
        sched_raw = str(form.get(f"scheduled_{wo_id}") or "").strip()
        partner_raw = form.get(f"partner_{wo_id}") or "0"
        try:
            pid = int(partner_raw)
        except (TypeError, ValueError):
            pid = 0
        await _apply_partner_and_schedule_from_form(
            db, wo, partner_id=pid, scheduled_date=sched_raw
        )
        if getattr(wo, "approval_requested", False):
            already += 1
            continue
        if not wo.partner_id:
            no_partner += 1
            continue
        if not wo.scheduled_date:
            no_date += 1
            continue
        wo.approval_requested = True
        wo.approval_requested_by = label
        wo.approval_requested_at = now
        ok += 1

    await db.commit()
    parts = []
    if ok:
        parts.append(f"{ok}건 작업허가 승인요청")
    if no_date:
        parts.append(f"{no_date}건 예정일 미지정 제외")
    if no_partner:
        parts.append(f"{no_partner}건 업체 미지정 제외")
    if already:
        parts.append(f"{already}건 이미 요청됨")
    if ok:
        return _d1_redirect(
            d1_board=d1_board,
            d1_partner_id=d1_partner_id,
            filter_status=filter_status,
            page=page,
            message=" · ".join(parts),
        )
    return _d1_redirect(
        d1_board=d1_board,
        d1_partner_id=d1_partner_id,
        filter_status=filter_status,
        page=page,
        error=" · ".join(parts) if parts else "작업허가 승인요청된 항목이 없습니다.",
    )


@app.post("/admin/work-orders/{wo_id}/advance")
async def work_order_advance(
    wo_id: int,
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    """다음 단계로 진행: 정비의뢰 → 정비중 → 정비완료."""
    wo = await db.get(WorkOrder, wo_id)
    if not wo or not wo.is_active:
        raise HTTPException(404)
    step = _wo_process_step(wo.status)
    if step == 1:
        wo.status = WorkOrderStatus.in_progress
    elif step == 2:
        wo.status = WorkOrderStatus.completed
        wo.completed_at = datetime.utcnow()
        await _ensure_maintenance_history(db, wo)
    await db.commit()
    return RedirectResponse(f"/admin/work-orders/{wo_id}", status_code=303)


# ── D-1 Plans ─────────────────────────────────────────────────────────


def _risk_page_context(user, **extra):
    from risk_assessment import (
        ai_ready_for_user,
        list_majors,
        list_presets,
        mask_api_key,
        user_openai_credentials,
    )

    majors = list_majors()
    presets = list_presets()
    major_id_map = {m["name"]: m["id"] for m in majors if m.get("name")}
    _key, _model = user_openai_credentials(user)
    ready = ai_ready_for_user(user)
    ctx = {
        "user": user,
        "majors": majors,
        "presets": presets,
        "presets_json": json.dumps(presets, ensure_ascii=False),
        "major_id_map_json": json.dumps(major_id_map, ensure_ascii=False),
        "ai_ready": ready,
        "ai_key_masked": mask_api_key(_key),
        "ai_model": _model or "gpt-4o-mini",
        "use_ai": False,
        "five_m": {},
        "meta": {},
        "selected_major": "",
        "preset_name": "",
        "preset_name_json": '""',
        "work_name": "",
        "error_msg": "",
        "info_msg": "",
        "command_result": "",
        "rows_json": "[]",
        "register_msg": "",
    }
    ctx.update(extra)
    if "preset_name" in extra or "preset_name_json" not in extra:
        ctx["preset_name_json"] = json.dumps(ctx.get("preset_name") or "", ensure_ascii=False)
    if "ai_ready" not in extra:
        ctx["ai_ready"] = ready
        ctx["ai_key_masked"] = mask_api_key(_key)
        ctx["ai_model"] = _model or "gpt-4o-mini"
    return ctx


@app.get("/admin/risk-assessment")
async def risk_assessment_page(
    request: Request,
    user: User = Depends(require_login),
):
    try:
        return templates.TemplateResponse(
            request, "risk_assessment.html", _risk_page_context(user)
        )
    except Exception as e:
        print(f"[risk] page failed: {e}", flush=True)
        return templates.TemplateResponse(
            request,
            "error.html",
            {
                "user": user,
                "status_code": 500,
                "message": "위험성평가 화면을 불러오지 못했습니다.",
                "detail": str(e)[:500],
            },
            status_code=500,
        )


@app.post("/admin/risk-assessment/assess")
async def risk_assessment_run(
    request: Request,
    work_name: str = Form(...),
    Man: str = Form(""),
    Machine: str = Form(""),
    Material: str = Form(""),
    Method: str = Form(""),
    Management: str = Form(""),
    Environment: str = Form(""),
    major_name: str = Form(""),
    preset_name: str = Form(""),
    use_ai: str = Form("0"),
    department: str = Form(""),
    evaluator: str = Form(""),
    assessment_no: str = Form(""),
    apply_type: str = Form("정기평가"),
    user: User = Depends(require_can_edit),
):
    from risk_assessment import assess, get_preset

    five_m = {
        "Man": Man.strip(),
        "Machine": Machine.strip(),
        "Material": Material.strip(),
        "Method": Method.strip(),
        "Management": Management.strip(),
        "Environment": Environment.strip(),
    }
    # 프리셋만 고르고 5M이 비면 자동 채움
    if preset_name.strip() and not any(five_m.values()):
        p = get_preset(name=preset_name.strip())
        if p and p.get("five_m_one_e"):
            five_m = {k: (p["five_m_one_e"].get(k) or "") for k in five_m}

    meta = {
        "department": department.strip(),
        "evaluator": evaluator.strip() or user.name,
        "assessment_no": assessment_no.strip(),
        "apply_type": apply_type.strip() or "정기평가",
    }
    try:
        from risk_assessment import user_openai_credentials

        api_key, openai_model = user_openai_credentials(user)
        result = assess(
            work_name=work_name.strip(),
            five_m=five_m,
            use_ai=(use_ai == "1"),
            major_name=major_name.strip(),
            meta=meta,
            api_key=api_key,
            openai_model=openai_model,
        )
    except Exception as e:
        print(f"[risk] assess failed: {e}", flush=True)
        return templates.TemplateResponse(
            request,
            "risk_assessment.html",
            _risk_page_context(
                user,
                work_name=work_name.strip(),
                five_m=five_m,
                meta=meta,
                selected_major=major_name.strip(),
                preset_name=preset_name.strip(),
                use_ai=(use_ai == "1"),
                error_msg=f"평가 중 오류: {e}",
            ),
            status_code=500,
        )

    # 세션에는 가벼운 메타만 보관 (쿠키 용량 초과 방지 — 내보내기는 폼 POST 사용)
    request.session["risk_last"] = {
        "work_name": result["work_name"],
        "five_m": five_m,
        "major_name": major_name.strip(),
        "meta": {**(meta or {}), "mode": result["mode"]},
    }

    return templates.TemplateResponse(
        request,
        "risk_assessment.html",
        _risk_page_context(
            user,
            work_name=result["work_name"],
            five_m=five_m,
            meta=meta,
            selected_major=major_name.strip(),
            preset_name=preset_name.strip() or result["work_name"],
            use_ai=(result["mode"] == "ai"),
            form_rows=result["form_rows"],
            result_rows=result["rows"],
            rows_json=json.dumps(result["rows"], ensure_ascii=False),
            report_text=result["report_text"],
            mode_label=result["mode_label"],
            error_msg=result.get("error") or "",
            register_msg=result.get("register_msg") or "",
            info_msg=result.get("register_msg") or "",
        ),
    )


@app.post("/admin/risk-assessment/ai-settings")
async def risk_assessment_ai_settings(
    request: Request,
    openai_api_key: str = Form(""),
    openai_model: str = Form("gpt-4o-mini"),
    clear_key: str = Form("0"),
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    """로그인한 사용자 계정에만 OpenAI 키를 저장 (다른 계정과 공유하지 않음)."""
    from risk_assessment import mask_api_key

    db_user = await db.get(User, user.id)
    if not db_user or not db_user.is_active:
        raise HTTPException(401, detail="login_required")

    key_in = (openai_api_key or "").strip()
    model_in = (openai_model or "").strip() or "gpt-4o-mini"

    if clear_key == "1":
        db_user.openai_api_key = None
        db_user.openai_model = model_in
        await db.commit()
        await db.refresh(db_user)
        return templates.TemplateResponse(
            request,
            "risk_assessment.html",
            _risk_page_context(
                db_user,
                info_msg="내 계정의 OpenAI 키를 삭제했습니다.",
                use_ai=False,
            ),
        )

    if key_in:
        # 마스킹 표시값이 다시 저장되지 않게
        if set(key_in) <= {"•", "*"} or key_in.endswith("…"):
            pass
        else:
            db_user.openai_api_key = key_in
    db_user.openai_model = model_in
    await db.commit()
    await db.refresh(db_user)

    has_key = bool((db_user.openai_api_key or "").strip())
    msg = (
        f"내 계정에 OpenAI 키 저장됨 ({mask_api_key(db_user.openai_api_key)}, 모델: {db_user.openai_model})"
        if has_key
        else "OpenAI API 키를 입력하세요. (계정별로만 저장·사용됩니다)"
    )
    return templates.TemplateResponse(
        request,
        "risk_assessment.html",
        _risk_page_context(
            db_user,
            info_msg=msg if has_key else "",
            error_msg="" if has_key else msg,
            use_ai=has_key,
        ),
    )


@app.post("/admin/risk-assessment/learn")
async def risk_assessment_learn(
    request: Request,
    major_name: str = Form(...),
    allow_update: str = Form("0"),
    files: list[UploadFile] = File(default=[]),
    user: User = Depends(require_can_edit),
):
    """위험성평가 문서 업로드 → 소분류 학습 등록."""
    import tempfile
    from pathlib import Path

    from risk_assessment import learn_documents

    raw_files = files
    if raw_files is None:
        raw_files = []
    elif not isinstance(raw_files, list):
        raw_files = [raw_files]
    uploads = [f for f in raw_files if f and getattr(f, "filename", None)]
    if not uploads:
        return templates.TemplateResponse(
            request,
            "risk_assessment.html",
            _risk_page_context(
                user,
                selected_major=major_name.strip(),
                error_msg="학습할 문서 파일을 선택하세요. (xlsx/docx/pptx/pdf)",
            ),
        )

    tmp_paths: list[Path] = []
    try:
        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            for f in uploads:
                safe = Path(f.filename or "doc").name
                dest = td_path / safe
                data = await f.read()
                dest.write_bytes(data)
                tmp_paths.append(dest)

            result = learn_documents(
                tmp_paths,
                major_name.strip(),
                allow_update=(allow_update == "1"),
            )
    except Exception as e:
        print(f"[risk] learn failed: {e}", flush=True)
        return templates.TemplateResponse(
            request,
            "risk_assessment.html",
            _risk_page_context(
                user,
                selected_major=major_name.strip(),
                error_msg=f"문서 학습 실패: {e}",
            ),
            status_code=500,
        )

    ok_n = len(result.get("registered") or [])
    err_n = len(result.get("errors") or [])
    parts = [f"문서 학습: 파싱 {result.get('parsed', 0)}건 → 등록 {ok_n}건"]
    if result.get("registered"):
        names = ", ".join(r["name"] for r in result["registered"][:8])
        parts.append(f"등록: {names}")
    info = " · ".join(parts)
    err = ""
    if result.get("errors"):
        err = "일부 실패: " + " / ".join(result["errors"][:5])

    return templates.TemplateResponse(
        request,
        "risk_assessment.html",
        _risk_page_context(
            user,
            selected_major=major_name.strip(),
            info_msg=info,
            error_msg=err,
        ),
    )


@app.post("/admin/risk-assessment/export-html")
async def risk_assessment_export_html(
    request: Request,
    work_name: str = Form(""),
    report_text: str = Form(""),
    Man: str = Form(""),
    Machine: str = Form(""),
    Material: str = Form(""),
    Method: str = Form(""),
    Management: str = Form(""),
    Environment: str = Form(""),
    major_name: str = Form(""),
    department: str = Form(""),
    evaluator: str = Form(""),
    assessment_no: str = Form(""),
    apply_type: str = Form("정기평가"),
    user: User = Depends(require_login),
):
    from datetime import datetime as _dt
    from html import escape
    from io import BytesIO
    from urllib.parse import quote

    from risk_assessment.web_bridge import assess

    job = (work_name or "").strip() or "위험성평가"
    body = (report_text or "").strip()
    if not body:
        five_m = {
            "Man": Man.strip(),
            "Machine": Machine.strip(),
            "Material": Material.strip(),
            "Method": Method.strip(),
            "Management": Management.strip(),
            "Environment": Environment.strip(),
        }
        rebuilt = assess(
            job,
            five_m,
            use_ai=False,
            major_name=major_name.strip(),
            meta={
                "department": department.strip(),
                "evaluator": evaluator.strip() or user.name,
                "assessment_no": assessment_no.strip(),
                "apply_type": apply_type.strip() or "정기평가",
            },
        )
        job = rebuilt.get("work_name") or job
        body = (rebuilt.get("report_text") or "").strip()

    if not body:
        return JSONResponse({"ok": False, "error": "내보낼 평가 결과가 없습니다."}, status_code=400)

    html_body = escape(body).replace("\n", "<br>\n")
    stamp = _dt.now().strftime("%Y%m%d_%H%M")
    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<title>P-WIDE 위험성평가 - {escape(job)}</title>
<style>
body {{ font-family: 'Malgun Gothic', sans-serif; max-width: 1200px; margin: 2rem auto; padding: 0 1rem; line-height: 1.6; }}
h1 {{ color: #2b5797; border-bottom: 2px solid #2b5797; padding-bottom: 0.5rem; }}
</style>
</head>
<body>
<h1>P-WIDE 위험성평가 도우미 V3</h1>
<p><strong>작업/설비:</strong> {escape(job)}</p>
<p><strong>생성일시:</strong> {_dt.now():%Y-%m-%d %H:%M:%S}</p>
<hr>
<div>{html_body}</div>
</body>
</html>"""
    filename = quote(f"{job}_{stamp}.html")
    return StreamingResponse(
        BytesIO(html.encode("utf-8")),
        media_type="text/html; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@app.post("/admin/risk-assessment/export-excel")
async def risk_assessment_export_excel(
    request: Request,
    work_name: str = Form(""),
    rows_json: str = Form(""),
    Man: str = Form(""),
    Machine: str = Form(""),
    Material: str = Form(""),
    Method: str = Form(""),
    Management: str = Form(""),
    Environment: str = Form(""),
    major_name: str = Form(""),
    department: str = Form(""),
    evaluator: str = Form(""),
    assessment_no: str = Form(""),
    apply_type: str = Form("정기평가"),
    user: User = Depends(require_login),
):
    from datetime import datetime as _dt
    from io import BytesIO
    from urllib.parse import quote

    from risk_assessment.web_bridge import assess, export_excel_bytes

    job = (work_name or "").strip() or "위험성평가"
    meta = {
        "department": department.strip(),
        "evaluator": evaluator.strip() or user.name,
        "assessment_no": assessment_no.strip(),
        "apply_type": apply_type.strip() or "정기평가",
    }

    rows = []
    if (rows_json or "").strip():
        try:
            parsed = json.loads(rows_json)
            if isinstance(parsed, list):
                rows = parsed
        except json.JSONDecodeError:
            rows = []

    try:
        if not rows:
            five_m = {
                "Man": Man.strip(),
                "Machine": Machine.strip(),
                "Material": Material.strip(),
                "Method": Method.strip(),
                "Management": Management.strip(),
                "Environment": Environment.strip(),
            }
            rebuilt = assess(
                job,
                five_m,
                use_ai=False,
                major_name=major_name.strip(),
                meta=meta,
            )
            rows = rebuilt.get("rows") or []
            job = rebuilt.get("work_name") or job
            meta = {**meta, "mode": rebuilt.get("mode") or "local"}
        if not rows:
            return JSONResponse({"ok": False, "error": "내보낼 평가 결과가 없습니다."}, status_code=400)
        data = export_excel_bytes(job, rows, meta)
    except Exception as e:
        print(f"[risk] export-excel failed: {e}", flush=True)
        return JSONResponse({"ok": False, "error": f"Excel 저장 실패: {e}"}, status_code=500)

    stamp = _dt.now().strftime("%Y%m%d_%H%M")
    filename = quote(f"{job}_{stamp}.xlsx")
    return StreamingResponse(
        BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@app.post("/admin/risk-assessment/command")
async def risk_assessment_command(
    request: Request,
    command_num: int = Form(...),
    work_name: str = Form(...),
    major_name: str = Form(""),
    Man: str = Form(""),
    Machine: str = Form(""),
    Material: str = Form(""),
    Method: str = Form(""),
    Management: str = Form(""),
    Environment: str = Form(""),
    report_text: str = Form(""),
    user_question: str = Form(""),
    user: User = Depends(require_can_edit),
):
    from risk_assessment.web_bridge import run_additional

    five_m = {
        "Man": Man.strip(),
        "Machine": Machine.strip(),
        "Material": Material.strip(),
        "Method": Method.strip(),
        "Management": Management.strip(),
        "Environment": Environment.strip(),
    }
    last = request.session.get("risk_last") or {}
    try:
        result_text = run_additional(
            command_num,
            work_name.strip(),
            five_m,
            report_text or last.get("report_text") or "",
            major_name=major_name.strip(),
            user_question=user_question.strip(),
        )
    except Exception as e:
        result_text = f"추가 명령 실행 오류: {e}"

    # 직전 평가 결과가 있으면 함께 다시 표시
    form_rows = None
    result_rows = None
    mode_label = ""
    if last.get("work_name"):
        try:
            from risk_assessment import assess

            again = assess(
                work_name=last["work_name"],
                five_m=last.get("five_m") or five_m,
                use_ai=False,
                major_name=last.get("major_name") or major_name,
                meta=last.get("meta") or {},
            )
            form_rows = again["form_rows"]
            result_rows = again["rows"]
            mode_label = again["mode_label"]
            report_text = again["report_text"]
        except Exception:
            report_text = report_text
    else:
        report_text = report_text

    return templates.TemplateResponse(
        request,
        "risk_assessment.html",
        _risk_page_context(
            user,
            work_name=work_name.strip(),
            five_m=five_m,
            meta=last.get("meta") or {},
            selected_major=major_name.strip(),
            preset_name=work_name.strip(),
            form_rows=form_rows,
            result_rows=result_rows,
            rows_json=json.dumps(result_rows or [], ensure_ascii=False),
            report_text=report_text,
            mode_label=mode_label or "로컬 전용",
            command_result=result_text,
        ),
    )


@app.get("/admin/d1")
async def d1_list(
    request: Request,
    status: list[str] = Query(default=[]),
    view: str = "",
    partner_id: int = 0,
    board: list[str] = Query(default=[]),
    page: int = Query(1),
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    today = _today_kst()
    tomorrow = today + timedelta(days=1)
    status_vals = _parse_wo_status_filters(status)
    view_partner = (view or "").strip().lower() == "partner"
    try:
        partner_sel = int(partner_id or 0)
    except (TypeError, ValueError):
        partner_sel = 0

    open_statuses = [
        WorkOrderStatus.received,
        WorkOrderStatus.assigned,
        WorkOrderStatus.in_progress,
    ]
    done_statuses = [
        WorkOrderStatus.completed,
        WorkOrderStatus.verified,
        WorkOrderStatus.closed,
    ]

    open_works = (
        await db.execute(
            select(WorkOrder)
            .where(
                WorkOrder.is_active == True,  # noqa: E712
                WorkOrder.status.in_(open_statuses),
                *_wo_d1_sql_gate(),
            )
            .options(
                selectinload(WorkOrder.equipment),
                selectinload(WorkOrder.partner),
            )
            .order_by(
                WorkOrder.scheduled_date.asc().nullslast(),
                WorkOrder.priority.desc(),
                WorkOrder.id.asc(),
            )
        )
    ).scalars().unique().all()

    completed_raw = (
        await db.execute(
            select(WorkOrder)
            .where(
                WorkOrder.is_active == True,  # noqa: E712
                WorkOrder.status.in_(done_statuses),
                *_wo_d1_sql_gate(),
            )
            .options(
                selectinload(WorkOrder.equipment),
                selectinload(WorkOrder.partner),
            )
            .order_by(
                WorkOrder.completed_at.desc().nullslast(),
                WorkOrder.id.desc(),
            )
            .limit(300)
        )
    ).scalars().unique().all()

    # 이전 잘못된 자동 승인요청 반영분 정리
    clear_changed = False
    for _wo in list(open_works) + list(completed_raw):
        if _wo_clear_auto_approval_request(_wo):
            clear_changed = True
    if clear_changed:
        await db.commit()

    def _apply_status(items: list) -> list:
        return [w for w in items if _wo_matches_status_filter(w, status_vals)]

    def _apply_partner(items: list) -> list:
        if not partner_sel:
            return items
        if partner_sel == -1:
            return [w for w in items if not w.partner_id]
        return [w for w in items if int(w.partner_id or 0) == partner_sel]

    today_works = _apply_partner(
        _apply_status([w for w in open_works if w.scheduled_date == today])
    )
    tomorrow_works = _apply_partner(
        _apply_status([w for w in open_works if w.scheduled_date == tomorrow])
    )
    scheduled_works = _apply_partner(
        _apply_status(
            [w for w in open_works if w.scheduled_date not in (today, tomorrow)]
        )
    )
    completed_works = _apply_partner(_apply_status(list(completed_raw)))
    receipt_works = _apply_partner(
        _apply_status(
            [w for w in open_works if _wo_is_today_partner_receipt(w, today)]
        )
    )
    receipt_works = _sort_orders_today_receipt_first(receipt_works, today)

    status_pool = _apply_status(list(open_works) + list(completed_raw))
    filtered_all = today_works + tomorrow_works + scheduled_works + completed_works
    today_receipt_ids = {int(w.id) for w in receipt_works}

    boards, board_mode = _d1_resolve_boards(request, board, partner_id=partner_sel)
    board_map = {
        "receipt": receipt_works,
        "today": today_works,
        "tomorrow": tomorrow_works,
        "scheduled": scheduled_works,
        "completed": completed_works,
    }
    board_titles = {
        "receipt": "접수된 항목 오늘접수내용",
        "today": f"오늘 작업 ({today})",
        "tomorrow": f"내일 작업 ({tomorrow})",
        "scheduled": "정비 예정항목",
        "completed": "정비완료항목",
    }
    board_select_urls = _d1_board_select_urls(status_vals, partner_sel)

    selected_works: list = []
    seen_ids: set[int] = set()
    for key in boards:
        for w in board_map.get(key, []):
            wid = int(w.id)
            if wid in seen_ids:
                continue
            seen_ids.add(wid)
            selected_works.append(w)

    if board_mode == "all":
        board_title = "전체항목"
    elif boards:
        board_title = board_titles.get(boards[0], boards[0])
    else:
        board_title = "선택된 작업 구분 없음"

    partners = (
        await db.execute(
            select(Partner).where(Partner.is_active == True).order_by(Partner.name)
        )
    ).scalars().all()

    selected_partner_name = ""
    if partner_sel == -1:
        selected_partner_name = "미지정"
    elif partner_sel > 0:
        for p in partners:
            if p.id == partner_sel:
                selected_partner_name = p.name
                break
        if not selected_partner_name:
            selected_partner_name = f"업체 #{partner_sel}"

    partner_groups: list[dict] = []
    pager = None
    board_orders: list = []
    board_sections: list[dict] = []
    if view_partner:
        by_partner: dict[str, list] = {}
        for w in filtered_all:
            name = w.partner.name if w.partner else "미지정"
            by_partner.setdefault(name, []).append(w)
        for name in sorted(by_partner.keys(), key=lambda n: (n == "미지정", n)):
            items = by_partner[name]
            if partner_sel != 0:
                items = _sort_orders_today_receipt_first(items, today)
            partner_groups.append(
                {
                    "name": name,
                    "orders": items,
                    "today_count": sum(
                        1 for w in items if _wo_is_today_partner_receipt(w, today)
                    ),
                }
            )
        pager = {
            "page": 1,
            "total": len(filtered_all),
            "total_pages": 1,
            "page_numbers": [1],
            "has_prev": False,
            "has_next": False,
            "items": filtered_all,
            "per_page": max(1, len(filtered_all) or 1),
        }
        board_orders = filtered_all
    elif boards:
        pager = _paginate(selected_works, page)
        board_orders = pager["items"]
        # 페이지에 포함된 항목만 구분해 섹션 표시 (전체항목일 때 유형별 묶음)
        if board_mode == "all":
            by_key: dict[str, list] = {k: [] for k in boards}
            id_to_key: dict[int, str] = {}
            for key in boards:
                for w in board_map.get(key, []):
                    id_to_key[int(w.id)] = key
            for w in board_orders:
                key = id_to_key.get(int(w.id))
                if key:
                    by_key[key].append(w)
            for key in boards:
                board_sections.append(
                    {
                        "key": key,
                        "title": board_titles.get(key, key),
                        "orders": by_key.get(key, []),
                        "hint": (
                        "당일 등록·협력사 지정된 정비의뢰"
                        if key == "receipt"
                        else (
                            "오늘·내일 외 미완료 정비(예정일 미지정·과거·모레 이후)"
                            if key == "scheduled"
                            else ""
                        )
                    ),
                    "total": len(board_map.get(key, [])),
                }
            )
        else:
            key = boards[0]
            board_sections.append(
                {
                    "key": key,
                    "title": board_titles.get(key, key),
                    "orders": board_orders,
                    "hint": (
                        "당일 등록·협력사 지정된 정비의뢰"
                        if key == "receipt"
                        else (
                            "오늘·내일 외 미완료 정비(예정일 미지정·과거·모레 이후)"
                            if key == "scheduled"
                            else ""
                        )
                    ),
                    "total": len(selected_works),
                }
            )
    else:
        pager = _paginate([], page)
        board_orders = []

    partner_counts: dict[int, int] = {}
    for w in status_pool:
        pid = int(w.partner_id or 0)
        partner_counts[pid] = partner_counts.get(pid, 0) + 1

    partner_btns: list[dict] = []
    if view_partner:
        for p in partners:
            cnt = partner_counts.get(p.id, 0)
            if cnt > 0 or partner_sel == p.id:
                partner_btns.append({"id": p.id, "name": p.name, "count": cnt})
        none_cnt = partner_counts.get(0, 0)
        if none_cnt > 0 or partner_sel == -1:
            partner_btns.append({"id": -1, "name": "미지정", "count": none_cnt})

    return templates.TemplateResponse(
        request,
        "d1_plans.html",
        {
            "user": user,
            "today": today,
            "tomorrow": tomorrow,
            "today_works": today_works,
            "tomorrow_works": tomorrow_works,
            "scheduled_works": scheduled_works,
            "completed_works": completed_works,
            "board_orders": board_orders,
            "board_sections": board_sections,
            "board_title": board_title,
            "board_counts": {
                "receipt": len(receipt_works),
                "today": len(today_works),
                "tomorrow": len(tomorrow_works),
                "scheduled": len(scheduled_works),
                "completed": len(completed_works),
            },
            "board_select_urls": board_select_urls,
            "filtered_count": len(filtered_all),
            "partner_groups": partner_groups,
            "partner_btns": partner_btns,
            "today_receipt_ids": today_receipt_ids,
            "selected_partner_name": selected_partner_name,
            "pager": pager,
            "filters": {
                "statuses": status_vals,
                "view_partner": view_partner,
                "partner_id": partner_sel,
                "boards": boards,
                "board": board_mode,
                "board_mode": board_mode,
                "page": pager["page"] if pager else 1,
            },
            "partners": partners,
            "flash_message": request.query_params.get("message") or "",
            "flash_error": request.query_params.get("error") or "",
        },
    )


@app.get("/admin/d1/export")
async def d1_export(
    request: Request,
    status: list[str] = Query(default=[]),
    view: str = "",
    partner_id: int = 0,
    board: list[str] = Query(default=[]),
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    """D-1 화면의 현재 필터 결과를 Excel로 내보내기."""
    today = _today_kst()
    tomorrow = today + timedelta(days=1)
    status_vals = _parse_wo_status_filters(status)
    view_partner = (view or "").strip().lower() == "partner"
    try:
        partner_sel = int(partner_id or 0)
    except (TypeError, ValueError):
        partner_sel = 0
    open_statuses = [
        WorkOrderStatus.received,
        WorkOrderStatus.assigned,
        WorkOrderStatus.in_progress,
    ]
    done_statuses = [
        WorkOrderStatus.completed,
        WorkOrderStatus.verified,
        WorkOrderStatus.closed,
    ]

    open_works = (
        await db.execute(
            select(WorkOrder)
            .where(
                WorkOrder.is_active == True,
                WorkOrder.status.in_(open_statuses),
                *_wo_d1_sql_gate(),
            )
            .options(
                selectinload(WorkOrder.equipment),
                selectinload(WorkOrder.partner),
            )
            .order_by(
                WorkOrder.scheduled_date.asc().nullslast(),
                WorkOrder.priority.desc(),
                WorkOrder.id.asc(),
            )
        )
    ).scalars().unique().all()
    completed = (
        await db.execute(
            select(WorkOrder)
            .where(
                WorkOrder.is_active == True,
                WorkOrder.status.in_(done_statuses),
                *_wo_d1_sql_gate(),
            )
            .options(
                selectinload(WorkOrder.equipment),
                selectinload(WorkOrder.partner),
            )
            .order_by(
                WorkOrder.completed_at.desc().nullslast(),
                WorkOrder.id.desc(),
            )
            .limit(300)
        )
    ).scalars().unique().all()

    def _ok(w: WorkOrder) -> bool:
        if not _wo_matches_status_filter(w, status_vals):
            return False
        pid = int(w.partner_id or 0)
        if partner_sel == -1 and pid != 0:
            return False
        if partner_sel > 0 and pid != partner_sel:
            return False
        return True

    today_o = [w for w in open_works if w.scheduled_date == today and _ok(w)]
    tomorrow_o = [w for w in open_works if w.scheduled_date == tomorrow and _ok(w)]
    scheduled_o = [
        w for w in open_works if w.scheduled_date not in (today, tomorrow) and _ok(w)
    ]
    completed_o = [w for w in completed if _ok(w)]
    receipt_o = [
        w
        for w in open_works
        if _wo_is_today_partner_receipt(w, today) and _ok(w)
    ]

    if view_partner:
        orders = today_o + tomorrow_o + scheduled_o + completed_o
    else:
        boards, _board_mode = _d1_resolve_boards(
            request, board, partner_id=partner_sel
        )
        board_map = {
            "receipt": receipt_o,
            "today": today_o,
            "tomorrow": tomorrow_o,
            "scheduled": scheduled_o,
            "completed": completed_o,
        }
        orders = []
        seen: set[int] = set()
        for key in boards:
            for w in board_map.get(key, []):
                wid = int(w.id)
                if wid in seen:
                    continue
                seen.add(wid)
                orders.append(w)

    return _work_orders_excel_response(orders, "D1작업")


def _facility_sql_gate():
    """시설섹션: 승인요청된 항목."""
    return (
        WorkOrder.partner_id.is_not(None),
        WorkOrder.approval_requested.is_(True),
        WorkOrder.is_active.is_(True),
    )


@app.get("/admin/facility-section")
async def facility_section_list(
    request: Request,
    board: str = Query("day_before"),
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    """정비관리(시설섹션): 승인요청된 하루전·당일 작업."""
    today = _today_kst()
    tomorrow = today + timedelta(days=1)
    board_key = (board or "day_before").strip().lower()
    if board_key not in ("day_before", "today", "all"):
        board_key = "day_before"

    open_statuses = [
        WorkOrderStatus.received,
        WorkOrderStatus.assigned,
        WorkOrderStatus.in_progress,
    ]
    rows = (
        await db.execute(
            select(WorkOrder)
            .where(
                WorkOrder.status.in_(open_statuses),
                *_facility_sql_gate(),
            )
            .options(
                selectinload(WorkOrder.equipment),
                selectinload(WorkOrder.partner),
            )
            .order_by(
                WorkOrder.scheduled_date.asc().nullslast(),
                WorkOrder.priority.desc(),
                WorkOrder.id.asc(),
            )
        )
    ).scalars().unique().all()

    day_before = [w for w in rows if w.scheduled_date == tomorrow]
    today_works = [w for w in rows if w.scheduled_date == today]

    return templates.TemplateResponse(
        request,
        "facility_section.html",
        {
            "user": user,
            "today": today,
            "tomorrow": tomorrow,
            "board": board_key,
            "day_before_works": day_before,
            "today_works": today_works,
            "flash_message": request.query_params.get("message") or "",
            "flash_error": request.query_params.get("error") or "",
            "partners": (
                await db.execute(
                    select(Partner)
                    .where(Partner.is_active == True)  # noqa: E712
                    .order_by(Partner.name)
                )
            ).scalars().all(),
        },
    )


def _facility_redirect(*, board: str = "day_before", message: str = "", error: str = ""):
    from urllib.parse import urlencode

    params: list[tuple[str, str]] = []
    b = (board or "day_before").strip() or "day_before"
    if b != "day_before":
        params.append(("board", b))
    if message.strip():
        params.append(("message", message.strip()))
    if error.strip():
        params.append(("error", error.strip()))
    qs = f"?{urlencode(params)}" if params else ""
    return RedirectResponse(f"/admin/facility-section{qs}", status_code=303)


@app.post("/admin/facility-section/{wo_id}/permit")
async def facility_permit_work(
    wo_id: int,
    board: str = Form("day_before"),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    """시설섹션: 개별 작업허가 (하루전 항목)."""
    today = _today_kst()
    tomorrow = today + timedelta(days=1)
    wo = await db.get(WorkOrder, wo_id)
    if not wo or not wo.is_active:
        raise HTTPException(404)
    if not getattr(wo, "approval_requested", False):
        return _facility_redirect(board=board, error="승인요청된 항목이 아닙니다.")
    if wo.scheduled_date != tomorrow:
        return _facility_redirect(
            board=board,
            error="작업허가는 작업일 하루전 항목만 가능합니다.",
        )
    if getattr(wo, "work_permitted", False):
        return _facility_redirect(board=board, message="이미 작업허가된 항목입니다.")
    wo.work_permitted = True
    wo.work_permitted_by = _wo_approver_label(user)
    wo.work_permitted_at = datetime.utcnow()
    await db.commit()
    return _facility_redirect(board=board, message="작업허가 처리되었습니다.")


@app.post("/admin/facility-section/permit-bulk")
async def facility_permit_work_bulk(
    request: Request,
    board: str = Form("day_before"),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    """시설섹션: 선택 항목 일괄 작업허가."""
    form = await request.form()
    today = _today_kst()
    tomorrow = today + timedelta(days=1)
    raw_ids = form.getlist("wo_ids")
    ids: list[int] = []
    for raw in raw_ids:
        try:
            ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    if not ids:
        return _facility_redirect(board=board, error="작업허가할 항목을 선택하세요.")

    label = _wo_approver_label(user)
    now = datetime.utcnow()
    ok = 0
    skip = 0
    for wo_id in ids:
        wo = await db.get(WorkOrder, wo_id)
        if (
            not wo
            or not wo.is_active
            or not getattr(wo, "approval_requested", False)
            or wo.scheduled_date != tomorrow
            or getattr(wo, "work_permitted", False)
        ):
            skip += 1
            continue
        wo.work_permitted = True
        wo.work_permitted_by = label
        wo.work_permitted_at = now
        ok += 1
    await db.commit()
    if ok:
        msg = f"{ok}건 작업허가"
        if skip:
            msg += f" · {skip}건 제외"
        return _facility_redirect(board=board, message=msg)
    return _facility_redirect(
        board=board,
        error="작업허가할 수 있는 하루전 항목이 없습니다.",
    )


@app.post("/admin/facility-section/{wo_id}/unpermit")
async def facility_unpermit_work(
    wo_id: int,
    board: str = Form("day_before"),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    """시설섹션: 작업허가(승인) 취소."""
    wo = await db.get(WorkOrder, wo_id)
    if not wo or not wo.is_active:
        raise HTTPException(404)
    wo.work_permitted = False
    wo.work_permitted_by = None
    wo.work_permitted_at = None
    await db.commit()
    return _facility_redirect(board=board, message="승인이 취소되었습니다.")


@app.post("/admin/d1")
async def d1_create(user: User = Depends(require_can_create)):
    """D-1 JSA/TBM 등록 UI 제거 — 목록으로 리다이렉트."""
    return RedirectResponse("/admin/d1", status_code=303)


@app.get("/admin/d1/{plan_id}")
async def d1_detail(
    plan_id: int,
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    plan = (
        await db.execute(
            select(D1Plan)
            .where(D1Plan.id == plan_id)
            .options(
                selectinload(D1Plan.site),
                selectinload(D1Plan.building),
                selectinload(D1Plan.equipment),
                selectinload(D1Plan.partner),
            )
        )
    ).scalar_one_or_none()
    if not plan:
        raise HTTPException(404)
    return templates.TemplateResponse(
        request, "d1_detail.html", {"user": user, "plan": plan}
    )


@app.post("/admin/d1/{plan_id}/advance")
async def d1_advance(
    plan_id: int,
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    plan = await db.get(D1Plan, plan_id)
    if not plan:
        raise HTTPException(404)
    flow = [
        D1Status.draft,
        D1Status.review,
        D1Status.approved,
        D1Status.jsa_pending,
        D1Status.tbm_pending,
        D1Status.permit_pending,
        D1Status.in_progress,
        D1Status.completed,
    ]
    try:
        idx = flow.index(plan.status)
        if idx < len(flow) - 1:
            plan.status = flow[idx + 1]
            if plan.status == D1Status.permit_pending:
                plan.permit_no = f"WP-{plan.id:05d}-{date.today().strftime('%Y%m%d')}"
            if plan.status == D1Status.completed:
                plan.completed_at = datetime.utcnow()
    except ValueError:
        pass
    await db.commit()
    return RedirectResponse(f"/admin/d1/{plan_id}", status_code=303)


# ── PM & Partners ─────────────────────────────────────────────────────


def _pm_eq_options():
    return (
        selectinload(PMSchedule.equipment)
        .selectinload(Equipment.zone)
        .selectinload(Zone.floor)
        .selectinload(Floor.building)
    )


def _pm_match_filters(
    schedule: PMSchedule,
    *,
    q: str = "",
    building_id: int | None = None,
    equipment_id: int | None = None,
    due_only: bool = False,
    today: date | None = None,
) -> bool:
    eq = schedule.equipment
    if equipment_id and (not eq or eq.id != equipment_id):
        return False
    bld = _equipment_building(eq)
    if building_id and (not bld or bld.id != building_id):
        return False
    if due_only:
        today = today or _today_kst()
        if not schedule.next_due or schedule.next_due > today:
            return False
    if q:
        needle = q.strip().lower()
        hay = " ".join(
            [
                schedule.title or "",
                schedule.assignee_name or "",
                eq.code if eq else "",
                eq.name if eq else "",
                eq.category if eq else "",
                bld.name if bld else "",
            ]
        ).lower()
        if needle not in hay:
            return False
    return True


async def _pm_filtered_schedules(
    db: AsyncSession,
    *,
    q: str = "",
    building_id: int | None = None,
    equipment_id: int | None = None,
    due_only: bool = False,
) -> list[PMSchedule]:
    today = _today_kst()
    schedules = (
        await db.execute(
            select(PMSchedule)
            .where(PMSchedule.is_active == True)  # noqa: E712
            .options(_pm_eq_options(), selectinload(PMSchedule.inspections))
            .order_by(PMSchedule.next_due.asc().nullslast())
        )
    ).scalars().unique().all()
    return [
        s
        for s in schedules
        if _pm_match_filters(
            s,
            q=q,
            building_id=building_id,
            equipment_id=equipment_id,
            due_only=due_only,
            today=today,
        )
    ]


def _pm_excel_response(schedules: list[PMSchedule]):
    from urllib.parse import quote

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "예방점검"
    ws.append(
        [
            "건물",
            "설비코드",
            "설비명",
            "분류",
            "점검명",
            "주기",
            "주기일수",
            "담당",
            "다음점검",
            "마지막점검",
            "최근결과",
            "등록내용",
            "점검자",
            "점검일시",
        ]
    )
    for pm in schedules:
        eq = pm.equipment
        bld = _equipment_building(eq)
        latest = None
        if pm.inspections:
            latest = max(pm.inspections, key=lambda x: x.inspected_at or datetime.min)
        ws.append(
            [
                bld.name if bld else "",
                eq.code if eq else "",
                eq.name if eq else "",
                eq.category if eq else "",
                pm.title,
                _pm_freq_label(pm.frequency),
                _pm_cycle_days(pm.frequency, pm.custom_days),
                pm.assignee_name or "",
                pm.next_due.isoformat() if pm.next_due else "",
                pm.last_done.isoformat() if pm.last_done else "",
                _pm_result_label(latest.result) if latest else "",
                (latest.note or "") if latest else "",
                (latest.inspector_name or "") if latest else "",
                _fmt_kst(latest.inspected_at) if latest and latest.inspected_at else "",
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now(KST).strftime("%Y%m%d_%H%M")
    filename = quote(f"예방점검_{stamp}.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


async def _create_pm_work_order(
    db: AsyncSession,
    *,
    eq: Equipment,
    result: PMResult,
    note: str,
    inspector_name: str,
) -> WorkOrder:
    site_id = None
    bld = _equipment_building(eq)
    if bld:
        site_id = bld.site_id
    result_label = _pm_result_label(result)
    wo = WorkOrder(
        title=f"[PM점검-{result_label}] {eq.code} {eq.name}",
        description=(
            note.strip()
            or f"예방점검 결과: {result_label}"
        ),
        priority="high" if result == PMResult.fault else "normal",
        assignee_name=inspector_name.strip() or None,
        equipment_id=eq.id,
        site_id=site_id,
        status=WorkOrderStatus.received,
        work_type="정비",
    )
    db.add(wo)
    await db.flush()
    return wo


async def _get_or_create_pm_schedule(db: AsyncSession, eq: Equipment) -> PMSchedule:
    """활성 점검주기가 없으면 QR 등 임시 점검용 일정을 생성."""
    active = next((s for s in (eq.pm_schedules or []) if s.is_active), None)
    if active:
        return active
    schedule = PMSchedule(
        equipment_id=eq.id,
        title=f"{eq.code} 예방점검",
        frequency=PMFrequency.monthly,
        is_active=True,
    )
    db.add(schedule)
    await db.flush()
    if eq.pm_schedules is None:
        eq.pm_schedules = []
    eq.pm_schedules.append(schedule)
    schedule.equipment = eq
    return schedule


async def _record_pm_inspection(
    db: AsyncSession,
    schedule: PMSchedule,
    *,
    result_raw: str,
    note: str = "",
    inspector_name: str = "",
    create_work_order: bool = False,
) -> tuple[PMInspection, WorkOrder | None]:
    try:
        result = PMResult(result_raw)
    except ValueError:
        result = PMResult.normal
    eq = schedule.equipment
    if not eq:
        raise HTTPException(404, detail="설비를 찾을 수 없습니다.")

    wo = None
    if create_work_order and result in (PMResult.caution, PMResult.fault):
        wo = await _create_pm_work_order(
            db,
            eq=eq,
            result=result,
            note=note,
            inspector_name=inspector_name,
        )

    insp = PMInspection(
        schedule_id=schedule.id,
        equipment_id=eq.id,
        result=result,
        note=note.strip() or None,
        inspector_name=inspector_name.strip() or None,
        inspected_at=datetime.utcnow(),
        work_order_id=wo.id if wo else None,
    )
    db.add(insp)
    _pm_advance_schedule(schedule, _today_kst())
    return insp, wo


# ── 점검일지 ──────────────────────────────────────────────────────────


async def _ensure_inspection_log_tables() -> None:
    from sqlalchemy import text

    url = (os.environ.get("DATABASE_URL") or os.environ.get("DATABASE_INTERNAL_URL") or "").lower()
    is_pg = "postgresql" in url or "postgres" in url
    async with engine.begin() as conn:
        if is_pg:
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS inspection_log_buildings (
                        id SERIAL PRIMARY KEY,
                        building_id INTEGER NOT NULL UNIQUE REFERENCES buildings(id),
                        created_at TIMESTAMP WITHOUT TIME ZONE
                    )
                    """
                )
            )
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS inspection_log_files (
                        id SERIAL PRIMARY KEY,
                        building_id INTEGER NOT NULL REFERENCES buildings(id),
                        title VARCHAR(200) NOT NULL,
                        original_name VARCHAR(300),
                        stored_name VARCHAR(300) NOT NULL,
                        content_type VARCHAR(100),
                        file_data BYTEA,
                        uploaded_by VARCHAR(100),
                        created_at TIMESTAMP WITHOUT TIME ZONE
                    )
                    """
                )
            )
        else:
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS inspection_log_buildings (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        building_id INTEGER NOT NULL UNIQUE,
                        created_at DATETIME
                    )
                    """
                )
            )
            await conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS inspection_log_files (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        building_id INTEGER NOT NULL,
                        title VARCHAR(200) NOT NULL,
                        original_name VARCHAR(300),
                        stored_name VARCHAR(300) NOT NULL,
                        content_type VARCHAR(100),
                        file_data BLOB,
                        uploaded_by VARCHAR(100),
                        created_at DATETIME
                    )
                    """
                )
            )
        try:
            if is_pg:
                await conn.execute(
                    text(
                        "ALTER TABLE inspection_log_files "
                        "ADD COLUMN IF NOT EXISTS last_edit_pos JSONB"
                    )
                )
                await conn.execute(
                    text(
                        "ALTER TABLE inspection_log_files "
                        "ADD COLUMN IF NOT EXISTS file_size INTEGER"
                    )
                )
            else:
                await conn.execute(
                    text(
                        "ALTER TABLE inspection_log_files "
                        "ADD COLUMN last_edit_pos TEXT"
                    )
                )
                await conn.execute(
                    text(
                        "ALTER TABLE inspection_log_files "
                        "ADD COLUMN file_size INTEGER"
                    )
                )
        except Exception:
            pass


def _inspection_log_excel_ok(filename: str) -> bool:
    name = (filename or "").lower()
    return name.endswith((".xls", ".xlsx", ".xlsm"))


async def _load_inspection_log_bytes(
    doc: InspectionLogFile,
    building_id: int,
    db: AsyncSession,
) -> bytes | None:
    data = doc.file_data
    if data:
        return bytes(data)
    file_path = (
        Path("static")
        / "uploads"
        / "buildings"
        / str(building_id)
        / "inspection_logs"
        / doc.stored_name
    )
    if file_path.is_file():
        raw = file_path.read_bytes()
        if raw:
            doc.file_data = raw
            doc.file_size = len(raw)
            await db.commit()
            return raw
    return None


async def _persist_inspection_log_bytes(
    doc: InspectionLogFile,
    raw: bytes,
    *,
    uploaded_by: str | None = None,
) -> None:
    """점검일지 바이너리를 DB에 저장 (OnlyOffice/간단편집기 공용)."""
    orig = doc.original_name or doc.stored_name or "inspection.xlsx"
    stem = Path(orig).stem or "inspection"
    # OnlyOffice 저장본은 보통 xlsx
    lower = (orig or "").lower()
    if lower.endswith((".xlsx", ".xlsm", ".xls")):
        new_name = orig[:300]
        if lower.endswith(".xls") and not lower.endswith(".xlsx") and not lower.endswith(".xlsm"):
            new_name = f"{stem}.xlsx"[:300]
            doc.content_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        elif lower.endswith(".xlsx"):
            doc.content_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        elif lower.endswith(".xlsm"):
            doc.content_type = "application/vnd.ms-excel.sheet.macroEnabled.12"
    else:
        new_name = f"{stem}.xlsx"[:300]
        doc.content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    doc.file_data = raw
    doc.file_size = len(raw)
    doc.original_name = new_name
    if uploaded_by:
        doc.uploaded_by = uploaded_by


def _inspection_media_type(filename: str, content_type: str | None = None) -> str:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if name.endswith(".xlsm"):
        return "application/vnd.ms-excel.sheet.macroEnabled.12"
    if name.endswith(".xls"):
        return "application/vnd.ms-excel"
    return (content_type or "").strip() or "application/octet-stream"

@app.get("/admin/inspection-logs")
async def inspection_logs_page(
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    try:
        await _ensure_inspection_log_tables()
    except Exception as e:
        print(f"[inspection_logs] ensure tables: {e}", flush=True)

    selected_rows = (
        await db.execute(
            select(InspectionLogBuilding, Building)
            .join(Building, Building.id == InspectionLogBuilding.building_id)
            .where(Building.is_active == True)  # noqa: E712
            .options(selectinload(Building.site))
            .order_by(InspectionLogBuilding.id.desc())
        )
    ).all()
    selected_buildings = _sort_buildings([b for _, b in selected_rows])
    selected_ids = {b.id for b in selected_buildings}

    all_buildings = _sort_buildings(
        list(
            (
                await db.execute(
                    select(Building)
                    .where(Building.is_active == True)  # noqa: E712
                    .options(selectinload(Building.site))
                )
            ).scalars().all()
        )
    )
    available_buildings = [b for b in all_buildings if b.id not in selected_ids]
    building_groups = group_buildings_by_site(selected_buildings)

    return templates.TemplateResponse(
        request,
        "inspection_logs.html",
        {
            "user": user,
            "selected_buildings": selected_buildings,
            "building_groups": building_groups,
            "available_buildings": available_buildings,
            "error": request.query_params.get("error"),
            "message": request.query_params.get("message"),
        },
    )


@app.post("/admin/inspection-logs/buildings")
async def inspection_logs_add_building(
    building_id: int = Form(...),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    try:
        await _ensure_inspection_log_tables()
    except Exception:
        pass

    building = await db.get(Building, building_id)
    if not building or not building.is_active:
        return RedirectResponse(
            "/admin/inspection-logs?error=" + quote("건물을 찾을 수 없습니다."),
            status_code=303,
        )
    exists = (
        await db.execute(
            select(InspectionLogBuilding).where(
                InspectionLogBuilding.building_id == building_id
            )
        )
    ).scalar_one_or_none()
    if exists:
        return RedirectResponse(
            "/admin/inspection-logs?message=" + quote("이미 등록된 건물입니다."),
            status_code=303,
        )
    db.add(InspectionLogBuilding(building_id=building_id))
    await db.commit()
    return RedirectResponse(
        "/admin/inspection-logs?message="
        + quote(f"「{building.name}」 건물이 추가되었습니다."),
        status_code=303,
    )


@app.post("/admin/inspection-logs/buildings/{building_id}/remove")
async def inspection_logs_remove_building(
    building_id: int,
    user: User = Depends(require_can_delete),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    row = (
        await db.execute(
            select(InspectionLogBuilding).where(
                InspectionLogBuilding.building_id == building_id
            )
        )
    ).scalar_one_or_none()
    if row:
        await db.delete(row)
        await db.commit()
    return RedirectResponse(
        "/admin/inspection-logs?message=" + quote("건물 등록이 해제되었습니다."),
        status_code=303,
    )


@app.get("/admin/inspection-logs/{building_id}")
async def inspection_log_building_detail(
    building_id: int,
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    try:
        await _ensure_inspection_log_tables()
    except Exception:
        pass

    registered = (
        await db.execute(
            select(InspectionLogBuilding).where(
                InspectionLogBuilding.building_id == building_id
            )
        )
    ).scalar_one_or_none()
    if not registered:
        from urllib.parse import quote

        return RedirectResponse(
            "/admin/inspection-logs?error="
            + quote("점검일지에 등록되지 않은 건물입니다. 먼저 건물을 추가하세요."),
            status_code=303,
        )

    building = (
        await db.execute(
            select(Building)
            .where(Building.id == building_id, Building.is_active == True)  # noqa: E712
            .options(selectinload(Building.site))
        )
    ).scalar_one_or_none()
    if not building:
        raise HTTPException(404, detail="건물을 찾을 수 없습니다.")

    files = (
        await db.execute(
            select(InspectionLogFile)
            .where(InspectionLogFile.building_id == building_id)
            .order_by(InspectionLogFile.id.desc())
        )
    ).scalars().all()
    await _backfill_attachment_sizes(db, files)

    return templates.TemplateResponse(
        request,
        "inspection_log_detail.html",
        {
            "user": user,
            "building": building,
            "files": files,
            "upload_max_mb": UPLOAD_MAX_FILE_MB,
            "error": request.query_params.get("error"),
            "message": request.query_params.get("message"),
        },
    )


@app.post("/admin/inspection-logs/{building_id}/upload")
async def inspection_log_upload(
    building_id: int,
    request: Request,
    title: str = Form(""),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote
    import uuid

    try:
        await _ensure_inspection_log_tables()
    except Exception:
        pass

    registered = (
        await db.execute(
            select(InspectionLogBuilding).where(
                InspectionLogBuilding.building_id == building_id
            )
        )
    ).scalar_one_or_none()
    building = await db.get(Building, building_id)
    if not registered or not building or not building.is_active:
        return RedirectResponse(
            "/admin/inspection-logs?error=" + quote("등록된 건물이 아닙니다."),
            status_code=303,
        )

    form = await request.form()
    uploads = form.getlist("files") if hasattr(form, "getlist") else []
    if not uploads:
        one = form.get("file")
        uploads = [one] if one else []

    saved = 0
    errors: list[str] = []
    for item in uploads[:UPLOAD_MAX_FILES_PER_REQUEST]:
        if not hasattr(item, "filename") or not item.filename:
            continue
        fname = item.filename
        if not _inspection_log_excel_ok(fname):
            errors.append(f"{fname}: 엑셀 파일만 업로드 가능합니다.")
            continue
        raw = await item.read()
        if not raw:
            errors.append(f"{fname}: 빈 파일입니다.")
            continue
        if len(raw) > UPLOAD_MAX_FILE_BYTES:
            errors.append(f"{fname}: {UPLOAD_MAX_FILE_MB}MB 초과")
            continue
        ext = Path(fname).suffix.lower() or ".xlsx"
        stored = f"{uuid.uuid4().hex}{ext}"
        display_title = (title or "").strip() or Path(fname).stem
        db.add(
            InspectionLogFile(
                building_id=building_id,
                title=display_title[:200],
                original_name=fname[:300],
                stored_name=stored,
                content_type=getattr(item, "content_type", None) or "",
                file_data=raw,
                file_size=len(raw),
                uploaded_by=user.name or user.username,
            )
        )
        saved += 1

    if saved:
        await db.commit()
    if saved and not errors:
        msg = f"엑셀 파일 {saved}개가 업로드되었습니다."
        return RedirectResponse(
            f"/admin/inspection-logs/{building_id}?message=" + quote(msg),
            status_code=303,
        )
    if saved and errors:
        msg = f"{saved}개 저장, 일부 실패: " + "; ".join(errors[:3])
        return RedirectResponse(
            f"/admin/inspection-logs/{building_id}?message=" + quote(msg),
            status_code=303,
        )
    err = errors[0] if errors else "업로드할 엑셀 파일을 선택하세요."
    return RedirectResponse(
        f"/admin/inspection-logs/{building_id}?error=" + quote(err),
        status_code=303,
    )


@app.get("/admin/inspection-logs/{building_id}/files/{file_id}/file")
async def inspection_log_file_download(
    building_id: int,
    file_id: int,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    doc = await db.get(InspectionLogFile, file_id)
    if not doc or doc.building_id != building_id:
        raise HTTPException(404, detail="파일을 찾을 수 없습니다.")
    data = await _load_inspection_log_bytes(doc, building_id, db)
    if not data:
        raise HTTPException(404, detail="파일 데이터가 없습니다. 다시 업로드해 주세요.")

    filename = doc.original_name or doc.stored_name or "inspection.xlsx"
    ascii_name = filename.encode("ascii", "ignore").decode("ascii") or "inspection.xlsx"
    media = _inspection_media_type(filename, doc.content_type)

    return Response(
        content=data,
        media_type=media,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{quote(filename)}'
            ),
            "Cache-Control": "private, max-age=3600",
        },
    )


@app.get("/admin/inspection-logs/{building_id}/files/{file_id}/edit")
async def inspection_log_file_edit(
    building_id: int,
    file_id: int,
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
    editor: str | None = Query(None),
):
    """점검일지 편집. OnlyOffice 설정 시 Docs 편집기, ?editor=legacy 로 간단 편집기."""
    import json as _json

    registered = (
        await db.execute(
            select(InspectionLogBuilding).where(
                InspectionLogBuilding.building_id == building_id
            )
        )
    ).scalar_one_or_none()
    building = (
        await db.execute(
            select(Building)
            .where(Building.id == building_id, Building.is_active == True)  # noqa: E712
            .options(selectinload(Building.site))
        )
    ).scalar_one_or_none()
    doc = await db.get(InspectionLogFile, file_id)
    if not registered or not building or not doc or doc.building_id != building_id:
        raise HTTPException(404, detail="파일을 찾을 수 없습니다.")

    can_save = can_edit(user)
    file_url = f"/admin/inspection-logs/{building_id}/files/{file_id}/file"
    legacy_url = (
        f"/admin/inspection-logs/{building_id}/files/{file_id}/edit?editor=legacy"
    )
    use_legacy = (editor or "").lower() in ("legacy", "simple", "js")
    want_oo = oo.onlyoffice_enabled() and not use_legacy
    cursor_url = f"/admin/inspection-logs/{building_id}/files/{file_id}/cursor"
    last_edit_pos = getattr(doc, "last_edit_pos", None) or None

    if want_oo:
        oo_error = None
        editor_config = None
        data = await _load_inspection_log_bytes(doc, building_id, db)
        if not data:
            oo_error = "파일 데이터가 없습니다. 다시 업로드해 주세요."
        else:
            try:
                editor_config = oo.build_editor_config(
                    request=request,
                    building_id=building_id,
                    file_id=file_id,
                    filename=doc.original_name or doc.stored_name or "inspection.xlsx",
                    title=doc.title or doc.original_name or "점검일지",
                    file_bytes=data,
                    user_id=user.id,
                    user_name=user.name or user.username or f"user-{user.id}",
                    can_edit=can_save,
                )
            except Exception as e:
                oo_error = f"OnlyOffice 설정 생성 실패: {e}"

        return templates.TemplateResponse(
            request,
            "inspection_log_onlyoffice.html",
            {
                "user": user,
                "building": building,
                "doc": doc,
                "file_url": file_url,
                "legacy_edit_url": legacy_url,
                "can_edit": can_save,
                "oo_enabled": True,
                "oo_error": oo_error,
                "onlyoffice_api_js": f"{oo.onlyoffice_url()}/web-apps/apps/api/documents/api.js",
                "editor_config_json": _json.dumps(
                    editor_config or {}, ensure_ascii=False
                ),
                "cursor_url": cursor_url,
                "last_edit_pos": last_edit_pos,
                "file_id": file_id,
            },
        )

    return templates.TemplateResponse(
        request,
        "inspection_log_edit.html",
        {
            "user": user,
            "building": building,
            "doc": doc,
            "file_url": file_url,
            "save_url": f"/admin/inspection-logs/{building_id}/files/{file_id}/save",
            "can_save": can_save,
            "onlyoffice_available": oo.onlyoffice_enabled(),
            "onlyoffice_edit_url": (
                f"/admin/inspection-logs/{building_id}/files/{file_id}/edit"
            ),
            "cursor_url": cursor_url,
            "last_edit_pos": last_edit_pos,
            "file_id": file_id,
        },
    )


@app.post("/admin/inspection-logs/{building_id}/files/{file_id}/cursor")
async def inspection_log_file_cursor(
    building_id: int,
    file_id: int,
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    """마지막 편집 셀 위치 저장 (다시 열 때 복원용)."""
    doc = await db.get(InspectionLogFile, file_id)
    if not doc or doc.building_id != building_id:
        raise HTTPException(404, detail="파일을 찾을 수 없습니다.")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, detail="잘못된 요청입니다.")
    if not isinstance(body, dict):
        raise HTTPException(400, detail="잘못된 요청입니다.")
    pos = {
        "sheet": str(body.get("sheet") or "")[:80] or None,
        "sheetIndex": body.get("sheetIndex"),
        "cell": str(body.get("cell") or "")[:20] or None,
        "x": body.get("x"),
        "y": body.get("y"),
        "user_id": user.id,
        "updated_at": datetime.utcnow().isoformat() + "Z",
    }
    try:
        if pos["sheetIndex"] is not None:
            pos["sheetIndex"] = int(pos["sheetIndex"])
        if pos["x"] is not None:
            pos["x"] = int(pos["x"])
        if pos["y"] is not None:
            pos["y"] = int(pos["y"])
    except (TypeError, ValueError):
        raise HTTPException(400, detail="위치 값이 올바르지 않습니다.")
    doc.last_edit_pos = pos
    await db.commit()
    return JSONResponse({"ok": True})


@app.get("/oo/inspection-logs/{building_id}/files/{file_id}/content")
async def onlyoffice_file_content(
    building_id: int,
    file_id: int,
    token: str = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """OnlyOffice Document Server가 문서를 내려받는 URL (세션 쿠키 불필요)."""
    from urllib.parse import quote

    try:
        oo.verify_access_token(
            token, building_id=building_id, file_id=file_id, purpose="content"
        )
    except ValueError as e:
        raise HTTPException(403, detail=str(e)) from e

    doc = await db.get(InspectionLogFile, file_id)
    if not doc or doc.building_id != building_id:
        raise HTTPException(404, detail="파일을 찾을 수 없습니다.")
    data = await _load_inspection_log_bytes(doc, building_id, db)
    if not data:
        raise HTTPException(404, detail="파일 데이터가 없습니다.")

    filename = doc.original_name or doc.stored_name or "inspection.xlsx"
    ascii_name = filename.encode("ascii", "ignore").decode("ascii") or "inspection.xlsx"
    return Response(
        content=data,
        media_type=_inspection_media_type(filename, doc.content_type),
        headers={
            "Content-Disposition": (
                f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{quote(filename)}'
            ),
            "Cache-Control": "no-store",
        },
    )


@app.post("/oo/inspection-logs/{building_id}/files/{file_id}/callback")
async def onlyoffice_file_callback(
    building_id: int,
    file_id: int,
    request: Request,
    token: str = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """OnlyOffice 저장 콜백. status 2/6 일 때 서버에 파일 반영."""
    import requests as _requests

    try:
        payload_tok = oo.verify_access_token(
            token, building_id=building_id, file_id=file_id, purpose="callback"
        )
    except ValueError as e:
        raise HTTPException(403, detail=str(e)) from e

    try:
        body = await request.json()
    except Exception:
        body = {}

    try:
        oo.verify_callback_jwt(request.headers.get("Authorization"), body)
    except ValueError as e:
        if oo.jwt_enabled():
            raise HTTPException(403, detail=str(e)) from e

    status = int(body.get("status") or 0)
    if status in (2, 6):
        if not payload_tok.get("edit"):
            return JSONResponse({"error": 1})
        url = (body.get("url") or "").strip()
        if not url:
            return JSONResponse({"error": 1})
        doc = await db.get(InspectionLogFile, file_id)
        if not doc or doc.building_id != building_id:
            return JSONResponse({"error": 1})
        try:
            resp = _requests.get(url, timeout=120)
            resp.raise_for_status()
            raw = resp.content
        except Exception as e:
            print(f"[onlyoffice] download save failed: {e}", flush=True)
            return JSONResponse({"error": 1})
        if not raw:
            return JSONResponse({"error": 1})
        if len(raw) > UPLOAD_MAX_FILE_BYTES:
            print("[onlyoffice] file too large", flush=True)
            return JSONResponse({"error": 1})
        await _persist_inspection_log_bytes(
            doc,
            raw,
            uploaded_by=f"onlyoffice:{payload_tok.get('uid') or ''}",
        )
        await db.commit()
        print(
            f"[onlyoffice] saved building={building_id} file={file_id} bytes={len(raw)} status={status}",
            flush=True,
        )
    return JSONResponse({"error": 0})


@app.post("/admin/inspection-logs/{building_id}/files/{file_id}/save")
async def inspection_log_file_save(
    building_id: int,
    file_id: int,
    request: Request,
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    """브라우저에서 수정한 엑셀(base64 xlsx)을 저장."""
    import base64

    doc = await db.get(InspectionLogFile, file_id)
    if not doc or doc.building_id != building_id:
        raise HTTPException(404, detail="파일을 찾을 수 없습니다.")

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, detail="잘못된 요청입니다.")

    b64 = (body.get("data_base64") or "").strip()
    if not b64:
        raise HTTPException(400, detail="저장할 데이터가 없습니다.")
    if "," in b64 and b64.lower().startswith("data:"):
        b64 = b64.split(",", 1)[1]
    try:
        raw = base64.b64decode(b64)
    except Exception:
        raise HTTPException(400, detail="파일 디코딩에 실패했습니다.")
    if not raw:
        raise HTTPException(400, detail="빈 파일입니다.")
    if len(raw) > UPLOAD_MAX_FILE_BYTES:
        raise HTTPException(
            400, detail=f"파일이 너무 큽니다. (최대 {UPLOAD_MAX_FILE_MB}MB)"
        )

    await _persist_inspection_log_bytes(
        doc, raw, uploaded_by=user.name or user.username
    )
    await db.commit()
    new_name = doc.original_name or "inspection.xlsx"
    return JSONResponse(
        {
            "ok": True,
            "message": "저장되었습니다.",
            "filename": new_name,
        }
    )


@app.post("/admin/inspection-logs/{building_id}/files/{file_id}/delete")
async def inspection_log_file_delete(
    building_id: int,
    file_id: int,
    user: User = Depends(require_can_delete),
    db: AsyncSession = Depends(get_db),
):
    from urllib.parse import quote

    doc = await db.get(InspectionLogFile, file_id)
    if not doc or doc.building_id != building_id:
        raise HTTPException(404)
    await db.delete(doc)
    await db.commit()
    return RedirectResponse(
        f"/admin/inspection-logs/{building_id}?message=" + quote("파일이 삭제되었습니다."),
        status_code=303,
    )


@app.get("/admin/pm")
async def pm_list(
    request: Request,
    q: str = Query(""),
    building_id: int | None = Query(None),
    equipment_id: int | None = Query(None),
    due: str = Query(""),
    tab: str = Query("list"),
    page: int = Query(1),
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    today = _today_kst()
    due_only = due in ("1", "due", "overdue")
    active_tab = tab if tab in ("list", "settings") else "list"

    # 점검목록: 건물 선택 전에는 목록을 조회·표시하지 않음
    if active_tab == "list" and not building_id:
        all_schedules: list = []
    else:
        all_schedules = await _pm_filtered_schedules(
            db,
            q=q,
            building_id=building_id,
            equipment_id=equipment_id,
            due_only=due_only,
        )

    pager = _paginate(list(all_schedules), page)
    schedules = pager["items"]

    buildings = _sort_buildings(
        (
            await db.execute(
                select(Building).where(Building.is_active == True).order_by(Building.name)  # noqa: E712
            )
        ).scalars().all()
    )

    eq_q = (
        select(Equipment)
        .where(Equipment.is_active == True)  # noqa: E712
        .options(
            selectinload(Equipment.zone)
            .selectinload(Zone.floor)
            .selectinload(Floor.building),
            selectinload(Equipment.pm_schedules),
        )
        .order_by(Equipment.code)
    )
    if building_id:
        eq_q = (
            eq_q.join(Zone, Equipment.zone_id == Zone.id)
            .join(Floor, Zone.floor_id == Floor.id)
            .where(Floor.building_id == building_id)
        )
    equipment_list = (await db.execute(eq_q)).scalars().unique().all()

    # 건물별·설비별 그룹 (목록 탭) — 현재 페이지 일정만
    grouped: dict[str, dict] = {}
    for pm in schedules:
        eq = pm.equipment
        bld = _equipment_building(eq)
        bname = bld.name if bld else "미지정 건물"
        bid = bld.id if bld else 0
        bucket = grouped.setdefault(
            str(bid),
            {"building_id": bid, "building_name": bname, "equipment": {}},
        )
        ek = str(eq.id) if eq else "0"
        e_bucket = bucket["equipment"].setdefault(
            ek,
            {
                "equipment": eq,
                "schedules": [],
            },
        )
        e_bucket["schedules"].append(pm)

    grouped_list = sorted(
        grouped.values(),
        key=lambda g: _building_sort_key(g["building_name"]),
    )
    for g in grouped_list:
        g["equipment_list"] = sorted(
            g["equipment"].values(),
            key=lambda e: (e["equipment"].code if e["equipment"] else ""),
        )

    return templates.TemplateResponse(
        request,
        "pm.html",
        {
            "user": user,
            "schedules": schedules,
            "grouped_list": grouped_list,
            "buildings": buildings,
            "equipment_list": equipment_list,
            "today": today,
            "tab": active_tab,
            "pager": pager,
            "filters": {
                "q": q,
                "building_id": building_id,
                "equipment_id": equipment_id,
                "due": due_only,
                "page": pager["page"],
            },
            "freq_choices": [
                (PMFrequency.daily, "매일"),
                (PMFrequency.weekly, "매주"),
                (PMFrequency.monthly, "매월"),
                (PMFrequency.quarterly, "분기"),
                (PMFrequency.semi_annual, "반기"),
                (PMFrequency.annual, "연간"),
                (PMFrequency.custom, "사용자지정(일)"),
            ],
            "message": request.query_params.get("msg", ""),
            "error": request.query_params.get("error", ""),
        },
    )


@app.get("/admin/pm/export")
async def pm_export(
    q: str = Query(""),
    building_id: int | None = Query(None),
    equipment_id: int | None = Query(None),
    due: str = Query(""),
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    due_only = due in ("1", "due", "overdue")
    schedules = await _pm_filtered_schedules(
        db,
        q=q,
        building_id=building_id,
        equipment_id=equipment_id,
        due_only=due_only,
    )
    return _pm_excel_response(schedules)


@app.post("/admin/pm/schedules")
async def pm_schedule_upsert(
    equipment_id: int = Form(...),
    title: str = Form(""),
    frequency: str = Form("monthly"),
    custom_days: str = Form(""),
    assignee_name: str = Form(""),
    next_due: str = Form(""),
    building_id: str = Form(""),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    eq = (
        await db.execute(
            select(Equipment)
            .where(Equipment.id == equipment_id, Equipment.is_active == True)  # noqa: E712
            .options(selectinload(Equipment.pm_schedules))
        )
    ).scalar_one_or_none()
    if not eq:
        return RedirectResponse("/admin/pm?tab=settings&error=설비를+찾을+수+없습니다", status_code=303)

    freq = _parse_pm_frequency(frequency)
    days_val = None
    if freq == PMFrequency.custom:
        try:
            days_val = max(1, int(custom_days.strip() or "30"))
        except ValueError:
            days_val = 30
    else:
        days_val = _pm_cycle_days(freq, None)

    due_date = None
    if next_due.strip():
        try:
            due_date = date.fromisoformat(next_due.strip())
        except ValueError:
            due_date = None
    if due_date is None:
        due_date = _today_kst() + timedelta(days=days_val)

    existing = next((s for s in (eq.pm_schedules or []) if s.is_active), None)
    pm_title = title.strip() or f"{eq.code} 예방점검"
    if existing:
        existing.title = pm_title
        existing.frequency = freq
        existing.custom_days = days_val if freq == PMFrequency.custom else None
        existing.assignee_name = assignee_name.strip() or None
        existing.next_due = due_date
        existing.is_active = True
    else:
        db.add(
            PMSchedule(
                equipment_id=eq.id,
                title=pm_title,
                frequency=freq,
                custom_days=days_val if freq == PMFrequency.custom else None,
                assignee_name=assignee_name.strip() or None,
                next_due=due_date,
                is_active=True,
            )
        )
    await db.commit()

    qs = ["tab=settings", "msg=점검주기가+저장되었습니다"]
    if building_id.strip():
        qs.append(f"building_id={building_id.strip()}")
    return RedirectResponse(f"/admin/pm?{'&'.join(qs)}", status_code=303)


@app.post("/admin/pm/schedules/{schedule_id}/deactivate")
async def pm_schedule_deactivate(
    schedule_id: int,
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    schedule = await db.get(PMSchedule, schedule_id)
    if not schedule:
        raise HTTPException(404)
    schedule.is_active = False
    await db.commit()
    return RedirectResponse("/admin/pm?tab=settings&msg=주기가+해제되었습니다", status_code=303)


@app.post("/admin/pm/schedules/{schedule_id}/inspect")
async def pm_inspect(
    schedule_id: int,
    result: str = Form(...),
    note: str = Form(""),
    inspector_name: str = Form(""),
    request_work_order: str = Form("0"),
    redirect_to: str = Form(""),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    schedule = (
        await db.execute(
            select(PMSchedule)
            .where(PMSchedule.id == schedule_id, PMSchedule.is_active == True)  # noqa: E712
            .options(
                selectinload(PMSchedule.equipment)
                .selectinload(Equipment.zone)
                .selectinload(Zone.floor)
                .selectinload(Floor.building)
            )
        )
    ).scalar_one_or_none()
    if not schedule:
        raise HTTPException(404)

    insp, wo = await _record_pm_inspection(
        db,
        schedule,
        result_raw=result,
        note=note,
        inspector_name=inspector_name or user.name,
        create_work_order=request_work_order.strip().lower() in ("1", "true", "yes"),
    )
    await db.commit()

    if wo:
        return RedirectResponse(f"/admin/work-orders/{wo.id}", status_code=303)
    target = redirect_to.strip() or "/admin/pm?msg=점검결과가+저장되었습니다"
    return RedirectResponse(target, status_code=303)


MATERIAL_DEFAULT_GROUPS = ("소모품", "수공구", "비품", "안전용품", "풍수해 자재")


def _mat_qs(popup: int = 0, **extra) -> str:
    from urllib.parse import urlencode

    params = {k: v for k, v in extra.items() if v not in (None, "", 0, False)}
    if popup:
        params["popup"] = "1"
    return f"?{urlencode(params)}" if params else ""


async def _material_record_log(
    db: AsyncSession, action: str, name: str, quantity: int = 0, reason: str = ""
) -> None:
    db.add(
        MaterialLog(
            action=action,
            name=name,
            quantity=quantity,
            reason=reason or None,
        )
    )


async def _list_material_groups(db: AsyncSession) -> list[str]:
    """조회 전용: DB·자재·기본 그룹을 합쳐 정렬 (GET에서 commit 하지 않음)."""
    existing = {
        g.name
        for g in (await db.execute(select(MaterialGroup))).scalars().all()
        if g.name
    }
    item_groups = {
        g
        for g in (await db.execute(select(MaterialItem.group_name).distinct())).scalars().all()
        if g
    }
    return sorted(set(MATERIAL_DEFAULT_GROUPS) | existing | item_groups)


async def _ensure_material_groups(db: AsyncSession) -> list[str]:
    """쓰기 경로용: 기본/자재 그룹을 DB에 반영 후 목록 반환."""
    existing = {
        g.name
        for g in (await db.execute(select(MaterialGroup))).scalars().all()
        if g.name
    }
    item_groups = {
        g
        for g in (await db.execute(select(MaterialItem.group_name).distinct())).scalars().all()
        if g
    }
    for name in list(MATERIAL_DEFAULT_GROUPS) + list(item_groups):
        if name and name not in existing:
            db.add(MaterialGroup(name=name))
            existing.add(name)
    await db.flush()
    return await _list_material_groups(db)


def _material_item_dict(it: MaterialItem) -> dict:
    return {
        "id": it.id,
        "name": it.name,
        "quantity": int(it.quantity or 0),
        "spec": it.spec or "",
        "remarks": it.remarks or "",
        "group_name": it.group_name or "",
        "location": it.location or "",
    }


def _material_log_dict(lg: MaterialLog) -> dict:
    return {
        "id": lg.id,
        "action": lg.action,
        "name": lg.name,
        "quantity": int(lg.quantity or 0),
        "reason": lg.reason or "",
        "created_at": lg.created_at,
        "created_at_fmt": _fmt_kst(lg.created_at),
    }


@app.get("/admin/materials")
async def materials_page(
    request: Request,
    q: str = "",
    group: str = "",
    popup: int = 0,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    # GET은 읽기만 — commit 금지(템플릿 렌더 시 만료된 user ORM 접근 → greenlet 500 방지)
    try:
        groups = await _list_material_groups(db)

        q_val = (q or "").strip()
        group_val = (group or "").strip()
        stmt = select(MaterialItem).order_by(MaterialItem.group_name, MaterialItem.name)
        if group_val:
            stmt = stmt.where(MaterialItem.group_name == group_val)
        if q_val:
            like = f"%{q_val}%"
            stmt = stmt.where(
                or_(
                    MaterialItem.name.ilike(like),
                    MaterialItem.spec.ilike(like),
                    MaterialItem.remarks.ilike(like),
                    MaterialItem.group_name.ilike(like),
                    MaterialItem.location.ilike(like),
                )
            )

        items = [_material_item_dict(it) for it in (await db.execute(stmt)).scalars().all()]
        all_items = [
            _material_item_dict(it)
            for it in (
                await db.execute(select(MaterialItem).order_by(MaterialItem.name))
            ).scalars().all()
        ]
        items_json = json.dumps(all_items, ensure_ascii=False)
        logs = [
            _material_log_dict(lg)
            for lg in (
                await db.execute(
                    select(MaterialLog)
                    .order_by(MaterialLog.created_at.desc())
                    .limit(300)
                )
            ).scalars().all()
        ]
    except Exception as e:
        print(f"[materials] page failed: {e}", flush=True)
        await db.rollback()
        return templates.TemplateResponse(
            request,
            "error.html",
            {
                "user": user,
                "status_code": 500,
                "message": "자재관리 화면을 불러오지 못했습니다.",
                "detail": str(e)[:500],
            },
            status_code=500,
        )

    return templates.TemplateResponse(
        request,
        "materials.html",
        {
            "user": user,
            "items": items,
            "all_items": all_items,
            "items_json": items_json,
            "logs": logs,
            "groups": groups,
            "q": q_val,
            "selected_group": group_val,
            "popup": bool(popup),
            "message": request.query_params.get("message") or "",
            "error": request.query_params.get("error") or "",
            "open_dlg": request.query_params.get("open") or "",
        },
    )


@app.post("/admin/materials/add")
async def materials_add(
    name: str = Form(...),
    quantity: str = Form("0"),
    group_name: str = Form(...),
    spec: str = Form(""),
    remarks: str = Form(""),
    location: str = Form(""),
    popup: int = Form(0),
    filter_group: str = Form(""),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    nm = name.strip()
    grp = group_name.strip()
    if not nm or not grp:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error='자재 이름과 그룹을 입력하세요.', group=filter_group)}",
            status_code=303,
        )
    groups = await _ensure_material_groups(db)
    if grp not in groups:
        db.add(MaterialGroup(name=grp))
        await db.flush()
    try:
        qty = int(str(quantity).strip() or 0)
    except ValueError:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error='수량은 숫자로 입력하세요.', group=filter_group)}",
            status_code=303,
        )
    existing = (
        await db.execute(select(MaterialItem).where(MaterialItem.name == nm))
    ).scalar_one_or_none()
    if existing:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error=f'이미 등록된 자재입니다: {nm}', group=filter_group)}",
            status_code=303,
        )
    db.add(
        MaterialItem(
            name=nm,
            quantity=qty,
            spec=spec.strip() or None,
            remarks=remarks.strip() or None,
            group_name=grp,
            location=location.strip() or None,
        )
    )
    await _material_record_log(db, "등록", nm, qty)
    await db.commit()
    return RedirectResponse(
        f"/admin/materials{_mat_qs(popup, message=f'{nm} 자재가 추가되었습니다. 수량: {qty}', group=filter_group or grp)}",
        status_code=303,
    )


@app.post("/admin/materials/stock-in")
async def materials_stock_in(
    name: str = Form(...),
    quantity: str = Form(...),
    spec: str = Form(""),
    remarks: str = Form(""),
    location: str = Form(""),
    group_name: str = Form(""),
    popup: int = Form(0),
    filter_group: str = Form(""),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    nm = name.strip()
    try:
        qty = int(str(quantity).strip())
    except ValueError:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error='수량은 정수로 입력하세요.', group=filter_group)}",
            status_code=303,
        )
    if qty <= 0:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error='입고 수량은 1 이상이어야 합니다.', group=filter_group)}",
            status_code=303,
        )
    item = (
        await db.execute(select(MaterialItem).where(MaterialItem.name == nm))
    ).scalar_one_or_none()
    if not item:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error=f'{nm}이(가) 자재 목록에 없습니다.', group=filter_group)}",
            status_code=303,
        )
    item.quantity = int(item.quantity or 0) + qty
    if spec.strip():
        item.spec = spec.strip()
    if remarks.strip():
        item.remarks = remarks.strip()
    if location.strip():
        item.location = location.strip()
    if group_name.strip():
        item.group_name = group_name.strip()
    await _material_record_log(db, "입고", nm, qty)
    await db.commit()
    return RedirectResponse(
        f"/admin/materials{_mat_qs(popup, message=f'{nm} 입고 완료. 총 수량: {item.quantity}', group=filter_group)}",
        status_code=303,
    )


@app.post("/admin/materials/stock-out")
async def materials_stock_out(
    name: str = Form(...),
    quantity: str = Form(...),
    reason: str = Form(""),
    popup: int = Form(0),
    filter_group: str = Form(""),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    nm = name.strip()
    try:
        qty = int(str(quantity).strip())
    except ValueError:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error='수량은 정수로 입력하세요.', group=filter_group)}",
            status_code=303,
        )
    if qty <= 0:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error='출고 수량은 1 이상이어야 합니다.', group=filter_group)}",
            status_code=303,
        )
    reason_text = (reason or "").strip()
    if not reason_text:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error='출고 사유를 입력하세요.', group=filter_group)}",
            status_code=303,
        )
    item = (
        await db.execute(select(MaterialItem).where(MaterialItem.name == nm))
    ).scalar_one_or_none()
    if not item:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error=f'{nm}이(가) 자재 목록에 없습니다.', group=filter_group)}",
            status_code=303,
        )
    current = int(item.quantity or 0)
    if qty > current:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error=f'재고 부족: 현재 {current}개', group=filter_group)}",
            status_code=303,
        )
    item.quantity = current - qty
    await _material_record_log(db, "출고", nm, qty, reason_text)
    await db.commit()
    return RedirectResponse(
        f"/admin/materials{_mat_qs(popup, message=f'{nm} 출고 완료. 잔여: {item.quantity}', group=filter_group)}",
        status_code=303,
    )


@app.post("/admin/materials/delete")
async def materials_delete(
    name: str = Form(...),
    popup: int = Form(0),
    filter_group: str = Form(""),
    user: User = Depends(require_can_delete),
    db: AsyncSession = Depends(get_db),
):
    nm = name.strip()
    item = (
        await db.execute(select(MaterialItem).where(MaterialItem.name == nm))
    ).scalar_one_or_none()
    if not item:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error=f'{nm}이(가) 자재 목록에 없습니다.', group=filter_group)}",
            status_code=303,
        )
    qty = int(item.quantity or 0)
    await db.delete(item)
    await _material_record_log(db, "삭제", nm, qty)
    await db.commit()
    return RedirectResponse(
        f"/admin/materials{_mat_qs(popup, message=f'{nm} 자재가 삭제되었습니다.', group=filter_group)}",
        status_code=303,
    )


@app.post("/admin/materials/reset")
async def materials_reset(
    popup: int = Form(0),
    user: User = Depends(require_can_delete),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import delete

    await db.execute(delete(MaterialItem))
    await db.execute(delete(MaterialLog))
    await _material_record_log(db, "초기화", "(전체)", 0, "전체 데이터 초기화")
    await db.commit()
    return RedirectResponse(
        f"/admin/materials{_mat_qs(popup, message='전체 자재·로그가 초기화되었습니다.')}",
        status_code=303,
    )


@app.post("/admin/materials/groups/add")
async def materials_group_add(
    group_name: str = Form(...),
    popup: int = Form(0),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    name = group_name.strip()
    if not name:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error='그룹 이름을 입력하세요.', open='group-add')}",
            status_code=303,
        )
    existing = (
        await db.execute(select(MaterialGroup).where(MaterialGroup.name == name))
    ).scalar_one_or_none()
    if existing:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error=f'이미 있는 그룹입니다: {name}')}",
            status_code=303,
        )
    db.add(MaterialGroup(name=name))
    await db.commit()
    return RedirectResponse(
        f"/admin/materials{_mat_qs(popup, message=f'그룹 추가: {name}', group=name)}",
        status_code=303,
    )


@app.post("/admin/materials/groups/delete")
async def materials_group_delete(
    group_name: str = Form(...),
    popup: int = Form(0),
    user: User = Depends(require_can_delete),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import delete, update

    name = group_name.strip()
    grp = (
        await db.execute(select(MaterialGroup).where(MaterialGroup.name == name))
    ).scalar_one_or_none()
    if not grp:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error=f'그룹이 없습니다: {name}')}",
            status_code=303,
        )
    # 해당 그룹 자재는 소모품으로 이동(없으면 생성)
    fallback = "소모품"
    if name == fallback:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error='기본 그룹(소모품)은 삭제할 수 없습니다.')}",
            status_code=303,
        )
    fb = (
        await db.execute(select(MaterialGroup).where(MaterialGroup.name == fallback))
    ).scalar_one_or_none()
    if not fb:
        db.add(MaterialGroup(name=fallback))
    await db.execute(
        update(MaterialItem)
        .where(MaterialItem.group_name == name)
        .values(group_name=fallback)
    )
    await db.execute(delete(MaterialGroup).where(MaterialGroup.name == name))
    await db.commit()
    return RedirectResponse(
        f"/admin/materials{_mat_qs(popup, message=f'그룹 삭제: {name} (자재는 {fallback}으로 이동)')}",
        status_code=303,
    )


@app.post("/admin/materials/groups/rename")
async def materials_group_rename(
    old_name: str = Form(...),
    new_name: str = Form(...),
    popup: int = Form(0),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import update

    old = old_name.strip()
    new = new_name.strip()
    if not old or not new:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error='기존/새 그룹 이름을 입력하세요.', open='group-rename')}",
            status_code=303,
        )
    grp = (
        await db.execute(select(MaterialGroup).where(MaterialGroup.name == old))
    ).scalar_one_or_none()
    if not grp:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error=f'그룹이 없습니다: {old}')}",
            status_code=303,
        )
    clash = (
        await db.execute(select(MaterialGroup).where(MaterialGroup.name == new))
    ).scalar_one_or_none()
    if clash and new != old:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error=f'이미 있는 그룹 이름입니다: {new}')}",
            status_code=303,
        )
    grp.name = new
    await db.execute(
        update(MaterialItem)
        .where(MaterialItem.group_name == old)
        .values(group_name=new)
    )
    await db.commit()
    return RedirectResponse(
        f"/admin/materials{_mat_qs(popup, message=f'그룹 이름 변경: {old} → {new}', group=new)}",
        status_code=303,
    )


@app.get("/admin/materials/export")
async def materials_export(
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    from io import BytesIO
    from urllib.parse import quote

    from openpyxl import Workbook

    items = (
        await db.execute(
            select(MaterialItem).order_by(MaterialItem.group_name, MaterialItem.name)
        )
    ).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "자재현황"
    ws.append(["자재 이름", "수량", "사양", "비고", "그룹", "저장 위치"])
    for it in items:
        ws.append(
            [
                it.name,
                it.quantity,
                it.spec or "",
                it.remarks or "",
                it.group_name,
                it.location or "",
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now(KST).strftime("%Y%m%d_%H%M")
    filename = quote(f"자재현황_{stamp}.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@app.get("/admin/materials/export-logs")
async def materials_export_logs(
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    from io import BytesIO
    from urllib.parse import quote

    from openpyxl import Workbook

    logs = (
        await db.execute(select(MaterialLog).order_by(MaterialLog.created_at.desc()))
    ).scalars().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "로그"
    ws.append(["일시", "구분", "자재명", "수량", "사유"])
    for lg in logs:
        ws.append(
            [_fmt_kst(lg.created_at), lg.action, lg.name, lg.quantity, lg.reason or ""]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now(KST).strftime("%Y%m%d_%H%M")
    filename = quote(f"자재로그_{stamp}.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@app.post("/admin/materials/import")
async def materials_import(
    file: UploadFile = File(...),
    popup: int = Form(0),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    from io import BytesIO

    from openpyxl import load_workbook

    if not file.filename:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error='파일을 선택하세요.')}",
            status_code=303,
        )
    raw = await file.read()
    try:
        wb = load_workbook(BytesIO(raw), data_only=True)
        ws = wb.active
    except Exception as e:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error=f'Excel 읽기 실패: {e}')}",
            status_code=303,
        )

    created = 0
    updated = 0
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return RedirectResponse(
            f"/admin/materials{_mat_qs(popup, error='빈 파일입니다.')}",
            status_code=303,
        )

    header = [str(c or "").strip() for c in rows[0]]
    start = 1
    colmap = {"name": 0, "qty": 1, "spec": 2, "remarks": 3, "group": 4, "location": 5}
    joined = "|".join(header)
    if any(k in joined for k in ("자재", "그룹", "수량", "저장")):
        for i, h in enumerate(header):
            hl = h.lower()
            if "저장" in h or "location" in hl or "위치" in h:
                colmap["location"] = i
            elif "그룹" in h or "group" in hl:
                colmap["group"] = i
            elif "자재" in h or "이름" in h or "name" in hl or "품명" in h:
                colmap["name"] = i
            elif "수량" in h or "qty" in hl:
                colmap["qty"] = i
            elif "사양" in h or "spec" in hl:
                colmap["spec"] = i
            elif "비고" in h or "remark" in hl:
                colmap["remarks"] = i
    else:
        start = 0

    def cell(row, idx):
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return "" if v is None else str(v).strip()

    await _ensure_material_groups(db)
    known = {
        g.name
        for g in (await db.execute(select(MaterialGroup))).scalars().all()
    }

    for row in rows[start:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        nm = cell(row, colmap["name"])
        if not nm:
            continue
        grp = cell(row, colmap["group"]) or "소모품"
        if grp not in known:
            db.add(MaterialGroup(name=grp))
            known.add(grp)
        try:
            qty = int(float(cell(row, colmap["qty"]) or 0))
        except ValueError:
            qty = 0
        spec = cell(row, colmap["spec"]) or None
        remarks = cell(row, colmap["remarks"]) or None
        location = cell(row, colmap.get("location")) or None

        existing = (
            await db.execute(select(MaterialItem).where(MaterialItem.name == nm))
        ).scalar_one_or_none()
        if existing:
            existing.quantity = qty
            existing.spec = spec
            existing.remarks = remarks
            existing.group_name = grp
            existing.location = location
            updated += 1
        else:
            db.add(
                MaterialItem(
                    name=nm,
                    quantity=qty,
                    spec=spec,
                    remarks=remarks,
                    group_name=grp,
                    location=location,
                )
            )
            created += 1

    await _material_record_log(
        db, "등록", "(Excel가져오기)", created + updated, f"신규 {created} / 갱신 {updated}"
    )
    await db.commit()
    return RedirectResponse(
        f"/admin/materials{_mat_qs(popup, message=f'Excel 반영: 신규 {created}건, 갱신 {updated}건')}",
        status_code=303,
    )


@app.get("/admin/partners")
async def partners_list(
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    partners = (
        await db.execute(
            select(Partner)
            .where(Partner.is_active == True)
            .order_by(Partner.name)
        )
    ).scalars().all()
    return templates.TemplateResponse(
        request, "partners.html", {"user": user, "partners": partners}
    )


@app.post("/admin/partners")
async def partner_create(
    name: str = Form(...),
    code: str = Form(...),
    contact_name: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
    contract_end: str = Form(""),
    user: User = Depends(require_can_create),
    db: AsyncSession = Depends(get_db),
):
    code_val = code.strip()
    name_val = name.strip()
    if not name_val or not code_val:
        return RedirectResponse("/admin/partners?error=required", status_code=303)

    existing = (
        await db.execute(select(Partner).where(Partner.code == code_val))
    ).scalar_one_or_none()
    if existing and existing.is_active:
        return RedirectResponse("/admin/partners?error=code", status_code=303)
    if existing and not existing.is_active:
        existing.is_active = True
        existing.name = name_val
        existing.contact_name = contact_name.strip() or None
        existing.phone = phone.strip() or None
        existing.email = email.strip() or None
        existing.contract_end = (
            date.fromisoformat(contract_end) if contract_end.strip() else None
        )
        await db.commit()
        return RedirectResponse("/admin/partners", status_code=303)

    end_date = None
    if contract_end.strip():
        try:
            end_date = date.fromisoformat(contract_end.strip())
        except ValueError:
            end_date = None

    db.add(
        Partner(
            name=name_val,
            code=code_val,
            contact_name=contact_name.strip() or None,
            phone=phone.strip() or None,
            email=email.strip() or None,
            contract_end=end_date,
        )
    )
    await db.commit()
    return RedirectResponse("/admin/partners", status_code=303)


@app.get("/admin/partners/{partner_id}/edit")
async def partner_edit_page(
    partner_id: int,
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    partner = await db.get(Partner, partner_id)
    if not partner or not partner.is_active:
        raise HTTPException(404)
    return templates.TemplateResponse(
        request, "partner_edit.html", {"user": user, "partner": partner}
    )


@app.post("/admin/partners/{partner_id}/edit")
async def partner_edit(
    partner_id: int,
    name: str = Form(...),
    code: str = Form(...),
    contact_name: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
    contract_end: str = Form(""),
    user: User = Depends(require_can_edit),
    db: AsyncSession = Depends(get_db),
):
    partner = await db.get(Partner, partner_id)
    if not partner or not partner.is_active:
        raise HTTPException(404)

    code_val = code.strip()
    name_val = name.strip()
    if not name_val or not code_val:
        return RedirectResponse(
            f"/admin/partners/{partner_id}/edit?error=required", status_code=303
        )

    dup = (
        await db.execute(
            select(Partner).where(Partner.code == code_val, Partner.id != partner_id)
        )
    ).scalar_one_or_none()
    if dup and dup.is_active:
        return RedirectResponse(
            f"/admin/partners/{partner_id}/edit?error=code", status_code=303
        )

    end_date = None
    if contract_end.strip():
        try:
            end_date = date.fromisoformat(contract_end.strip())
        except ValueError:
            end_date = None

    partner.name = name_val
    partner.code = code_val
    partner.contact_name = contact_name.strip() or None
    partner.phone = phone.strip() or None
    partner.email = email.strip() or None
    partner.contract_end = end_date
    await db.commit()
    return RedirectResponse("/admin/partners", status_code=303)


@app.post("/admin/partners/{partner_id}/delete")
async def partner_delete(
    partner_id: int,
    user: User = Depends(require_can_delete),
    db: AsyncSession = Depends(get_db),
):
    partner = await db.get(Partner, partner_id)
    if not partner:
        raise HTTPException(404)
    partner.is_active = False
    await db.commit()
    return RedirectResponse("/admin/partners", status_code=303)


# ── QR / Mobile Equipment View ────────────────────────────────────────


def _public_base_url(request: Request | None = None) -> str:
    env = (os.environ.get("PUBLIC_BASE_URL") or "").strip().rstrip("/")
    if env:
        return env
    if request is not None:
        return str(request.base_url).rstrip("/")
    return "http://127.0.0.1:8000"


def _equipment_mobile_url(code: str, request: Request | None = None) -> str:
    return f"{_public_base_url(request)}/eq/{code}"


def _safe_qr_filename(code: str) -> str:
    import re

    safe = re.sub(r"[^\w가-힣.\-]+", "_", (code or "").strip()) or "equipment"
    return f"{safe}.png"


def _qr_png_bytes(url: str) -> bytes:
    from io import BytesIO

    import qrcode

    img = qrcode.make(url)
    buf = BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


@app.get("/admin/equipment/building/{building_id}/qr-zip")
async def equipment_building_qr_zip(
    building_id: int,
    request: Request,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    """해당 건물 설비 QR을 코드명.png로 ZIP 압축 다운로드."""
    from io import BytesIO
    from urllib.parse import quote
    import zipfile

    building = await db.get(Building, building_id)
    if not building or not building.is_active:
        raise HTTPException(404, detail="건물을 찾을 수 없습니다.")

    equipment = (
        await db.execute(
            select(Equipment)
            .join(Zone)
            .join(Floor)
            .where(
                Floor.building_id == building_id,
                Equipment.is_active == True,
            )
            .order_by(Equipment.code)
        )
    ).scalars().unique().all()
    if not equipment:
        raise HTTPException(404, detail="해당 건물에 등록된 설비가 없습니다.")

    buf = BytesIO()
    used_names: set[str] = set()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for eq in equipment:
            fname = _safe_qr_filename(eq.code)
            base, ext = fname.rsplit(".", 1) if "." in fname else (fname, "png")
            candidate = fname
            n = 2
            while candidate.lower() in used_names:
                candidate = f"{base}_{n}.{ext}"
                n += 1
            used_names.add(candidate.lower())
            zf.writestr(candidate, _qr_png_bytes(_equipment_mobile_url(eq.code, request)))

    buf.seek(0)
    stamp = datetime.now(KST).strftime("%Y%m%d")
    zip_name = quote(f"{building.name or building.code or 'building'}_QR_{stamp}.zip")
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{zip_name}"},
    )


@app.get("/admin/equipment/{eq_id}/qr.png")
async def equipment_qr_png(
    eq_id: int,
    request: Request,
    download: int = 0,
    user: User = Depends(require_login),
    db: AsyncSession = Depends(get_db),
):
    """설비 QR PNG (화면 표시 또는 파일 다운로드)."""
    from urllib.parse import quote

    eq = await db.get(Equipment, eq_id)
    if not eq or not eq.is_active:
        raise HTTPException(404)
    data = _qr_png_bytes(_equipment_mobile_url(eq.code, request))
    filename = _safe_qr_filename(eq.code)
    headers = {
        "Content-Disposition": (
            f"attachment; filename*=UTF-8''{quote(filename)}"
            if download
            else f"inline; filename*=UTF-8''{quote(filename)}"
        )
    }
    return StreamingResponse(iter([data]), media_type="image/png", headers=headers)


@app.get("/eq/{code}")
async def equipment_mobile(
    code: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Equipment)
        .where(Equipment.code == code)
        .options(
            selectinload(Equipment.zone).selectinload(Zone.floor).selectinload(Floor.building),
            selectinload(Equipment.pm_schedules).selectinload(PMSchedule.inspections),
            selectinload(Equipment.pm_inspections).selectinload(PMInspection.schedule),
            selectinload(Equipment.maintenance_records),
            selectinload(Equipment.equipment_type),
        )
    )
    eq = result.scalar_one_or_none()
    if not eq:
        raise HTTPException(404, detail="설비를 찾을 수 없습니다.")
    active_pms = [s for s in (eq.pm_schedules or []) if s.is_active]
    all_inspections = sorted(
        eq.pm_inspections or [],
        key=lambda x: x.inspected_at or datetime.min,
        reverse=True,
    )
    history = sorted(
        eq.maintenance_records or [],
        key=lambda x: (x.work_date or date.min, x.id or 0),
        reverse=True,
    )
    location_parts = []
    if eq.zone and eq.zone.floor and eq.zone.floor.building:
        location_parts = [
            eq.zone.floor.building.name,
            eq.zone.floor.name,
            eq.zone.name,
        ]
    sheet_fields = get_category_fields(eq.category or "", [eq]) if eq.category else []
    history_json = json.dumps(
        [
            {
                "id": h.id,
                "title": h.title or "",
                "work_date": h.work_date.isoformat() if h.work_date else "",
                "worker_name": h.worker_name or "",
                "cause": h.cause or "",
                "action": h.action or "",
                "parts_used": h.parts_used or "",
                "work_hours": h.work_hours,
                "cost": h.cost,
                "note": h.note or "",
                "is_manual": bool(h.is_manual),
                "work_order_id": h.work_order_id,
            }
            for h in history
        ],
        ensure_ascii=False,
    )
    msg = request.query_params.get("msg", "")
    error = request.query_params.get("error", "")
    return templates.TemplateResponse(
        request,
        "mobile_equipment.html",
        {
            "eq": eq,
            "active_pms": active_pms,
            "all_inspections": all_inspections,
            "history": history,
            "history_json": history_json,
            "location_text": " / ".join(p for p in location_parts if p) or "-",
            "sheet_fields": sheet_fields,
            "today": _today_kst(),
            "message": msg,
            "error": error,
        },
    )


@app.post("/eq/{code}/pm-inspect")
async def equipment_mobile_pm_inspect(
    code: str,
    schedule_id: int = Form(...),
    result: str = Form(...),
    note: str = Form(""),
    inspector_name: str = Form(""),
    request_work_order: str = Form("0"),
    db: AsyncSession = Depends(get_db),
):
    eq = (
        await db.execute(
            select(Equipment)
            .where(Equipment.code == code, Equipment.is_active == True)  # noqa: E712
            .options(
                selectinload(Equipment.zone)
                .selectinload(Zone.floor)
                .selectinload(Floor.building),
                selectinload(Equipment.pm_schedules),
            )
        )
    ).scalar_one_or_none()
    if not eq:
        raise HTTPException(404, detail="설비를 찾을 수 없습니다.")

    schedule = next(
        (s for s in (eq.pm_schedules or []) if s.id == schedule_id and s.is_active),
        None,
    )
    if not schedule:
        schedule = await _get_or_create_pm_schedule(db, eq)
    schedule.equipment = eq

    insp, wo = await _record_pm_inspection(
        db,
        schedule,
        result_raw=result,
        note=note,
        inspector_name=inspector_name,
        create_work_order=request_work_order.strip().lower() in ("1", "true", "yes"),
    )
    await db.commit()

    if wo:
        return RedirectResponse(
            f"/eq/{code}?msg=점검+저장+및+정비의뢰+생성(#{wo.id})",
            status_code=303,
        )
    return RedirectResponse(
        f"/eq/{code}?msg=점검결과가+저장되었습니다",
        status_code=303,
    )


@app.get("/")
async def root():
    return RedirectResponse("/admin/dashboard", status_code=303)
